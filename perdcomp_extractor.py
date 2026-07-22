#!/usr/bin/env python3
"""
PERDCOMP Extractor  v5.0  -  AgriTax Tributário & Contábil
100% LOCAL · SEM API · SEM CUSTO
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, re, os, sys, io, base64, unicodedata
from pathlib import Path
from datetime import datetime

try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    import pdf2image
    PDF2IMAGE_OK = True
except ImportError:
    PDF2IMAGE_OK = False

try:
    import pytesseract
    PYTESSERACT_OK = True
except ImportError:
    PYTESSERACT_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAFcAAABGCAYAAAC5QghSAAAcqUlEQVR4nO2ceXxV1bn3v2tPZ8g8ARkgDAIKUhWxSKviAFbberlardbbV6219Tq1VlH7drpv1Uq1el+v1rmtQ337Xjp5tU5ULYqKQQYBEyVBhswJnJwkJ8OZ9l7r/rHP3jkJQVES7Od++vAJOdnD2mv91rOe57ee59lHKKUU/5BxEe3T7sD/ZPkHuOMo/wB3HMX4dB6rQDnuR2GAkpnP/7PmWvz9OTQFiE+7E2Mih1hVMvNo9yF3P4BsfARkGtXxHHLLdzLXCIY0++9s3j+mHFqzoBQIgbP1ClTH0yDTiHA1qvVPqL2voKZ9C9XzDtrkC11zcWg7h1ISIfQxa/HQaa6yAYVs+wOqdSXouWCEIVjh218VeQPV8gdU/wfIHfeDHfNuHosOAMr9pyRKOUM/SECMKbBwqDRXySFNTO5BTDoLjBxEycmIvCOh/MuISV9ElJ2K1MOIUBXOrl+BFkSb9g0X/H00WbmQ+y5D+VMwZLEFQojMETF0Tuxr05N2jF17/0rK6eczVRejlEIcpIM9BOAqlwWk9qD2rkJMWIo27Spky28RZUuQ7f+F6ngWbdaNqFgtNP8eihdinPIGKtEBCBQClOO7OnfQYh+gPswNOjKJ7cRJOYOk7BjxdJSBZAe98Sa6Bz6ge3AHkb46TptzZ+aZ8qDd6viCqyQIDbXnGeTOX8DgTkTl11F6Mc6G76PP+y4qHkW2PAnCRPR9gGp/DSdvFkrPQXW+hHn8SkTOVGA4eI5MkbRjJNLdJNLdxNNREulukukeEnYvKTtGyu4nZfeTdgZIO4PYMo7tJLBlAkemkMoGQBMGuhYgL1hJecFxmWcdPGMZR3Bd50VqD7LhxyATrgZbZYjCExETZiMKF0CoB5FsR0y9FLXzV6C9hm4VQ+82iG7Abn+Ogapl9PRvoyfeRG98N32JNgZTe0mme0k7gzgyiVR2xjC4+i0QCKEh0PzfCPezJkx0w2IIQkHaGSQ/PJ2C0BT3yBhw7vEFFw3MUkTp6SB0RM7hiElfAacffeEziJxZqHgLmtOPFiyHYx/BqTibSLicDs2hc3A20b1/ZGDP/ydt9yCVzIBmoAkdTRgIoWEaOWTrmhrWB++XGnbco/cK0ISOLRNUFC1ECB2pHLQxcG7jBG4GWJlEtj+BSu2G+A7E5EtRPetwNi1DWWWYC55H++Bu0q2/p33PCzSXn0VrbDO9/fWklY0IFqHLAXQEllHgMWD3f6X8zyqzw/uknEKh0DWLqSWnAmNjEmBcwPU2Cj04tV9HRVeDkQfxTlTXy8i2P4BMYybaSLQ+yXa7m+3504mSxGn+DYaw0PUgIWGglA3o7mLPOLSxFoGG7cQpyZlNWd6RwMGzBE/GHlwlQejIxjtRkefBmgBOAlH5L1D1LQzHxonvohaT2sgr9AoNQ7MwEVhmkctBUUifGwwRLIHI8mpZ57K0+OOKEAJbJphRdgaaMMbMJMB4gCsEKBvV/QaEZiIKFqJVXooyS9B61hIt+yJvdr1GW7wJU+iElByi914wJ8NPBcIn/VKlkTKN9Em/GnJWGfvrgvJxlrTAkWnCVhmHTfxypvtjt68aY3AlZLyyNuVK0AKI0mXI7tfQ1p/MNi2XGn0CaeUQ0gIuaFl3ezskKdOknQRSpQGBpecQtsrIsSYQskqwjFwEOo5KkrL7XSqW6iJh9/j290BEExpxu5fZk84mbJWhlDOmu7QxBDfjxBIfINvuR+1dCUKgWeVo9dezQStkk1mJqVJYwsgsewCBJjSksknZMZSShKwSJuYfRXnhcUwqmE9BaCq6ZpJI99CXaCEWb6Y/2c5gKpLRdoGuWQh0FO7fByJSOQSMAo6s/Do+dRxDGRtwlXQ7NlCLU7sEUp1gFKGwER2PUmM7bDYmElIp9/KMPdWEhi1TJO0BgmYh1SWnMK10KdUli7GMfADaet5mS/Nv6OjdRCzeRNKOZZF/PcssuKbhQIHVhEE83cVRk79JfmhyJmgztqGWMdJcCRjIPY9BshOsUpTdg6EH2RiPsFmECOP4G0ohdKRME7d7yQtWMq/y68yc+E8UhqcB0NW/jffb76Ep+ip9iTZAoWtBDM0iYOZnbDF8UocmcJ1YXrCSoyd/053sMdZaGFOzIMEsA0AZRRh5x9BgW2yI7SCkWUjf2wuS6V6CVhELqi7miIrzCFsTAGjtfot3W56gtWcdtkxg6mGCZpHbvMpwCCUPmpIJoZG2Bzhp1k8JmkVjbmv954xJJiITQwCQ0RfQ8o6lKz3AM1suQ8h+yHh1R6awZYLDys7k2KlXURCqBlxN3dT4II1dq1FITD03Y4ddVjCW4pqDKDMnnsWph98+bsDCmGhuJuqVqMdp+QkMvEs6NJM1AwU4qQ5MswCBRtLuJWyVsXj6zcyYcCbgRqq2NP+GrS2Pk7L7CZj5gEApB+nTsrETTeiknH6Kc2by+cN+mDEH4xfSPkhw3SAzyZ0475+GTLZi6gabBmLscSoImfkoBPF0lMqi41k86xbygpUAdA98wOvbb6a9dz0Bo4CAWZDFc8deRCZ+YOm5LDniLgJG/rg4sWw5OHCVO/Oy/d9RyVYMq4yIA7VqCgHdXdCJdJQ5FV/j84f9IOPNYefeVby+/WZSdh8hs8TdGIwjsJowSDtxdM3k9Ln3UJRz2LiaA08ODlzPw8ZrQYJwutlkH4GNIIAkkY5x3NTvML/6Cp/cb256hLd3/wemFsYy8nxaNV6iCYOkHSNkFbN0zt1MzD/6kAALY6K5oIq/imFOpE2bQlPragI6JNIxjp9xA5+pugSpbDRh8NaO29nS/Cghq9j1/uOqrW7ocDAVobzgWE4+/DYKQtWHDFgYI7agACET/LX+32iKrEJKm89Ov5ajJ1+GI9Pomskb22+htvX/EbZKM85q7GNcblBcIJVDyu7D1HM4svJfmF99BbpmjbuNHSkHqbkOSuiI9jvY0/wLWtMzcGSKoyZfmgE2ha5ZvLXjdmpbnyRslY2pGXB5s4bA3cqmnTiOTBIw8pk58Sw+M/lSSnJmuV3l0AILY2Vzo//J9rTFgJPiiLLTOX768ozGWmxu+hVbmh89CGCzE5HCDzQq5eCoFI6TQioHS8+hNPcIphSfxPSy0ykMT4fMdX6a5xDLQWpuJk1W9Qt2RL/DhPBUTpx1K0o56JrJzr2reHvX/yVkFY8C7Mjs7cg0jZuKUbicV0oHKe1MVlZg6mFyAuWU5MxiUsF8yguOoyR3Fl5sQSmJuyE8NPZ1NPlIcL3gtSfZgWSJQgOa0zqDEr50+AosIxdweeyahn9z81tqKB3uNiqRSKSykdIrzhh6jsgUaOiahaGHCOj5hKxicgPl5IeqKQpPpyhnBvnByRh6aER/HRDikJuA0eQjwRUZmzbqucyZ+rYn+ezUqynJne1mYZXDq/U/wpFJLCMPWyZxnCSOdKNihhbA1HMIWKUEzSJCZjFBs5iQWUTQKs787R4PmIUEjHxMPTxqH7xJ8bO9n6KmjpT9gut51pbutTRH12SiUkE+M/kbGFrAHZDQ6E92YBkFHDPl2z7lWr/7Pvb0bcXUQwymIgTNQkry5lGaO4eS3MMpCE8l15pE0CzcR/M+rmRP/sclPmIcImHZsn/NzTx4U+MDNEXXEDDySTkDTMw/isqiRShlI4RJ7+Bujpp8KULoCGBvXy0bGu/D0nOZmH8000pPp6poEfmZeoDRRGUCNKOVI7lpI3ciY4lmalueRAgNy8jjmMnfQtNM/FqFcQbr48qo4Hpa2zO4k+hAA/nByb6t3RV5mcqiRb73Lc2bQ8DIBxSOTPNa/Y+ZXPR5jp16NeUFx45odyhHhhgqyPDs4/6g8WAfTHayteUxhNDJCUzkqKpvoGEC4DgOsVhsGMBKDdUnCLEv+Pn5+WjaIY4tuDYMdnetJmnHCJrF2DKJrgVo6V5L2hnM2ECVCYC4dKexazVzKs5nTsUFQy1lwpGekxprcatSBY2NjVx99dXDwNJ1HV13n2nbNlJm6huUQtM0HnzwQSorKzNFd4coWK4JDaUkjV2rM7kpgaYFEFLQl2ihvWc9U0oW+xouhI5CUVX0OSwjD1B+yFAgMuxgKLPq5b1EltYOsQXl89LhDALC1gTmVlwIQhAw8jOryTUJhmFQUlKCpmk+WIODg8RibhlqQUEBoVDI12RN03zgx0v22f56gEX6t/HM5v+FlCkmFRxLeeECNjU+CMCsictYPPuWUbeTXoDmYKmQxwAO+HqlfM10HAdd13n22We5++67EUJw0003sWTJEv8cMO7g7qO5vkmIvIwjEyglqShcyOxJ/8zmpl+hCZ22nnWk7D5fS4espcrYUkFXfz27u16mL9GOAPJDU5g96RxSdh8f7HkWTRjkhyZz2AS3XmDX3r/SPbgDqRyqS06mLO9IOnrfoTm6hpQzwOdmfJ/B1F4aOv8L0LCMHI4o/6ofxhRC+GAJIXzN9HTHMxHeORiakNGOaZq2X/s90k57kzry3D7gemnupugaNGGBUFQWLSQnMJGS3MPp6q+nP9lOa08N00qXDit19wqGNzbez+amX+HIJGknjlQ2jkyyrf2PlObO4YM9z+GoNNUli31wGzqfYXvnX9zlnIrwXtt/8l7bSlJOH7mBCj5/2A/oS7Tx1o7bEcIgNzCJ2RP/GU03RkzwECXLXpSjHcueEA+kbHPhmRfv9/7scjag2fZ7GLjeMt/bV0v3wAcIISgITac0dw6gqC45mT19W9HQ2RV5mWmlS/1Befduaf4N63b+OzmBCQihMbV0CWV5c0k7A+zc+1d2Rv5KXqiSRLo3o/muWEYuAbOQkFVMQ8dT2DJJZdHxFIankhuo9Cc+ZJYghE7ALGT//OLDxdPCpqYmampqqK+vp6Ojg3Q6TX5+PkceeSTnnnsu+fluer+9vZ17770XgJKSEq699lp/VQghuOeee+jo6MBxHK688kqmTp2KUmqk5rqzuivyslvtomBy8QnomgVAdcnJbG56BBC096wnke7OZE9dYPsSbbzT9Ag5gQmk7H6OmfItFky9xm99XtXFvPDuv9IzuAuUHFYd436WpO0BinNmsuiw71NesGBE7yRSOW7w5hPGgj1AIpEIl112GXv37qWgoIDc3FwSiQTJZJKamhrWrl3LXXfdRX5+PuXl5VRUVPD4449jmiaFhYVcdtllCCF48sknWblyJalUivPOO48pU6b4KyALXIUQOo5M0hx9A10LIJXDlOKT/JcyCsPTKcqZSbS/noHUXlq613LYhC8hlY0uLBq7VpOyY5h6DkXhGRxbfZVfA6aUJGQWc/ikr/DGBzejC2ufgQs00s4AMyf+E+UFC7KAlJnNwsGLt2QLCgo4++yzmT17NjNmzCA3N5doNMqDDz7Itm3beP/993nqqae45JJLcByHq6++ml27drF161Z+97vfsWjRIizL4rHHHiMcDnP00Udz/fXXDzMdvrHwtKgjtpne+G6UUpTmHsGkgvmZihYTTRgcNuFL2DKJJnR2R14Z1uGu/m1owsCW8UwhsUun3CJlPbNlHnr5Yz/DR6pUFuvQx4UfW5bFVVddxZIlS5g2bRplZWXMnj2bW2+9FdM0sSyLd999d6hXQnD99deTk5MDwB133MHPfvYzlFJYlsXy5csxDGP/Nhdg996XkDKNZQRwZII1DT/JWr6CRLo7U8kN7b0bGUztJWy5xSBJuxeEQElFbqB8nwEdOLUSmYkZvzRQNgjt7e00NTURjUZJJpOEQiFisRjJZBJwmYbjOFRWVnL11VezYsUKOjs7AdcJXn755UyfPn0YzQMfXNckpJ1BmrvX+sGUPX3v0tazflintIwzsfRcBlN7aY6+wexJZ7uQZGmYFy8YZVhjgc1BiQdsY2MjjzzyCO+88w6dnZ0+JSspKfG10BMP4KVLl/L666/z5ptvomkaCxYsYNmyZfsACxlwPTrV3rOevkQLph4iaBYxp+KCYRF8b7cUHdhOW886dGGyO/KKD65bhunyxr5ECyOXv5tq+XRDgh5gnZ2dLF++nLa2NkpKSrjqqqs46qij0DSN22+/nba2tmH2U0qJrus0Njby3nvvYVkWQgi2b99OQ0MDs2bN8h2ZJxnNdRvZFXkJgJQzwBHl5/PZad8btYOxRAt/2nA2hh6kM7aZWKKF/GAVZXlzUcrBMnJp7X4LJxOPcGO8Ck2YJNOxUds8VCKlxDAMVq1aRXt7O7m5uVxzzTUsXbrUv8Y0zVE5cjqd5o477iASifi2t7e3l9tvv5377rsPy7KGmRvN28sn7RitPesw9BC6MKkuORmpnMz7Wk7Wj01eoILSvLk40iaR7qEl+gYAk4tPJCcwCQX0JVtZ+8EKP8arayZdA/XUtf0OS8/J1IEdevEGvmfPHh/EadOm+ecdx8Fxhtt6TyN//etfU1dXh5SSiy66iG9+85s4jsOOHTu4//770TRt2G7NkEpmtrQ1DCT3YOhBCsLTKMub5wZGRixjmYmATSk+ifbejRh6gMauV5lTcQEhs5jPTv8er7y3nKBZQEPnM0QHtlOSezgpZ4DGyN+wZYKAWQBOfPig/ffGPvwrAYTQ/J8DAdLb8nqgelpYXV1NIpEgGAzyy1/+knPPPZdUKsWLL75IT08PoVDIX+K6rrN+/Xr+/Oc/o+s6xxxzDOeffz4A69atY/369Tz//PMsWLCAk046yZ8MzU/VdDyNIxPEU11UFX0OXTNHLYbzrp9Ssth9QUPatPbU0D2wA1DMnPBlTjl8BboWwJEpWntqeKfpYWpbnqA0bw7zqi4ibffvwxxsmSRtD5C2BzLl+vuKUnLorUh74CPBTSaTdHd3093dTSrlppi8qNmZZ57JokWLiMVi1NTU8N3vfpfrrruOHTt2IKUkEonQ29sLQGNjIz/+8Y/p6ekhJyeHG2+80d/l3XDDDRQVFdHX18dPf/pTtm/f7mtwJiqm6IxtcV/owKEkZxZBs5iRe/YRQ2VPbKtrNnAoDs8kZJX4b8MMpvbQFH2dvngrCEFhaBozJpzBjj3P87dt/xtDs6gs+hxnHHkfCkXPwA7i6S6UUhSEppAbrMiKjLn9SNkx9vTVAgJDs5iYf/SoDtKzezt27GD9+vUIIVi4cKG/LYXMWzy2zerVq6mvr0dKycyZMznhhBN48803iUajlJWVsXTpUrZs2UJdXR26rjNnzhzmzZvnL39N03j//ffZunUrjuMwa9YsFixY4PZhbL8p5KO/5WNT4wNs2H0fmmYypfgkTp/7H4e8EgaG89yDuffD2hm2Q/O+h+BAuejQPV5QWyCVza7Iy6Nfj6Kx6zVMPYSUKYq8wg3vzTP/+w/293w1dM0BOESllO+gRuqQF+3yzmdf533OzlxkH8sGUwiBlHKfeyBrh/ZJNGfoHrfjiXQ3q+quob1nA9PLvsDh5V+hKHwYmtDpT3bwbssTRAcaMLQAhh5i5sSz3HbcXMcBBLk+XqpoZEjxQM+PPPZR7ewvDzdG70S4L+MZWpD84GQ6xEZ27l1FU9erBM1ChNBJ2X04MoUQOgm7hxNn/h8Kw9M/FZNwqMS3uaMtm/3JR9mrbe1/5L32lfQO7iblDGSiWgaGFqIgNJX51f/KtNLTPjGwnqf+sMytd82HBbk/zvMOxMZmO0v4mCWkB+YEPNvr0NW3jb5kC2lnAEMLkR+anHl5mVGBPZD2x+qaD7sXOCAwP+p5Qiml0uk0LS0tBINBlFKk02mqqqqGZmDE7Le0tFBWVoZpmv45jzirTFGzpo1ucbqj3VhBk5xwrm/8R7Y/Wm4rexA9PT1s376d+fPnD8sIZEt9fT2xWIy5c+cSDodHzYGNPDayneznGYZBbm4ura2tFBUVEQwG9+l3f38//f39TJo0yW0ToK+vjxdeeIHbbruNO++8k1WrVjE4OOjuMjKJunfeeYcnnngCgFtuuYXe3t5hA/c+u4AYmUloxrZT/makuzvKY088Rnww4S9rr/2NGzf67XsOxBuwN1BvYm+++WbWrFnDihUrsO2h6klvUu666y6efvppamtrefjhh0mn034fs/vsTV5239vb2/0NhxCCeDzOo48+6m8ofv7zn9Pa2jpsq/vwww+zc+dOtm3b5qeDpJSuQyssLOTKK6/koYceIi8vjwsvvBDHcdi5cyd9fX1MmDCB0tJSqqqqAAgEAmzcuJGCggIWLVoEQENDA9XV1USjUXRdx7IsbrzxJs455xzOOeccot1RampqOH7h8ZSWlqKU8tufOHEiZWVlVFa6ubJIJMLatWuZOXMmRxxxxDAb293djWmafPWrX+VHP/oRhuFOpBfy+8tf/kIkEmHFihUAJBIJDMPwJ7C/v58TTjgBXdfZvXs3PT099Pf3s3DhQgYHB/nhD3/I4sWLufDCC4nH49TU1DB//ny/b8FgkNraWjo7O1m4cCG6rjNlyhRycnKIRCLDFS57xm3bJp1O+0vytttu48UXX8S2bXbv3s2aNWsAGBgYoLW1lZdffpn7778fgDvvvJPOzk6ee+45nnvuOZ8zplIpP2dl2zZPPvkkf/rTnxBCcMstt7Bq1Sps22bXrl2sXbuWeDzOT37yE5LJJA899BBvvvmmb3bArZx59913ufbaa7n11lupq6sjkUj4y3PTpk1+fUIymfRDg7/97W954YUXeO+993zg7733XtasWcOzzz7L73//ex+HdDqNEIKuri5SqRTPPPMMjz76KADxeJyWlhZqamq46667APjb3/5GZ2enHxUbBq63ZLyl6C3HcDjMxRdfzOTJk0kmk/6sWJbFxRdfzE033cSGDRsACIfDhMNhP2NaUlJCeXk5ixcvRgiBaZokk0mKi4tpbm4GICcnh0suuYSqqioSiQS5ubk0NDQQDoc577zzOO2003jttdcAl3tGIhHuvfde7rnnHhYuXMjzzz/PypUryZZgMEhnZye6rhMIBPw+v/7661x++eVcccUVNDQ0kEqlCIfDLFu2jCVLltDR0UFhYSEVFRWceOKJGIaBaZrE4/FhfdZ1nfPPP5/ly5dTV1cHMCy9s4/mejI4OEgikfD/7uvro6enx3dyAwNusKS7u5u3336bF198kcLCQn9Ga2traWhoIJ1Oo5Sit7eXrVu3AvDAAw/Q2dlJRUUFg4ODfvu9vb0opbBtm2g0SmVlJZFIhLq6OtatW8fMmTOBoR1VMpkkmUzyhS98gYceeghd1zEMAyklSinOOussnn76aWpqamhpaWHTpk0AVFRUsHr1ajZu3EggEMAwDHp7e+nt7SUajZJKpVBK0dfXx5YtW1BK8fjjj7N9+3aqq6vp7+8H3Pjthg0bePXVVwmH3ZrhWCyGbdvYtu1jBJlNhLekjjvuOAKBgD+YM888k9LSUoQQTJ8+3Vf5s88+m/r6erq7u/ne99yA+kUXXcTmzZupqqpi3rx5CCH42te+Rl1dHWeccQbf/va3eemll7Asi5NOOgmAL37xixQXFyOEYMaMGW5StLSUK664glWrVjF//nyWLVvmP7esrIzrrruOp556ivz8fB544AHq6upoaWlh6tSpSCmZO3cuP/jBD3jllVfQNI1p06Yxf/58rrnmGlauXElzczM33HADmqZxyimnMGGC+1K3N84LLriAt956C9u2ueiii3j22WdJpVJ+MH3ZsmV0dHTQ1dXF8uXLATj11FOZNGkSUkoWL17s4/d39xWvB1txOF4Vi59EhpHR7DAauB7Yo0rZHttxHP+Yd95blt792U7Ioy3ZBN1rZ7T2s68duaf3znn3ZD/Pa3vkNR4Xzg66+DHXrM1CNr3y/E42F8/u20e1A/Df67bMCeqSvbYAAAAASUVORK5CYII="

C_GREEN       = "#7AB82E"
C_GREEN_DARK  = "#5A8A1E"
C_GREEN_LIGHT = "#EAF4D3"
C_GREEN_MID   = "#C5E08A"
C_YELLOW      = "#F5A623"
C_YELLOW_LIGHT= "#FEF3DC"
C_GRAY_DARK   = "#3D3D3D"
C_GRAY        = "#6B6B6B"
C_GRAY_LIGHT  = "#F2F4F0"
C_WHITE       = "#FFFFFF"
C_BORDER      = "#D0DDB8"
C_RED         = "#C0392B"
C_RED_LIGHT   = "#FDEDEC"
C_BLUE        = "#2563EB"
C_BLUE_LIGHT  = "#EFF6FF"

# Verificação de dependência Anthropic (usado pelo DCTF Extractor)
try:
    import anthropic as _anthropic_lib
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False

APP_TITLE = "AgriTax Audit  v6.0"

# Aba 1 - Detalhamento (crédito e débitos na mesma linha)
ABA1_COLS = [
    # -- Dados Cadastrais --
    ("cnpj",                    "CNPJ",                       150),
    ("razao_social",            "Razão Social",               210),
    ("tipo_pedido",             "Tipo de Pedido",             180),
    ("numero_perdcomp",         "Nº PERDCOMP",                175),
    ("numero_pedido_vinculado", "Nº PER/DCOMP Vinculado",     175),
    ("data_transmissao",        "Dt. Transmissão",            115),
    ("situacao_perdcomp",       "Situação PERDCOMP",          150),
    # -- Crédito --
    ("tipo_credito",            "Tipo de Crédito",            220),
    ("periodo_credito",         "Competência / Período Créd.",145),
    ("comp_teste_credito",      "Competência Teste (Créd.)",  140),
    ("valor_total_credito",     "Vl. Total do Crédito",       155),
    ("valor_utilizado_doc",     "Vl. Utilizado neste Doc.",   155),
    # -- Débito (repetido por linha) --
    ("codigo_receita_debito",   "Cód. Receita Déb.",          115),
    ("tipo_debito",             "Tipo de Débito",             220),
    ("periodo_debito",          "Período de Apuração Déb.",   140),
    ("comp_teste_debito",       "Competência Teste (Déb.)",   140),
    ("valor_principal",         "Vl. Principal",              130),
    ("valor_multa",             "Vl. Multa",                  115),
    ("valor_juros",             "Vl. Juros",                  115),
    ("valor_total_debito",      "Vl. Total Débito",           130),
]
ABA1_KEYS   = [c[0] for c in ABA1_COLS]
ABA1_LABELS = [c[1] for c in ABA1_COLS]
ABA1_WIDTHS = [c[2] for c in ABA1_COLS]

# Aba 3 / Aba 5 - formato legado (linha por crédito OU por débito)
DETAIL_COLS = [
    ("cnpj",                    "CNPJ",                        150),
    ("razao_social",            "Razão Social",                210),
    ("tipo_pedido",             "Tipo de Pedido",              190),
    ("numero_perdcomp",         "Nº PERDCOMP",                 175),
    ("numero_pedido_vinculado", "Nº PER/DCOMP Vinculado",      175),
    ("data_transmissao",        "Dt. Transmissão",             115),
    ("tipo_registro",           "Tipo de Registro",             95),
    ("tipo",                    "Tipo / Grupo de Tributo",     260),
    ("periodo_apuracao",        "Período de Apuração",         160),
    ("competencia_teste",       "Competência Teste",           140),
    ("valor_original",          "Vl. Original / Principal",    155),
    ("valor_utilizado",         "Vl. Utilizado / Total",       155),
    ("valor_multa",             "Vl. Multa",                   120),
    ("valor_juros",             "Vl. Juros",                   120),
    ("valor_total",             "Vl. Total Débito",            130),
]
DETAIL_KEYS   = [c[0] for c in DETAIL_COLS]
DETAIL_LABELS = [c[1] for c in DETAIL_COLS]
DETAIL_WIDTHS = [c[2] for c in DETAIL_COLS]

CTRL_COLS = [
    ("cnpj",                  "CNPJ",                      150),
    ("razao_social",          "Razão Social",              210),
    ("numero_perdcomp",       "Nº PER/DCOMP Origem",       175),
    ("situacao_perdcomp",     "Situação PERDCOMP",         150),
    ("tipo_pedido_display",   "Tipo de Pedido",            180),
    ("tipo_credito",          "Tipo de Crédito",           220),
    ("competencia",           "Competência / Período",     155),
    ("competencia_teste",     "Competência Teste",         140),
    ("valor_total_credito",   "Vl. Total do Crédito",      150),
    ("valor_total_correcao",  "Vl. Total da Correção",     150),
    ("total_compensado",      "Total Compensado",          145),
    ("saldo_disponivel",      "Saldo Disponível",          145),
    ("ultimo_retificador",    "Último Retificador",        175),
]
CTRL_KEYS   = [c[0] for c in CTRL_COLS]
CTRL_LABELS = [c[1] for c in CTRL_COLS]

# Aba 4 - Na planilha de status, sem PDF importado
ABA4_COLS = [
    ("numero_perdcomp",  "Nº PERDCOMP",        180),
    ("cnpj",             "CNPJ",               150),
    ("tipo_documento",   "Tipo de Documento",  200),
    ("tipo_credito",     "Tipo de Crédito",    230),
    ("data_transmissao", "Dt. Transmissão",    115),
    ("situacao",         "Situação",           160),
]
ABA4_KEYS   = [c[0] for c in ABA4_COLS]
ABA4_LABELS = [c[1] for c in ABA4_COLS]
ABA4_WIDTHS = [c[2] for c in ABA4_COLS]

# Aba 5 - PDF importado, não consta na planilha de status
ABA5_COLS = [
    ("numero_perdcomp",  "Nº PERDCOMP",        180),
    ("cnpj",             "CNPJ",               150),
    ("razao_social",     "Razão Social",        220),
    ("tipo_pedido",      "Tipo de Pedido",      190),
    ("data_transmissao", "Dt. Transmissão",     115),
    ("arquivo",          "Arquivo PDF",         240),
]
ABA5_KEYS   = [c[0] for c in ABA5_COLS]
ABA5_LABELS = [c[1] for c in ABA5_COLS]
ABA5_WIDTHS = [c[2] for c in ABA5_COLS]

# Aba 6 - Pedidos de Ressarcimento com código detalhado de crédito
ABA6_COLS = [
    ("cnpj",                "CNPJ",                   150),
    ("razao_social",        "Razão Social",            210),
    ("numero_perdcomp",     "Nº PER/DCOMP",           175),
    ("data_transmissao",    "Dt. Transmissão",         115),
    ("tipo_credito",        "Tipo de Crédito",         230),
    ("codigo_credito",      "Cód./Detalhe Crédito",    230),
    ("competencia",         "Competência / Período",   140),
    ("competencia_teste",   "Competência Teste",       140),
    ("valor_total_credito", "Vl. Total do Crédito",    155),
    ("total_compensado",    "Total Compensado",        155),
    ("saldo_disponivel",    "Saldo Disponível",        155),
]
ABA6_KEYS   = [c[0] for c in ABA6_COLS]
ABA6_LABELS = [c[1] for c in ABA6_COLS]
ABA6_WIDTHS = [c[2] for c in ABA6_COLS]

# -----------------------------------------------------------------------------
# Competência Teste — código canônico padronizado para todos os módulos
# -----------------------------------------------------------------------------

def format_competencia_teste(p) -> str:
    """
    Converte um período/competência bruto em código canônico no formato:
      AAAA.MM   — competência mensal (01..12)
      AAAA.NT   — competência trimestral (1T..4T)
      AAAA      — competência anual

    Entradas suportadas (e resultado):
      '28/02/2026'            → '2026.02'   (DARF/DAS: DD/MM/AAAA — descarta dia)
      '02/2026'               → '2026.02'   (DAS: MM/AAAA)
      'Fevereiro de 2026'     → '2026.02'   (DCOMP débito: "Mês de AAAA")
      'Fevereiro/2026'        → '2026.02'
      '1º Trimestre/2024'     → '2024.1T'   (PERDCOMP crédito trimestral)
      '3º Trimestre 2024'     → '2024.3T'
      'Trimestre 3/2024'      → '2024.3T'
      '2024'                  → '2024'      (crédito anual)

    Retorna string vazia se `p` for vazio/None ou não reconhecido.
    """
    if not p:
        return ""

    s = str(p).strip().upper()
    if not s:
        return ""

    # Remove preposição "DE" solta entre palavras (ex.: "FEVEREIRO DE 2026")
    s = re.sub(r"\s+DE\s+", " ", s)

    # ── Caso 1: DD/MM/AAAA (DARF) → AAAA.MM ──────────────────────────────────
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}.{int(m.group(2)):02d}"

    # ── Caso 2: MM/AAAA (DAS) → AAAA.MM ──────────────────────────────────────
    m = re.fullmatch(r"(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(2)}.{int(m.group(1)):02d}"

    # ── Caso 3: Nome do mês por extenso ou abreviado → AAAA.MM ───────────────
    # Nomes completos primeiro (mais específicos), depois abreviações.
    MESES = {
        "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "MARCO": "03",
        "ABRIL":   "04", "MAIO":      "05", "JUNHO":  "06",
        "JULHO":   "07", "AGOSTO":    "08", "SETEMBRO": "09",
        "OUTUBRO": "10", "NOVEMBRO":  "11", "DEZEMBRO": "12",
    }
    for nome, num in MESES.items():
        if nome in s:
            m_ano = re.search(r"(\d{4})", s)
            if m_ano:
                return f"{m_ano.group(1)}.{num}"
            return num   # sem ano (muito raro)

    # Abreviações de 3 letras (tipicamente em períodos diário/decendial):
    #   "21° Dia/Dez/2023", "3° Decendio/Dez/2023", "Jan/2024" etc.
    # Usa \b para evitar casar fragmentos dentro de outras palavras.
    MESES_ABREV = {
        "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
        "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
        "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12",
    }
    for abrev, num in MESES_ABREV.items():
        if re.search(rf"\b{abrev}\w*\b", s):  # aceita sufixos (Março, Dezembro, etc.)
            # Se já casou com o nome completo no loop acima, não chega aqui.
            # Mas casar abreviação curta ainda exige ano.
            m_ano = re.search(r"/(\d{4})\b", s) or re.search(r"(\d{4})\b", s)
            if m_ano:
                return f"{m_ano.group(1)}.{num}"

    # ── Caso 4: Trimestre → AAAA.NT ──────────────────────────────────────────
    # "1º TRIMESTRE/2024", "3º TRIMESTRE 2024", "3 TRIMESTRE 2024"
    m = re.search(r"(\d)[ºO°]?\s*TRIMESTRE\s*[/\s]*(\d{4})", s)
    if m:
        return f"{m.group(2)}.{m.group(1)}T"
    # "TRIMESTRE 3/2024"
    m = re.search(r"TRIMESTRE\s+(\d)\s*[/\s]*(\d{4})", s)
    if m:
        return f"{m.group(2)}.{m.group(1)}T"
    # Sem ano: "3º TRIMESTRE"
    m = re.search(r"(\d)[ºO°]?\s*TRIMESTRE\b", s)
    if m:
        return f"{m.group(1)}T"

    # ── Caso 5: apenas o ano ──────────────────────────────────────────────────
    m = re.fullmatch(r"(\d{4})", s)
    if m:
        return s

    # Fallback — retorna a string limpa
    return s


# -----------------------------------------------------------------------------
# Parser da planilha de status (exportada do eCAC)
# -----------------------------------------------------------------------------

_PERDCOMP_NUM_RE = re.compile(r'\d{5}\.\d{5}\.\d{6}\.\d\.\d\.\d{2}-\d{4}')

def parse_status_excel(path: str) -> dict:
    """
    Lê planilha de status exportada do eCAC e retorna:
      { numero_perdcomp: { situacao, cnpj, tipo_documento, tipo_credito, data_transmissao } }

    Detecta automaticamente as colunas pelo nome do cabeçalho.
    Situações consideradas "cancelado":  contêm "cancelad".
    Situações consideradas "retificado": contêm "retificad".
    """
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    # Localiza linha de cabeçalho (primeira com ≥ 3 células preenchidas)
    hdr_idx = 0
    for i, row in enumerate(rows[:15]):
        if sum(1 for c in row if c is not None) >= 3:
            hdr_idx = i
            break

    headers = [str(c).strip().lower() if c else "" for c in rows[hdr_idx]]

    def _find(*kws):
        """Retorna índice da coluna cujo header contenha qualquer uma das palavras-chave."""
        for kw in kws:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    col_num  = _find("número do documento", "nº perdcomp", "num perdcomp",
                     "numero do documento", "perdcomp", "número", "numero", "nº", "n°")
    col_sit  = _find("situação", "situacao", "status")
    col_cnpj = _find("cnpj")
    col_tipo = _find("tipo de documento", "tipo doc")
    col_cred = _find("tipo de crédito", "tipo de credito", "crédito", "credito")
    col_data = _find("data de transmissão", "data transmissao", "transmissão", "transmissao", "data")

    # Fallback: detecta a coluna do Nº PERDCOMP pelo formato dos valores
    if col_num is None:
        for row in rows[hdr_idx + 1 : hdr_idx + 6]:
            for j, cell in enumerate(row):
                if cell and _PERDCOMP_NUM_RE.match(str(cell).strip()):
                    col_num = j
                    break
            if col_num is not None:
                break

    if col_num is None:
        return {}

    result = {}
    for row in rows[hdr_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        raw_num = str(row[col_num]).strip() if row[col_num] is not None else ""
        # Remove espaços internos que às vezes aparecem no Excel
        num = re.sub(r'\s+', '', raw_num)
        if not _PERDCOMP_NUM_RE.match(num):
            continue

        def _val_at(col):
            return str(row[col]).strip() if col is not None and col < len(row) and row[col] is not None else ""

        result[num] = {
            "situacao":       _val_at(col_sit),
            "cnpj":           _val_at(col_cnpj),
            "tipo_documento": _val_at(col_tipo),
            "tipo_credito":   _val_at(col_cred),
            "data_transmissao": _val_at(col_data),
        }
    return result


def _is_cancelled(num: str, status_map: dict) -> bool:
    """True se o PERDCOMP estiver cancelado na planilha de status."""
    if not num or not status_map:
        return False
    return "cancelad" in status_map.get(num, {}).get("situacao", "").lower()


def _is_retified(num: str, status_map: dict) -> bool:
    """True se o PERDCOMP estiver retificado na planilha de status do eCAC.

    Cobre os casos em que a planilha indica 'Retificado' mas o PDF do
    retificador ainda não foi importado — sem esse helper, esses casos
    seriam tratados como vigentes no Controle de Créditos.
    """
    if not num or not status_map:
        return False
    return "retificad" in status_map.get(num, {}).get("situacao", "").lower()


# -----------------------------------------------------------------------------
# Parser de PDF
# -----------------------------------------------------------------------------

def _get(label, text):
    pat = re.escape(label) + r'[ \t]+([^\n\r]+)'
    m = re.search(pat, text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def _get_first(text, *labels):
    for label in labels:
        v = _get(label, text)
        if v: return v
    return ""

def _val(label, block):
    """Extrai valor numérico (começa com dígito) após rótulo."""
    pat = re.escape(label) + r'[ \t]+([\d][^\n\r]*)'
    m = re.search(pat, block, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def parse_brl(val):
    if not val: return 0.0
    clean = re.sub(r'[R$\s]', "", str(val).strip())
    clean = clean.replace(".", "").replace(",", ".")
    try: return float(clean)
    except: return 0.0

def format_brl(value):
    s = f"R$ {value:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def parse_pdf(path):
    with pdfplumber.open(path) as pdf:
        pages = [p.extract_text(x_tolerance=3, y_tolerance=3) or "" for p in pdf.pages]
    txt = "\n".join(pages)

    # CNPJ e Nº PERDCOMP ficam na mesma linha do cabeçalho de cada página
    hm = re.search(
        r'CNPJ\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\s+'
        r'(\d{5}\.\d{5}\.\d{6}\.\d\.\d\.\d{2}-\d{4})',
        txt)
    cnpj            = hm.group(1) if hm else ""
    numero_perdcomp = hm.group(2) if hm else ""

    nome     = _get("Nome Empresarial", txt)
    data_tx  = _get("Data de Transmissão", txt)
    tipo_doc = _get("Tipo de Documento", txt)
    tipo_cred= _get("Tipo de Crédito", txt)

    # Pedido de origem (Nº do PER/DCOMP Inicial)
    num_vinc = _get("Nº do PER/DCOMP Inicial", txt).strip()

    # Retificação: este documento é uma retificação de outro?
    retificador = _get("PER/DCOMP Retificador", txt)            # "Sim" ou "Não"
    num_retificado = ""
    if retificador.lower() == "sim":
        # Número do PER/DCOMP que está sendo retificado
        num_retificado = _get_first(txt,
            "Nº do PER/DCOMP que está sendo Retificado",
            "Nº do PER/DCOMP sendo Retificado",
            "PER/DCOMP a ser Retificado",
            "Número do PER/DCOMP Retificado",
            "PER/DCOMP Retificado",
            "Nº do PER/DCOMP Inicial")    # fallback: em alguns casos coincide

    # Valores do crédito - prioridade: mais específico primeiro
    # ATENÇÃO: Ressarcimento e Restituição têm nomes de campos DIFERENTES no eCAC!
    valor_cred = _get_first(txt,
        "Valor do Pedido de Ressarcimento",            # PER Ressarcimento PIS/COFINS (LIQUIDO)
        "Valor do Pedido de Restituição",              # PER Restituição (Retenção)
        "Crédito Passível de Ressarcimento",           # PER Ressarcimento fallback
        "Crédito Passível de Restituição Apurado no",  # DCOMP Retenção
        "Valor Original do Crédito Inicial",            # PIS/COFINS DCOMP
        "Crédito Passível de Restituição",              # PER Restituição fallback
        "Valor Original do Crédito",
        "Valor a Compensar",
        "Valor Total do Crédito",
        "Valor Bruto do Crédito",
        "Valor do Ressarcimento",
        "Valor do Crédito",                             # BRUTO - usar só como último recurso
        "Crédito Original na Data da Entrega",          # "da" - Retenção
        "Crédito Original na Data de Entrega")          # "de" - PIS/COFINS

    valor_util = _get_first(txt,
        "Total do Crédito Original Utilizado neste Documento",
        "Valor Utilizado nesta DCOMP",
        "Valor Compensado",
        "Total dos Débitos deste Documento")

    # "Total dos Débitos deste Documento" — capturado SEPARADAMENTE para que
    # o controle de créditos use o valor real dos débitos compensados, não o
    # crédito original consumido (que difere quando há atualização SELIC /
    # correções monetárias).
    valor_total_debitos = _get_first(txt,
        "Total dos Débitos deste Documento")

    # Período do crédito: Trimestre+Ano, Mês+Ano, Competência (Retenção) ou Período de Apuração
    m_trim = re.search(r'\nTrimestre[ \t]+([^\n]+)', txt)
    m_ano  = re.search(r'\nAno[ \t]+(\d{4})', txt)
    m_mes  = re.search(r'\nMês[ \t]+([^\n]+)', txt)
    m_comp = re.search(r'\nCompetência[ \t]+([^\n]+)', txt)   # Retenção Lei 9.711/98
    trimestre  = m_trim.group(1).strip() if m_trim else ""
    mes        = m_mes.group(1).strip()  if m_mes  else ""
    ano        = m_ano.group(1)          if m_ano  else ""
    competencia= m_comp.group(1).strip() if m_comp else ""
    if trimestre and ano:
        periodo_cred = f"{trimestre}/{ano}"
    elif mes and ano:
        periodo_cred = f"{mes}/{ano}"
    elif competencia:
        periodo_cred = competencia                             # ex: "Abril de 2025"
    elif ano:
        periodo_cred = ano
    else:
        m_pa = re.search(
            r'(?:PIS|COFINS|CSLL|IRPJ|RETEN|CONTRIB)[^\n]*\n'
            r'(?:[^\n]*\n){0,15}?Período de Apuração[ \t]+([^\n]+)',
            txt, re.IGNORECASE)
        periodo_cred = m_pa.group(1).strip() if m_pa else ""

    # Código/Natureza do crédito (ressarcimento PIS/COFINS: 101, 201, 310 etc.)
    # Padrões: "101 - Aquisição de bens para revenda", "Natureza do Crédito: 101"
    codigo_credito = _get_first(txt,
        "Natureza da Base de Cálculo",
        "Natureza do Crédito",
        "Código do Crédito")
    if not codigo_credito:
        # Tenta extrair padrão "NNN - Descrição" na seção de crédito (antes dos débitos)
        trecho = txt[:txt.find("\n001. Débito")] if "\n001. Débito" in txt else txt
        m_cod = re.search(r'\n(\d{3})\s*[-]\s*([^\n]{5,60})', trecho)
        if m_cod:
            codigo_credito = f"{m_cod.group(1)} - {m_cod.group(2).strip()}"

    credito = {
        "tipo_credito":    tipo_cred,
        "periodo_apuracao": periodo_cred,
        "valor_original":  valor_cred,
        "valor_utilizado": valor_util,
        "valor_total_debitos": valor_total_debitos,
        "codigo_credito":  codigo_credito,
    }

    # Débitos: cada bloco inicia com "NNN. Débito TIPO"
    debitos = []
    debt_iters = list(re.finditer(r'\n(\d{3})\.\s+Débito\s+([^\n]+)', txt))
    for i, dm in enumerate(debt_iters):
        s   = dm.start()
        e   = debt_iters[i+1].start() if i+1 < len(debt_iters) else len(txt)
        blk = txt[s:e]

        grupo  = _get("Grupo de Tributo", blk)
        codigo = _get("Código da Receita/Denominação", blk)
        # Período de Apuração - precisa do "de" para não pegar "Período Apuração DCTFWeb"
        pm_pa  = re.search(r'Período de Apuração[ \t]+([^\n]+)', blk, re.IGNORECASE)
        periodo = pm_pa.group(1).strip() if pm_pa else ""

        principal = _val("Principal", blk)
        multa     = _val("Multa",     blk)
        juros     = _val("Juros",     blk)
        total     = _val("Total",     blk)   # lowercase "Total", diferente de "TOTAL"

        tipo_deb = grupo if grupo else dm.group(2).strip()
        if codigo and codigo[:40] not in tipo_deb:
            tipo_deb = f"{tipo_deb} | {codigo[:50]}"

        # Extrai apenas o código NNNN-NN (ou NNNN) da string "codigo"
        # Ex.: "0481-01 - IRRF sobre Rendimentos" → "0481-01"
        cod_receita_deb = ""
        if codigo:
            m_cod = re.match(r'(\d{3,4}(?:-\d{1,2})?)', codigo.strip())
            if m_cod:
                cod_receita_deb = m_cod.group(1)

        debitos.append({
            "tipo_debito":       tipo_deb,
            "codigo_receita_debito": cod_receita_deb,
            "periodo_apuracao":  periodo,
            "valor_principal":   principal,
            "valor_multa":       multa,
            "valor_juros":       juros,
            "valor_total":       total,
            "valor_compensado":  total,
        })

    # Cria entrada de crédito se há pedido vinculado OU se há valor de crédito/utilizado
    tem_credito = bool(num_vinc or valor_cred or valor_util)
    return [{
        "cnpj":                    cnpj,
        "razao_social":            nome,
        "tipo_pedido":             tipo_doc,
        "numero_perdcomp":         numero_perdcomp,
        "numero_pedido_vinculado": num_vinc,
        "data_transmissao":        data_tx,
        "retificador":             retificador,           # "Sim" / "Não"
        "numero_perdcomp_retificado": num_retificado,     # Nº do doc que este retifica
        "creditos":                [credito] if tem_credito else [],
        "debitos":                 debitos,
    }]

# -----------------------------------------------------------------------------
# Transformações
# -----------------------------------------------------------------------------

def flatten_rows(perdcomps, filename):
    rows = []
    for p in perdcomps:
        hdr = {k: p.get(k,"") for k in [
            "cnpj","razao_social","tipo_pedido","numero_perdcomp",
            "numero_pedido_vinculado","data_transmissao",
            "retificador","numero_perdcomp_retificado"]}
        hdr["_source"] = filename
        for c in p.get("creditos", []):
            rows.append({**hdr,
                "tipo_registro":    "Crédito",
                "tipo":             c.get("tipo_credito",""),
                "periodo_apuracao": c.get("periodo_apuracao",""),
                "competencia_teste": format_competencia_teste(c.get("periodo_apuracao","")),
                "valor_original":   c.get("valor_original",""),
                "valor_utilizado":  c.get("valor_utilizado",""),
                "valor_total_debitos": c.get("valor_total_debitos",""),
                "codigo_credito":   c.get("codigo_credito",""),
                "valor_multa":"","valor_juros":"","valor_total":""})
        for d in p.get("debitos", []):
            rows.append({**hdr,
                "tipo_registro":    "Débito",
                "tipo":             d.get("tipo_debito",""),
                "codigo_receita_debito": d.get("codigo_receita_debito",""),
                "periodo_apuracao": d.get("periodo_apuracao",""),
                "competencia_teste": format_competencia_teste(d.get("periodo_apuracao","")),
                "valor_original":   d.get("valor_principal",""),
                "valor_utilizado":  d.get("valor_compensado",""),
                "valor_multa":      d.get("valor_multa",""),
                "valor_juros":      d.get("valor_juros",""),
                "valor_total":      d.get("valor_total","")})
        if not p.get("creditos") and not p.get("debitos"):
            rows.append({**hdr,
                "tipo_registro":"","tipo":"","periodo_apuracao":"",
                "competencia_teste":"",
                "valor_original":"","valor_utilizado":"",
                "valor_multa":"","valor_juros":"","valor_total":""})
    return rows


def combine_rows_for_aba1(rows: list, status_map: dict = None) -> list:
    """
    Transforma as linhas internas (uma por crédito OU por débito) no formato
    da Aba 1: uma linha por débito com dados cadastrais e de crédito repetidos.

    Regras:
    - Agrupa por (source, numero_perdcomp)
    - Se há débitos: uma linha por débito, crédito repetido em cada
    - Se não há débitos (ex: Pedido de Restituição): uma linha só, débito vazio
    - Se `status_map` for fornecido (planilha de status do eCAC), preenche
      `situacao_perdcomp` em cada linha com a situação do PERDCOMP.
    """
    from collections import OrderedDict

    if status_map is None:
        status_map = {}

    groups: dict = OrderedDict()

    for r in rows:
        key = (r.get("_source",""), r.get("numero_perdcomp",""))
        if key not in groups:
            groups[key] = {
                "hdr": {k: r.get(k,"") for k in [
                    "cnpj","razao_social","tipo_pedido",
                    "numero_perdcomp","numero_pedido_vinculado",
                    "data_transmissao","_source"]},
                "credito": None,
                "debitos": [],
            }
        tr = r.get("tipo_registro","")
        if tr == "Crédito" and groups[key]["credito"] is None:
            groups[key]["credito"] = r
        elif tr == "Débito":
            groups[key]["debitos"].append(r)

    _empty_db = {k:"" for k in [
        "codigo_receita_debito",
        "tipo_debito","periodo_debito","comp_teste_debito",
        "valor_principal","valor_multa","valor_juros","valor_total_debito"]}

    result = []
    for key, g in groups.items():
        hdr = g["hdr"]
        cr  = g["credito"] or {}
        dbs = g["debitos"]

        # Situação vinda da planilha de status do eCAC (se importada).
        # Em branco quando não há planilha ou o nº não foi encontrado.
        num_perdcomp = hdr.get("numero_perdcomp", "").strip()
        situacao = status_map.get(num_perdcomp, {}).get("situacao", "")
        hdr["situacao_perdcomp"] = situacao

        cr_fields = {
            "tipo_credito":        cr.get("tipo",""),
            "periodo_credito":     cr.get("periodo_apuracao",""),
            "comp_teste_credito":  format_competencia_teste(cr.get("periodo_apuracao","")),
            "valor_total_credito": cr.get("valor_original",""),
            "valor_utilizado_doc": cr.get("valor_utilizado",""),
        }

        if not dbs:
            result.append({**hdr, **cr_fields, **_empty_db})
        else:
            for db in dbs:
                result.append({
                    **hdr, **cr_fields,
                    "codigo_receita_debito": db.get("codigo_receita_debito",""),
                    "tipo_debito":       db.get("tipo",""),
                    "periodo_debito":    db.get("periodo_apuracao",""),
                    "comp_teste_debito": format_competencia_teste(db.get("periodo_apuracao","")),
                    "valor_principal":   db.get("valor_original",""),
                    "valor_multa":       db.get("valor_multa",""),
                    "valor_juros":       db.get("valor_juros",""),
                    "valor_total_debito":db.get("valor_total",""),
                })
    return result


def _parse_date(dt_str):
    """Converte 'DD/MM/AAAA' em tupla comparável (AAAA, MM, DD)."""
    if not dt_str:
        return (0, 0, 0)
    try:
        d, m, y = dt_str.strip().split('/')
        return (int(y), int(m), int(d))
    except Exception:
        return (0, 0, 0)


def build_credit_control(rows, status_map=None):
    """
    Aba 2 — Controle de Créditos Tributários.

    REGRAS:
    1. Exibir SOMENTE Pedido de Restituição e Pedido de Ressarcimento.
       DCOMPs, Pedidos de Reembolso e outros tipos NÃO entram.

    2. Excluir pedidos com situação "Cancelado" na planilha de status.

    3. Pedidos retificados:
       - A linha exibe o Nº do pedido ORIGINAL (raiz da cadeia).
       - O valor do crédito e o período vêm do ÚLTIMO retificador importado.
       - A coluna "Último Retificador" mostra o Nº da retificação mais recente.
       - O total compensado considera apenas as DCOMPs vigentes (não-retificadas).

    status_map: { numero_perdcomp: {situacao, ...} }
    """
    if status_map is None:
        status_map = {}

    # ── Passo 1: cadeia de retificações ──────────────────────────────────────
    # retificado_por  : {num_retificado -> row do retificador mais recente}
    # original_de     : {num_qualquer   -> Nº raiz da cadeia}
    # retified_set    : conjunto de todos os Nº que foram retificados
    retificado_por: dict = {}
    retified_set:  set  = set()

    for r in rows:
        if r.get("tipo_registro") != "Crédito":
            continue
        orig = r.get("numero_perdcomp_retificado", "").strip()
        if not orig:
            continue
        retified_set.add(orig)
        ex = retificado_por.get(orig)
        if not ex or _parse_date(r.get("data_transmissao","")) > \
                     _parse_date(ex.get("data_transmissao","")):
            retificado_por[orig] = r

    def _raiz(num: str) -> str:
        """Sobe a cadeia até encontrar o pedido original (que não retifica nenhum outro)."""
        visited = set()
        while True:
            # Verifica se este num retifica outro
            for r in rows:
                if r.get("tipo_registro") != "Crédito": continue
                if r.get("numero_perdcomp","").strip() != num: continue
                retificado = r.get("numero_perdcomp_retificado","").strip()
                if retificado and retificado not in visited:
                    visited.add(num)
                    num = retificado
                    break
            else:
                return num   # não retifica nenhum -> é a raiz

    def _ultimo_retificador(num_original: str) -> str:
        """Desce a cadeia a partir do original para achar o retificador mais recente."""
        atual = num_original
        while atual in retificado_por:
            proximo = retificado_por[atual].get("numero_perdcomp","").strip()
            if not proximo or proximo == atual:
                break
            atual = proximo
        return atual if atual != num_original else ""

    # ── Passo 2: coleta apenas PERs (Restituição e Ressarcimento) ────────────
    # per_map: {num_original -> melhor row de crédito (mais recente)}
    per_map:  dict[str, dict] = {}   # {raiz -> row mais recente com valor}
    tipo_map: dict[str, str]  = {}   # {raiz -> tipo_pedido display}
    tco_map:  dict[str, float]= {}   # {raiz -> total compensado pelas DCOMPs}
    tcr_map:  dict[str, float]= {}   # {raiz -> total das correções SELIC acumuladas}

    TIPOS_ACEITOS = ("restitu", "ressarc")   # Regra 1

    for r in rows:
        if r.get("tipo_registro") != "Crédito":
            continue
        tp = r.get("tipo_pedido", "").lower()
        if not any(k in tp for k in TIPOS_ACEITOS):
            continue

        num = r.get("numero_perdcomp", "").strip()
        if not num:
            continue

        # Regra 2: cancela se situação = Cancelado
        if _is_cancelled(num, status_map):
            continue

        # Regra 3: resolve a raiz da cadeia
        raiz = _raiz(num)

        # Regra 2: cancela se a raiz também está cancelada
        if _is_cancelled(raiz, status_map):
            continue

        # Decide se esta row é mais recente que a já registrada para a raiz
        dt_this = _parse_date(r.get("data_transmissao",""))
        existing = per_map.get(raiz)
        if existing is None or dt_this >= _parse_date(existing.get("data_transmissao","")):
            per_map[raiz] = r

        # Tipo do pedido: prefer display do pedido original (raiz)
        if raiz not in tipo_map or num == raiz:
            raw_tp = r.get("tipo_pedido","")
            if "ressarc" in raw_tp.lower():
                tipo_map[raiz] = "Pedido de Ressarcimento"
            else:
                tipo_map[raiz] = "Pedido de Restituição"

    # ── Passo 3: acumula compensado das DCOMPs vigentes ──────────────────────
    for r in rows:
        if r.get("tipo_registro") != "Crédito":
            continue
        if "compensa" not in r.get("tipo_pedido","").lower():
            continue
        num_dcomp = r.get("numero_perdcomp","").strip()
        if num_dcomp in retified_set:          # DCOMP retificada por outro PDF: ignora
            continue
        vinc = r.get("numero_pedido_vinculado","").strip()
        if not vinc:
            continue
        # Descobre se o vinculo é raiz direta ou um retificador posterior
        raiz = _raiz(vinc) if vinc in retified_set else vinc
        if raiz not in per_map:                # não é PER que nos interessa
            continue
        # Exclui DCOMPs canceladas ou retificadas conforme planilha de status
        # do eCAC (cobre também os casos em que o PDF retificador não foi
        # importado, mas o eCAC já marcou a DCOMP original como retificada).
        if _is_cancelled(num_dcomp, status_map):
            continue
        if _is_retified(num_dcomp, status_map):
            continue
        # Para o controle de créditos, usa "Total dos Débitos deste Documento"
        # — esse é o valor real consumido do crédito quando há correção SELIC,
        # diferente do "Total do Crédito Original Utilizado neste Documento".
        # Fallback: para PERDCOMPs antigos sem o campo separado, mantém o
        # valor utilizado original.
        v_debitos  = parse_brl(r.get("valor_total_debitos",""))
        v_utilizado= parse_brl(r.get("valor_utilizado",""))
        valor_a_compensar = v_debitos if v_debitos > 0 else v_utilizado
        tco_map[raiz] = tco_map.get(raiz, 0.0) + valor_a_compensar
        # A diferença é a correção SELIC acumulada sobre o crédito utilizado.
        # Só faz sentido somar quando ambos os campos estão presentes.
        if v_debitos > 0 and v_utilizado > 0:
            tcr_map[raiz] = tcr_map.get(raiz, 0.0) + (v_debitos - v_utilizado)

    # ── Passo 4: monta resumo ─────────────────────────────────────────────────
    summary = []
    for raiz, best_row in per_map.items():
        tc  = parse_brl(best_row.get("valor_original",""))
        tco = tco_map.get(raiz, 0.0)
        tcr = tcr_map.get(raiz, 0.0)
        # Saldo = (crédito original + correção SELIC acumulada) − total dos débitos compensados
        sd  = (tc + tcr) - tco if tc > 0 else 0.0

        ult_retif = _ultimo_retificador(raiz)

        competencia_raw = best_row.get("periodo_apuracao","")
        summary.append({
            "cnpj":                best_row.get("cnpj",""),
            "razao_social":        best_row.get("razao_social",""),
            "numero_perdcomp":     raiz,
            "situacao_perdcomp":   status_map.get(raiz, {}).get("situacao", ""),
            "tipo_pedido_display": tipo_map.get(raiz,""),
            "tipo_credito":        best_row.get("tipo",""),
            "competencia":         competencia_raw,
            "competencia_teste":   format_competencia_teste(competencia_raw),
            "valor_total_credito": format_brl(tc) if tc > 0 else "Não localizado",
            "valor_total_correcao":format_brl(tcr) if tcr != 0 else "",
            "total_compensado":    format_brl(tco),
            "saldo_disponivel":    format_brl(sd)  if tc > 0 else "Verificar",
            "ultimo_retificador":  ult_retif,
            "_raw_credito":    tc,
            "_raw_correcao":   tcr,
            "_raw_compensado": tco,
            "_raw_saldo":      sd,
        })

    return summary


def build_unlinked_compensations(rows, status_map=None):
    """Compensações sem vínculo, excluindo canceladas."""
    if status_map is None:
        status_map = {}
    return [r for r in rows
            if "compensa" in r.get("tipo_pedido","").lower()
            and not r.get("numero_pedido_vinculado","").strip()
            and not _is_cancelled(r.get("numero_perdcomp","").strip(), status_map)]


def build_missing_from_excel(status_map: dict, rows: list) -> list:
    """
    Aba 4 - PERDCOMPs que estão na planilha de status mas cujo PDF não foi importado.
    """
    if not status_map:
        return []
    imported = {r.get("numero_perdcomp","").strip() for r in rows if r.get("numero_perdcomp")}
    result = []
    for num, info in status_map.items():
        if num not in imported:
            result.append({
                "numero_perdcomp":  num,
                "cnpj":             info.get("cnpj",""),
                "tipo_documento":   info.get("tipo_documento",""),
                "tipo_credito":     info.get("tipo_credito",""),
                "data_transmissao": info.get("data_transmissao",""),
                "situacao":         info.get("situacao",""),
            })
    result.sort(key=lambda r: r["numero_perdcomp"])
    return result


def build_missing_from_pdfs(status_map: dict, rows: list) -> list:
    """
    Aba 5 - PDFs importados cujo Nº PERDCOMP não consta na planilha de status.
    Só relevante quando status_map está preenchido.
    """
    if not status_map:
        return []
    seen: set = set()
    result = []
    for r in rows:
        num = r.get("numero_perdcomp","").strip()
        if not num or num in seen:
            continue
        seen.add(num)
        if num not in status_map:
            result.append({
                "numero_perdcomp":  num,
                "cnpj":             r.get("cnpj",""),
                "razao_social":     r.get("razao_social",""),
                "tipo_pedido":      r.get("tipo_pedido",""),
                "data_transmissao": r.get("data_transmissao",""),
                "arquivo":          r.get("_source",""),
            })
    result.sort(key=lambda r: r["numero_perdcomp"])
    return result

    result.sort(key=lambda r: r["numero_perdcomp"])
    return result


def build_ressarcimento_aba6(rows: list, status_map: dict = None) -> list:
    """
    Aba 6 - Pedidos de Ressarcimento com codigo detalhado do credito.
    Filtra entradas cujo tipo_pedido contem 'ressarcimento' ou
    tipo_credito contem 'ressarc'.
    """
    ctrl = build_credit_control(rows, status_map)

    per_info: dict = {}
    for r in rows:
        if r.get("tipo_registro") != "Crédito":
            continue
        if "ressarc" not in r.get("tipo_pedido","").lower():
            continue
        num = r.get("numero_perdcomp","").strip()
        if num and num not in per_info:
            per_info[num] = {
                "data_transmissao": r.get("data_transmissao",""),
                "codigo_credito":   r.get("codigo_credito",""),
            }

    result = []
    for cr in ctrl:
        num = cr.get("numero_perdcomp","")
        if num not in per_info and "ressarc" not in cr.get("tipo_credito","").lower():
            continue
        info = per_info.get(num, {})
        result.append({
            "cnpj":                cr["cnpj"],
            "razao_social":        cr["razao_social"],
            "numero_perdcomp":     num,
            "data_transmissao":    info.get("data_transmissao",""),
            "tipo_credito":        cr.get("tipo_credito",""),
            "codigo_credito":      info.get("codigo_credito",""),
            "competencia":         cr.get("competencia",""),
            "competencia_teste":   cr.get("competencia_teste",""),
            "valor_total_credito": cr.get("valor_total_credito",""),
            "total_compensado":    cr.get("total_compensado",""),
            "saldo_disponivel":    cr.get("saldo_disponivel",""),
            "_raw_credito":        cr.get("_raw_credito",0.),
            "_raw_compensado":     cr.get("_raw_compensado",0.),
            "_raw_saldo":          cr.get("_raw_saldo",0.),
        })
    return result


# -----------------------------------------------------------------------------
# Exportação Excel
# -----------------------------------------------------------------------------

def _bd():
    s = Side(style="thin", color="D0DDB8")
    return Border(bottom=s, left=s, right=s)

def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = border
    return c

def export_excel(rows, ctrl_rows, unlinked_rows, aba4_rows, aba5_rows, aba6_rows, path,
                 status_map=None):
    wb      = openpyxl.Workbook()
    left_al = Alignment(horizontal="left",   vertical="center")
    ctr_al  = Alignment(horizontal="center", vertical="center")
    rgt_al  = Alignment(horizontal="right",  vertical="center")
    bd      = _bd()

    hdr_fill = PatternFill("solid", fgColor="5A8A1E")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cr_fill  = PatternFill("solid", fgColor="EAF4D3")
    db_fill  = PatternFill("solid", fgColor="FDEDEC")
    cr_font  = Font(name="Calibri", color="3D6B0A", size=9)
    db_font  = Font(name="Calibri", color="C0392B", size=9)
    nm_font  = Font(name="Calibri", size=9, color="3D3D3D")
    src_fill = PatternFill("solid", fgColor="F5F5F0")
    src_font = Font(name="Calibri", italic=True, color="7AB82E", size=8, bold=True)
    pos_font = Font(name="Calibri", bold=True, color="3D6B0A", size=9)
    neg_font = Font(name="Calibri", bold=True, color="C0392B", size=9)
    zer_font = Font(name="Calibri", bold=True, color="6B6B6B", size=9)
    ttl_fill = PatternFill("solid", fgColor="C5E08A")
    ttl_font = Font(name="Calibri", bold=True, color="3D3D3D", size=10)

    def write_detail(ws, title_txt, clr_hex, data_rows, hf=None, hfl=None):
        n = len(DETAIL_COLS)
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        tc = ws["A1"]; tc.value = title_txt
        tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        tc.fill = PatternFill("solid", fgColor=clr_hex); tc.alignment = ctr_al
        ws.row_dimensions[1].height = 30
        hf  = hf  or hdr_font
        hfl = hfl or hdr_fill
        for ci, label in enumerate(DETAIL_LABELS, 1):
            _cell(ws, 2, ci, label, font=hf, fill=hfl, alignment=ctr_al, border=bd)
        ws.row_dimensions[2].height = 22
        dr, last = 3, None
        for r in data_rows:
            src = r.get("_source","")
            if src != last:
                last = src
                ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=n)
                _cell(ws, dr, 1, f"   {src}", font=src_font, fill=src_fill, alignment=left_al)
                ws.row_dimensions[dr].height = 14; dr += 1
            tr   = r.get("tipo_registro","")
            rf   = cr_fill if tr=="Crédito" else (db_fill if tr=="Débito" else None)
            rf2  = cr_font if tr=="Crédito" else (db_font if tr=="Débito" else nm_font)
            for ci, key in enumerate(DETAIL_KEYS, 1):
                _cell(ws, dr, ci, r.get(key,""), font=rf2, fill=rf, alignment=left_al, border=bd)
            ws.row_dimensions[dr].height = 17; dr += 1
        for ci, w in enumerate(DETAIL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w / 7.5
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"

    # ABA 1 - Detalhamento (crédito + débitos na mesma linha)
    ws1 = wb.active; ws1.title = "Detalhamento"
    aba1_data = combine_rows_for_aba1(rows, status_map)
    n1 = len(ABA1_COLS)

    # Título
    ws1.merge_cells(f"A1:{get_column_letter(n1)}1")
    tc1 = ws1["A1"]
    tc1.value = "DETALHAMENTO DE PERDCOMPS - AgriTax Tributário & Contábil"
    tc1.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc1.fill  = PatternFill("solid", fgColor="7AB82E"); tc1.alignment = ctr_al
    ws1.row_dimensions[1].height = 30

    # Linha 2: seções (Cadastral / Crédito / Débito)
    # Cadastral: cols 1-7, Crédito: 8-12, Débito: 13-19
    sec_cfg = [
        (1,   7, "DADOS CADASTRAIS",  "3A5A10"),
        (8,  12, "CRÉDITO",           "1A5276"),
        (13, 19, "DÉBITO",            "7B241C"),
    ]
    for c1, c2, label, color in sec_cfg:
        ws1.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        _cell(ws1, 2, c1, label,
              font=Font(name="Calibri", bold=True, color="FFFFFF", size=10),
              fill=PatternFill("solid", fgColor=color),
              alignment=ctr_al)
    ws1.row_dimensions[2].height = 18

    # Linha 3: cabeçalhos das colunas
    hdr_fills = (
        [PatternFill("solid", fgColor="5A8A1E")] * 7 +   # cadastral: verde
        [PatternFill("solid", fgColor="2471A3")] * 5 +   # crédito: azul (5 cols)
        [PatternFill("solid", fgColor="A93226")] * 7     # débito: vermelho (7 cols)
    )
    for ci, (label, fill) in enumerate(zip(ABA1_LABELS, hdr_fills), 1):
        _cell(ws1, 3, ci, label,
              font=Font(name="Calibri", bold=True, color="FFFFFF", size=9),
              fill=fill, alignment=ctr_al, border=bd)
    ws1.row_dimensions[3].height = 22

    # Dados
    cr_bg   = PatternFill("solid", fgColor="EAF4D3")  # fundo crédito: verde claro
    db_bg   = PatternFill("solid", fgColor="FDECEA")  # fundo débito: vermelho claro
    cad_bg  = PatternFill("solid", fgColor="F7FBF0")  # fundo cadastral: neutro
    cr_fnt  = Font(name="Calibri", size=9, color="1A5276")
    db_fnt  = Font(name="Calibri", size=9, color="7B241C")
    cad_fnt = Font(name="Calibri", size=9, color="3D3D3D")
    src1_fill = PatternFill("solid", fgColor="F5F5F0")
    src1_font = Font(name="Calibri", italic=True, color="7AB82E", size=8, bold=True)

    dr, last_src = 4, None
    for r in aba1_data:
        src = r.get("_source","")
        if src != last_src:
            last_src = src
            ws1.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=n1)
            _cell(ws1, dr, 1, f"   {src}", font=src1_font, fill=src1_fill, alignment=left_al)
            ws1.row_dimensions[dr].height = 14; dr += 1

        has_debit = bool(r.get("tipo_debito",""))
        for ci, key in enumerate(ABA1_KEYS, 1):
            if ci <= 7:
                fnt, fill = cad_fnt, cad_bg
            elif ci <= 12:
                fnt, fill = cr_fnt, cr_bg
            else:
                fnt, fill = (db_fnt, db_bg) if has_debit else (cad_fnt, cad_bg)
            _cell(ws1, dr, ci, r.get(key,""), font=fnt, fill=fill,
                  alignment=left_al, border=bd)
        ws1.row_dimensions[dr].height = 17; dr += 1

    for ci, w in enumerate(ABA1_WIDTHS, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w / 7.5
    ws1.freeze_panes = "A4"
    ws1.auto_filter.ref = f"A3:{get_column_letter(n1)}3"

    # ABA 2 - Controle de Créditos
    ws2 = wb.create_sheet("Controle de Créditos")
    n2 = len(CTRL_COLS)
    ws2.merge_cells(f"A1:{get_column_letter(n2)}1")
    tc2 = ws2["A1"]; tc2.value = "CONTROLE DE CRÉDITOS TRIBUTÁRIOS - PERDCOMP"
    tc2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc2.fill = PatternFill("solid", fgColor="5A8A1E"); tc2.alignment = ctr_al
    ws2.row_dimensions[1].height = 30
    h2_fill = PatternFill("solid", fgColor="3A5A10")
    h2_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    for ci, label in enumerate(CTRL_LABELS, 1):
        _cell(ws2, 2, ci, label, font=h2_font, fill=h2_fill, alignment=ctr_al, border=bd)
    ws2.row_dimensions[2].height = 22
    r2_fill   = PatternFill("solid", fgColor="F7FBF0")
    r2_font   = Font(name="Calibri", size=9, color="3D3D3D")
    money_keys= {"valor_total_credito","valor_total_correcao","total_compensado","saldo_disponivel"}
    for ri, cr in enumerate(ctrl_rows, 3):
        sd  = cr.get("_raw_saldo", 0.)
        sf  = pos_font if sd > 0 else (neg_font if sd < 0 else zer_font)
        for ci, key in enumerate(CTRL_KEYS, 1):
            fnt = sf if key == "saldo_disponivel" else r2_font
            al  = rgt_al if key in money_keys else left_al
            _cell(ws2, ri, ci, cr.get(key,""), font=fnt, fill=r2_fill, alignment=al, border=bd)
        ws2.row_dimensions[ri].height = 18
    if ctrl_rows:
        tot = 3 + len(ctrl_rows)
        # 13 colunas: merge 1..n2-5 (=8) para "TOTAIS"; últimas 5 = vl_cred, correção, comp, saldo, ult_retif
        ws2.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=n2-5)
        _cell(ws2, tot, 1, "TOTAIS", font=ttl_font, fill=ttl_fill, alignment=ctr_al)
        tc  = sum(r.get("_raw_credito",    0.) for r in ctrl_rows)
        tcr = sum(r.get("_raw_correcao",   0.) for r in ctrl_rows)
        tco = sum(r.get("_raw_compensado", 0.) for r in ctrl_rows)
        tsd = sum(r.get("_raw_saldo",      0.) for r in ctrl_rows)
        for ci, val in zip([n2-4, n2-3, n2-2, n2-1], [tc, tcr, tco, tsd]):  # n2 = ult_retificador (vazio)
            sf = pos_font if val > 0 else (neg_font if val < 0 else zer_font)
            _cell(ws2, tot, ci, format_brl(val),
                  font=sf, fill=ttl_fill, alignment=rgt_al, border=bd)
        ws2.row_dimensions[tot].height = 22
    # 13 colunas: cnpj | razao | num | situação | tipo_pedido | tipo_cred | comp | comp_teste | vl_cred | correção | comp | saldo | ult_retif
    ctrl_widths = [140, 200, 185, 130, 165, 220, 130, 120, 155, 155, 155, 155, 185]
    for ci, w in enumerate(ctrl_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w / 7.5
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:{get_column_letter(n2)}2"

    # ABA 3 - Compensações Sem Vínculo
    ws3    = wb.create_sheet("Comp. Sem Vínculo")
    h3_fill= PatternFill("solid", fgColor="E09A1E")
    h3_font= Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    title3 = "COMPENSAÇÕES SEM PER/DCOMP DE ORIGEM VINCULADO"
    if unlinked_rows:
        write_detail(ws3, title3, "F5A623", unlinked_rows, hf=h3_font, hfl=h3_fill)
    else:
        n = len(DETAIL_COLS)
        ws3.merge_cells(f"A1:{get_column_letter(n)}1")
        tc3 = ws3["A1"]; tc3.value = title3
        tc3.font = Font(name="Calibri",bold=True,color="FFFFFF",size=12)
        tc3.fill = PatternFill("solid",fgColor="F5A623"); tc3.alignment = ctr_al
        ws3.row_dimensions[1].height = 30
        for ci,label in enumerate(DETAIL_LABELS,1):
            _cell(ws3,2,ci,label,font=h3_font,fill=h3_fill,alignment=ctr_al,border=bd)
        ws3.merge_cells(f"A3:{get_column_letter(n)}3")
        _cell(ws3,3,1,"Nenhuma compensação sem vínculo encontrada.",
              font=Font(name="Calibri",italic=True,color="6B6B6B",size=9),
              fill=PatternFill("solid",fgColor="FEF9F0"),alignment=ctr_al)
        ws3.row_dimensions[3].height=22
        for ci,w in enumerate(DETAIL_WIDTHS,1):
            ws3.column_dimensions[get_column_letter(ci)].width=w/7.5
    # ════════════════════════════════════════════════════════════
    # ABA 4 - Na planilha de status, sem PDF importado
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Sem PDF (na planilha)")
    h4_fill = PatternFill("solid", fgColor="1A365D")
    h4_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    r4_fill = PatternFill("solid", fgColor="EBF4FF")
    r4_font = Font(name="Calibri", size=9, color="1A365D")
    canc_fill = PatternFill("solid", fgColor="FFF3CD")
    canc_font = Font(name="Calibri", size=9, color="856404")

    n4 = len(ABA4_COLS)
    ws4.merge_cells(f"A1:{get_column_letter(n4)}1")
    tc4 = ws4["A1"]
    tc4.value = "PERDCOMPS NA PLANILHA DE STATUS - SEM PDF IMPORTADO"
    tc4.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc4.fill  = PatternFill("solid", fgColor="1A365D"); tc4.alignment = ctr_al
    ws4.row_dimensions[1].height = 30
    for ci, label in enumerate(ABA4_LABELS, 1):
        _cell(ws4, 2, ci, label, font=h4_font, fill=h4_fill, alignment=ctr_al, border=bd)
    ws4.row_dimensions[2].height = 22

    if aba4_rows:
        for ri, r in enumerate(aba4_rows, 3):
            sit = r.get("situacao","").lower()
            rf  = canc_fill if "cancelad" in sit else r4_fill
            rf2 = canc_font if "cancelad" in sit else r4_font
            for ci, key in enumerate(ABA4_KEYS, 1):
                _cell(ws4, ri, ci, r.get(key,""), font=rf2, fill=rf, alignment=left_al, border=bd)
            ws4.row_dimensions[ri].height = 17
    else:
        ws4.merge_cells(f"A3:{get_column_letter(n4)}3")
        _cell(ws4, 3, 1,
              "Todos os PERDCOMPs da planilha têm PDF importado (ou planilha não foi carregada).",
              font=Font(name="Calibri", italic=True, color="6B6B6B", size=9),
              fill=r4_fill, alignment=ctr_al)
        ws4.row_dimensions[3].height = 22

    for ci, w in enumerate(ABA4_WIDTHS, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w / 7.5
    ws4.freeze_panes = "A3"
    ws4.auto_filter.ref = f"A2:{get_column_letter(n4)}2"

    # ════════════════════════════════════════════════════════════
    # ABA 5 - PDF importado, não consta na planilha de status
    # ════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Sem Planilha (PDF importado)")
    h5_fill = PatternFill("solid", fgColor="4A235A")
    h5_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    r5_fill = PatternFill("solid", fgColor="F8F0FF")
    r5_font = Font(name="Calibri", size=9, color="4A235A")

    n5 = len(ABA5_COLS)
    ws5.merge_cells(f"A1:{get_column_letter(n5)}1")
    tc5 = ws5["A1"]
    tc5.value = "PDFs IMPORTADOS - NÃO CONSTAM NA PLANILHA DE STATUS"
    tc5.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc5.fill  = PatternFill("solid", fgColor="4A235A"); tc5.alignment = ctr_al
    ws5.row_dimensions[1].height = 30
    for ci, label in enumerate(ABA5_LABELS, 1):
        _cell(ws5, 2, ci, label, font=h5_font, fill=h5_fill, alignment=ctr_al, border=bd)
    ws5.row_dimensions[2].height = 22

    if aba5_rows:
        for ri, r in enumerate(aba5_rows, 3):
            for ci, key in enumerate(ABA5_KEYS, 1):
                _cell(ws5, ri, ci, r.get(key,""), font=r5_font, fill=r5_fill,
                      alignment=left_al, border=bd)
            ws5.row_dimensions[ri].height = 17
    else:
        ws5.merge_cells(f"A3:{get_column_letter(n5)}3")
        _cell(ws5, 3, 1,
              "Todos os PDFs importados constam na planilha (ou planilha não foi carregada).",
              font=Font(name="Calibri", italic=True, color="6B6B6B", size=9),
              fill=r5_fill, alignment=ctr_al)
        ws5.row_dimensions[3].height = 22

    for ci, w in enumerate(ABA5_WIDTHS, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w / 7.5
    ws5.freeze_panes = "A3"
    ws5.auto_filter.ref = f"A2:{get_column_letter(n5)}2"

    # ========================================================
    # ABA 6 - Pedidos de Ressarcimento (código crédito detalh.)
    # ========================================================
    ws6 = wb.create_sheet("Ressarcimentos")
    h6_fill  = PatternFill("solid", fgColor="1B4332")
    h6_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    r6_fill  = PatternFill("solid", fgColor="D8F3DC")
    r6_font  = Font(name="Calibri", size=9, color="1B4332")
    n6 = len(ABA6_COLS)
    ws6.merge_cells(f"A1:{get_column_letter(n6)}1")
    tc6 = ws6["A1"]
    tc6.value = "PEDIDOS DE RESSARCIMENTO - DETALHAMENTO POR TIPO DE CRÉDITO"
    tc6.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc6.fill  = PatternFill("solid", fgColor="1B4332"); tc6.alignment = ctr_al
    ws6.row_dimensions[1].height = 30
    for ci, label in enumerate(ABA6_LABELS, 1):
        _cell(ws6, 2, ci, label, font=h6_font, fill=h6_fill, alignment=ctr_al, border=bd)
    ws6.row_dimensions[2].height = 22
    money6 = {"valor_total_credito","total_compensado","saldo_disponivel"}
    if aba6_rows:
        for ri, r in enumerate(aba6_rows, 3):
            sd  = r.get("_raw_saldo", 0.)
            sf  = pos_font if sd > 0 else (neg_font if sd < 0 else zer_font)
            for ci, key in enumerate(ABA6_KEYS, 1):
                fnt = sf if key == "saldo_disponivel" else r6_font
                al  = rgt_al if key in money6 else left_al
                _cell(ws6, ri, ci, r.get(key,""), font=fnt, fill=r6_fill,
                      alignment=al, border=bd)
            ws6.row_dimensions[ri].height = 18
        tot6 = 3 + len(aba6_rows)
        ws6.merge_cells(start_row=tot6, start_column=1, end_row=tot6, end_column=n6-3)
        _cell(ws6, tot6, 1, "TOTAIS", font=ttl_font, fill=ttl_fill, alignment=ctr_al)
        t6c  = sum(r.get("_raw_credito",0.)    for r in aba6_rows)
        t6co = sum(r.get("_raw_compensado",0.) for r in aba6_rows)
        t6sd = sum(r.get("_raw_saldo",0.)      for r in aba6_rows)
        for ci, val in zip([n6-2, n6-1, n6], [t6c, t6co, t6sd]):
            sf = pos_font if val>0 else (neg_font if val<0 else zer_font)
            _cell(ws6, tot6, ci, format_brl(val), font=sf, fill=ttl_fill,
                  alignment=rgt_al, border=bd)
        ws6.row_dimensions[tot6].height = 22
    else:
        ws6.merge_cells(f"A3:{get_column_letter(n6)}3")
        _cell(ws6, 3, 1, "Nenhum Pedido de Ressarcimento encontrado.",
              font=Font(name="Calibri",italic=True,color="6B6B6B",size=9),
              fill=r6_fill, alignment=ctr_al)
        ws6.row_dimensions[3].height = 22
    for ci, w in enumerate(ABA6_WIDTHS, 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w / 7.5
    ws6.freeze_panes = "A3"
    ws6.auto_filter.ref = f"A2:{get_column_letter(n6)}2"

    wb.save(path)


# -----------------------------------------------------------------------------
# DataStore — Repositório centralizado de dados importados
# -----------------------------------------------------------------------------

class DataStore:
    """
    Repositório único de dados importados em memória.

    Todos os módulos de Auditoria leem desta estrutura, que é populada
    pela Central de Importação. Notifica os listeners (módulos) quando
    novos dados são importados, permitindo que recalculem suas análises.

    Arquitetura de vida da sessão:
      - Única instância por execução do programa (singleton por conveniência)
      - Dados vivem só enquanto o programa está aberto (sem persistência)
      - Listeners são módulos que assinam eventos de mudança
    """

    def __init__(self):
        # Dados brutos por tipo de documento
        # perdcomp_raw: lista de dicts retornados por parse_pdf() (cabeçalho + créditos + débitos)
        self.perdcomp_raw: list = []
        # perdcomp_rows: resultado de flatten_rows (uma linha por crédito OU débito)
        self.perdcomp_rows: list = []
        # darf_rows: resultado de parse_darf_pdf (consolidado)
        self.darf_rows: list = []
        # dctf_rows: resultado de extract_dctf (uma linha por tributo declarado)
        self.dctf_rows: list = []
        # dctfweb_rows: resultado de extract_dctfweb (uma linha por tributo da DCTFWeb)
        self.dctfweb_rows: list = []
        # efd_rows: resultado de extract_efd_contribuicoes (PIS/COFINS por código)
        self.efd_rows: list = []
        # Planilha de status do eCAC
        self.status_map: dict = {}
        self.status_path: str = ""

        # Rastreamento de arquivos processados (para mostrar na Central)
        # Cada item: {"path": "...", "nome": "...", "status": "✓|⏳|✗", "erro": "..."}
        self.perdcomp_files: list = []
        self.darf_files:    list = []
        self.dctf_files:    list = []
        self.dctfweb_files: list = []
        self.efd_files:     list = []

        # Observadores (módulos de auditoria) — chamados quando dados mudam
        # Formato: {"perdcomp": [callback, ...], "darf": [...], "dctf": [...], "status": [...]}
        self._listeners: dict = {
            "perdcomp": [], "darf": [], "dctf": [], "dctfweb": [],
            "efd": [], "status": [], "any": [],
        }

    # ── Publicação / assinatura de mudanças ────────────────────────────────
    def subscribe(self, tipo: str, callback):
        """Registra callback para ser chamado quando dados do tipo mudarem.

        tipo pode ser: 'perdcomp', 'darf', 'dctf', 'status', 'any'.
        """
        if tipo not in self._listeners:
            self._listeners[tipo] = []
        if callback not in self._listeners[tipo]:
            self._listeners[tipo].append(callback)

    def _notify(self, tipo: str):
        """Notifica os listeners (e os que assinam 'any')."""
        for cb in list(self._listeners.get(tipo, [])) + list(self._listeners.get("any", [])):
            try:
                cb()
            except Exception as e:
                # Erro num listener não deve quebrar os outros
                print(f"[DataStore] Erro em listener de '{tipo}': {e}")

    # ── API de mutação (chamada pela Central de Importação) ────────────────
    # ── Chaves naturais para deduplicação ─────────────────────────────────
    # Cada tipo de documento tem uma chave única que identifica unicamente o
    # documento. Se um documento com a mesma chave já foi importado, o novo é
    # IGNORADO (não duplicado). O usuário recebe aviso da quantidade ignorada.
    @staticmethod
    def _key_perdcomp(r: dict) -> str:
        return (r.get("numero_perdcomp", "") or "").strip()

    @staticmethod
    def _key_darf(r: dict) -> str:
        # Combinação CNPJ + nº doc + código + período é única para cada item
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        return (f"{cnpj}|{r.get('numero_doc','')}|{r.get('codigo','')}"
                f"|{r.get('periodo','')}|{r.get('total_item','')}")

    @staticmethod
    def _key_dctf(r: dict) -> str:
        # Nº da declaração + código de receita (uma DCTF pode ter vários tributos)
        return f"{r.get('numero_declaracao','')}|{r.get('codigo_receita','')}"

    @staticmethod
    def _key_dctfweb(r: dict) -> str:
        # Nº do recibo + código + CNO/CNPJ-Prest (um recibo pode ter vários tributos)
        return (f"{r.get('numero_recibo','')}|{r.get('codigo_receita','')}"
                f"|{r.get('cno','')}|{r.get('cnpj_prest','')}")

    @staticmethod
    def _key_efd(r: dict) -> str:
        # CNPJ + período + código de receita (uma EFD por mês por contribuinte)
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        return f"{cnpj}|{r.get('periodo','')}|{r.get('codigo_receita','')}|{r.get('tributo','')}"

    def _dedup_and_add(self, kind: str, target_list: list, files_list: list,
                        new_rows: list, file_records: list, key_fn) -> dict:
        """Filtra duplicados pelas chaves, adiciona apenas linhas novas, e marca
        os file_records cujos arquivos foram TODOS rejeitados como duplicados.

        Retorna: {"adicionadas": N, "ignoradas": N, "arquivos_dup": N,
                   "chaves_dup": [...]}.
        """
        chaves_existentes = {key_fn(r) for r in target_list if key_fn(r)}
        rows_aceitas = []
        chaves_ignoradas = set()
        sources_ignorados = set()
        sources_aproveitados = set()

        for r in new_rows:
            chave = key_fn(r)
            src = r.get("_source", "")
            if chave and chave in chaves_existentes:
                chaves_ignoradas.add(chave)
                if src: sources_ignorados.add(src)
                continue
            rows_aceitas.append(r)
            if src: sources_aproveitados.add(src)
            if chave: chaves_existentes.add(chave)

        # Marca arquivos cujas linhas foram TODAS ignoradas (duplicado total)
        arq_dup_count = 0
        for fr in file_records:
            nome = fr.get("nome", "")
            if (nome in sources_ignorados
                    and nome not in sources_aproveitados
                    and fr.get("status") == "✓"):
                fr["status"] = "⚠"
                fr["erro"] = "Documento já importado (duplicado)"
                arq_dup_count += 1

        target_list.extend(rows_aceitas)
        files_list.extend(file_records)

        return {
            "adicionadas":   len(rows_aceitas),
            "ignoradas":     len(chaves_ignoradas),
            "arquivos_dup":  arq_dup_count,
            "chaves_dup":    sorted(chaves_ignoradas)[:10],   # amostra
        }

    def add_perdcomps(self, new_raw: list, file_records: list) -> dict:
        """Adiciona PERDCOMPs com deduplicação por nº da PER/DCOMP.

        Retorna dict com {adicionadas, ignoradas, arquivos_dup, chaves_dup}.
        """
        # PERDCOMP é especial: new_raw contém dicts brutos (1 por PDF), mas
        # o que vai pra perdcomp_rows é o flatten. Dedup deve ser feito no raw.
        chaves_existentes = {self._key_perdcomp(p) for p in self.perdcomp_raw
                              if self._key_perdcomp(p)}
        raws_aceitos = []
        chaves_ignoradas = set()
        sources_ignorados = set()
        sources_aproveitados = set()

        for p in new_raw:
            chave = self._key_perdcomp(p)
            src = p.get("_source", "")
            if chave and chave in chaves_existentes:
                chaves_ignoradas.add(chave)
                if src: sources_ignorados.add(src)
                continue
            raws_aceitos.append(p)
            if src: sources_aproveitados.add(src)
            if chave: chaves_existentes.add(chave)

        arq_dup_count = 0
        for fr in file_records:
            nome = fr.get("nome", "")
            if (nome in sources_ignorados
                    and nome not in sources_aproveitados
                    and fr.get("status") == "✓"):
                fr["status"] = "⚠"
                fr["erro"] = "PERDCOMP já importada (duplicado)"
                arq_dup_count += 1

        self.perdcomp_raw.extend(raws_aceitos)
        # Recalcula flatten consolidado (todos os raws, não só os novos)
        self.perdcomp_rows = []
        for p in self.perdcomp_raw:
            src = p.get("_source", "")
            self.perdcomp_rows.extend(flatten_rows([p], src))
        self.perdcomp_files.extend(file_records)
        self._notify("perdcomp")

        return {
            "adicionadas":  len(raws_aceitos),
            "ignoradas":    len(chaves_ignoradas),
            "arquivos_dup": arq_dup_count,
            "chaves_dup":   sorted(chaves_ignoradas)[:10],
        }

    def add_darfs(self, new_rows: list, file_records: list) -> dict:
        info = self._dedup_and_add("darf", self.darf_rows, self.darf_files,
                                    new_rows, file_records, self._key_darf)
        self._notify("darf")
        return info

    def add_dctfs(self, new_rows: list, file_records: list) -> dict:
        info = self._dedup_and_add("dctf", self.dctf_rows, self.dctf_files,
                                    new_rows, file_records, self._key_dctf)
        self._notify("dctf")
        return info

    def add_dctfwebs(self, new_rows: list, file_records: list) -> dict:
        info = self._dedup_and_add("dctfweb", self.dctfweb_rows, self.dctfweb_files,
                                    new_rows, file_records, self._key_dctfweb)
        self._notify("dctfweb")
        return info

    def add_efds(self, new_rows: list, file_records: list) -> dict:
        info = self._dedup_and_add("efd", self.efd_rows, self.efd_files,
                                    new_rows, file_records, self._key_efd)
        self._notify("efd")
        return info

    def set_status_map(self, status_map: dict, path: str):
        self.status_map = status_map or {}
        self.status_path = path or ""
        self._notify("status")

    def clear_all(self):
        self.perdcomp_raw.clear()
        self.perdcomp_rows.clear()
        self.darf_rows.clear()
        self.dctf_rows.clear()
        self.dctfweb_rows.clear()
        self.efd_rows.clear()
        self.status_map.clear()
        self.status_path = ""
        self.perdcomp_files.clear()
        self.darf_files.clear()
        self.dctf_files.clear()
        self.dctfweb_files.clear()
        self.efd_files.clear()
        self._notify("any")

    def clear_perdcomps(self):
        self.perdcomp_raw.clear()
        self.perdcomp_rows.clear()
        self.perdcomp_files.clear()
        self._notify("perdcomp")

    def clear_darfs(self):
        self.darf_rows.clear()
        self.darf_files.clear()
        self._notify("darf")

    def clear_dctfs(self):
        self.dctf_rows.clear()
        self.dctf_files.clear()
        self._notify("dctf")

    def clear_dctfwebs(self):
        self.dctfweb_rows.clear()
        self.dctfweb_files.clear()
        self._notify("dctfweb")

    def clear_efds(self):
        self.efd_rows.clear()
        self.efd_files.clear()
        self._notify("efd")

    # ── Resumos (para a Central mostrar contadores) ────────────────────────
    def summary(self) -> dict:
        return {
            "perdcomp_files":     len(self.perdcomp_files),
            "perdcomp_rows":      len(self.perdcomp_rows),
            "darf_files":         len(self.darf_files),
            "darf_rows":          len(self.darf_rows),
            "dctf_files":         len(self.dctf_files),
            "dctf_rows":          len(self.dctf_rows),
            "dctfweb_files":      len(self.dctfweb_files),
            "dctfweb_rows":       len(self.dctfweb_rows),
            "efd_files":          len(self.efd_files),
            "efd_rows":           len(self.efd_rows),
            "status_loaded":      bool(self.status_map),
            "status_registros":   len(self.status_map),
        }


# Instância global única — acessível por qualquer módulo da app
_DATA_STORE: "DataStore | None" = None

def get_datastore() -> DataStore:
    """Retorna a instância singleton do DataStore da sessão atual."""
    global _DATA_STORE
    if _DATA_STORE is None:
        _DATA_STORE = DataStore()
    return _DATA_STORE


# -----------------------------------------------------------------------------
# Interface gráfica - Popup de filtro estilo Excel
# -----------------------------------------------------------------------------

class ColFilterPopup(tk.Toplevel):
    """
    Popup de filtro estilo Excel: lista de valores únicos com checkboxes
    e campo de busca, ativado ao clicar no cabeçalho de uma coluna.
    """
    def __init__(self, parent, col_label, all_values, active_set, on_apply, px, py):
        super().__init__(parent)
        self._on_apply   = on_apply
        self._all_values = sorted(str(v) for v in all_values if str(v).strip())
        self._active_set = set(active_set) if active_set is not None else set(self._all_values)

        # Aparência sem borda de janela
        self.overrideredirect(True)
        self.configure(bg="#F0F6E8", relief="solid", bd=1)
        self.resizable(False, False)

        # Título
        tk.Label(self, text=f"  Filtrar: {col_label[:22]}",
                 bg="#3A5A10", fg="white",
                 font=("Segoe UI", 8, "bold"),
                 anchor="w").pack(fill="x")

        # Campo de busca
        sf = tk.Frame(self, bg="#F0F6E8")
        sf.pack(fill="x", padx=6, pady=(6, 2))
        tk.Label(sf, text="🔍", bg="#F0F6E8", font=("Segoe UI", 8)).pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._on_search)
        tk.Entry(sf, textvariable=self._search_var,
                 font=("Segoe UI", 8), width=20, relief="flat",
                 highlightthickness=1, highlightbackground="#C5E08A",
                 bg="white").pack(side="left", fill="x", expand=True, ipady=3, padx=(2, 0))

        # Linha separadora
        tk.Frame(self, bg="#C5E08A", height=1).pack(fill="x", pady=2)

        # "Selecionar Tudo"
        self._sel_all = tk.BooleanVar(value=len(self._active_set) == len(self._all_values))
        tk.Checkbutton(self, text="  (Selecionar Tudo)", variable=self._sel_all,
                       command=self._toggle_all,
                       bg="#EAF4D3", fg="#3A5A10",
                       font=("Segoe UI", 8, "bold"),
                       selectcolor="white", anchor="w",
                       activebackground="#EAF4D3").pack(fill="x", padx=4, pady=(0, 1))

        tk.Frame(self, bg="#C5E08A", height=1).pack(fill="x")

        # Lista de checkboxes rolável
        lf = tk.Frame(self, bg="white")
        lf.pack(fill="both", expand=True, padx=4, pady=(1, 0))
        vsb = ttk.Scrollbar(lf, orient="vertical")
        self._canvas = tk.Canvas(lf, bg="white", highlightthickness=0,
                                  yscrollcommand=vsb.set, width=230, height=200)
        vsb.configure(command=self._canvas.yview)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._inner = tk.Frame(self._canvas, bg="white")
        self._inner_id = self._canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>",
            lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<MouseWheel>",
            lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._check_vars: dict = {}
        self._populate(self._all_values)

        # Botões
        bf = tk.Frame(self, bg="#F0F6E8")
        bf.pack(fill="x", padx=6, pady=6)
        tk.Button(bf, text="  OK  ", bg="#3A5A10", fg="white",
                  font=("Segoe UI", 8, "bold"), relief="flat",
                  padx=8, pady=4, cursor="hand2",
                  command=self._apply,
                  activebackground="#4A7018").pack(side="left")
        tk.Button(bf, text="Limpar Filtro", bg="#E8E8E8",
                  font=("Segoe UI", 8), relief="flat",
                  padx=8, pady=4, cursor="hand2",
                  command=self._clear,
                  activebackground="#D0D0D0").pack(side="left", padx=(4, 0))
        tk.Button(bf, text="✕", bg="#F0F6E8", fg="#888",
                  font=("Segoe UI", 8), relief="flat", padx=6, pady=4,
                  cursor="hand2", command=self.destroy).pack(side="right")

        # Posiciona próximo ao cursor, dentro da tela
        self.update_idletasks()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x = min(px, sw - w - 10)
        y = min(py, sh - h - 10)
        self.geometry(f"+{max(0,x)}+{max(0,y)}")
        self.focus_set()
        self.bind("<Escape>", lambda e: self.destroy())
        self.bind("<FocusOut>", self._on_focus_out)

    def _on_focus_out(self, event):
        try:
            fw = self.focus_get()
            if fw is None or not (str(fw) == str(self) or str(fw).startswith(str(self) + ".")):
                self.after(150, lambda: self.destroy() if self.winfo_exists() else None)
        except Exception:
            pass

    def _populate(self, values):
        for w in self._inner.winfo_children():
            w.destroy()
        self._check_vars = {}
        for val in values:
            active = (val in self._active_set)
            v = tk.BooleanVar(value=active)
            v.trace_add("write", lambda *a: self._update_sel_all())
            display = val[:38] + "…" if len(val) > 38 else val
            cb = tk.Checkbutton(self._inner, text=f"  {display}",
                                 variable=v, bg="white",
                                 font=("Segoe UI", 8), anchor="w",
                                 selectcolor="white",
                                 activebackground="#EAF4D3")
            cb.pack(fill="x", anchor="w", padx=2, pady=1)
            self._check_vars[val] = v
        self._update_sel_all()

    def _update_sel_all(self):
        if not self._check_vars:
            return
        all_on  = all(v.get() for v in self._check_vars.values())
        self._sel_all.set(all_on)

    def _toggle_all(self):
        state = self._sel_all.get()
        for v in self._check_vars.values():
            v.set(state)

    def _on_search(self, *args):
        q = self._search_var.get().strip().lower()
        filtered = [v for v in self._all_values if not q or q in v.lower()]
        self._populate(filtered)

    def _apply(self):
        selected = {k for k, v in self._check_vars.items() if v.get()}
        # Preserve selections for items hidden by search
        q = self._search_var.get().strip().lower()
        if q:
            hidden_active = {v for v in self._active_set if q not in v.lower()}
            selected |= hidden_active
        # None means "no filter" (all values)
        result = None if selected >= set(self._all_values) else selected
        self._on_apply(result)
        self.destroy()

    def _clear(self):
        self._on_apply(None)
        self.destroy()


class App:
    def __init__(self, root):
        self.root = root; self.rows = []; self.files = []; self._logo_img = None
        self.status_map: dict = {}
        # Filtros por coluna para cada aba: {col_key: set_de_valores_a_mostrar | None}
        self.col_filt_1: dict = {}
        self.col_filt_2: dict = {}
        self.col_filt_3: dict = {}
        self.col_filt_4: dict = {}
        self.col_filt_5: dict = {}
        self.col_filt_6: dict = {}
        self._setup_window(); self._apply_styles(); self._build_ui(); self._check_deps()

    def _setup_window(self):
        self.root.title(APP_TITLE)
        self.root.geometry("1280x780")
        self.root.minsize(1000, 620)
        self.root.configure(bg=C_GRAY_LIGHT)
        # Se for Toplevel (aberto do launcher), configura como janela independente
        if isinstance(self.root, tk.Toplevel):
            self.root.protocol("WM_DELETE_WINDOW", self.root.destroy)
        try: self.root.state("zoomed")
        except: pass

    def _apply_styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("TNotebook", background=C_GRAY_LIGHT, borderwidth=0, tabmargins=[0,4,0,0])
        s.configure("TNotebook.Tab", background="#D8E8B8", foreground=C_GRAY_DARK,
                    padding=[16,7], font=("Segoe UI",9), borderwidth=0)
        s.map("TNotebook.Tab", background=[("selected",C_WHITE)],
              foreground=[("selected",C_GREEN_DARK)])
        s.configure("Treeview", background=C_WHITE, foreground=C_GRAY_DARK,
                    fieldbackground=C_WHITE, rowheight=24, font=("Segoe UI",8))
        s.configure("Treeview.Heading", background=C_GREEN_DARK, foreground=C_WHITE,
                    font=("Segoe UI",8,"bold"), relief="flat")
        s.map("Treeview", background=[("selected","#C5E08A")], foreground=[("selected",C_GRAY_DARK)])
        s.map("Treeview.Heading", background=[("active","#4A7018")])
        s.configure("TProgressbar", troughcolor=C_GREEN_LIGHT, background=C_GREEN, bordercolor=C_BORDER)

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg=C_WHITE, height=72)
        hdr.pack(fill="x", side="top"); hdr.pack_propagate(False)
        tk.Frame(self.root, bg=C_GREEN, height=3).pack(fill="x", side="top")

        if PIL_OK:
            try:
                raw = base64.b64decode(LOGO_B64)
                img = Image.open(io.BytesIO(raw))
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(hdr, image=self._logo_img, bg=C_WHITE).pack(side="left", padx=18, pady=10)
            except: pass

        tf = tk.Frame(hdr, bg=C_WHITE); tf.pack(side="left", pady=12)
        tk.Label(tf, text="PERDCOMP Extractor", bg=C_WHITE, fg=C_GREEN_DARK,
                 font=("Segoe UI",15,"bold")).pack(anchor="w")
        tk.Label(tf, text="eCAC · Receita Federal  |  Local · Sem API · Gratuito",
                 bg=C_WHITE, fg=C_GRAY, font=("Segoe UI",8)).pack(anchor="w")

        bf = tk.Frame(hdr, bg=C_WHITE); bf.pack(side="right", padx=16, pady=14)
        self.btn_clear = tk.Button(bf, text="✕  Limpar",
            bg="#E8E8E8", fg=C_GRAY, relief="flat", cursor="hand2",
            font=("Segoe UI",9), padx=12, pady=6, command=self._clear, state="disabled",
            activebackground="#D0D0D0", activeforeground=C_GRAY_DARK)
        self.btn_clear.pack(side="right", padx=(6,0))
        self.btn_export = tk.Button(bf, text="⬇  Exportar Excel",
            bg=C_GREEN, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI",9,"bold"), padx=14, pady=6, command=self._export, state="disabled",
            activebackground=C_GREEN_DARK, activeforeground=C_WHITE)
        self.btn_export.pack(side="right")

        main = tk.Frame(self.root, bg=C_GRAY_LIGHT); main.pack(fill="both", expand=True)

        left = tk.Frame(main, bg=C_WHITE, width=260,
                        highlightthickness=1, highlightbackground=C_BORDER)
        left.pack(fill="y", side="left", padx=(10,0), pady=10); left.pack_propagate(False)

        sec1 = tk.Frame(left, bg=C_WHITE); sec1.pack(fill="x", padx=14, pady=(18,0))
        self.btn_add = tk.Button(sec1, text="📄  Adicionar PDFs",
            bg=C_GREEN, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI",10,"bold"), pady=9, command=self._add_files,
            activebackground=C_GREEN_DARK, activeforeground=C_WHITE)
        self.btn_add.pack(fill="x")
        self.btn_proc = tk.Button(sec1, text="▶  Processar",
            bg=C_YELLOW, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI",10,"bold"), pady=9, command=self._process, state="disabled",
            activebackground="#D4891A", activeforeground=C_WHITE)
        self.btn_proc.pack(fill="x", pady=(8,0))

        tk.Frame(left, bg=C_BORDER, height=1).pack(fill="x", padx=14, pady=10)

        # Importar planilha de status
        tk.Label(left, text="PLANILHA DE STATUS (eCAC)",
                 bg=C_WHITE, fg=C_GRAY, font=("Segoe UI",7,"bold")).pack(anchor="w", padx=14, pady=(0,4))
        self.btn_status = tk.Button(left, text="📊  Importar Planilha",
            bg="#1A365D", fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI",9,"bold"), pady=7, command=self._import_status,
            activebackground="#0F2040", activeforeground=C_WHITE)
        self.btn_status.pack(fill="x", padx=14)
        self.lbl_status_info = tk.Label(left, text="Nenhuma planilha carregada",
            bg=C_WHITE, fg=C_GRAY, font=("Segoe UI",7), wraplength=220, justify="left")
        self.lbl_status_info.pack(anchor="w", padx=14, pady=(3,0))

        tk.Frame(left, bg=C_BORDER, height=1).pack(fill="x", padx=14, pady=12)
        tk.Label(left, text="ARQUIVOS NA FILA", bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI",7,"bold")).pack(anchor="w", padx=14, pady=(0,4))

        lbf = tk.Frame(left, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        lbf.pack(fill="both", expand=True, padx=14)
        self.file_lb = tk.Listbox(lbf, bg=C_GRAY_LIGHT, fg=C_GRAY_DARK,
            selectbackground=C_GREEN_MID, selectforeground=C_GRAY_DARK,
            relief="flat", font=("Segoe UI",8), activestyle="none", borderwidth=0)
        self.file_lb.pack(fill="both", expand=True, padx=4, pady=4)

        tk.Frame(left, bg=C_BORDER, height=1).pack(fill="x", padx=14, pady=8)
        self.lbl_stats = tk.Label(left, text="", bg=C_WHITE, fg=C_GRAY,
                                   font=("Segoe UI",8), justify="left")
        self.lbl_stats.pack(anchor="w", padx=14, pady=(0,14))

        right = tk.Frame(main, bg=C_GRAY_LIGHT)
        right.pack(fill="both", expand=True, side="left", padx=10, pady=10)
        self.nb = ttk.Notebook(right); self.nb.pack(fill="both", expand=True)

        t1 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t1, text="  📋  Detalhamento  ")
        self.cf1 = self._make_col_filter_bar(t1, ABA1_COLS, self._rebuild_aba1)
        self.tree1 = self._make_tree(t1, ABA1_KEYS, ABA1_COLS,
            extra_tags={"credito":("#EAF4D3","#1A5276"),
                        "debito": ("#FDECEA","#7B241C"),
                        "source": ("#F2F4F0","#7AB82E")})

        t2 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t2, text="  📊  Controle de Créditos  ")
        self.cf2 = self._make_col_filter_bar(t2, CTRL_COLS, self._refresh_ctrl)
        self.tree2 = self._make_tree(t2, CTRL_KEYS, CTRL_COLS,
            extra_tags={"positivo":("#EAF4D3","#3D6B0A"),
                        "negativo":("#FDEDEC","#C0392B"),
                        "totals":  ("#C5E08A","#3D3D3D")})

        t3 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t3, text="  ⚠  Comp. Sem Vínculo  ")
        self.cf3 = self._make_col_filter_bar(t3, DETAIL_COLS, self._refresh_unlinked)
        self.tree3 = self._make_tree(t3, DETAIL_KEYS, DETAIL_COLS,
            extra_tags={"source":("#FEF9F0","#C87A00")})

        t4 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t4, text="  🔍  Sem PDF (na planilha)  ")
        self.cf4 = self._make_col_filter_bar(t4, ABA4_COLS, self._refresh_aba4)
        self.tree4 = self._make_tree(t4, ABA4_KEYS, ABA4_COLS,
            extra_tags={"cancelado":("#FFF3CD","#856404"),
                        "normal":   ("#EBF4FF","#1A365D")})

        t5 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t5, text="  📥  Sem Planilha (PDF importado)  ")
        self.cf5 = self._make_col_filter_bar(t5, ABA5_COLS, self._refresh_aba5)
        self.tree5 = self._make_tree(t5, ABA5_KEYS, ABA5_COLS,
            extra_tags={"item":("#F8F0FF","#4A235A")})

        t6 = tk.Frame(self.nb, bg=C_WHITE); self.nb.add(t6, text="  🌿  Ressarcimentos  ")
        self.cf6 = self._make_col_filter_bar(t6, ABA6_COLS, self._refresh_aba6)
        self.tree6 = self._make_tree(t6, ABA6_KEYS, ABA6_COLS,
            extra_tags={"positivo":("#D8F3DC","#1B4332"),
                        "negativo":("#FDEDEC","#C0392B"),
                        "totals":  ("#A8D5B5","#1B4332")})

        # Liga os filtros Excel a cada aba (adiado para após criação das árvores)
        self.root.after(100, self._wire_all_col_filters)

        sb = tk.Frame(right, bg=C_GREEN_LIGHT, highlightthickness=1,
                      highlightbackground=C_BORDER, height=26)
        sb.pack(fill="x", side="bottom", pady=(4,0)); sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Pronto. Adicione os PDFs para começar.")
        tk.Label(sb, textvariable=self.status_var, bg=C_GREEN_LIGHT, fg=C_GRAY,
                 font=("Segoe UI",8), anchor="w").pack(fill="x", padx=10, pady=4)
        self.progress = ttk.Progressbar(right, mode="indeterminate")
        self.progress.pack(fill="x", side="bottom", pady=(0,2))

    def _make_col_filter_bar(self, parent, cols, callback):
        """
        Cria barra compacta com indicadores de filtros ativos e botão 'Limpar Tudo'.
        O filtro real é acionado clicando no cabeçalho da coluna (▼).
        Retorna dict {col_key: StringVar} vazio — mantém compatibilidade de interface.
        O callback é ignorado pois os filtros são gerenciados pelo col_filt_N dicts.
        """
        bar = tk.Frame(parent, bg="#EEF6E6", height=22)
        bar.pack(fill="x", side="top"); bar.pack_propagate(False)
        tk.Label(bar, text="  ▼ Clique nos cabeçalhos das colunas para filtrar",
                 bg="#EEF6E6", fg="#5A7A2A",
                 font=("Segoe UI", 7)).pack(side="left", pady=2)
        # Botão limpar todos os filtros desta aba — wired after tab creation
        self._last_clear_btn_parent = bar
        return {}    # vazio — col_filt_N é o estado real

    def _add_clear_btn(self, bar, tab_id):
        """Adiciona botão 'Limpar filtros' à barra do tab."""
        def clear():
            getattr(self, f"col_filt_{tab_id}").clear()
            getattr(self, f"_refresh_{tab_id}_dispatch")()
            self._update_headings(getattr(self, f"tree{tab_id}"),
                                  getattr(self, f"col_filt_{tab_id}"),
                                  getattr(self, f"_cols_{tab_id}"))
        tk.Button(bar, text="✕ Limpar filtros", bg="#EEF6E6", fg="#888",
                  relief="flat", font=("Segoe UI", 7), padx=8, cursor="hand2",
                  command=clear,
                  activebackground="#D8ECC8").pack(side="right", padx=4, pady=2)

    def _make_tree(self, parent, keys, cols, extra_tags=None):
        frame = tk.Frame(parent, bg=C_WHITE)
        frame.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        col_ids = [c[0] for c in cols]
        tree = ttk.Treeview(frame, columns=col_ids, show="headings",
                            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for c in cols:
            key, label = c[0], c[1]
            w = c[2] if len(c) > 2 else 140
            tree.heading(key, text=f"{label}  ▼")
            tree.column(key, width=w, minwidth=60, stretch=False)
        tree.tag_configure("credito", background="#EAF4D3", foreground="#3D6B0A")
        tree.tag_configure("debito",  background="#FDEDEC", foreground="#C0392B")
        tree.tag_configure("source",  background="#F2F4F0", foreground="#7AB82E")
        if extra_tags:
            for tag, (bg, fg) in extra_tags.items():
                tree.tag_configure(tag, background=bg, foreground=fg)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    def _wire_col_filters(self, tree, tab_id, cols, get_data_fn, refresh_fn):
        """
        Liga os cabeçalhos de uma árvore ao popup de filtro Excel.
        get_data_fn: callable() -> list[dict]  — dados SEM filtros
        refresh_fn : callable()               — reconstrói a árvore
        """
        col_filt = getattr(self, f"col_filt_{tab_id}")
        # Guarda referência às colunas para atualizar headings
        setattr(self, f"_cols_{tab_id}", cols)
        setattr(self, f"_refresh_{tab_id}_dispatch", refresh_fn)

        for col in cols:
            key, label = col[0], col[1]

            def make_cmd(k=key, l=label):
                def cmd():
                    # Coleta valores únicos desta coluna
                    data = get_data_fn()
                    all_vals = sorted({str(r.get(k,"")) for r in data if str(r.get(k,"")).strip()})
                    if not all_vals:
                        return
                    active = col_filt.get(k)
                    px, py = self.root.winfo_pointerxy()
                    # Pequeno offset para parecer um dropdown
                    py += 4

                    def on_apply(sel_set):
                        if sel_set is None:
                            col_filt.pop(k, None)
                        else:
                            col_filt[k] = sel_set
                        refresh_fn()
                        self._update_headings(tree, col_filt, cols)

                    ColFilterPopup(self.root, l, all_vals, active, on_apply, px, py)
                return cmd

            tree.heading(key, command=make_cmd())

    @staticmethod
    def _update_headings(tree, col_filt, cols):
        """Muda texto do cabeçalho para indicar filtro ativo (🔽) ou inativo (▼)."""
        for col in cols:
            key, label = col[0], col[1]
            icon = " 🔽" if key in col_filt else "  ▼"
            tree.heading(key, text=f"{label}{icon}")

    def _wire_all_col_filters(self):
        """Liga os cabeçalhos das 6 abas aos popups de filtro Excel."""
        snap = list(self.rows)
        sm   = self.status_map
        self._wire_col_filters(self.tree1, 1, ABA1_COLS,
            lambda: combine_rows_for_aba1(list(self.rows), self.status_map),
            self._rebuild_aba1)
        self._wire_col_filters(self.tree2, 2, CTRL_COLS,
            lambda: build_credit_control(list(self.rows), self.status_map),
            self._refresh_ctrl)
        self._wire_col_filters(self.tree3, 3, DETAIL_COLS,
            lambda: build_unlinked_compensations(list(self.rows), self.status_map),
            self._refresh_unlinked)
        self._wire_col_filters(self.tree4, 4, ABA4_COLS,
            lambda: build_missing_from_excel(self.status_map, list(self.rows)),
            self._refresh_aba4)
        self._wire_col_filters(self.tree5, 5, ABA5_COLS,
            lambda: build_missing_from_pdfs(self.status_map, list(self.rows)),
            self._refresh_aba5)
        self._wire_col_filters(self.tree6, 6, ABA6_COLS,
            lambda: build_ressarcimento_aba6(list(self.rows), self.status_map),
            self._refresh_aba6)

    def _check_deps(self):
        missing = []
        if not PDFPLUMBER_OK: missing.append("pdfplumber")
        if not OPENPYXL_OK:   missing.append("openpyxl")
        if missing:
            messagebox.showerror("Dependências",
                f"Instale:\n\n  pip install {' '.join(missing)}\n\nReinicie.")

    def _add_files(self):
        paths = filedialog.askopenfilenames(title="Selecionar PDFs",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        for p in paths:
            if p not in self.files:
                self.files.append(p); self.file_lb.insert("end", Path(p).name)
        if self.files: self.btn_proc.config(state="normal")
        self._update_stats()

    def _process(self):
        if not self.files: return
        self.btn_proc.config(state="disabled"); self.btn_add.config(state="disabled")
        self.progress.start(10)
        def worker():
            errors = []
            duplicates = []
            for path in list(self.files):
                fname = Path(path).name
                self._set_status(f"Processando: {fname} …")
                try:
                    perdcomps = parse_pdf(path)
                    new_rows  = flatten_rows(perdcomps, fname)

                    # Deduplicação: ignora PERDCOMPs já carregados
                    existing = {r.get("numero_perdcomp","").strip()
                                for r in self.rows if r.get("numero_perdcomp")}
                    deduped  = [r for r in new_rows
                                if not r.get("numero_perdcomp","").strip()
                                or r.get("numero_perdcomp","").strip() not in existing]
                    skipped  = {r.get("numero_perdcomp","")
                                for r in new_rows
                                if r.get("numero_perdcomp","").strip() in existing}
                    if skipped:
                        duplicates.extend([f"{fname}: Nº {n} já carregado — ignorado"
                                           for n in sorted(skipped)])

                    if deduped:
                        self.rows.extend(deduped)
                        self.root.after(0, lambda r=deduped, f=fname: self._append(r, f))
                except Exception as e:
                    errors.append(f"{fname}: {e}")

            if duplicates:
                msg = "\n".join(duplicates)
                self.root.after(0, lambda m=msg: messagebox.showwarning(
                    "PDFs duplicados ignorados", m))
            self.root.after(0, lambda: self._done(errors))
        threading.Thread(target=worker, daemon=True).start()

    def _append(self, new_rows, source):
        """Chamado via root.after a cada PDF processado — atualiza todas as árvores."""
        self._refresh_all()

    def _refresh_all(self):
        """Reconstrói todas as abas a partir de self.rows (estado completo)."""
        try: self._rebuild_aba1()
        except Exception: pass
        try: self._refresh_ctrl()
        except Exception: pass
        try: self._refresh_unlinked()
        except Exception: pass
        try: self._refresh_aba4()
        except Exception: pass
        try: self._refresh_aba5()
        except Exception: pass
        try: self._refresh_aba6()
        except Exception: pass

    @staticmethod
    def _match(q: str, vals) -> bool:
        if not q: return True
        return any(q in str(v).lower() for v in vals)

    @staticmethod
    def _col_match(col_filt: dict, row: dict) -> bool:
        """
        Retorna True se a linha passa em TODOS os filtros ativos (AND).
        col_filt: {col_key: set_de_valores} — inclui linha apenas se valor está no set.
        """
        for col_key, val_set in col_filt.items():
            if not val_set:
                continue
            if str(row.get(col_key, "")) not in val_set:
                return False
        return True

    def _rebuild_aba1(self):
        for i in self.tree1.get_children(): self.tree1.delete(i)
        cf   = self.col_filt_1
        snap = list(self.rows)
        last = None
        for r in combine_rows_for_aba1(snap, self.status_map):
            src = r.get("_source","")
            if cf and not self._col_match(cf, r):
                continue
            if src != last:
                last = src
                self.tree1.insert("","end",
                    values=([f"   {src}"]+[""]*(len(ABA1_COLS)-1)), tags=("source",))
            tag = "debito" if r.get("tipo_debito") else "credito"
            self.tree1.insert("","end",
                values=tuple(r.get(k,"") for k in ABA1_KEYS), tags=(tag,))

    def _refresh_ctrl(self):
        for i in self.tree2.get_children(): self.tree2.delete(i)
        cf   = self.col_filt_2
        snap = list(self.rows)
        ctrl = build_credit_control(snap, self.status_map)
        tc = tcr = tco = tsd = 0.
        for cr in ctrl:
            sd  = cr.get("_raw_saldo", 0.)
            tag = "positivo" if sd > 0 else ("negativo" if sd < 0 else "")
            if not cf or self._col_match(cf, cr):
                self.tree2.insert("","end",
                    values=tuple(cr.get(k,"") for k in CTRL_KEYS), tags=(tag,))
            tc  += cr.get("_raw_credito",    0.)
            tcr += cr.get("_raw_correcao",   0.)
            tco += cr.get("_raw_compensado", 0.)
            tsd += sd
        if ctrl:
            n   = len(CTRL_KEYS)
            tot = [""] * n
            # 13 cols: cnpj(0) razao(1) num(2) situação(3) tipo_pedido(4) tipo_cred(5)
            #          comp(6) comp_teste(7) vl_cred(8) correção(9) compensado(10) saldo(11) ult_retif(12)
            tot[0]   = "TOTAIS"
            tot[n-5] = format_brl(tc)    # Vl. Total do Crédito
            tot[n-4] = format_brl(tcr)   # Vl. Total da Correção
            tot[n-3] = format_brl(tco)   # Total Compensado
            tot[n-2] = format_brl(tsd)   # Saldo Disponível
            # tot[n-1] = "" (Último Retificador - não soma)
            self.tree2.insert("","end", values=tot, tags=("totals",))

    def _refresh_unlinked(self):
        for i in self.tree3.get_children(): self.tree3.delete(i)
        cf   = self.col_filt_3
        snap = list(self.rows)
        last = None
        for r in build_unlinked_compensations(snap, self.status_map):
            src = r.get("_source","")
            if cf and not self._col_match(cf, r):
                continue
            if src != last:
                last = src
                self.tree3.insert("","end",
                    values=([f"   {src}"]+[""]*(len(DETAIL_COLS)-1)), tags=("source",))
            tag = ("credito" if r.get("tipo_registro")=="Crédito"
                   else "debito" if r.get("tipo_registro")=="Débito" else "")
            self.tree3.insert("","end",
                values=tuple(r.get(k,"") for k in DETAIL_KEYS), tags=(tag,))

    def _import_status(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha de status do eCAC",
            filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not path:
            return
        try:
            self.status_map = parse_status_excel(path)
            n = len(self.status_map)
            fname = Path(path).name
            self.lbl_status_info.config(
                text=f"✓ {fname}\n{n} PERDCOMP(s) carregado(s)",
                fg="#1A365D")
            self._set_status(f"Planilha carregada: {n} PERDCOMPs.")
            self._refresh_all()
            self._update_stats()
        except Exception as e:
            messagebox.showerror("Erro ao ler planilha", str(e))

    def _refresh_aba4(self):
        for i in self.tree4.get_children(): self.tree4.delete(i)
        cf   = self.col_filt_4
        snap = list(self.rows)
        for r in build_missing_from_excel(self.status_map, snap):
            if cf and not self._col_match(cf, r): continue
            sit = r.get("situacao","").lower()
            tag = "cancelado" if "cancelad" in sit else "normal"
            self.tree4.insert("","end",
                values=tuple(r.get(k,"") for k in ABA4_KEYS), tags=(tag,))

    def _refresh_aba5(self):
        for i in self.tree5.get_children(): self.tree5.delete(i)
        cf   = self.col_filt_5
        snap = list(self.rows)
        for r in build_missing_from_pdfs(self.status_map, snap):
            if cf and not self._col_match(cf, r): continue
            self.tree5.insert("","end",
                values=tuple(r.get(k,"") for k in ABA5_KEYS), tags=("item",))

    def _refresh_aba6(self):
        for i in self.tree6.get_children(): self.tree6.delete(i)
        cf   = self.col_filt_6
        snap = list(self.rows)
        aba6 = build_ressarcimento_aba6(snap, self.status_map)
        tc = tco = tsd = 0.
        for r in aba6:
            sd  = r.get("_raw_saldo", 0.)
            tag = "positivo" if sd > 0 else ("negativo" if sd < 0 else "")
            if not cf or self._col_match(cf, r):
                self.tree6.insert("","end",
                    values=tuple(r.get(k,"") for k in ABA6_KEYS), tags=(tag,))
            tc  += r.get("_raw_credito",    0.)
            tco += r.get("_raw_compensado", 0.)
            tsd += sd
        if aba6:
            n   = len(ABA6_KEYS)
            tot = [""] * n
            tot[0] = "TOTAIS"; tot[n-3] = format_brl(tc)
            tot[n-2] = format_brl(tco); tot[n-1] = format_brl(tsd)
            self.tree6.insert("","end", values=tot, tags=("totals",))

    def _done(self, errors):
        self.progress.stop()
        self.btn_add.config(state="normal")
        self.btn_proc.config(state="normal" if self.files else "disabled")
        self.btn_export.config(state="normal" if self.rows else "disabled")
        self.btn_clear.config(state="normal" if self.rows else "disabled")
        # Garante estado final correto independente de timing dos _append callbacks
        self._refresh_all()
        self._update_stats()
        if errors: messagebox.showerror("Erros", "\n".join(errors))
        self._set_status(f"Concluído! {len(self.rows)} linha(s) extraída(s)." if not errors
                         else f"Concluído com {len(errors)} erro(s).")

    def _export(self):
        if not self.rows: return
        path = filedialog.asksaveasfilename(title="Salvar Excel",
            defaultextension=".xlsx",
            initialfile=f"perdcomps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel","*.xlsx"),("Todos","*.*")])
        if not path: return
        try:
            sm = self.status_map
            export_excel(
                self.rows,
                build_credit_control(self.rows, sm),
                build_unlinked_compensations(self.rows, sm),
                build_missing_from_excel(sm, self.rows),
                build_missing_from_pdfs(sm, self.rows),
                build_ressarcimento_aba6(self.rows, sm),
                path,
                status_map=sm)
            messagebox.showinfo("Exportado!", f"Planilha salva em:\n{path}")
            if sys.platform == "win32": os.startfile(path)
        except Exception as e: messagebox.showerror("Erro", str(e))

    def _clear(self):
        if messagebox.askyesno("Limpar","Remover todos os dados?"):
            self.rows.clear(); self.files.clear(); self.file_lb.delete(0,"end")
            for t in [self.tree1,self.tree2,self.tree3,self.tree4,self.tree5,self.tree6]:
                for i in t.get_children(): t.delete(i)
            for cf in [self.col_filt_1,self.col_filt_2,self.col_filt_3,
                       self.col_filt_4,self.col_filt_5,self.col_filt_6]:
                cf.clear()
            for b in [self.btn_export,self.btn_clear,self.btn_proc]:
                b.config(state="disabled")
            self._update_stats(); self._set_status("Dados limpos.")

    def _set_status(self, msg):
        self.status_var.set(msg); self.root.update_idletasks()

    def _update_stats(self):
        snap  = list(self.rows)                      # snapshot thread-safe
        sm    = self.status_map
        total = len(snap)
        cred  = sum(1 for r in snap if r.get("tipo_registro")=="Crédito")
        deb   = sum(1 for r in snap if r.get("tipo_registro")=="Débito")
        src   = len({r.get("_source") for r in snap if r.get("_source")})
        ctrl  = build_credit_control(snap, sm)       # única chamada pesada
        unl   = len({r.get("numero_perdcomp","")
                     for r in build_unlinked_compensations(snap, sm)
                     if r.get("numero_perdcomp","")})
        n_st  = len(sm)
        n_a4  = len(build_missing_from_excel(sm, snap))
        n_a5  = len(build_missing_from_pdfs(sm, snap))
        # Aba 6: filtra ctrl já calculado, sem recalcular build_credit_control
        n_a6  = sum(1 for c in ctrl if "ressarc" in c.get("tipo_credito","").lower())
        self.lbl_stats.config(text=(
            f"Arquivos processados : {src}\n"
            f"Total de linhas      : {total}\n"
            f"  Créditos           : {cred}\n"
            f"  Débitos            : {deb}\n"
            f"Pedidos no controle  : {len(ctrl)}\n"
            f"  Ressarcimentos     : {n_a6}\n"
            f"Comp. sem vínculo    : {unl}\n"
            f"Na planilha          : {n_st}\n"
            f"  Sem PDF (Aba 4)    : {n_a4}\n"
            f"  Sem Planilha(Aba 5): {n_a5}\n"
            f"Arquivos na fila     : {len(self.files)}"))

    def _set_status(self, msg):
        self.status_var.set(msg); self.root.update_idletasks()

    def _update_stats(self):
        total = len(self.rows)
        cred  = sum(1 for r in self.rows if r.get("tipo_registro")=="Crédito")
        deb   = sum(1 for r in self.rows if r.get("tipo_registro")=="Débito")
        src   = len({r.get("_source") for r in self.rows if r.get("_source")})
        sm    = self.status_map
        ctrl  = build_credit_control(self.rows, sm)
        unl   = len({r.get("numero_perdcomp","")
                     for r in build_unlinked_compensations(self.rows, sm)
                     if r.get("numero_perdcomp","")})
        n_st  = len(sm)
        n_a4  = len(build_missing_from_excel(sm, self.rows))
        n_a5  = len(build_missing_from_pdfs(sm, self.rows))
        self.lbl_stats.config(text=(
            f"Arquivos processados : {src}\n"
            f"Total de linhas      : {total}\n"
            f"  Créditos           : {cred}\n"
            f"  Débitos            : {deb}\n"
            f"Pedidos no controle  : {len(ctrl)}\n"
            f"Comp. sem vínculo    : {unl}\n"
            f"Na planilha          : {n_st}\n"
            f"  Sem PDF (Aba 4)    : {n_a4}\n"
            f"  Sem Planilha(Aba 5): {n_a5}\n"
            f"Arquivos na fila     : {len(self.files)}"))



# =============================================================================
# DARF / DAS — Parser de comprovantes de arrecadação
# =============================================================================

DARF_COLS = [
    ("cnpj",           "CNPJ",               120),
    ("razao_social",   "Razão Social",        210),
    ("tipo_doc",       "Tipo",                 50),
    ("numero_doc",     "Nº Documento",        165),
    ("periodo",        "Período/Competência", 130),
    ("competencia_teste", "Competência Teste", 120),
    ("dt_vencimento",  "Dt. Vencimento",       98),
    ("dt_arrecadacao", "Dt. Arrecadação",      98),
    ("banco",          "Banco",               200),   # ← banco logo após datas
    ("agencia",        "Agência",              65),
    ("estabelecimento","Estabelecimento",       90),
    ("referencia",     "Referência",           100),
    ("codigo",         "Código",               55),
    ("descricao",      "Descrição",            250),
    ("principal",      "Principal",             88),
    ("multa",          "Multa",                 78),
    ("juros",          "Juros",                 78),
    ("total_item",     "Total Item",            88),
    ("total_doc",      "Total Documento",      108),
    ("_source",        "Arquivo PDF",          195),
]
DARF_KEYS = [c[0] for c in DARF_COLS]
DARF_MONEY_KEYS = {"principal", "multa", "juros", "total_item", "total_doc"}


def parse_darf_pdf(path: str) -> list:
    """
    Lê um PDF de comprovantes de arrecadação (DARF e/ou DAS).
    Retorna lista de dicts, um por linha de composição
    (headers repetidos para cada linha do documento).
    Documentos multi-página são consolidados pelo Número do Documento.
    """
    import pdfplumber, re

    def _first(pat, txt, default=""):
        m = re.search(pat, txt)
        return m.group(1).strip() if m else default

    def _brl(v):
        if not v or v.strip() in ("-", ""):
            return ""
        v = re.sub(r"[R$\s]", "", v.strip()).replace(".", "").replace(",", ".")
        try:
            return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return v

    # ── 1. Agrupa páginas por Número do Documento ───────────────────────────
    doc_map: dict = {}   # {numero_doc: {meta, items: []}}

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(x_tolerance=3, y_tolerance=3) or ""

            # Tipo (DARF ou DAS)
            if "arrecadação de DAS" in txt:
                tipo = "DAS"
            else:
                tipo = "DARF"

            # Número do documento (chave de agrupamento)
            num = _first(r"Número do Documento\s*\n?\s*(\d{17})", txt) or \
                  _first(r"(\d{17})", txt)
            if not num:
                continue

            if num not in doc_map:
                # Extrai cabeçalho
                cnpj_rs = re.search(
                    r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\s+(.+?)(?:\n|$)", txt)
                cnpj = cnpj_rs.group(1).strip() if cnpj_rs else ""
                razao = cnpj_rs.group(2).strip() if cnpj_rs else ""

                # Período/Competência e Data de Vencimento
                # PDF layout: "DD/MM/AAAA  DD/MM/AAAA  NNNNNNNNNNNNNNNNN"
                # DAS layout: "MM/AAAA  DD/MM/AAAA  NNNNNNNNNNNNNNNNN"
                datas_m = re.search(
                    r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d{17}", txt)
                comp_m  = re.search(
                    r"(\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d{17}", txt)
                if datas_m:
                    periodo = datas_m.group(1)
                    dt_venc = datas_m.group(2)
                elif comp_m:
                    periodo = comp_m.group(1)
                    dt_venc = comp_m.group(2)
                else:
                    all_dates = re.findall(r"\d{2}/\d{2}/\d{4}", txt)
                    periodo = all_dates[0] if len(all_dates) > 0 else ""
                    dt_venc = all_dates[1] if len(all_dates) > 1 else ""

                # Banco e data de arrecadação
                # Trata "341 - BANCO ITAU S A 30/12/2024" e
                # "748 - BANCO COOPERATIVO SICREDI S/A - 29/11/2024"
                banco_m = re.search(
                    r"(\d{3})\s*-\s*([A-ZÀ-Ú][^\n]+?)\s+(\d{2}/\d{2}/\d{4})", txt)
                if banco_m:
                    # Remove " -" ou " S/A -" solto no final do nome
                    nome = re.sub(r"\s+-\s*$", "", banco_m.group(2).strip()).strip()
                    banco    = f"{banco_m.group(1)} - {nome}"
                    dt_arrec = banco_m.group(3)
                else:
                    pix_m = re.search(
                        r"(Documento pago via PIX)\s+(\d{2}/\d{2}/\d{4})", txt)
                    if pix_m:
                        banco    = pix_m.group(1)
                        dt_arrec = pix_m.group(2)
                    else:
                        banco    = ""
                        dt_arrec = ""

                # Agência / Estabelecimento / Referência
                # Linha dos valores: "0322  0671  0,00  10180218"
                # tokens: [agencia, estab, valor_reservado, referencia?]
                agencia = estab = ref = ""
                agencia_m = re.search(
                    r"Agência\s*Estabelecimento[^\n]*\n([^\n]+)", txt)
                if agencia_m:
                    tokens = agencia_m.group(1).split()
                    agencia = tokens[0] if len(tokens) > 0 else ""
                    estab   = tokens[1] if len(tokens) > 1 else ""
                    ref     = tokens[3] if len(tokens) > 3 else ""  # token 3 = referência

                doc_map[num] = {
                    "cnpj": cnpj, "razao_social": razao,
                    "tipo_doc": tipo, "numero_doc": num,
                    "periodo": periodo, "dt_vencimento": dt_venc,
                    "dt_arrecadacao": dt_arrec, "banco": banco,
                    "agencia": agencia, "estabelecimento": estab,
                    "referencia": ref, "total_doc": "",
                    "items": [],
                }

            meta = doc_map[num]

            # Atualiza campos bancários se estavam vazios na 1ª página
            # (cobre documentos multi-página onde o banco só aparece em página posterior)
            if not meta["banco"] and banco:
                meta["banco"] = banco
                meta["dt_arrecadacao"] = dt_arrec
            if not meta["agencia"] and agencia:
                meta["agencia"] = agencia
            if not meta["estabelecimento"] and estab:
                meta["estabelecimento"] = estab
            if not meta["referencia"] and ref:
                meta["referencia"] = ref

            # ── 2. Extrai itens da composição ─────────────────────────────
            in_comp = False
            lines = txt.split("\n")
            for line in lines:
                line = line.strip()
                if "Composição do Documento" in line:
                    in_comp = True
                    continue
                if not in_comp:
                    continue
                if line.startswith("Totais"):
                    # Extrai total geral
                    vals = re.findall(r"[\d\.]+,\d{2}", line)
                    if vals:
                        meta["total_doc"] = _brl(vals[-1].replace(".", "").replace(",", ".").replace(".", "").replace(",", "."))
                        meta["total_doc"] = vals[-1]   # já formatado
                    in_comp = False
                    continue
                if re.match(r"^Comprovante emitido", line):
                    in_comp = False
                    continue
                # Bloco bancário no rodapé ("Banco Data de Arrecadação",
                # "Agência Estabelecimento..."): fecha a composição para evitar
                # que a linha "NNNN NNNN 0,00" (agência/estabelecimento)
                # seja confundida com um item de arrecadação.
                if re.match(r"^(Banco\b|Agência\b|Age[nñ]cia\b)", line, re.IGNORECASE):
                    in_comp = False
                    continue

                # Linha de item (começa com código 4 dígitos)
                item_m = re.match(
                    r"^(\d{4})\s+(.+?)\s+([\d\.]+,\d{2}|-)\s+([\d\.]+,\d{2}|-)\s+([\d\.]+,\d{2}|-)\s+([\d\.]+,\d{2})$",
                    line
                )
                if item_m:
                    desc = item_m.group(2).strip()
                    # Descrição deve ter ao menos uma letra — senão é linha de
                    # agência/estabelecimento e não item de DARF.
                    if re.search(r"[A-Za-zÀ-ÿ]", desc):
                        meta["items"].append({
                            "codigo":    item_m.group(1),
                            "descricao": desc,
                            "principal": item_m.group(3) if item_m.group(3) != "-" else "",
                            "multa":     item_m.group(4) if item_m.group(4) != "-" else "",
                            "juros":     item_m.group(5) if item_m.group(5) != "-" else "",
                            "total_item":item_m.group(6),
                        })
                    continue

                # Linha de item sem juros/multa (só total à direita)
                item_m2 = re.match(
                    r"^(\d{4})\s+(.+?)\s+([\d\.]+,\d{2})$", line)
                if item_m2:
                    desc = item_m2.group(2).strip()
                    # Mesma proteção — descrição precisa ter letra.
                    if re.search(r"[A-Za-zÀ-ÿ]", desc):
                        meta["items"].append({
                            "codigo":    item_m2.group(1),
                            "descricao": desc,
                            "principal": "", "multa": "", "juros": "",
                            "total_item": item_m2.group(3),
                        })

    # ── 3. Achata em linhas ──────────────────────────────────────────────────
    fname = Path(path).name
    rows = []
    for meta in doc_map.values():
        items = meta["items"]
        if not items:
            # Sem itens parseados: cria linha vazia para o documento aparecer
            items = [{"codigo": "", "descricao": "", "principal": "",
                      "multa": "", "juros": "", "total_item": ""}]
        for it in items:
            rows.append({
                "cnpj":           meta["cnpj"],
                "razao_social":   meta["razao_social"],
                "tipo_doc":       meta["tipo_doc"],
                "numero_doc":     meta["numero_doc"],
                "periodo":        meta["periodo"],
                "competencia_teste": format_competencia_teste(meta["periodo"]),
                "dt_vencimento":  meta["dt_vencimento"],
                "dt_arrecadacao": meta["dt_arrecadacao"],
                "banco":          meta["banco"],
                "agencia":        meta["agencia"],
                "estabelecimento":meta["estabelecimento"],
                "referencia":     meta["referencia"],
                "codigo":         it["codigo"],
                "descricao":      it["descricao"],
                "principal":      it["principal"],
                "multa":          it["multa"],
                "juros":          it["juros"],
                "total_item":     it["total_item"],
                "total_doc":      meta["total_doc"],
                "_source":        fname,
            })
    return rows


def export_darf_excel(all_rows: list, path: str):
    """Exporta lista de rows DARF/DAS para Excel formatado."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "DARF-DAS Arrecadação"

    # Cores
    H_BG  = "2C5F0A"; H_FG  = "FFFFFF"
    R1_BG = "EAF4D3"; R2_BG = "F7FCF0"
    TT_BG = "C5E08A"; TT_FG = "1B4332"

    bd = Border(
        left  =Side(style="thin", color="BBBBBB"),
        right =Side(style="thin", color="BBBBBB"),
        top   =Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"))
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=False)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    rgt = Alignment(horizontal="right",  vertical="center", wrap_text=False)
    h_font = Font(name="Calibri", bold=True, color=H_FG, size=10)
    d_font = Font(name="Calibri", size=9)
    t_font = Font(name="Calibri", bold=True, color=TT_FG, size=10)
    h_fill = PatternFill("solid", fgColor=H_BG)
    t_fill = PatternFill("solid", fgColor=TT_BG)

    labels = [c[1] for c in DARF_COLS]
    widths  = [c[2] for c in DARF_COLS]

    # Título
    n = len(DARF_COLS)
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    tc = ws["A1"]
    tc.value = "COMPROVANTES DE ARRECADAÇÃO — DARF / DAS"
    tc.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    tc.fill  = PatternFill("solid", fgColor=H_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabeçalho
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=2, column=ci, value=lbl)
        c.font = h_font; c.fill = h_fill
        c.alignment = ctr; c.border = bd
    ws.row_dimensions[2].height = 22

    # Dados
    valor_cols = {"principal", "multa", "juros", "total_item", "total_doc"}
    for ri, row in enumerate(all_rows, 3):
        fill = PatternFill("solid", fgColor=R1_BG if ri % 2 else R2_BG)
        for ci, col in enumerate(DARF_KEYS, 1):
            val = row.get(col, "")
            al  = rgt if col in valor_cols else lft
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.fill = fill
            c.alignment = al; c.border = bd
        ws.row_dimensions[ri].height = 16

    # Totais
    if all_rows:
        tr = 3 + len(all_rows)
        ws.merge_cells(f"A{tr}:{get_column_letter(n-5)}{tr}")
        ws.cell(row=tr, column=1, value="TOTAIS").font = t_font
        ws.cell(row=tr, column=1).fill = t_fill
        ws.cell(row=tr, column=1).alignment = ctr
        def sum_col(key):
            s = 0.0
            for r in all_rows:
                v = r.get(key,"").replace(".","").replace(",",".")
                try: s += float(v)
                except: pass
            return f"{s:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        for col_name, ci in [("principal", n-4), ("multa", n-3),
                               ("juros", n-2), ("total_item", n-1)]:
            c = ws.cell(row=tr, column=ci+1, value=sum_col(col_name))  # offset
            c.font = t_font; c.fill = t_fill
            c.alignment = rgt; c.border = bd
        ws.row_dimensions[tr].height = 20

    # Larguras das colunas
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w / 7.5

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"
    wb.save(path)


# =============================================================================
# DarfApp — Módulo DARF Extractor (janela independente)
# =============================================================================

class DarfApp(tk.Toplevel):

    def __init__(self, master):
        super().__init__(master)
        self.title("🧾  DARF Extractor — AgriTax Audit")
        self.geometry("1340x780")
        self.configure(bg=C_GRAY_LIGHT)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self._logo_img = None
        self.rows:  list = []
        self.files: list = []
        self.col_filt: dict = {}
        self._build_ui()
        self.after(100, self._wire_filters)

    def _build_ui(self):
        # ── Cabeçalho ──────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C_GREEN_DARK, height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        try:
            import base64, io
            from PIL import Image, ImageTk
            img_data = base64.b64decode(LOGO_B64)
            img = Image.open(io.BytesIO(img_data)).resize((44, 44), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(hdr, image=self._logo_img, bg=C_GREEN_DARK).pack(side="left", padx=12)
        except Exception:
            pass
        tk.Label(hdr, text="🧾  DARF Extractor",
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 16, "bold")).pack(side="left", pady=8)
        tk.Label(hdr, text="Comprovantes de Arrecadação (DARF / DAS)",
                 bg=C_GREEN_DARK, fg="#C5E08A",
                 font=("Segoe UI", 9)).pack(side="left", padx=16)

        # ── Corpo principal ────────────────────────────────────────────────
        main = tk.Frame(self, bg=C_GRAY_LIGHT)
        main.pack(fill="both", expand=True, padx=8, pady=8)

        # ── Painel esquerdo ────────────────────────────────────────────────
        left = tk.Frame(main, bg=C_WHITE, width=230,
                        highlightthickness=1, highlightbackground=C_BORDER)
        left.pack(side="left", fill="y", padx=(0,8))
        left.pack_propagate(False)

        tk.Label(left, text="📥  PDFs de DARF/DAS",
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 9, "bold")).pack(fill="x", ipady=6)

        bf = tk.Frame(left, bg=C_WHITE)
        bf.pack(fill="x", padx=8, pady=6)

        def btn(parent, text, cmd, color=C_GREEN):
            return tk.Button(parent, text=text, command=cmd,
                             bg=color, fg=C_WHITE, relief="flat",
                             font=("Segoe UI", 8, "bold"), cursor="hand2",
                             activebackground=C_GREEN_DARK,
                             pady=5).pack(fill="x", pady=2)

        btn(bf, "➕  Adicionar PDFs", self._add_files)
        self.btn_proc   = tk.Button(bf, text="⚙  Processar", state="disabled",
                                    command=self._process, bg=C_GREEN, fg=C_WHITE,
                                    relief="flat", font=("Segoe UI",8,"bold"),
                                    cursor="hand2", activebackground=C_GREEN_DARK, pady=5)
        self.btn_proc.pack(fill="x", pady=2)
        self.btn_export = tk.Button(bf, text="📊  Exportar Excel", state="disabled",
                                    command=self._export, bg="#1A6B6B", fg=C_WHITE,
                                    relief="flat", font=("Segoe UI",8,"bold"),
                                    cursor="hand2", activebackground="#134F4F", pady=5)
        self.btn_export.pack(fill="x", pady=2)
        tk.Button(bf, text="🗑  Limpar", command=self._clear,
                  bg=C_GRAY, fg=C_WHITE, relief="flat",
                  font=("Segoe UI",8,"bold"), cursor="hand2",
                  activebackground="#444", pady=5).pack(fill="x", pady=2)

        tk.Frame(left, bg=C_BORDER, height=1).pack(fill="x", pady=4)
        tk.Label(left, text="Arquivos adicionados:",
                 bg=C_WHITE, fg=C_GRAY, font=("Segoe UI", 8)).pack(anchor="w", padx=8)

        lb_frame = tk.Frame(left, bg=C_WHITE)
        lb_frame.pack(fill="both", expand=True, padx=4, pady=(0,4))
        vsb = ttk.Scrollbar(lb_frame, orient="vertical")
        self.file_lb = tk.Listbox(lb_frame, bg=C_WHITE, fg=C_GRAY_DARK,
                                   font=("Segoe UI", 8), relief="flat",
                                   yscrollcommand=vsb.set, selectmode="extended",
                                   activestyle="none")
        vsb.config(command=self.file_lb.yview)
        vsb.pack(side="right", fill="y")
        self.file_lb.pack(fill="both", expand=True)

        # Contador
        self.lbl_stats = tk.Label(left, text="0 documentos | 0 itens",
                                   bg=C_GREEN_LIGHT, fg=C_GRAY,
                                   font=("Segoe UI", 7), anchor="w")
        self.lbl_stats.pack(fill="x", padx=4, pady=(0,4))

        # ── Painel direito ──────────────────────────────────────────────────
        right = tk.Frame(main, bg=C_GRAY_LIGHT)
        right.pack(fill="both", expand=True)

        # Barra de filtros hint
        hint = tk.Frame(right, bg="#EEF6E6", height=22)
        hint.pack(fill="x"); hint.pack_propagate(False)
        tk.Label(hint, text="  ▼ Clique nos cabeçalhos das colunas para filtrar por valor",
                 bg="#EEF6E6", fg="#5A7A2A",
                 font=("Segoe UI", 7)).pack(side="left", pady=2)

        # Treeview
        t_frame = tk.Frame(right, bg=C_WHITE)
        t_frame.pack(fill="both", expand=True, pady=(0,0))
        vsb2 = ttk.Scrollbar(t_frame, orient="vertical")
        hsb2 = ttk.Scrollbar(t_frame, orient="horizontal")
        col_ids = [c[0] for c in DARF_COLS]
        self.tree = ttk.Treeview(t_frame, columns=col_ids, show="headings",
                                  yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        for c in DARF_COLS:
            self.tree.heading(c[0], text=f"{c[1]}  ▼")
            self.tree.column(c[0], width=c[2], minwidth=40, stretch=False)
        self.tree.tag_configure("darf", background="#EDF5FF", foreground="#1A3A6B")
        self.tree.tag_configure("das",  background="#FFF8E8", foreground="#7A4A00")
        self.tree.tag_configure("src",  background="#F2F4F0", foreground="#7AB82E")
        vsb2.config(command=self.tree.yview)
        hsb2.config(command=self.tree.xview)
        vsb2.pack(side="right", fill="y")
        hsb2.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # Barra de status + progress
        sb = tk.Frame(right, bg=C_GREEN_LIGHT, height=24)
        sb.pack(fill="x", pady=(4,0)); sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Adicione os PDFs de DARF/DAS para começar.")
        tk.Label(sb, textvariable=self.status_var, bg=C_GREEN_LIGHT,
                 fg=C_GRAY, font=("Segoe UI", 7), anchor="w").pack(fill="x", padx=8)
        self.progress = ttk.Progressbar(right, mode="indeterminate")
        self.progress.pack(fill="x", pady=(0,2))

    def _wire_filters(self):
        for col in DARF_COLS:
            key, label = col[0], col[1]
            def make_cmd(k=key, l=label):
                def cmd():
                    all_vals = sorted({str(r.get(k,"")) for r in self.rows if str(r.get(k,"")).strip()})
                    if not all_vals: return
                    active = self.col_filt.get(k)
                    px, py = self.winfo_pointerxy()
                    def on_apply(sel):
                        if sel is None: self.col_filt.pop(k, None)
                        else:           self.col_filt[k] = sel
                        self._refresh()
                        self._update_headings()
                    ColFilterPopup(self, l, all_vals, active, on_apply, px, py+4)
                return cmd
            self.tree.heading(key, command=make_cmd())

    def _update_headings(self):
        for col in DARF_COLS:
            k, lbl = col[0], col[1]
            icon = " 🔽" if k in self.col_filt else "  ▼"
            self.tree.heading(k, text=f"{lbl}{icon}")

    @staticmethod
    def _col_match(col_filt, row):
        for k, vs in col_filt.items():
            if vs and str(row.get(k,"")) not in vs:
                return False
        return True

    def _refresh(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        cf = self.col_filt
        for r in self.rows:
            if cf and not self._col_match(cf, r):
                continue
            tag = "das" if r.get("tipo_doc") == "DAS" else "darf"
            self.tree.insert("", "end",
                values=tuple(r.get(k, "") for k in DARF_KEYS),
                tags=(tag,))

    def _update_stats(self):
        docs = len({r["numero_doc"] for r in self.rows if r.get("numero_doc")})
        self.lbl_stats.config(text=f"{docs} documento(s)  |  {len(self.rows)} item(s)")

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar PDFs de DARF/DAS",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.file_lb.insert("end", Path(p).name)
        if self.files:
            self.btn_proc.config(state="normal")

    def _process(self):
        if not self.files: return
        self.btn_proc.config(state="disabled")
        self.progress.start(10)
        self.status_var.set("Processando PDFs…")

        import threading
        def worker():
            errors = []
            new_rows = []
            seen = {r["_source"] for r in self.rows}
            for path in self.files:
                fname = Path(path).name
                if fname in seen:
                    continue
                try:
                    new_rows.extend(parse_darf_pdf(path))
                except Exception as e:
                    errors.append(f"{fname}: {e}")
            self.after(0, lambda: self._done(new_rows, errors))
        threading.Thread(target=worker, daemon=True).start()

    def _done(self, new_rows, errors):
        self.progress.stop()
        self.rows.extend(new_rows)
        self._refresh()
        self._update_stats()
        self.btn_proc.config(state="normal" if self.files else "disabled")
        if self.rows:
            self.btn_export.config(state="normal")
        n = len({r["numero_doc"] for r in new_rows if r.get("numero_doc")})
        self.status_var.set(f"Processados: {n} documento(s), {len(new_rows)} item(s) novos.")
        if errors:
            messagebox.showerror("Erros", "\n".join(errors[:10]), parent=self)

    def _export(self):
        if not self.rows: return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile="darf_das_arrecadacao.xlsx")
        if not path: return
        try:
            export_darf_excel(self.rows, path)
            messagebox.showinfo("Exportado",
                f"Arquivo salvo em:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e), parent=self)

    def _clear(self):
        if not self.rows and not self.files: return
        if messagebox.askyesno("Limpar","Remover todos os dados?", parent=self):
            self.rows.clear(); self.files.clear()
            self.file_lb.delete(0,"end")
            self.col_filt.clear()
            for i in self.tree.get_children(): self.tree.delete(i)
            self.btn_proc.config(state="disabled")
            self.btn_export.config(state="disabled")
            self._update_stats()
            self.status_var.set("Dados limpos.")



# =============================================================================
# DCTF Extractor  v3.1  —  100% Local · Sem API · Sem Custo
# =============================================================================
# PDFs de DCTF do eCAC impressos pelo browser (Microsoft Print to PDF) armazenam
# o texto como VETORES — pdfplumber retorna 0 chars. Solucao: renderizar cada
# pagina como imagem PNG (resolucao 300 DPI) e aplicar OCR via pytesseract.
#
# Dependencias: pip install pdfplumber pytesseract openpyxl
# + Tesseract instalado:
#   Windows : https://github.com/UB-Mannheim/tesseract/wiki
#   Linux   : sudo apt install tesseract-ocr
#   Mac     : brew install tesseract
# =============================================================================

try:
    import pytesseract as _pytesseract
    PYTESSERACT_OK = True
except ImportError:
    PYTESSERACT_OK = False

try:
    import pdf2image as _pdf2image
    PDF2IMAGE_OK = True
except ImportError:
    PDF2IMAGE_OK = False


# ── Diagnóstico do ambiente OCR ───────────────────────────────────────────────
def _probe_ocr_environment() -> dict:
    """
    Verifica o estado do ambiente OCR (Tesseract + idiomas + Poppler).
    Tenta autodetectar o Tesseract no Windows mesmo quando não está no PATH,
    buscando nos caminhos de instalação padrão.

    Retorna:
        {
          "pytesseract_ok": bool,
          "tesseract_exe":  str | None,   # caminho detectado
          "tesseract_ver":  str | None,   # ex: "5.3.1"
          "langs":          list[str],    # ex: ["eng", "por"]
          "has_por":        bool,         # True se português disponível
          "pdf2image_ok":   bool,
          "poppler_ok":     bool,         # True se Poppler responde
          "errors":         list[str],
          "recommendations":list[str],
        }
    """
    import os as _os, shutil as _shutil, subprocess as _subprocess
    info = {
        "pytesseract_ok": PYTESSERACT_OK,
        "tesseract_exe":  None,
        "tesseract_ver":  None,
        "langs":          [],
        "has_por":        False,
        "pdf2image_ok":   PDF2IMAGE_OK,
        "poppler_ok":     False,
        "errors":         [],
        "recommendations":[],
    }

    if not PYTESSERACT_OK:
        info["errors"].append("Biblioteca Python 'pytesseract' não instalada.")
        info["recommendations"].append(
            "Instale com: pip install pytesseract pdf2image pdfplumber Pillow")
        return info

    # ── Tenta localizar o binário do Tesseract ───────────────────────────────
    candidates = []
    # 1) PATH
    in_path = _shutil.which("tesseract") or _shutil.which("tesseract.exe")
    if in_path:
        candidates.append(in_path)
    # 2) Caminhos padrão de instalação no Windows
    for pth in [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        _os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
        _os.path.expandvars(r"%USERPROFILE%\AppData\Local\Tesseract-OCR\tesseract.exe"),
        # 3) Caminhos comuns no Linux/Mac (em geral já estão no PATH, mas por garantia)
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/opt/homebrew/bin/tesseract",
    ]:
        if pth and _os.path.isfile(pth) and pth not in candidates:
            candidates.append(pth)

    if not candidates:
        info["errors"].append("Tesseract OCR não encontrado no sistema.")
        info["recommendations"].append(
            "Baixe e instale: https://github.com/UB-Mannheim/tesseract/wiki\n"
            "Durante a instalação, marque 'Add to PATH' e inclua o idioma Português.")
        return info

    tesseract_exe = candidates[0]
    info["tesseract_exe"] = tesseract_exe
    # Configura pytesseract para usar o binário detectado (cobre o caso em que
    # o Tesseract está instalado mas não foi adicionado ao PATH do Windows).
    try:
        _pytesseract.pytesseract.tesseract_cmd = tesseract_exe
    except Exception:
        pass

    # ── Obtém versão ─────────────────────────────────────────────────────────
    try:
        out = _subprocess.run([tesseract_exe, "--version"],
                              capture_output=True, text=True, timeout=5)
        first_line = (out.stdout or out.stderr or "").splitlines()[0] if (out.stdout or out.stderr) else ""
        # formato "tesseract 5.3.1" ou "tesseract v5.3.1.20230401"
        import re as _re
        m = _re.search(r"tesseract\s+v?(\d+\.\d+\.\d+)", first_line, _re.IGNORECASE)
        if m:
            info["tesseract_ver"] = m.group(1)
    except Exception as e:
        info["errors"].append(f"Tesseract encontrado em {tesseract_exe} mas não executa: {e}")

    # ── Lista idiomas disponíveis ────────────────────────────────────────────
    try:
        langs = _pytesseract.get_languages(config="")
        info["langs"] = sorted(langs) if langs else []
        info["has_por"] = "por" in info["langs"]
        if not info["has_por"]:
            info["recommendations"].append(
                "Instale o pacote de idioma Português do Tesseract para melhor acurácia "
                "com acentos e caracteres especiais. No Windows, reinstale o Tesseract e "
                "marque 'Portuguese' em 'Additional language data'.")
    except Exception as e:
        info["errors"].append(f"Não foi possível listar idiomas do Tesseract: {e}")

    # ── Verifica Poppler (necessário pelo pdf2image) ─────────────────────────
    if PDF2IMAGE_OK:
        poppler_bin = _shutil.which("pdftoppm") or _shutil.which("pdftoppm.exe")
        # Tenta caminhos padrão no Windows
        if not poppler_bin:
            for pth in [
                r"C:\poppler\Library\bin\pdftoppm.exe",
                r"C:\Program Files\poppler\Library\bin\pdftoppm.exe",
                r"C:\Program Files\poppler\bin\pdftoppm.exe",
            ]:
                if _os.path.isfile(pth):
                    poppler_bin = pth
                    break
        info["poppler_ok"] = bool(poppler_bin)
        if not info["poppler_ok"]:
            info["recommendations"].append(
                "Poppler (requerido pelo pdf2image) não detectado. No Windows baixe em "
                "https://github.com/oschwartz10612/poppler-windows/releases, extraia em "
                "C:\\poppler\\ e adicione C:\\poppler\\Library\\bin ao PATH.")

    return info


# Cache do diagnóstico — probe é executado uma vez por sessão
_OCR_ENV_CACHE = None

def _get_ocr_env():
    global _OCR_ENV_CACHE
    if _OCR_ENV_CACHE is None:
        _OCR_ENV_CACHE = _probe_ocr_environment()
    return _OCR_ENV_CACHE


def _format_ocr_diagnostic(env: dict) -> str:
    """Monta mensagem amigável a partir do diagnóstico, sem repetir o óbvio."""
    lines = ["Nao foi possivel extrair texto deste PDF.\n"]
    lines.append("── Diagnóstico do ambiente OCR ──")
    if env["pytesseract_ok"]:
        lines.append("  ✓ pytesseract (Python) OK")
    else:
        lines.append("  ✗ pytesseract (Python) NÃO instalado")
    if env["tesseract_exe"]:
        ver = env["tesseract_ver"] or "versão desconhecida"
        lines.append(f"  ✓ Tesseract OCR  {ver}  em: {env['tesseract_exe']}")
        if env["has_por"]:
            lines.append(f"  ✓ Idioma Português instalado")
        else:
            langs = ", ".join(env["langs"][:6]) if env["langs"] else "(nenhum)"
            lines.append(f"  ⚠ Idioma Português AUSENTE  (idiomas atuais: {langs})")
    else:
        lines.append("  ✗ Tesseract OCR NÃO encontrado no sistema")
    if env["pdf2image_ok"]:
        lines.append(f"  {'✓' if env['poppler_ok'] else '⚠'} pdf2image OK  "
                     f"{'(Poppler detectado)' if env['poppler_ok'] else '(Poppler NÃO detectado)'}")
    else:
        lines.append("  ⚠ pdf2image (Python) não instalado — é um fallback opcional")

    if env["recommendations"]:
        lines.append("\n── Como resolver ──")
        for i, rec in enumerate(env["recommendations"], 1):
            lines.append(f"  {i}. {rec}")
    return "\n".join(lines)


# ── Colunas ───────────────────────────────────────────────────────────────────
DCTF_DETAIL_COLS = [
    ("cnpj",                 "CNPJ",                      150),
    ("nome_empresarial",     "Nome Empresarial",          230),
    ("periodo_competencia",  "Periodo de Competencia",    145),
    ("numero_declaracao",    "Numero da Declaracao",      210),
    ("numero_recibo",        "Numero do Recibo",          175),
    ("data_recepcao",        "Data de Recepcao",          110),
    ("data_processamento",   "Data de Processamento",     130),
    ("situacao_declaracao",  "Situacao da Declaracao",    120),
    ("retificadora",         "Decl. Retificadora",        110),
    ("grupo_tributo",        "Grupo do Tributo",          290),
    ("codigo_receita",       "Codigo da Receita",          95),
    ("periodicidade",        "Periodicidade",              90),
    ("periodo_apuracao",     "Periodo de Apuracao",       130),
    ("competencia_teste",    "Competencia Teste",         130),
    ("debito_apurado",       "Debito Apurado",            120),
    ("credito_pagamento",    "Credito - Pagamento",       130),
    ("credito_compensacoes", "Credito - Compensacoes",    145),
    ("credito_parcelamento", "Credito - Parcelamento",    145),
    ("credito_suspensao",    "Credito - Suspensao",       130),
    ("soma_creditos",        "Soma dos Creditos Vinc.",   155),
    ("saldo_pagar",          "Saldo a Pagar",             110),
    ("valor_total_debito",   "Valor Total do Debito",     140),
    ("total_contribuicao",   "Total da Contribuicao",     140),
    ("pagamento_darf",       "Pagamento com DARF",        140),
    ("darf_pa",              "DARF - PA",                  95),
    ("darf_codigo_receita",  "DARF - Cod. Receita",        95),
    ("darf_vencimento",      "DARF - Vencimento",         110),
    ("darf_principal",       "DARF - Principal",          120),
    ("darf_multa",           "DARF - Multa",              100),
    ("darf_juros",           "DARF - Juros",              100),
    ("darf_total",           "DARF - Total",              110),
    ("darf_pago",            "DARF - Valor Pago",         120),
]

DCTF_RESUMO_COLS = [
    ("codigo_receita",     "Codigo da Receita",    120),
    ("grupo_tributo",      "Grupo do Tributo",      290),
    ("qtd_declaracoes",    "Qtd. Declaracoes",      100),
    ("total_debito",       "Total Debito Apurado",  150),
    ("total_pagamento",    "Total Pagamento",       140),
    ("total_compensacao",  "Total Compensacoes",    145),
    ("total_parcelamento", "Total Parcelamento",    140),
    ("total_suspensao",    "Total Suspensao",       130),
    ("total_saldo",        "Total Saldo a Pagar",   140),
    ("total_darf_pago",    "Total DARF Pago",       130),
]

DCTF_MONEY_KEYS = {
    "debito_apurado", "credito_pagamento", "credito_compensacoes",
    "credito_parcelamento", "credito_suspensao", "soma_creditos",
    "saldo_pagar", "valor_total_debito", "total_contribuicao",
    "pagamento_darf", "darf_principal", "darf_multa", "darf_juros",
    "darf_total", "darf_pago",
    "total_debito", "total_pagamento", "total_compensacao",
    "total_parcelamento", "total_suspensao", "total_saldo", "total_darf_pago",
}

# Palavras-chave que provam que o OCR produziu texto util de DCTF.
# Incluimos variantes com e sem acento porque Tesseract com lang='por'
# preserva acentos (DEBITO / DÉBITO, DECLARACAO / DECLARAÇÃO).
_DCTF_KEYWORDS = ["CNPJ", "DCTF", "DEBITO", "DÉBITO", "TRIBUTO",
                  "MINISTERIO", "MINISTÉRIO", "DECLARACAO", "DECLARAÇÃO"]


def _dctf_strip_accents(s: str) -> str:
    """Remove acentos para comparacao tolerante. 'DÉBITO' -> 'DEBITO'."""
    if not s:
        return s
    import unicodedata
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


# ── Helpers ───────────────────────────────────────────────────────────────────
def _dctf_brl(s) -> float:
    """Converte valor BRL com artefatos OCR para float. Ex: '1.118,21' -> 1118.21"""
    if isinstance(s, (int, float)):
        return float(s)
    s = re.sub(r"\s+", "", str(s).strip())
    if re.match(r"^[\d.]+,\d{1,2}$", s):        # formato BR
        return float(s.replace(".", "").replace(",", "."))
    if re.match(r"^[\d,]+\.\d{1,2}$", s):        # formato EN
        return float(s.replace(",", ""))
    s = re.sub(r"[^\d.]", "", s.replace(",", "."))
    try:
        return float(s)
    except ValueError:
        return 0.0


def _dctf_re(pattern, text, default=""):
    """Busca regex case-insensitive; retorna grupo 1 ou default."""
    m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    return m.group(1).strip() if m else default


def _dctf_val(pattern, text) -> float:
    """Busca regex e converte resultado para float BRL."""
    m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    return _dctf_brl(m.group(1)) if m else 0.0


def _dctf_last_brl(line: str) -> float:
    """Pega o ultimo valor monetario BR da linha. Ex: 'Total ...: 242,77'"""
    nums = re.findall(r"\d[\d.]*,\d{2}", line)
    return _dctf_brl(nums[-1]) if nums else 0.0


def _dctf_text_valido(pages: list) -> bool:
    """
    Valida se a lista de textos contem conteudo util de DCTF.
    Evita aceitar texto de baixa qualidade OCR como se fosse valido.
    Requer: pelo menos 200 caracteres totais E pelo menos 1 palavra-chave DCTF.
    """
    full = " ".join(pages)
    if len(full.strip()) < 200:
        return False
    return any(kw in full.upper() for kw in _DCTF_KEYWORDS)


# ── Extracao de texto das paginas ─────────────────────────────────────────────
def _is_cid_text(text: str) -> bool:
    """
    Detecta se o texto extraido eh majoritariamente glifos CID (Character ID)
    sem mapeamento Unicode — situacao tipica dos PDFs da DCTF gerados via
    "Microsoft Print to PDF" a partir do eCAC. Esses PDFs contem o texto
    visualmente (glifos posicionados) mas sem a tabela ToUnicode que traduz
    os IDs dos glifos para caracteres Unicode reais, entao o texto sai como
    '(cid:131)(cid:132)(cid:133)' em vez de palavras legiveis.

    Heuristica: texto eh considerado CID se a contagem de '(cid:' representar
    mais de 10% do conteudo.
    """
    if not text:
        return False
    cid_count = text.count("(cid:")
    # Proporcao alta de CIDs em relacao ao tamanho do texto
    return cid_count > 20 and (cid_count * 8) > len(text) * 0.1


def _dctf_get_pages_text(pdf_path: str) -> list:
    """
    Extrai texto de cada pagina do PDF. Tenta tres estrategias em ordem:
      1. pdfplumber extrai texto diretamente (PDFs com texto selecionavel)
      2. pdfplumber.to_image(300 DPI) + pytesseract (PDFs vetoriais — caso DCTF eCAC)
      3. pdf2image(300 DPI) + pytesseract (fallback quando estrategia 2 nao disponivel)

    Caso especial: os PDFs da DCTF gerados no eCAC via "Microsoft Print to PDF"
    contem texto embutido mas usam fontes sem mapeamento Unicode (ToUnicode CMap
    ausente), o que faz o texto sair como '(cid:XXX)' em vez de letras. Nesses
    casos a estrategia 1 produz texto tecnicamente presente mas ilegivel, e a
    unica saida eh rasterizar + OCR. A mensagem de erro explica isso ao usuario.

    Lanca RuntimeError com diagnóstico detalhado de cada estratégia se todas
    falharem.
    """
    tentativas = []   # ("nome", "resultado descritivo")

    # Estrategia 1 — texto direto
    cid_detectado = False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [page.extract_text(x_tolerance=3, y_tolerance=3) or ""
                     for page in pdf.pages]
        full_len = sum(len(p) for p in pages)
        if _dctf_text_valido(pages):
            return pages
        cid_detectado = any(_is_cid_text(p) for p in pages)
        tentativas.append(("pdfplumber direto",
            f"{len(pages)} pág., {full_len} chars"
            + (" (texto CID — PDF vetorial sem ToUnicode)" if cid_detectado else " (texto insuficiente ou sem keywords DCTF)")))
    except Exception as e:
        tentativas.append(("pdfplumber direto", f"ERRO: {type(e).__name__}: {e}"))

    # PDFs de DCTF do eCAC precisam de OCR
    env = _get_ocr_env()
    if not env["pytesseract_ok"] or not env["tesseract_exe"]:
        diag = _format_ocr_diagnostic(env)
        if cid_detectado:
            diag = (
                "Este PDF foi gerado pelo eCAC via 'Imprimir como PDF'.\n"
                "O texto existe no arquivo MAS esta embutido como glifos sem\n"
                "mapeamento Unicode (ToUnicode CMap ausente).\n\n"
            ) + diag
        raise RuntimeError(diag)

    ocr_lang = "por" if env["has_por"] else "eng"
    ocr_cfg  = "--psm 6 --oem 3"

    # Estrategia 2 — pdfplumber.to_image (200 DPI) + OCR
    # 200 DPI é suficiente para texto gerado por computador e ~2.25x mais rápido que 300 DPI
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = []
            for page in pdf.pages:
                pil_img = page.to_image(resolution=200).original
                pages.append(_pytesseract.image_to_string(pil_img, lang=ocr_lang, config=ocr_cfg))
        full_len = sum(len(p) for p in pages)
        if _dctf_text_valido(pages):
            return pages
        tentativas.append(("pdfplumber.to_image(200) + OCR",
            f"{len(pages)} pág., {full_len} chars (texto insuficiente ou sem keywords DCTF)"))
    except Exception as e:
        tentativas.append(("pdfplumber.to_image(200) + OCR", f"ERRO: {type(e).__name__}: {e}"))

    # Estrategia 3 — pdf2image (200 DPI) + OCR
    if PDF2IMAGE_OK:
        try:
            images = _pdf2image.convert_from_path(pdf_path, dpi=200)
            pages = [_pytesseract.image_to_string(img, lang=ocr_lang, config=ocr_cfg)
                     for img in images]
            full_len = sum(len(p) for p in pages)
            if _dctf_text_valido(pages):
                return pages
            tentativas.append(("pdf2image(200) + OCR",
                f"{len(pages)} pág., {full_len} chars (texto insuficiente ou sem keywords DCTF)"))
            # Fallback extremo: se o texto tem algum tamanho mas não tem as keywords,
            # retorna mesmo assim — pode ser DCTF com layout diferente.
            # Melhor tentar parser do que falhar silenciosamente.
            if full_len > 100:
                return pages
        except Exception as e:
            tentativas.append(("pdf2image(300) + OCR", f"ERRO: {type(e).__name__}: {e}"))
    else:
        tentativas.append(("pdf2image(300) + OCR", "pdf2image não disponível"))

    # Monta diagnóstico detalhado — mostra o que cada estratégia retornou
    det = "\n".join(f"  [{i}] {nome}: {res}" for i, (nome, res) in enumerate(tentativas, 1))
    diag = _format_ocr_diagnostic(env)
    raise RuntimeError(
        f"Nenhuma estratégia de extração funcionou para este PDF.\n\n"
        f"--- Tentativas ---\n{det}\n\n"
        f"--- Ambiente OCR ---\n{diag}"
    )


# ── Parser do cabecalho ───────────────────────────────────────────────────────
def _dctf_parse_cabecalho(full_text: str) -> dict:
    """Extrai campos do cabecalho a partir do texto OCR completo do PDF.

    Normaliza acentos e dashes antes de aplicar regex — cobre tanto OCR
    em Portugues (preserva acentos) quanto em Ingles (perde acentos).
    """
    # Normaliza acentos e dashes — tolerante a ambas variantes do OCR
    full_text = _dctf_strip_accents(full_text)
    full_text = full_text.replace("—", "-").replace("–", "-")

    r = _dctf_re

    # CNPJ — formato XX.XXX.XXX/XXXX-XX
    cnpj = r(r"CNPJ[:\s]+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", full_text)

    # Periodo de competencia — "Dezembro/2024" logo apos o CNPJ
    comp = r(r"CNPJ[:\s]+[\d./-]+\s+((?:Jan|Fev|Mar|Abr|Mai|Jun|Jul|Ago|Set|Out|Nov|Dez)\w*/\d{4})",
             full_text)
    if not comp:
        comp = r(r"\b((?:Janeiro|Fevereiro|Mar[cç]o|Abril|Maio|Junho|Julho|Agosto|"
                  r"Setembro|Outubro|Novembro|Dezembro)/\d{4})\b", full_text)

    # Numero da declaracao — "100.2024.2025.1861637518"
    # OCR varia: "Declaragao", "Declaracgeo", "Declaragdo", etc.
    num_decl = r(r"N[^\s]{0,12}mero\s+da\s+Declara[^\s:]{0,15}[:\s]+([\d.]{8,})", full_text)

    # Numero do recibo
    num_recibo = r(r"N[^\s]{0,12}mero\s+do\s+Recibo[:\s]+([\d./-]+)", full_text)

    # Datas — OCR: "Recepcdo", "Recepgdo", "2zocessamento", etc.
    data_rec  = r(r"Data\s+de\s+Recep[^\s:]{0,12}[:\s]+(\d{2}/\d{2}/\d{4})", full_text)
    data_proc = r(r"Data\s+de\s+[Pp2][^\s:]{0,14}[:\s]+(\d{2}/\d{2}/\d{4})", full_text)

    # Situacao da declaracao
    situacao = r(r"Situa[^\s:]{0,12}[:\s]+(Normal|Retificad[ao]|Ativa|Em\s+An[aai]lise)",
                  full_text)
    if not situacao:
        situacao = r(r"Situa[^\s:]{0,12}[:\s]+([^\n]{1,25})", full_text)
    situacao = situacao.strip()[:30]

    # Declaracao retificadora — OCR: "N&o", "Nao", "Sim", etc.
    retif_raw = r(r"Declara[^\s]{0,12}\s+Retificadora[:\s]+(\S+)", full_text)
    retif = "Sim" if re.match(r"[Ss]", retif_raw.strip()) else "Nao"

    # Nome empresarial
    nome = r(r"Nome\s+Empresarial[:\s]+([^\n]+)", full_text).strip()

    return {
        "cnpj":                cnpj,
        "nome_empresarial":    nome,
        "periodo_competencia": comp,
        "numero_declaracao":   num_decl,
        "numero_recibo":       num_recibo,
        "data_recepcao":       data_rec,
        "data_processamento":  data_proc,
        "situacao_declaracao": situacao,
        "retificadora":        retif,
    }


# ── Parser de uma pagina de tributo ──────────────────────────────────────────
def _dctf_parse_tributo(page_text: str) -> dict:
    """
    Extrai campos de uma pagina que contem um tributo declarado.
    Layout OCR tipico:
      GRUPO DO TRIBUTO . PIS/PASEP - CONTRIB. P/PROGRAMA...
      CODIGO RECEITA : 6912-01
      PERIODICIDADE: Mensal  PERIODO DE APURACAO: Dezembro/2024
      DEBITO APURADO  242,77
      - PAGAMENTO  242,77
      ...
      PA: 31/12/2024  ...  Codigo da Receita: 6912
      Data do Vencimento  24/01/2025
      Valor do Principal: 242,77 / Valor da Multa: 0,80 / ...

    IMPORTANTE: normaliza o texto removendo acentos no inicio, porque
    Tesseract com lang='por' produz 'DÉBITO APURADO' e sem 'por' produz
    'DEBITO APURADO'. Os regex abaixo assumem texto SEM acento.
    Tambem normaliza em-dash/en-dash para hyphen simples.
    """
    # Normaliza acentos e dashes antes de aplicar qualquer regex
    page_text = _dctf_strip_accents(page_text)
    page_text = page_text.replace("—", "-").replace("–", "-")

    r = _dctf_re
    v = lambda pat: _dctf_val(pat, page_text)
    N = r"([\d.,]+)"

    # --- Grupo do tributo ---
    # Linha OCR: "GRUPO DO TRIBUTO . PIS/PASEP - CONTRIB. ..."
    # Separador entre "GRUPO DO TRIBUTO" e o nome pode ser: . : ; | ' ` > 1 I , espaco
    # (OCR varia conforme a fonte e o DPI — cobre todos os artefatos conhecidos)
    grupo = ""
    m_grp = re.search(r"GRUPO\s+DO\s+TRIBUTO\s*[.:;|'`1I>,\s]\s*(.+)",
                       page_text, re.IGNORECASE)
    if m_grp:
        # Limpa separadores do início do sufixo. NUNCA remove caracteres que
        # poderiam ser o primeiro char de um nome de tributo válido (I de IRRF,
        # 1 seria letra/dígito, etc.). Só remove sinais de pontuação OCR e espaços.
        _sep_clean = lambda txt: re.sub(r"^[.:;|'`>,\s]+", "", txt).strip()
        sufixo = _sep_clean(m_grp.group(1))
        # A linha ANTERIOR pode conter o inicio do nome do tributo
        antes = page_text[:m_grp.start()].rstrip()
        linha_ant = antes.split("\n")[-1].strip() if "\n" in antes else antes.strip()
        eh_nome = (
            len(linha_ant) > 5
            and re.search(r"[A-Z]{3}", linha_ant)
            and not re.search(
                r"MINISTERIO|SECRETARIA|INFORMACAO|CNPJ|DCTF|"
                r"DEBITO\s+APURADO|FIM\s+DE|GRUPO|FISCAL",
                linha_ant, re.IGNORECASE)
        )
        if eh_nome and sufixo:
            linha_ant = re.sub(r"^[_\-\s']+", "", linha_ant).strip()
            grupo = (linha_ant + " " + sufixo).strip()
        elif eh_nome:
            grupo = re.sub(r"^[_\-\s']+", "", linha_ant).strip()
        elif sufixo:
            grupo = sufixo
            # Verifica continuacao na linha seguinte
            resto = page_text[m_grp.end():]
            nl = re.match(r"[ \t]*([A-Z][A-Z0-9/. -]{3,})\n", resto)
            if nl and not re.search(
                    r"^(CODIGO|PERIODICIDADE|PERIODO|DEBITO|CREDITO|SOMA|SALDO|Valor|Total|Pag)",
                    nl.group(1).strip(), re.IGNORECASE):
                grupo = (grupo + " " + nl.group(1).strip()).strip()

    # --- Codigo da Receita ---
    # OCR: "CODIGO RECEITA : 6912-01" ou "CODIGO RECEITA > 6912-01" ou "CODIGO RECEITA 1 6912-01"
    cod = r(r"CODIGO\s+RECEITA\s*[^\d]?\s*(\d{4}-\d{2})", page_text)
    if not cod:
        cod = r(r"\b(\d{4}-\d{2})\b", page_text)

    # --- Periodicidade e Periodo de Apuracao ---
    # Cobre variações do OCR: "PERIODO" ou "PERTODO" (I→T), "APURACAO" ou "APURAC4O".
    period = r(r"PERIODICIDADE[:\s]+(\w+)", page_text)
    # Chave "PERIODO DE APURACAO" — aceita PERIODO / PERTODO etc. (1 char variável)
    _key_pa = r"PER.ODO\s+DE\s+APURAC.O"
    # Mensal: "Dezembro/2024"
    pa = r(_key_pa + r"[:\s]+((?:Jan|Fev|Mar|Abr|Mai|Jun|Jul|Ago|Set|Out|Nov|Dez)\w*/\d{4})",
            page_text)
    # Trimestral: "4° Trimestre/2024"  (o ° às vezes vira "o" no OCR)
    if not pa:
        pa = r(_key_pa + r"[:\s]+(\d[°ºo]?\s*Trimestre/\d{4})", page_text)
    # Decendial: "3° Decendio/Dez/2023"
    if not pa:
        pa = r(_key_pa + r"[:\s]+(\d[°ºo]?\s*Decendio/\w+/\d{4})", page_text)
    # Diária: "21° Dia/Dez/2023"
    if not pa:
        pa = r(_key_pa + r"[:\s]+(\d{1,2}[°ºo]?\s*Dia/\w+/\d{4})", page_text)
    # Quinzenal: "2° Quinzena/Dez/2023"
    if not pa:
        pa = r(_key_pa + r"[:\s]+(\d[°ºo]?\s*Quinzena/\w+/\d{4})", page_text)
    # Semanal: "5° Semana/Dez/2023"
    if not pa:
        pa = r(_key_pa + r"[:\s]+(\d[°ºo]?\s*Semana/\w+/\d{4})", page_text)
    # Anual / Semestral / fallback genérico: "Algo/2024"
    if not pa:
        pa = r(_key_pa + r"[:\s]+([\w°ºo° /]+?/\d{4})", page_text)

    # --- Debito apurado e creditos vinculados ---
    debito    = v(r"DEBITO\s+APURADO\s+" + N)
    pagamento = v(r"[-–—]\s*PAGAMENTO\s+" + N)
    compens   = v(r"[-–—]\s*COMPENSACOES\s+" + N)
    parcel    = v(r"[-–—]\s*PARCELAMENTO\s+" + N)
    suspens   = v(r"[-–—]\s*SUSPENSAO\s+" + N)
    soma_cred = v(r"SOMA\s+DOS\s+CREDITOS\s+VINCULADOS[:\s]+" + N)
    saldo     = v(r"SALDO\s+A\s+PAGAR\s+DO\s+DEBITO[:\s]+" + N)

    # Valor do Debito Total — OCR converte "R$" em "RS"
    vl_total = v(r"Valor\s+do\s+D[eé][^\s]{0,5}\s*[-–]\s*R[S$]\s+Total[:\s]+" + N)

    # Total da Contribuicao — pega o ultimo valor monetario da linha
    linha_contr = r(r"(Total\s+da\s+Contribui[^\n]{5,120})", page_text)
    tot_contr = _dctf_last_brl(linha_contr) if linha_contr else 0.0

    # Pagamento com DARF
    pag_darf = v(r"Pagamento\s+com\s+DARF\s*[-–—]\s*R[S$]\s+Total[:\s]+" + N)

    # --- DARF vinculado ao debito ---
    darf_pa    = r(r"\bPA[:\s]+(\d{2}/\d{2}/\d{4})", page_text)
    darf_cod   = r(r"C[eé][^\s]{0,8}digo\s+da\s+Receita[:\s]+(\d{4})", page_text)
    darf_venc  = r(r"Data\s+do\s+Vencimento\s+(\d{2}/\d{2}/\d{4})", page_text)
    darf_princ = v(r"Valor\s+do\s+Principal[:\s]+" + N)
    darf_multa = v(r"Valor\s+da\s+Multa[:\s]+" + N)
    darf_juros = v(r"Valor\s+dos\s+Juros[:\s]+" + N)
    # DARF total: captura ate 12 chars para cobrir "1.121, 90" com espaco OCR
    m_dt = re.search(r"Valor\s+Total\s+do\s+DARF[:\s]+([\d.,\s]{3,12})",
                     page_text, re.IGNORECASE)
    darf_total = _dctf_brl(m_dt.group(1)) if m_dt else 0.0
    darf_pago  = v(r"Valor\s+Pago\s+do\s+D[eé][^\s]{0,5}[:\s]+" + N)

    return {
        "grupo_tributo":        grupo,
        "codigo_receita":       cod,
        "periodicidade":        period,
        "periodo_apuracao":     pa,
        "competencia_teste":    format_competencia_teste(pa),
        "debito_apurado":       debito,
        "credito_pagamento":    pagamento,
        "credito_compensacoes": compens,
        "credito_parcelamento": parcel,
        "credito_suspensao":    suspens,
        "soma_creditos":        soma_cred,
        "saldo_pagar":          saldo,
        "valor_total_debito":   vl_total,
        "total_contribuicao":   tot_contr,
        "pagamento_darf":       pag_darf,
        "darf_pa":              darf_pa,
        "darf_codigo_receita":  darf_cod,
        "darf_vencimento":      darf_venc,
        "darf_principal":       darf_princ,
        "darf_multa":           darf_multa,
        "darf_juros":           darf_juros,
        "darf_total":           darf_total,
        "darf_pago":            darf_pago,
    }


# ── Extracao principal ────────────────────────────────────────────────────────
def extract_dctf(pdf_path: str) -> list:
    """
    Processa um PDF de DCTF e retorna lista de dicts (uma linha por tributo).
    Lanca RuntimeError com mensagem clara se o PDF nao puder ser processado.
    """
    pages = _dctf_get_pages_text(pdf_path)
    full  = "\n".join(pages)

    cabecalho = _dctf_parse_cabecalho(full)
    cabecalho["_source"] = Path(pdf_path).name

    # Paginas de tributo: contem "GRUPO DO TRIBUTO" e "DEBITO APURADO"
    # Normaliza acentos para que funcione com OCR em PT (DÉBITO) ou EN (DEBITO).
    def _has_tributo(p: str) -> bool:
        norm = _dctf_strip_accents(p)
        return bool(re.search(r"GRUPO\s+DO\s+TRIBUTO", norm, re.IGNORECASE)
                    and re.search(r"DEBITO\s+APURADO", norm, re.IGNORECASE))

    tributo_pages = [p for p in pages if _has_tributo(p)]

    rows = [{**cabecalho, **_dctf_parse_tributo(p)} for p in tributo_pages]

    if not rows:
        rows.append({
            **cabecalho,
            "grupo_tributo":  "Nenhum tributo encontrado",
            "codigo_receita": "", "periodicidade": "", "periodo_apuracao": "",
            "competencia_teste": "",
        })

    return rows


# ── Resumo consolidado ────────────────────────────────────────────────────────
def build_dctf_resumo(rows: list) -> list:
    from collections import defaultdict
    acc = defaultdict(lambda: {
        "qtd_declaracoes": 0, "total_debito": 0.0, "total_pagamento": 0.0,
        "total_compensacao": 0.0, "total_parcelamento": 0.0,
        "total_suspensao": 0.0, "total_saldo": 0.0, "total_darf_pago": 0.0,
        "grupo_tributo": "",
    })
    for r in rows:
        k = r.get("codigo_receita") or "-"
        acc[k]["qtd_declaracoes"]    += 1
        acc[k]["total_debito"]       += float(r.get("debito_apurado",       0) or 0)
        acc[k]["total_pagamento"]    += float(r.get("credito_pagamento",    0) or 0)
        acc[k]["total_compensacao"]  += float(r.get("credito_compensacoes", 0) or 0)
        acc[k]["total_parcelamento"] += float(r.get("credito_parcelamento", 0) or 0)
        acc[k]["total_suspensao"]    += float(r.get("credito_suspensao",    0) or 0)
        acc[k]["total_saldo"]        += float(r.get("saldo_pagar",         0) or 0)
        acc[k]["total_darf_pago"]    += float(r.get("darf_pago",           0) or 0)
        if not acc[k]["grupo_tributo"]:
            acc[k]["grupo_tributo"] = r.get("grupo_tributo", "")
    return [{"codigo_receita": k, **v} for k, v in sorted(acc.items())]


# ── Exportacao Excel ─────────────────────────────────────────────────────────
def _dctf_border():
    t = Side(style="thin", color="D0DDB8")
    return Border(left=t, right=t, top=t, bottom=t)


def export_dctf_excel(rows: list, resumo_rows: list, path: str):
    from openpyxl import Workbook
    wb = Workbook()

    def _hdr(ws, ri, cols, title):
        ws.merge_cells(start_row=ri, start_column=1,
                       end_row=ri, end_column=len(cols))
        c = ws.cell(row=ri, column=1, value=title)
        c.font      = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        c.fill      = PatternFill("solid", fgColor="5A8A1E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 22
        for ci, (key, label, width) in enumerate(cols, 1):
            h = ws.cell(row=ri+1, column=ci, value=label)
            h.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            h.fill      = PatternFill("solid", fgColor="7AB82E")
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = width / 7
        ws.row_dimensions[ri+1].height = 30

    # Aba 1 — Detalhamento
    ws1 = wb.active; ws1.title = "Detalhamento"
    _hdr(ws1, 1, DCTF_DETAIL_COLS,
         "DCTF - DETALHAMENTO POR TRIBUTO  |  AgriTax Tributario & Contabil")
    cur_src, dr = None, 3
    for row in rows:
        src = row.get("_source", "")
        if src != cur_src:
            cur_src = src
            ws1.merge_cells(start_row=dr, start_column=1,
                            end_row=dr, end_column=len(DCTF_DETAIL_COLS))
            s = ws1.cell(row=dr, column=1, value=f"  {src}")
            s.font      = Font(bold=True, color="5A8A1E", size=9, name="Calibri")
            s.fill      = PatternFill("solid", fgColor="EAF4D3")
            s.alignment = Alignment(vertical="center")
            ws1.row_dimensions[dr].height = 16
            dr += 1
        bg = "FFFFFF" if dr % 2 == 0 else "F2F4F0"
        if float(row.get("saldo_pagar", 0) or 0) > 0:
            bg = "FDEDEC"
        for ci, (key, _, _w) in enumerate(DCTF_DETAIL_COLS, 1):
            v = row.get(key, "")
            if key in DCTF_MONEY_KEYS and v != "":
                v = float(v or 0)
            c = ws1.cell(row=dr, column=ci, value=v)
            c.font      = Font(size=9, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="right" if key in DCTF_MONEY_KEYS else "left",
                vertical="center")
            c.border = _dctf_border()
            if key in DCTF_MONEY_KEYS:
                c.number_format = "#.##0,00"
        dr += 1
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(DCTF_DETAIL_COLS))}2"

    # Aba 2 — Resumo
    ws2 = wb.create_sheet("Resumo por Tributo")
    _hdr(ws2, 1, DCTF_RESUMO_COLS, "DCTF - RESUMO CONSOLIDADO POR TRIBUTO")
    for ri, row in enumerate(resumo_rows, 3):
        bg = "FFFFFF" if ri % 2 == 0 else "F2F4F0"
        for ci, (key, _, _w) in enumerate(DCTF_RESUMO_COLS, 1):
            v = row.get(key, "")
            if key in DCTF_MONEY_KEYS:
                v = float(v or 0)
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font      = Font(size=9, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="right" if key in DCTF_MONEY_KEYS else "left",
                vertical="center")
            c.border = _dctf_border()
            if key in DCTF_MONEY_KEYS:
                c.number_format = "#.##0,00"
    tr = len(resumo_rows) + 3
    for ci, (key, _, _w) in enumerate(DCTF_RESUMO_COLS, 1):
        c = ws2.cell(row=tr, column=ci)
        if key == "codigo_receita":
            c.value = "TOTAL"
        elif key == "qtd_declaracoes":
            c.value = sum(r.get("qtd_declaracoes", 0) for r in resumo_rows)
        elif key in DCTF_MONEY_KEYS:
            c.value = sum(float(r.get(key, 0) or 0) for r in resumo_rows)
            c.number_format = "#.##0,00"
        c.font      = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="5A8A1E")
        c.border    = _dctf_border()
        c.alignment = Alignment(
            horizontal="right" if key in DCTF_MONEY_KEYS else "center",
            vertical="center")
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(DCTF_RESUMO_COLS))}2"
    wb.save(path)


# =============================================================================
# DCTFWeb — Declaração Completa (PDF de texto puro, sem OCR)
# =============================================================================
#
# Diferenças principais em relação à DCTF clássica:
#   - PDF nativo (texto extraível) — NÃO precisa OCR
#   - Tributos previdenciários (CP) e IRRF; sem PIS/COFINS/IPI
#   - Estrutura por blocos "Débito Apurado e Crédito Vinculado"
#   - Compensações vinculadas (lista de processos DCOMP) explícitas no PDF
#   - 3 categorias: 40 (Mensal), 41 (13º Anual), 44 (Aferição/Obra)
#   - Pode ter "Ausência de Fatos Geradores: Sim" → declaração negativa
# =============================================================================

DCTFWEB_DETAIL_COLS = [
    ("cnpj",                  "CNPJ",                       130),
    ("razao_social",          "Razão Social",               210),
    ("categoria",             "Categoria",                  140),
    ("classificacao_trib",    "Classif. Tributária",        180),
    ("periodo_apuracao_decl", "Período Decl.",              100),
    ("numero_recibo",         "Nº Recibo",                  140),
    ("dt_transmissao",        "Dt. Transmissão",            140),
    ("codigo_receita",        "Cód. Receita",                90),
    ("descricao",             "Descrição do Tributo",       260),
    ("grupo_tributo",         "Grupo do Tributo",           180),
    ("cno",                   "CNO",                        130),
    ("cnpj_prest",            "CNPJ Prest./Incorp.",        140),
    ("periodo",               "Período",                     90),
    ("competencia_teste",     "Competência Teste",          120),
    ("debito_apurado",        "Débito Apurado",             130),
    ("deducoes",              "Deduções (Sal. Família)",    140),
    ("cred_compensacao",      "Cred.Vinc.Compensação",      140),
    ("cred_pagamento",        "Cred.Vinc.Pagamento",        140),
    ("cred_suspensao",        "Cred.Vinc.Suspensão",        140),
    ("saldo_pagar",           "Saldo a Pagar",              130),
    ("qtd_compensacoes",      "Qtd. DCOMPs Vinculadas",     130),
    ("numeros_dcomp",         "Nºs DCOMPs Vinculadas",      280),
    ("ausencia_fatos",        "Aus. Fatos Geradores",       130),
]
DCTFWEB_KEYS = [c[0] for c in DCTFWEB_DETAIL_COLS]
DCTFWEB_MONEY_KEYS = {"debito_apurado", "deducoes", "cred_compensacao",
                       "cred_pagamento", "cred_suspensao", "saldo_pagar"}

DCTFWEB_RESUMO_COLS = [
    ("codigo_receita",     "Cód. Receita",            100),
    ("descricao",          "Descrição",               280),
    ("grupo_tributo",      "Grupo do Tributo",        180),
    ("qtd_declaracoes",    "Qtd. Declarações",        130),
    ("total_debito",       "Total Débito Apurado",    160),
    ("total_deducoes",     "Total Deduções",          140),
    ("total_compensacao",  "Total Compensações",      160),
    ("total_pagamento",    "Total Pagamento",         140),
    ("total_suspensao",    "Total Suspensão",         140),
    ("total_saldo",        "Total Saldo a Pagar",     160),
]
DCTFWEB_RESUMO_MONEY = {"total_debito", "total_deducoes", "total_compensacao",
                          "total_pagamento", "total_suspensao", "total_saldo"}


def extract_dctfweb(pdf_path: str) -> list:
    """Parser de DCTFWeb (Declaração Completa).

    Retorna uma lista de dicts, um por bloco "Débito Apurado e Crédito Vinculado".
    Para declarações com "Ausência de Fatos Geradores: Sim", retorna 1 linha
    representando a declaração negativa (sem tributos).
    """
    # PDF é texto puro — usa pdfplumber (sem OCR)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "\n".join(page.extract_text(layout=True) or ""
                                   for page in pdf.pages)
    except Exception:
        # Fallback para pdftotext se pdfplumber falhar
        try:
            import subprocess
            r = subprocess.run(["pdftotext", "-layout", pdf_path, "-"],
                                capture_output=True, text=True, timeout=60)
            full_text = r.stdout
        except Exception as e:
            raise RuntimeError(f"Falha ao extrair texto do PDF: {e}")

    if not full_text.strip():
        return []

    # ── Cabeçalho ──
    cabecalho = {}
    nome_arq = Path(pdf_path).name
    cabecalho["_source"] = nome_arq

    # CNPJ
    m = re.search(r'CNPJ\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', full_text)
    cabecalho["cnpj"] = m.group(1) if m else ""

    # Razão social — pega linha ANTES e linha DEPOIS de "Nome do Contribuinte",
    # filtrando ruído do cabeçalho ministerial
    razao = ""
    lines = full_text.split("\n")
    for i, ln in enumerate(lines):
        if "Nome do Contribuinte" in ln:
            parte1 = lines[i-1].strip() if i > 0 else ""
            parte2 = lines[i+1].strip() if i+1 < len(lines) else ""
            for ruido in ("MINISTÉRIO", "RELATÓRIO", "SECRETARIA"):
                if ruido in parte1.upper(): parte1 = ""
                if ruido in parte2.upper(): parte2 = ""
            if parte2.startswith("CNPJ") or "Período" in parte2:
                parte2 = ""
            razao = (parte1 + " " + parte2).strip()
            break
    cabecalho["razao_social"] = razao

    # Período da declaração
    m = re.search(r'Período apuração\s+(\d{2}/\d{4}|\d{4})', full_text)
    cabecalho["periodo_apuracao_decl"] = m.group(1) if m else ""

    # Número do Recibo
    m = re.search(r'Número do Recibo\s+(\S+)', full_text)
    cabecalho["numero_recibo"] = m.group(1) if m else ""

    # Data/Hora — primeira data DD/MM/AAAA HH:MM:SS
    m = re.search(r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})', full_text)
    cabecalho["dt_transmissao"] = m.group(1) if m else ""

    # Categoria via nome do arquivo: padrão *_NN_*.pdf onde NN é a categoria
    # Aceita também sufixo " (1)", " (2)" etc., ou número de aferição após o NN
    cat = ""
    m = re.search(r'_(\d{2})_(?:\d+)?(?:\s*\([^\)]*\))?\.pdf$', nome_arq)
    if m:
        cat = m.group(1)
    categoria_map = {
        "40": "Mensal",
        "41": "13º Salário (Anual)",
        "44": "Aferição (Obra/CNO)",
    }
    cabecalho["categoria"] = categoria_map.get(cat, f"Cat. {cat}" if cat else "")

    # Classificação Tributária
    m = re.search(r'Classificação Tributária\s+(.+?)(?:\n|$)', full_text)
    cabecalho["classificacao_trib"] = m.group(1).strip()[:100] if m else ""

    # Ausência de Fatos Geradores
    m = re.search(r'Ausência de Fatos Geradores\s+(Sim|Não)', full_text)
    cabecalho["ausencia_fatos"] = m.group(1) if m else "Não"

    # ── Blocos "Débito Apurado e Crédito Vinculado" ──
    blocos = re.split(r'Débito Apurado e Crédito Vinculado', full_text)[1:]

    debitos = []
    for bloco in blocos:
        mc = re.search(r'Código da Receita\s+(\d{4}-\d{2})', bloco)
        if not mc:
            continue
        codigo = mc.group(1)

        # Descrição (até quebra de linha ou novo campo)
        md = re.search(
            r'Descrição\s+(.+?)(?:\n\s*CNPJ Prest|\n\s*CNO\s|\n\s*Período Apuração|\n\s*Débito\s)',
            bloco, re.S)
        descricao = " ".join(md.group(1).split()) if md else ""

        # CNO (Cadastro Nacional de Obras)
        mcno = re.search(r'CNO\s+([\d\.]+/\d+|\-)', bloco)
        cno = mcno.group(1) if (mcno and mcno.group(1) != "-") else ""

        # CNPJ Prest/Incorp
        mcnpj = re.search(
            r'CNPJ Prest/Incorp\s*\n?.*?\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})',
            bloco, re.S)
        cnpj_prest = mcnpj.group(1) if mcnpj else ""

        # Período do tributo (pode diferir do período da declaração — ex: 13º)
        mp = re.search(r'Período Apuração\s*\n?\s*(\d{2}/\d{4}|\d{4})', bloco)
        periodo = mp.group(1) if mp else cabecalho.get("periodo_apuracao_decl", "")

        # Valores monetários
        def _money(label, txt):
            mm = re.search(rf'{label}\s+([\d\.]+,\d{{2}})', txt)
            if mm:
                try: return float(mm.group(1).replace(".", "").replace(",", "."))
                except ValueError: return 0.0
            return 0.0

        debito_apurado = _money(r'Débito Apurado', bloco)

        # Deduções (Salário Família) — linha "Deduções Salário Família: X,XX"
        m_ded = re.search(r'Deduções\s+(?:[^\d\n]*?)([\d\.]+,\d{2})', bloco)
        deducoes = float(m_ded.group(1).replace(".","").replace(",",".")) if m_ded else 0.0

        m_comp = re.search(r'Créditos Compensação:\s*([\d\.]+,\d{2})', bloco)
        cred_comp = float(m_comp.group(1).replace(".","").replace(",",".")) if m_comp else 0.0

        m_pag = re.search(r'Créditos Pagamento:\s*([\d\.]+,\d{2})', bloco)
        cred_pag = float(m_pag.group(1).replace(".","").replace(",",".")) if m_pag else 0.0

        m_susp = re.search(r'Créditos Suspensão:\s*([\d\.]+,\d{2})', bloco)
        cred_susp = float(m_susp.group(1).replace(".","").replace(",",".")) if m_susp else 0.0

        saldo = _money(r'Saldo a Pagar', bloco)

        # Compensações vinculadas (lista de DCOMPs)
        compensacoes = []
        for cm in re.finditer(
                r'Número do Processo\s+(\S+).+?Tipo\s+(\S+).+?Valor\s+([\d\.]+,\d{2})',
                bloco, re.S):
            compensacoes.append({
                "numero_processo": cm.group(1),
                "tipo": cm.group(2),
                "valor": float(cm.group(3).replace(".","").replace(",",".")),
            })

        # Grupo deriva da descrição
        desc_upper = descricao.upper()
        if descricao.startswith("CP"):
            grupo = "Contribuição Previdenciária"
        elif "IRRF" in desc_upper:
            grupo = "IRRF"
        elif "RET DE CONTRIB" in desc_upper or "CONTRIBUI" in desc_upper:
            grupo = "Outras Contribuições"
        else:
            grupo = "Outros"

        debitos.append({
            **cabecalho,
            "codigo_receita":   codigo,
            "descricao":        descricao[:180],
            "grupo_tributo":    grupo,
            "cno":              cno,
            "cnpj_prest":       cnpj_prest,
            "periodo":          periodo,
            "competencia_teste": format_competencia_teste(periodo),
            "debito_apurado":   debito_apurado,
            "deducoes":         deducoes,
            "cred_compensacao": cred_comp,
            "cred_pagamento":   cred_pag,
            "cred_suspensao":   cred_susp,
            "saldo_pagar":      saldo,
            "qtd_compensacoes": len(compensacoes),
            "numeros_dcomp":    " / ".join(c["numero_processo"] for c in compensacoes),
            "_compensacoes_raw": compensacoes,
        })

    # Se "Ausência de Fatos Geradores: Sim" e não capturamos nenhum bloco,
    # gera 1 linha representando a declaração negativa
    if not debitos and cabecalho.get("ausencia_fatos") == "Sim":
        debitos.append({
            **cabecalho,
            "codigo_receita":   "",
            "descricao":        "(Sem fatos geradores)",
            "grupo_tributo":    "—",
            "cno":              "",
            "cnpj_prest":       "",
            "periodo":          cabecalho.get("periodo_apuracao_decl", ""),
            "competencia_teste": format_competencia_teste(
                cabecalho.get("periodo_apuracao_decl", "")),
            "debito_apurado":   0.0,
            "deducoes":         0.0,
            "cred_compensacao": 0.0,
            "cred_pagamento":   0.0,
            "cred_suspensao":   0.0,
            "saldo_pagar":      0.0,
            "qtd_compensacoes": 0,
            "numeros_dcomp":    "",
            "_compensacoes_raw": [],
        })

    return debitos


def build_dctfweb_resumo(rows: list) -> list:
    """Resumo por código de receita (totalizadores)."""
    if not rows:
        return []
    grupos = {}
    for r in rows:
        cod = r.get("codigo_receita", "")
        if not cod: continue   # ignora linhas de "sem fatos geradores"
        if cod not in grupos:
            grupos[cod] = {
                "codigo_receita": cod,
                "descricao": r.get("descricao", ""),
                "grupo_tributo": r.get("grupo_tributo", ""),
                "qtd_declaracoes": 0,
                "total_debito": 0.0,
                "total_deducoes": 0.0,
                "total_compensacao": 0.0,
                "total_pagamento": 0.0,
                "total_suspensao": 0.0,
                "total_saldo": 0.0,
            }
        g = grupos[cod]
        g["qtd_declaracoes"] += 1
        g["total_debito"]      += float(r.get("debito_apurado", 0) or 0)
        g["total_deducoes"]    += float(r.get("deducoes", 0) or 0)
        g["total_compensacao"] += float(r.get("cred_compensacao", 0) or 0)
        g["total_pagamento"]   += float(r.get("cred_pagamento", 0) or 0)
        g["total_suspensao"]   += float(r.get("cred_suspensao", 0) or 0)
        g["total_saldo"]       += float(r.get("saldo_pagar", 0) or 0)
    return sorted(grupos.values(), key=lambda x: x["codigo_receita"])


def export_dctfweb_excel(rows: list, resumo_rows: list, path: str):
    """Exporta DCTFWeb para Excel: 2 abas (Detalhamento + Resumo por Tributo)."""
    from openpyxl import Workbook
    wb = Workbook()

    def _hdr(ws, ri, cols, title):
        ws.merge_cells(start_row=ri, start_column=1,
                        end_row=ri, end_column=len(cols))
        c = ws.cell(row=ri, column=1, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        c.fill = PatternFill("solid", fgColor="5A8A1E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 22
        for ci, (key, label, width) in enumerate(cols, 1):
            cc = ws.cell(row=ri+1, column=ci, value=label)
            cc.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cc.fill = PatternFill("solid", fgColor="3D6B0A")
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = width / 7

    brd = Border(left=Side(style="thin", color="CCCCCC"),
                  right=Side(style="thin", color="CCCCCC"),
                  top=Side(style="thin", color="CCCCCC"),
                  bottom=Side(style="thin", color="CCCCCC"))

    # ── Aba 1: Detalhamento ──
    ws1 = wb.active
    ws1.title = "Detalhamento"
    _hdr(ws1, 1, DCTFWEB_DETAIL_COLS,
          "DCTFWeb — Detalhamento por Tributo  |  AgriTax Audit")
    for ri, row in enumerate(rows, 3):
        # Cor por grupo + alerta para saldo a pagar
        saldo = float(row.get("saldo_pagar", 0) or 0)
        if saldo > 0.01:
            bg = "FDEDEC"
        elif row.get("cred_compensacao", 0):
            bg = "EAF4D3"   # tem compensação — verde claro
        else:
            bg = "F7FBED" if ri % 2 == 0 else "FFFFFF"
        for ci, (key, _, _w) in enumerate(DCTFWEB_DETAIL_COLS, 1):
            v = row.get(key, "")
            if key in DCTFWEB_MONEY_KEYS:
                v = float(v or 0)
            c = ws1.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="right" if key in DCTFWEB_MONEY_KEYS else "left",
                vertical="center")
            c.border = brd
            if key in DCTFWEB_MONEY_KEYS:
                c.number_format = '#,##0.00'
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(DCTFWEB_DETAIL_COLS))}2"

    # ── Aba 2: Resumo por Tributo ──
    ws2 = wb.create_sheet("Resumo por Tributo")
    _hdr(ws2, 1, DCTFWEB_RESUMO_COLS,
          "DCTFWeb — Resumo por Código de Receita  |  AgriTax Audit")
    for ri, row in enumerate(resumo_rows, 3):
        for ci, (key, _, _w) in enumerate(DCTFWEB_RESUMO_COLS, 1):
            v = row.get(key, "")
            if key in DCTFWEB_RESUMO_MONEY:
                v = float(v or 0)
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(
                horizontal="right" if (key in DCTFWEB_RESUMO_MONEY or key == "qtd_declaracoes") else "left",
                vertical="center")
            c.border = brd
            if key in DCTFWEB_RESUMO_MONEY:
                c.number_format = '#,##0.00'
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(DCTFWEB_RESUMO_COLS))}2"
    wb.save(path)


# =============================================================================
# EFD Contribuições — Arquivo SPED .txt (Ato COTEPE/ICMS 44/2018)
# =============================================================================
#
# Estrutura do arquivo SPED:
#   - Linhas iniciam e terminam por "|"
#   - Cada linha começa com REG (4 caracteres)
#   - Encoding: Latin-1 (ISO-8859-1) ou UTF-8 (varia)
#
# Registros relevantes para auditoria DCTF/DCTFWeb:
#   0000 — Cabeçalho: CNPJ, razão, período (DT_INI/DT_FIN), versão
#   0140 — Estabelecimentos (filiais)
#   M200 — Consolidação Contribuição PIS do período (débito + deduções + a recolher)
#   M205 — Detalhamento PIS por código de receita (NRO + VL_DEBITO + COD_REC)
#   M210 — Detalhamento PIS Não-Cumulativo (base de cálculo + alíquota + débito)
#   M400 — Receitas isentas/não tributadas/alíquota zero (PIS)
#   M600 — Consolidação Contribuição COFINS do período
#   M605 — Detalhamento COFINS por código de receita
#   M610 — Detalhamento COFINS Não-Cumulativo
#   M800 — Receitas isentas/não tributadas/alíquota zero (COFINS)
#
# Códigos de receita típicos:
#   PIS:    8109 (cumul.), 6912 (não-cumul. mercado interno), 4574 (importação)
#   COFINS: 2172 (cumul.), 5856 (não-cumul. mercado interno), 5442 (importação)
# =============================================================================

EFD_DETAIL_COLS = [
    ("cnpj",                "CNPJ",                       130),
    ("razao_social",        "Razão Social",               210),
    ("periodo",             "Período",                     90),
    ("competencia_teste",   "Competência Teste",          120),
    ("tributo",             "Tributo",                     90),  # PIS / COFINS
    ("codigo_receita",      "Cód. Receita",                90),
    ("descricao_codigo",    "Descrição Cód. Receita",     230),
    ("regime",              "Regime",                     130),  # Cumulativo / Não-Cumul.
    ("base_calculo",        "Base de Cálculo",            140),
    ("aliquota",            "Alíquota (%)",               100),
    ("debito_apurado",      "Débito Apurado",             140),
    ("ajuste_acrescimo",    "Ajustes Acréscimo",          140),
    ("ajuste_reducao",      "Ajustes Redução",            140),
    ("contrib_periodo",     "Contribuição do Período",    150),
    ("ded_credito",         "Deduções (Crédito)",         140),
    ("ded_outras",          "Outras Deduções",            140),
    ("contrib_a_recolher",  "Contribuição a Recolher",    150),
    ("_source",             "Arquivo Origem",             200),
]
EFD_KEYS = [c[0] for c in EFD_DETAIL_COLS]
EFD_MONEY_KEYS = {"base_calculo", "aliquota", "debito_apurado",
                   "ajuste_acrescimo", "ajuste_reducao", "contrib_periodo",
                   "ded_credito", "ded_outras", "contrib_a_recolher"}

EFD_RESUMO_COLS = [
    ("tributo",             "Tributo",                    100),
    ("codigo_receita",      "Cód. Receita",                90),
    ("descricao_codigo",    "Descrição",                  230),
    ("qtd_periodos",        "Qtd. Períodos",              130),
    ("total_base",          "Total Base de Cálculo",      170),
    ("total_debito",        "Total Débito Apurado",       170),
    ("total_deducoes",      "Total Deduções",             140),
    ("total_recolher",      "Total a Recolher",           150),
]
EFD_RESUMO_MONEY = {"total_base", "total_debito", "total_deducoes", "total_recolher"}

# Mapa código → descrição (códigos de receita típicos PIS/COFINS)
EFD_CODIGO_DESC = {
    # PIS
    "8109": "PIS - Faturamento (Cumulativo)",
    "6912": "PIS - Não-Cumulativo (Mercado Interno)",
    "4574": "PIS - Importação",
    "1921": "PIS - Folha de Salários",
    "8496": "PIS - Receitas Financeiras",
    # COFINS
    "2172": "COFINS - Faturamento (Cumulativo)",
    "5856": "COFINS - Não-Cumulativo (Mercado Interno)",
    "5442": "COFINS - Importação",
    "8645": "COFINS - Receitas Financeiras",
    "5960": "COFINS - Combustíveis",
}


def _efd_brl(s: str) -> float:
    """Converte string SPED ('1234,56' ou '0,00') para float."""
    s = (s or "").strip().replace(".", "").replace(",", ".")
    try: return float(s)
    except ValueError: return 0.0


def _efd_periodo(dt_ini: str) -> str:
    """Extrai período MM/AAAA de DT_INI no formato DDMMAAAA."""
    s = (dt_ini or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[2:4]}/{s[4:8]}"
    return ""


def _efd_decode_file(path: str) -> list:
    """Lê o arquivo SPED tentando Latin-1 e UTF-8 (encodings mais comuns)."""
    for enc in ("latin-1", "utf-8", "cp1252"):
        try:
            with open(path, "r", encoding=enc) as f:
                return [ln.rstrip("\r\n") for ln in f]
        except UnicodeDecodeError:
            continue
    # Último recurso: lê como bytes e decodifica com errors='replace'
    with open(path, "rb") as f:
        return [ln.decode("latin-1", errors="replace").rstrip("\r\n")
                for ln in f.readlines()]


def extract_efd_contribuicoes(path: str) -> list:
    """Parser de EFD Contribuições (.txt SPED).

    Retorna lista de dicts, uma linha por (período × código de receita)
    cruzando dados do M205/M210 (PIS) e M605/M610 (COFINS).
    """
    nome = Path(path).name
    lines = _efd_decode_file(path)

    # ── Cabeçalho (registro 0000) ──────────────────────────────────────────
    cabecalho = {"_source": nome, "cnpj": "", "razao_social": "",
                 "periodo": "", "competencia_teste": ""}
    for ln in lines:
        if ln.startswith("|0000|"):
            f = ln.split("|")
            # |0000|COD_VER|COD_FIN|DT_INI|DT_FIN|NOME|CNPJ|UF|...
            #  0    1       2       3      4      5    6    7
            if len(f) >= 8:
                dt_ini = f[4]
                cabecalho["periodo"] = _efd_periodo(dt_ini)
                cabecalho["competencia_teste"] = format_competencia_teste(
                    cabecalho["periodo"])
                cabecalho["razao_social"] = f[6][:200]
                cnpj_raw = re.sub(r"\D", "", f[7])
                if len(cnpj_raw) == 14:
                    cabecalho["cnpj"] = (
                        f"{cnpj_raw[0:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}"
                        f"/{cnpj_raw[8:12]}-{cnpj_raw[12:14]}")
                else:
                    cabecalho["cnpj"] = f[7]
            break

    # ── Indexa M205/M210 (PIS) e M605/M610 (COFINS) por código ─────────────
    # M200 — Consolidação PIS:
    #   |M200|VL_TOT_CONT_NC_PER|VL_TOT_CRED_DESC|VL_TOT_CRED_DESC_ANT|VL_TOT_CONT_NC_DEV|
    #         VL_RET_NC|VL_OUT_DED_NC|VL_CONT_NC_REC|
    #         VL_TOT_CONT_CUM_PER|VL_RET_CUM|VL_OUT_DED_CUM|VL_CONT_CUM_REC|
    # M205 — Detalhamento PIS por código:
    #   |M205|NUM_CAMPO|VL_DEBITO|COD_REC|
    # M210 — Detalhamento PIS Não-Cumulativo:
    #   |M210|COD_CONT|VL_REC_BRT|VL_BC_CONT|VL_AJUS_ACRES_BC_PIS|VL_AJUS_REDUC_BC_PIS|
    #         VL_BC_CONT_AJUS|ALIQ_PIS|QUANT_BC_PIS|ALIQ_PIS_QUANT|VL_CONT_APUR|...
    # (M600/M605/M610 análogos para COFINS)

    debitos = []
    i = 0
    n = len(lines)

    def _split(line):
        """Quebra linha em campos, removendo o '|' inicial e final."""
        parts = line.split("|")
        # |REG|f1|f2|...|fN|  → ['', 'REG', 'f1', ..., 'fN', '']
        return parts[1:-1] if len(parts) >= 2 else []

    # Coleta dados M210 / M610 (regime + base + alíquota) por código
    m210_por_codigo = {}   # PIS — não-cumulativo
    m610_por_codigo = {}   # COFINS — não-cumulativo

    while i < n:
        ln = lines[i]
        f = _split(ln)
        if not f:
            i += 1; continue
        reg = f[0]

        if reg == "M210":
            # PIS Não-Cumulativo
            # f: REG, COD_CONT, VL_REC_BRT, VL_BC_CONT, VL_AJUS_ACRES, VL_AJUS_RED,
            #    VL_BC_CONT_AJUS, ALIQ_PIS, ..., VL_CONT_APUR, ...
            try:
                cod = f[1] if len(f) > 1 else ""
                base    = _efd_brl(f[3]) if len(f) > 3 else 0.0
                acres   = _efd_brl(f[4]) if len(f) > 4 else 0.0
                reduc   = _efd_brl(f[5]) if len(f) > 5 else 0.0
                aliq    = _efd_brl(f[7]) if len(f) > 7 else 0.0
                debito  = _efd_brl(f[10]) if len(f) > 10 else 0.0
                m210_por_codigo[cod] = {
                    "base": base, "aliquota": aliq,
                    "ajus_acres": acres, "ajus_reduc": reduc,
                    "debito": debito,
                }
            except Exception:
                pass

        elif reg == "M610":
            # COFINS Não-Cumulativo (mesmo layout do M210)
            try:
                cod = f[1] if len(f) > 1 else ""
                base    = _efd_brl(f[3]) if len(f) > 3 else 0.0
                acres   = _efd_brl(f[4]) if len(f) > 4 else 0.0
                reduc   = _efd_brl(f[5]) if len(f) > 5 else 0.0
                aliq    = _efd_brl(f[7]) if len(f) > 7 else 0.0
                debito  = _efd_brl(f[10]) if len(f) > 10 else 0.0
                m610_por_codigo[cod] = {
                    "base": base, "aliquota": aliq,
                    "ajus_acres": acres, "ajus_reduc": reduc,
                    "debito": debito,
                }
            except Exception:
                pass

        i += 1

    # Soma de deduções totais do M200 / M600 (rateadas pelos códigos)
    def _parse_m200_m600(lines):
        """Retorna {'pis': dict, 'cofins': dict} com totais de débito/dedução."""
        out = {"pis": {}, "cofins": {}}
        for ln in lines:
            f = _split(ln)
            if not f: continue
            reg = f[0]
            if reg == "M200":
                # Mapeamento simplificado: usamos o total a recolher como referência
                try:
                    out["pis"] = {
                        "tot_nc":    _efd_brl(f[1]) if len(f)>1 else 0.0,
                        "ded_cred":  _efd_brl(f[2]) if len(f)>2 else 0.0,
                        "ded_outras": _efd_brl(f[6]) if len(f)>6 else 0.0,
                        "rec_nc":    _efd_brl(f[7]) if len(f)>7 else 0.0,
                        "tot_cum":   _efd_brl(f[8]) if len(f)>8 else 0.0,
                        "rec_cum":   _efd_brl(f[11]) if len(f)>11 else 0.0,
                    }
                except Exception: pass
            elif reg == "M600":
                try:
                    out["cofins"] = {
                        "tot_nc":    _efd_brl(f[1]) if len(f)>1 else 0.0,
                        "ded_cred":  _efd_brl(f[2]) if len(f)>2 else 0.0,
                        "ded_outras": _efd_brl(f[6]) if len(f)>6 else 0.0,
                        "rec_nc":    _efd_brl(f[7]) if len(f)>7 else 0.0,
                        "tot_cum":   _efd_brl(f[8]) if len(f)>8 else 0.0,
                        "rec_cum":   _efd_brl(f[11]) if len(f)>11 else 0.0,
                    }
                except Exception: pass
        return out

    consolidado = _parse_m200_m600(lines)

    # Agora processa M205 e M605 (detalhamento por código)
    for ln in lines:
        f = _split(ln)
        if not f: continue
        reg = f[0]

        if reg == "M205":
            # |M205|NUM_CAMPO|VL_DEBITO|COD_REC|
            num_campo = f[1] if len(f) > 1 else ""
            vl_debito = _efd_brl(f[2]) if len(f) > 2 else 0.0
            cod_rec   = (f[3] if len(f) > 3 else "").strip()
            if not cod_rec or vl_debito <= 0:
                continue
            # Junta com M210 (mesmo NUM_CAMPO) se houver
            m210 = m210_por_codigo.get(num_campo, {})
            base   = m210.get("base", 0.0)
            aliq   = m210.get("aliquota", 0.0)
            acres  = m210.get("ajus_acres", 0.0)
            reduc  = m210.get("ajus_reduc", 0.0)
            # Regime: cumulativo se NUM_CAMPO == "01" (campo do M200), senão não-cumulativo
            regime = "Cumulativo" if num_campo == "01" else "Não-Cumulativo"
            # Calcula recolher proporcional ao débito apurado
            pis_total = consolidado.get("pis", {})
            tot_deb = pis_total.get("tot_cum", 0) + pis_total.get("tot_nc", 0)
            ded_total = pis_total.get("ded_cred", 0) + pis_total.get("ded_outras", 0)
            rec_total = pis_total.get("rec_cum", 0) + pis_total.get("rec_nc", 0)
            if tot_deb > 0:
                frac = vl_debito / tot_deb
                ded_cred_alloc   = pis_total.get("ded_cred", 0) * frac
                ded_outras_alloc = pis_total.get("ded_outras", 0) * frac
            else:
                ded_cred_alloc = ded_outras_alloc = 0.0
            recolher = vl_debito - ded_cred_alloc - ded_outras_alloc

            debitos.append({
                **cabecalho,
                "tributo":            "PIS",
                "codigo_receita":     cod_rec,
                "descricao_codigo":   EFD_CODIGO_DESC.get(cod_rec, ""),
                "regime":             regime,
                "base_calculo":       base,
                "aliquota":           aliq,
                "debito_apurado":     vl_debito,
                "ajuste_acrescimo":   acres,
                "ajuste_reducao":     reduc,
                "contrib_periodo":    vl_debito + acres - reduc,
                "ded_credito":        round(ded_cred_alloc, 2),
                "ded_outras":         round(ded_outras_alloc, 2),
                "contrib_a_recolher": round(recolher, 2),
            })

        elif reg == "M605":
            num_campo = f[1] if len(f) > 1 else ""
            vl_debito = _efd_brl(f[2]) if len(f) > 2 else 0.0
            cod_rec   = (f[3] if len(f) > 3 else "").strip()
            if not cod_rec or vl_debito <= 0:
                continue
            m610 = m610_por_codigo.get(num_campo, {})
            base   = m610.get("base", 0.0)
            aliq   = m610.get("aliquota", 0.0)
            acres  = m610.get("ajus_acres", 0.0)
            reduc  = m610.get("ajus_reduc", 0.0)
            regime = "Cumulativo" if num_campo == "01" else "Não-Cumulativo"
            cof_total = consolidado.get("cofins", {})
            tot_deb = cof_total.get("tot_cum", 0) + cof_total.get("tot_nc", 0)
            if tot_deb > 0:
                frac = vl_debito / tot_deb
                ded_cred_alloc   = cof_total.get("ded_cred", 0) * frac
                ded_outras_alloc = cof_total.get("ded_outras", 0) * frac
            else:
                ded_cred_alloc = ded_outras_alloc = 0.0
            recolher = vl_debito - ded_cred_alloc - ded_outras_alloc

            debitos.append({
                **cabecalho,
                "tributo":            "COFINS",
                "codigo_receita":     cod_rec,
                "descricao_codigo":   EFD_CODIGO_DESC.get(cod_rec, ""),
                "regime":             regime,
                "base_calculo":       base,
                "aliquota":           aliq,
                "debito_apurado":     vl_debito,
                "ajuste_acrescimo":   acres,
                "ajuste_reducao":     reduc,
                "contrib_periodo":    vl_debito + acres - reduc,
                "ded_credito":        round(ded_cred_alloc, 2),
                "ded_outras":         round(ded_outras_alloc, 2),
                "contrib_a_recolher": round(recolher, 2),
            })

    return debitos


def build_efd_resumo(rows: list) -> list:
    """Resumo da EFD por código de receita."""
    if not rows:
        return []
    grupos = {}
    for r in rows:
        key = (r.get("tributo", ""), r.get("codigo_receita", ""))
        if key not in grupos:
            grupos[key] = {
                "tributo": r.get("tributo", ""),
                "codigo_receita": r.get("codigo_receita", ""),
                "descricao_codigo": r.get("descricao_codigo", ""),
                "qtd_periodos": 0,
                "total_base": 0.0,
                "total_debito": 0.0,
                "total_deducoes": 0.0,
                "total_recolher": 0.0,
            }
        g = grupos[key]
        g["qtd_periodos"] += 1
        g["total_base"]     += float(r.get("base_calculo", 0) or 0)
        g["total_debito"]   += float(r.get("debito_apurado", 0) or 0)
        g["total_deducoes"] += (float(r.get("ded_credito", 0) or 0)
                                 + float(r.get("ded_outras", 0) or 0))
        g["total_recolher"] += float(r.get("contrib_a_recolher", 0) or 0)
    return sorted(grupos.values(),
                   key=lambda x: (x["tributo"], x["codigo_receita"]))


def export_efd_excel(rows: list, resumo_rows: list, path: str):
    """Exporta EFD para Excel: 2 abas (Detalhamento + Resumo)."""
    from openpyxl import Workbook
    wb = Workbook()

    def _hdr(ws, ri, cols, title):
        ws.merge_cells(start_row=ri, start_column=1,
                        end_row=ri, end_column=len(cols))
        c = ws.cell(row=ri, column=1, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        c.fill = PatternFill("solid", fgColor="5A8A1E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 22
        for ci, (key, label, width) in enumerate(cols, 1):
            cc = ws.cell(row=ri+1, column=ci, value=label)
            cc.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cc.fill = PatternFill("solid", fgColor="3D6B0A")
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = width / 7

    brd = Border(left=Side(style="thin", color="CCCCCC"),
                  right=Side(style="thin", color="CCCCCC"),
                  top=Side(style="thin", color="CCCCCC"),
                  bottom=Side(style="thin", color="CCCCCC"))

    # ── Aba 1: Detalhamento ──
    ws1 = wb.active
    ws1.title = "Detalhamento EFD"
    _hdr(ws1, 1, EFD_DETAIL_COLS,
          "EFD Contribuições — Detalhamento PIS/COFINS  |  AgriTax Audit")
    for ri, row in enumerate(rows, 3):
        # Cor por tributo
        bg = "EFF6FF" if row.get("tributo") == "PIS" else "FEF3DC"  # azul / amarelo
        for ci, (key, _, _w) in enumerate(EFD_DETAIL_COLS, 1):
            v = row.get(key, "")
            if key in EFD_MONEY_KEYS:
                v = float(v or 0)
            c = ws1.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="right" if key in EFD_MONEY_KEYS else "left",
                vertical="center")
            c.border = brd
            if key in EFD_MONEY_KEYS:
                c.number_format = '#,##0.00'
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(EFD_DETAIL_COLS))}2"

    # ── Aba 2: Resumo ──
    ws2 = wb.create_sheet("Resumo por Código")
    _hdr(ws2, 1, EFD_RESUMO_COLS,
          "EFD Contribuições — Resumo por Código de Receita  |  AgriTax Audit")
    for ri, row in enumerate(resumo_rows, 3):
        for ci, (key, _, _w) in enumerate(EFD_RESUMO_COLS, 1):
            v = row.get(key, "")
            if key in EFD_RESUMO_MONEY:
                v = float(v or 0)
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(
                horizontal="right" if (key in EFD_RESUMO_MONEY or key == "qtd_periodos") else "left",
                vertical="center")
            c.border = brd
            if key in EFD_RESUMO_MONEY:
                c.number_format = '#,##0.00'
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(EFD_RESUMO_COLS))}2"
    wb.save(path)


# =============================================================================
# Confronto EFD × DCTF + DCTFWeb
# =============================================================================

CONFRONTO_EFD_COLS = [
    ("cnpj",                "CNPJ",                       130),
    ("razao_social",        "Razão Social",               210),
    ("periodo",             "Período",                     90),
    ("competencia_teste",   "Competência Teste",          120),
    ("tributo",             "Tributo",                     90),
    ("codigo_receita",      "Cód. Receita",                90),
    ("descricao_codigo",    "Descrição",                  220),
    # Lado EFD
    ("efd_debito",          "Débito EFD",                 140),
    ("efd_recolher",        "A Recolher EFD",             140),
    # Lado DCTF + DCTFWeb
    ("dctf_debito",         "Débito DCTF",                140),
    ("dctfweb_debito",      "Débito DCTFWeb",             140),
    ("total_decl",          "Total Declarado (DCTF+Web)", 165),
    # Análise
    ("diferenca",           "Diferença (EFD − Declar.)",  170),
    ("situacao",            "Situação",                   180),
    ("obs",                 "Observação",                 280),
]
CONFRONTO_EFD_KEYS  = [c[0] for c in CONFRONTO_EFD_COLS]
CONFRONTO_EFD_MONEY = {"efd_debito", "efd_recolher", "dctf_debito",
                        "dctfweb_debito", "total_decl", "diferenca"}

# Situações do confronto EFD × DCTF/DCTFWeb
SIT_E_OK         = "✓ Conforme"                   # EFD = DCTF + DCTFWeb
SIT_E_DIVERG     = "⚠ Divergente"                  # valores diferem (>R$0,05)
SIT_E_SO_EFD     = "⚠ Só na EFD"                  # EFD declarou mas não há DCTF/DCTFWeb
SIT_E_SO_DECL    = "⚠ Só na DCTF/DCTFWeb"        # DCTF/Web declarou mas não há EFD


def run_confronto_efd_dctf(efd_rows: list, dctf_rows: list,
                             dctfweb_rows: list) -> list:
    """Confronta o que a EFD apurou com o que foi declarado em DCTF + DCTFWeb.

    Chave: (CNPJ, código de receita, período).
    """
    from collections import defaultdict

    # Indexa EFD por chave (soma quando múltiplos registros)
    efd_idx: dict = defaultdict(lambda: {"debito":0.0, "recolher":0.0, "rows":[]})
    for r in efd_rows:
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(r.get("codigo_receita", ""))
        per  = _conc_norm_periodo(r.get("periodo", ""))
        if cnpj and cod and per:
            efd_idx[(cnpj, cod, per)]["debito"]   += float(r.get("debito_apurado", 0) or 0)
            efd_idx[(cnpj, cod, per)]["recolher"] += float(r.get("contrib_a_recolher", 0) or 0)
            efd_idx[(cnpj, cod, per)]["rows"].append(r)

    # Indexa DCTF clássica
    dctf_idx: dict = defaultdict(lambda: {"debito":0.0, "rows":[]})
    for r in dctf_rows:
        if not r.get("codigo_receita", "").strip():
            continue
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(r.get("codigo_receita", ""))
        per  = _conc_norm_periodo(r.get("periodo_apuracao", ""))
        if cnpj and cod and per:
            dctf_idx[(cnpj, cod, per)]["debito"] += float(r.get("debito_apurado", 0) or 0)
            dctf_idx[(cnpj, cod, per)]["rows"].append(r)

    # Indexa DCTFWeb (usa saldo a pagar — líquido de deduções)
    dctfweb_idx: dict = defaultdict(lambda: {"debito":0.0, "rows":[]})
    for r in dctfweb_rows:
        cod_raw = r.get("codigo_receita", "").strip()
        if not cod_raw: continue
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(cod_raw)
        per  = _conc_norm_periodo(r.get("periodo", ""))
        if cnpj and cod and per:
            dctfweb_idx[(cnpj, cod, per)]["debito"] += float(r.get("saldo_pagar", 0) or 0)
            dctfweb_idx[(cnpj, cod, per)]["rows"].append(r)

    all_keys = set(efd_idx.keys()) | set(dctf_idx.keys()) | set(dctfweb_idx.keys())
    TOL = 0.05
    result = []

    for key in sorted(all_keys):
        cnpj_raw, cod, per = key
        efd  = efd_idx.get(key)
        dctf = dctf_idx.get(key)
        web  = dctfweb_idx.get(key)

        efd_debito    = efd["debito"]   if efd  else 0.0
        efd_recolher  = efd["recolher"] if efd  else 0.0
        dctf_debito   = dctf["debito"]  if dctf else 0.0
        web_debito    = web["debito"]   if web  else 0.0
        total_decl    = dctf_debito + web_debito

        # Pega cabeçalho de qualquer um dos lados
        first = (efd["rows"][0] if efd else None) or \
                (dctf["rows"][0] if dctf else None) or \
                (web["rows"][0] if web else None) or {}
        cnpj_fmt = first.get("cnpj", "")
        razao    = (first.get("razao_social", "")
                    or first.get("nome_empresarial", ""))
        tributo  = first.get("tributo", "")
        if not tributo:
            # Infere pelo código se for PIS/COFINS conhecido
            if cod in ("8109","6912","4574","1921","8496"):  tributo = "PIS"
            elif cod in ("2172","5856","5442","8645","5960"): tributo = "COFINS"
            else: tributo = "—"
        descricao = (first.get("descricao_codigo", "")
                     or EFD_CODIGO_DESC.get(cod, ""))

        # Diferença + Situação
        diferenca = round(efd_debito - total_decl, 2)
        if efd_debito > 0 and total_decl == 0:
            situacao = SIT_E_SO_EFD
        elif efd_debito == 0 and total_decl > 0:
            situacao = SIT_E_SO_DECL
        elif abs(diferenca) <= TOL:
            situacao = SIT_E_OK
        else:
            situacao = SIT_E_DIVERG

        # Observação descritiva
        partes = []
        if efd:  partes.append(f"EFD R$ {efd_debito:,.2f}")
        if dctf: partes.append(f"DCTF R$ {dctf_debito:,.2f}")
        if web:  partes.append(f"DCTFWeb R$ {web_debito:,.2f}")
        if abs(diferenca) > TOL:
            partes.append(f"Δ R$ {abs(diferenca):,.2f}")
        obs = " | ".join(partes).replace(",","X").replace(".",",").replace("X",".")

        result.append({
            "cnpj": cnpj_fmt,
            "razao_social": razao,
            "periodo": per,
            "competencia_teste": format_competencia_teste(per),
            "tributo": tributo,
            "codigo_receita": cod,
            "descricao_codigo": descricao,
            "efd_debito":      efd_debito,
            "efd_recolher":    efd_recolher,
            "dctf_debito":     dctf_debito,
            "dctfweb_debito":  web_debito,
            "total_decl":      total_decl,
            "diferenca":       diferenca,
            "situacao":        situacao,
            "obs":             obs,
        })

    # Ordena: divergentes primeiro
    _ord = {SIT_E_DIVERG:0, SIT_E_SO_EFD:1, SIT_E_SO_DECL:2, SIT_E_OK:3}
    result.sort(key=lambda r: (_ord.get(r["situacao"],9),
                                 r["cnpj"], r["competencia_teste"],
                                 r["codigo_receita"]))
    return result


# ── Interface Grafica ─────────────────────────────────────────────────────────
class DctfApp(tk.Toplevel):
    """Modulo DCTF Extractor v3.1 - 100% local, sem API."""

    def __init__(self, master):
        super().__init__(master)
        self.title("DCTF Extractor v3.1 - AgriTax Audit  |  100% Local")
        self.configure(bg=C_GRAY_LIGHT)
        self.geometry("1350x780")
        self.minsize(900, 600)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.rows  = []
        self.files = []
        self._processing = False
        self._logo_img   = None
        # Filtros estilo Excel por coluna (por aba)
        self.col_filt_1: dict = {}   # Detalhamento
        self.col_filt_2: dict = {}   # Resumo por Tributo
        self._check_deps()
        self._build_ui()

    def _check_deps(self):
        """
        Verifica dependências Python e o ambiente OCR uma única vez, ao abrir
        o DCTF Extractor. Mostra um único diálogo consolidado (com diagnóstico
        completo) em vez de repetir a mensagem por PDF processado.
        """
        missing = []
        if not PDFPLUMBER_OK:   missing.append("pdfplumber")
        if not PYTESSERACT_OK:  missing.append("pytesseract")
        if not OPENPYXL_OK:     missing.append("openpyxl")

        # Diagnóstico do ambiente OCR (inclui detecção do Tesseract fora do PATH)
        env = _get_ocr_env()

        msgs = []
        if missing:
            msgs.append(
                "Bibliotecas Python ausentes: " + ", ".join(missing) + "\n"
                "Instale com: pip install " + " ".join(missing))
        if env["pytesseract_ok"] and not env["tesseract_exe"]:
            msgs.append(
                "Tesseract OCR não foi encontrado no sistema. Sem ele, PDFs "
                "da DCTF do eCAC (que são rasterizados) não poderão ser lidos.\n"
                "Baixe: https://github.com/UB-Mannheim/tesseract/wiki\n"
                "Durante a instalação, marque 'Add to PATH' e 'Portuguese'.")
        elif env["tesseract_exe"] and not env["has_por"]:
            msgs.append(
                f"Tesseract OCR {env['tesseract_ver'] or ''} detectado em:\n"
                f"  {env['tesseract_exe']}\n\n"
                "⚠ O idioma Português não está instalado — a acurácia com "
                "acentos pode ser reduzida. Reinstale o Tesseract marcando "
                "'Additional language data → Portuguese'.")

        if msgs:
            messagebox.showwarning(
                "DCTF Extractor — Verificação de Ambiente",
                "\n\n".join(msgs),
                parent=self)

    def _build_ui(self):
        # Barra de topo
        top = tk.Frame(self, bg=C_GREEN_DARK, height=56)
        top.pack(fill="x"); top.pack_propagate(False)
        try:
            import base64 as _b64, io as _io
            from PIL import Image as _PILImg, ImageTk as _ITk
            _img = _PILImg.open(_io.BytesIO(_b64.b64decode(LOGO_B64))).resize((42,42), _PILImg.LANCZOS)
            self._logo_img = _ITk.PhotoImage(_img)
            tk.Label(top, image=self._logo_img, bg=C_GREEN_DARK).pack(side="left", padx=12)
        except Exception:
            pass
        tk.Label(top, text="AgriTax",
                 font=("Segoe UI",16,"bold"), fg=C_WHITE,
                 bg=C_GREEN_DARK).pack(side="left", pady=10)
        tk.Label(top, text="  DCTF Extractor  v3.1",
                 font=("Segoe UI",11), fg=C_GREEN_MID,
                 bg=C_GREEN_DARK).pack(side="left", pady=10)
        tk.Label(top, text="  100% Local - Sem API - Sem Custo",
                 font=("Segoe UI",8), fg="#8fb84a",
                 bg=C_GREEN_DARK).pack(side="left", pady=10)

        body = tk.Frame(self, bg=C_GRAY_LIGHT)
        body.pack(fill="both", expand=True, padx=10, pady=8)

        left = tk.Frame(body, bg=C_WHITE, width=240,
                        highlightbackground=C_BORDER, highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0,8))
        left.pack_propagate(False)
        self._build_left(left)

        right = tk.Frame(body, bg=C_GRAY_LIGHT)
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

        sb = tk.Frame(self, bg=C_GREEN_DARK, height=24)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Pronto. Adicione PDFs de DCTF do eCAC.")
        tk.Label(sb, textvariable=self.status_var,
                 font=("Segoe UI",8), fg=C_GREEN_LIGHT,
                 bg=C_GREEN_DARK, anchor="w").pack(fill="x", padx=12, pady=3)

    def _build_left(self, parent):
        tk.Label(parent, text="Arquivos PDF",
                 font=("Segoe UI",10,"bold"), fg=C_GREEN_DARK,
                 bg=C_WHITE).pack(anchor="w", padx=12, pady=(12,4))

        lb_f = tk.Frame(parent, bg=C_WHITE)
        lb_f.pack(fill="both", expand=True, padx=8, pady=(0,6))
        vsb = ttk.Scrollbar(lb_f, orient="vertical")
        vsb.pack(side="right", fill="y")
        self.file_lb = tk.Listbox(lb_f, yscrollcommand=vsb.set,
                                   font=("Segoe UI",8), bg=C_GRAY_LIGHT,
                                   fg=C_GRAY_DARK, selectbackground=C_GREEN_MID,
                                   relief="flat", borderwidth=0)
        self.file_lb.pack(fill="both", expand=True)
        vsb.config(command=self.file_lb.yview)

        sf = tk.Frame(parent, bg=C_GREEN_LIGHT,
                      highlightbackground=C_BORDER, highlightthickness=1)
        sf.pack(fill="x", padx=8, pady=(0,8))
        self.lbl_stats = tk.Label(sf,
            text="Processados : 0\nTotal linhas : 0\nTributos     : 0\nNa fila      : 0",
            font=("Courier",8), fg=C_GREEN_DARK, bg=C_GREEN_LIGHT, justify="left")
        self.lbl_stats.pack(anchor="w", padx=10, pady=6)

        cfg = dict(font=("Segoe UI",9,"bold"), relief="flat", cursor="hand2", pady=5)
        self.btn_add = tk.Button(parent, text="Adicionar PDFs",
                                  bg=C_GREEN, fg=C_WHITE,
                                  command=self._add_files, **cfg)
        self.btn_add.pack(fill="x", padx=8, pady=(0,4))
        self.btn_proc = tk.Button(parent, text="Processar",
                                   bg=C_GREEN_DARK, fg=C_WHITE,
                                   command=self._process_files,
                                   state="disabled", **cfg)
        self.btn_proc.pack(fill="x", padx=8, pady=(0,4))
        self.btn_export = tk.Button(parent, text="Exportar Excel",
                                     bg=C_YELLOW, fg=C_WHITE,
                                     command=self._export_excel,
                                     state="disabled", **cfg)
        self.btn_export.pack(fill="x", padx=8, pady=(0,4))
        self.btn_clear = tk.Button(parent, text="Limpar Tudo",
                                    bg=C_RED, fg=C_WHITE,
                                    command=self._clear,
                                    state="disabled", **cfg)
        self.btn_clear.pack(fill="x", padx=8, pady=(0,10))
        self.progress = ttk.Progressbar(parent, mode="indeterminate", length=220)
        self.progress.pack(fill="x", padx=8, pady=(0,8))

    def _build_right(self, parent):
        nb = ttk.Notebook(parent)
        nb.pack(fill="both", expand=True)
        t1 = tk.Frame(nb, bg=C_WHITE); nb.add(t1, text="  Detalhamento  ")
        self.tree1 = self._make_tree(t1, DCTF_DETAIL_COLS)
        self._wire_dctf_filters(self.tree1, DCTF_DETAIL_COLS, self.col_filt_1, self._refresh_dctf)
        t2 = tk.Frame(nb, bg=C_WHITE); nb.add(t2, text="  Resumo por Tributo  ")
        self.tree2 = self._make_tree(t2, DCTF_RESUMO_COLS)
        self._wire_dctf_filters(self.tree2, DCTF_RESUMO_COLS, self.col_filt_2, self._refresh_dctf)

    def _wire_dctf_filters(self, tree, cols, col_filt, refresh_cb):
        """Liga os cabeçalhos da Treeview ao ColFilterPopup (com busca substring)."""
        # Armazena referência cruzada para permitir update de ícones
        if not hasattr(self, "_dctf_trees"):
            self._dctf_trees = []
        self._dctf_trees.append((tree, cols, col_filt))

        for key, label, _w in cols:
            def _make_cmd(tr=tree, k=key, l=label, cf=col_filt, rcb=refresh_cb):
                def cmd():
                    # Conjunto de valores únicos da coluna a partir das linhas atuais
                    src = self.rows
                    if tr is self.tree2:
                        src = build_dctf_resumo(self.rows)
                    vals = set()
                    for r in src:
                        v = r.get(k, "")
                        if k in DCTF_MONEY_KEYS:
                            try: v = self._fmt(v)
                            except Exception: v = str(v)
                        else:
                            v = str(v) if v is not None else ""
                        if v:
                            vals.add(v)
                    active = cf.get(k)
                    px = self.winfo_pointerx(); py = self.winfo_pointery()
                    def on_apply(sel):
                        if sel is None: cf.pop(k, None)
                        else:           cf[k] = sel
                        rcb()
                        self._update_dctf_headings()
                    ColFilterPopup(self, l, vals, active, on_apply, px, py+4)
                return cmd
            tree.heading(key, text=f"{label}  ▼", command=_make_cmd())

    def _update_dctf_headings(self):
        """Atualiza ícone ▼ / 🔽 dos cabeçalhos conforme filtro ativo."""
        for tree, cols, col_filt in getattr(self, "_dctf_trees", []):
            for key, label, _w in cols:
                icon = " 🔽" if key in col_filt else "  ▼"
                tree.heading(key, text=f"{label}{icon}")

    @staticmethod
    def _dctf_col_match(col_filt, row, money_keys):
        """Retorna True se row atende TODOS os filtros acumulativos."""
        def _norm(v, key):
            if key in money_keys:
                try: return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
                except Exception: return str(v)
            return str(v) if v is not None else ""
        for k, vs in col_filt.items():
            if not vs: continue
            if _norm(row.get(k, ""), k) not in vs:
                return False
        return True

    def _make_tree(self, parent, cols):
        frm = tk.Frame(parent, bg=C_WHITE)
        frm.pack(fill="both", expand=True, padx=4, pady=4)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [k for k,_,_ in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
                             yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for key, label, width in cols:
            tree.heading(key, text=label)
            tree.column(key, width=width, minwidth=50,
                        anchor="e" if key in DCTF_MONEY_KEYS else "w")
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        tree.tag_configure("even",  background="#F2F4F0")
        tree.tag_configure("odd",   background="#FFFFFF")
        tree.tag_configure("alert", background=C_RED_LIGHT)
        return tree

    def _fmt(self, v) -> str:
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v)

    def _add_files(self):
        paths = filedialog.askopenfilenames(parent=self,
            title="Selecione PDFs de DCTF",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        new = [p for p in paths if p not in self.files]
        self.files.extend(new)
        for p in new:
            self.file_lb.insert("end", Path(p).name)
        if self.files:
            self.btn_proc.config(state="normal")
            self.btn_clear.config(state="normal")
        self._update_stats()
        self._set_status(f"{len(new)} arquivo(s) adicionado(s). Total na fila: {len(self.files)}.")

    def _process_files(self):
        if self._processing or not self.files:
            return
        self._processing = True
        self.btn_proc.config(state="disabled")
        self.btn_add.config(state="disabled")
        self.progress.start(12)
        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        """Processa DCTFs em série (modo seguro).

        Paralelização com ThreadPoolExecutor causou crash do aplicativo no Windows
        (Tesseract / Poppler / pdfplumber.to_image não são 100% thread-safe).
        Este modo processa 1 PDF por vez, mas:
          - Salva log detalhado em %TEMP%\\agritax_dctf_log.txt
          - Um PDF problemático não derruba o processamento dos demais
          - OCR em 200 DPI já deixa o processamento razoavelmente rápido
        """
        import tempfile, traceback as _tb
        errors = []
        pdfs = list(self.files)
        total = len(pdfs)

        log_path = Path(tempfile.gettempdir()) / "agritax_dctf_log.txt"
        with open(log_path, "w", encoding="utf-8") as logf:
            logf.write(f"=== AgriTax DCTF Extractor — Log ===\n")
            logf.write(f"Início: {datetime.now().isoformat()}\n")
            logf.write(f"Total PDFs: {total}\n\n")
            logf.flush()

            for i, pdf in enumerate(pdfs, 1):
                pdf_name = Path(pdf).name
                logf.write(f"[{i}/{total}] INICIANDO: {pdf_name}\n")
                logf.flush()
                self.after(0, self._set_status, f"DCTF {i}/{total} — {pdf_name}")
                try:
                    new_rows = extract_dctf(pdf)
                    self.rows.extend(new_rows)
                    try:
                        self.files.remove(pdf)
                    except ValueError:
                        pass
                    logf.write(f"[{i}/{total}] OK: {pdf_name} — {len(new_rows)} tributo(s)\n")
                    self.after(0, self._refresh_trees)
                    self.after(0, self._update_stats)
                except Exception as e:
                    msg = f"{type(e).__name__}: {e}"
                    errors.append(f"{pdf_name}:\n{msg}")
                    logf.write(f"[{i}/{total}] ERRO: {pdf_name} — {msg}\n")
                    logf.write(_tb.format_exc() + "\n")
                logf.flush()

            logf.write(f"\nFim: {datetime.now().isoformat()}\n")
            logf.write(f"Total de linhas extraídas: {len(self.rows)}\n")
            logf.write(f"Erros: {len(errors)}\n")

        self._dctf_log_path = str(log_path)
        self.after(0, self._process_done, errors)

    def _process_done(self, errors):
        self._processing = False
        self.progress.stop()
        self.btn_add.config(state="normal")
        self.btn_proc.config(state="normal" if self.files else "disabled")
        if self.rows:
            self.btn_export.config(state="normal")
        if errors:
            # Deduplica: se todos os erros são da mesma causa (OCR indisponível
            # / texto inextraível), exibe UMA mensagem consolidada em vez de
            # repetir o diagnóstico para cada PDF.
            ocr_failures = [e for e in errors if "extrair texto deste PDF" in e]
            other_errs   = [e for e in errors if "extrair texto deste PDF" not in e]
            partes = []
            if ocr_failures:
                nomes = [e.split(":\n",1)[0] for e in ocr_failures]
                diag  = ocr_failures[0].split(":\n",1)[1] if ":\n" in ocr_failures[0] else ocr_failures[0]
                partes.append(
                    f"Falha ao extrair texto de {len(ocr_failures)} PDF(s):\n"
                    + "  • " + "\n  • ".join(nomes)
                    + "\n\n" + diag)
            if other_errs:
                partes.append("Outros erros:\n\n" + "\n\n".join(other_errs))
            messagebox.showwarning(
                f"Erros no processamento ({len(errors)} arquivo(s))",
                "\n\n═══════════════\n\n".join(partes), parent=self)
        srcs = len({r.get("_source") for r in self.rows})
        log_info = ""
        if hasattr(self, "_dctf_log_path"):
            log_info = f"  |  Log: {self._dctf_log_path}"
        self._set_status(
            f"Concluido. {len(self.rows)} linha(s) extraida(s) de {srcs} arquivo(s).{log_info}")
        self._refresh_trees()

    def _refresh_trees(self):
        for i in self.tree1.get_children(): self.tree1.delete(i)
        cf1 = self.col_filt_1
        idx_shown = 0
        for row in self.rows:
            if cf1 and not self._dctf_col_match(cf1, row, DCTF_MONEY_KEYS):
                continue
            vals = [self._fmt(row.get(k,"")) if k in DCTF_MONEY_KEYS
                    else row.get(k,"") for k,_,_ in DCTF_DETAIL_COLS]
            saldo = float(row.get("saldo_pagar", 0) or 0)
            tag = "alert" if saldo > 0 else ("even" if idx_shown%2==0 else "odd")
            self.tree1.insert("","end", values=vals, tags=(tag,))
            idx_shown += 1
        for i in self.tree2.get_children(): self.tree2.delete(i)
        cf2 = self.col_filt_2
        idx_shown = 0
        for row in build_dctf_resumo(self.rows):
            if cf2 and not self._dctf_col_match(cf2, row, DCTF_MONEY_KEYS):
                continue
            vals = [self._fmt(row.get(k,"")) if k in DCTF_MONEY_KEYS
                    else row.get(k,"") for k,_,_ in DCTF_RESUMO_COLS]
            self.tree2.insert("","end", values=vals,
                              tags=("even" if idx_shown%2==0 else "odd",))
            idx_shown += 1

    # Alias para o callback usado em _wire_dctf_filters
    _refresh_dctf = _refresh_trees

    def _export_excel(self):
        if not self.rows:
            messagebox.showinfo("Sem dados","Processe os PDFs primeiro.", parent=self)
            return
        path = filedialog.asksaveasfilename(parent=self,
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"DCTF_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path:
            return
        try:
            export_dctf_excel(self.rows, build_dctf_resumo(self.rows), path)
            self._set_status(f"Excel exportado: {Path(path).name}")
            if messagebox.askyesno("Exportado!",
                    f"Arquivo salvo:\n{path}\n\nDeseja abrir agora?", parent=self):
                import subprocess, platform
                if platform.system() == "Windows":
                    os.startfile(path)
                elif platform.system() == "Darwin":
                    subprocess.call(["open", path])
                else:
                    subprocess.call(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e), parent=self)

    def _clear(self):
        if messagebox.askyesno("Limpar","Remover todos os dados?", parent=self):
            self.rows.clear(); self.files.clear()
            self.file_lb.delete(0, "end")
            for t in [self.tree1, self.tree2]:
                for i in t.get_children(): t.delete(i)
            for b in [self.btn_export, self.btn_clear, self.btn_proc]:
                b.config(state="disabled")
            self._update_stats(); self._set_status("Dados limpos.")

    def _set_status(self, msg: str):
        self.status_var.set(msg)
        try:
            self.update_idletasks()
        except Exception:
            pass

    def _update_stats(self):
        srcs = len({r.get("_source") for r in self.rows if r.get("_source")})
        self.lbl_stats.config(text=(
            f"Processados : {srcs}\n"
            f"Total linhas : {len(self.rows)}\n"
            f"Tributos     : {len(self.rows)}\n"
            f"Na fila      : {len(self.files)}"
        ))




# =============================================================================
# LauncherApp — Tela principal do AgriTax Audit
# =============================================================================

class LauncherApp:
    """
    Tela principal do AgriTax Audit — layout split panel (Opção 2).
    Painel esquerdo: identidade visual (logo + branding).
    Painel direito: lista de módulos — cresce conforme novos módulos são adicionados.
    """

    # Definição dos módulos disponíveis — adicione aqui para incluir novos itens
    MODULES = [
        {
            "emoji": "📄",
            "title": "PERDCOMP Extractor",
            "subtitle": "Restituição, ressarcimento e compensação · eCAC",
            "tag": "perdcomp",
        },
        {
            "emoji": "🧾",
            "title": "DARF Extractor",
            "subtitle": "Comprovantes de arrecadação DARF e DAS",
            "tag": "darf",
        },
        {
            "emoji": "📊",
            "title": "DCTF Extractor",
            "subtitle": "DCTF eCAC - 100% Local, Sem API, Sem Custo",
            "tag": "dctf",
        },
        {
            "emoji": "🔍",
            "title": "Conciliação DARF × DCOMP",
            "subtitle": "Cruzamento de débitos pagos com compensações via PERDCOMP",
            "tag": "conciliacao",
        },
    ]

    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("AgriTax Audit  v6.0")
        root.resizable(False, True)   # altura flexível para futuras adições
        root.configure(bg=C_GREEN_DARK)
        self._logo_img = None
        self._build()

    def _build(self):
        root = self.root

        # ── Container principal (split) ────────────────────────────────────
        container = tk.Frame(root, bg=C_GREEN_DARK)
        container.pack(fill="both", expand=True)

        # ── PAINEL ESQUERDO — identidade visual ────────────────────────────
        LEFT_W = 220
        left = tk.Frame(container, bg=C_GREEN_DARK, width=LEFT_W)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)

        # Espaçador superior
        tk.Frame(left, bg=C_GREEN_DARK, height=40).pack()

        # Logo
        try:
            import base64, io
            from PIL import Image, ImageTk
            img_data = base64.b64decode(LOGO_B64)
            img = Image.open(io.BytesIO(img_data)).resize((80, 80), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            logo_frame = tk.Frame(left, bg="#2D5516", width=90, height=90,
                                  highlightthickness=2,
                                  highlightbackground="rgba(255,255,255,0)") 
            # Usa um frame arredondado simulado com cor levemente diferente
            logo_lbl = tk.Label(left, image=self._logo_img,
                                bg=C_GREEN_DARK, cursor="arrow")
            logo_lbl.pack(pady=(0, 14))
        except Exception:
            tk.Label(left, text="🌿", bg=C_GREEN_DARK,
                     font=("Segoe UI", 48)).pack(pady=(0, 14))

        # Linha divisória sutil
        tk.Frame(left, bg="#4A7018", height=1, width=80).pack(pady=(0, 14))

        # Nome do sistema
        tk.Label(left, text="AgriTax", bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 20, "bold")).pack()
        tk.Label(left, text="Audit", bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 20, "bold")).pack()

        tk.Frame(left, bg=C_GREEN_DARK, height=8).pack()

        tk.Label(left, text="Auditoria\nTributária", bg=C_GREEN_DARK,
                 fg="#9DCF5C", font=("Segoe UI", 9), justify="center").pack()

        tk.Frame(left, bg=C_GREEN_DARK, height=16).pack()

        # Badge versão
        ver_frame = tk.Frame(left, bg="#142A0B",
                             highlightthickness=1,
                             highlightbackground="#3A6B0E")
        ver_frame.pack(padx=30, ipadx=10, ipady=3)
        tk.Label(ver_frame, text="v 6.0", bg="#142A0B",
                 fg="#7AB82E", font=("Segoe UI", 9, "bold")).pack()

        # Preenchimento restante + rodapé no painel esquerdo
        spacer = tk.Frame(left, bg=C_GREEN_DARK)
        spacer.pack(fill="both", expand=True)
        tk.Label(left, text="© AgriTax", bg=C_GREEN_DARK,
                 fg="#4A7018", font=("Segoe UI", 8)).pack(pady=(0, 12))

        # ── SEPARADOR VERTICAL ─────────────────────────────────────────────
        tk.Frame(container, bg="#2D5516", width=1).pack(side="left", fill="y")

        # ── PAINEL DIREITO — lista de módulos ──────────────────────────────
        right = tk.Frame(container, bg=C_WHITE)
        right.pack(side="left", fill="both", expand=True)

        # Cabeçalho do painel direito
        rhdr = tk.Frame(right, bg=C_WHITE)
        rhdr.pack(fill="x", padx=28, pady=(28, 4))
        tk.Label(rhdr, text="Selecione o módulo",
                 bg=C_WHITE, fg=C_GRAY_DARK,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Label(rhdr, text="Clique em um módulo para abri-lo em nova janela",
                 bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI", 8)).pack(anchor="w")

        # Separador
        tk.Frame(right, bg="#E8EDE4", height=1).pack(fill="x", padx=28, pady=(8, 12))

        # Área de lista (scrollable futuramente)
        list_frame = tk.Frame(right, bg=C_WHITE)
        list_frame.pack(fill="both", expand=True, padx=18)

        for mod in self.MODULES:
            self._make_module_row(list_frame, mod)

        # Rodapé direito
        tk.Frame(right, bg="#E8EDE4", height=1).pack(fill="x", padx=28, pady=(12, 0))
        tk.Label(right, text="AgriTax Audit  v6.0  |  Plataforma de Auditoria Tributária",
                 bg=C_WHITE, fg="#BBBBBB",
                 font=("Segoe UI", 8)).pack(pady=(6, 12))

    def _make_module_row(self, parent, mod: dict):
        """Cria uma linha de módulo clicável no painel direito."""
        tag = mod["tag"]
        cmd = getattr(self, f"_open_{tag}", None)
        if cmd is None:
            return

        # Cores do ícone por módulo
        ico_colors = {
            "perdcomp": ("#EAF4D3", "#3A6B0E"),
            "darf":     ("#E0F4F0", "#0E5A4F"),
        }
        ico_bg, ico_fg_unused = ico_colors.get(tag, ("#F2F4F0", "#555"))

        row = tk.Frame(parent, bg=C_WHITE,
                       highlightthickness=1,
                       highlightbackground="#E8EDE4",
                       cursor="hand2")
        row.pack(fill="x", pady=5, ipady=2)

        # Ícone
        ico_box = tk.Frame(row, bg=ico_bg, width=46, height=46)
        ico_box.pack(side="left", padx=(12, 0), pady=10)
        ico_box.pack_propagate(False)
        tk.Label(ico_box, text=mod["emoji"], bg=ico_bg,
                 font=("Segoe UI", 20)).place(relx=0.5, rely=0.5, anchor="center")

        # Texto
        txt = tk.Frame(row, bg=C_WHITE)
        txt.pack(side="left", fill="x", expand=True, padx=14, pady=10)
        tk.Label(txt, text=mod["title"], bg=C_WHITE, fg=C_GREEN_DARK,
                 font=("Segoe UI", 11, "bold"), anchor="w").pack(fill="x")
        tk.Label(txt, text=mod["subtitle"], bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI", 8), anchor="w").pack(fill="x")

        # Seta
        arrow = tk.Label(row, text="›", bg=C_WHITE, fg="#BBBBBB",
                         font=("Segoe UI", 20, "bold"))
        arrow.pack(side="right", padx=16)

        # Hover — destaca a linha e a seta
        def on_enter(e):
            for w in (row, txt) + tuple(txt.winfo_children()) + (arrow,):
                try: w.config(bg="#F6FBF0")
                except: pass
            ico_box.config(bg=ico_colors.get(tag, ("#F2F4F0","#555"))[0])
            arrow.config(fg=C_GREEN, bg="#F6FBF0")
            row.config(highlightbackground=C_GREEN_MID)

        def on_leave(e):
            for w in (row, txt) + tuple(txt.winfo_children()) + (arrow,):
                try: w.config(bg=C_WHITE)
                except: pass
            ico_box.config(bg=ico_colors.get(tag, ("#F2F4F0","#555"))[0])
            arrow.config(fg="#BBBBBB", bg=C_WHITE)
            row.config(highlightbackground="#E8EDE4")

        def on_click(e):
            cmd()

        for w in (row, ico_box, txt, arrow) + tuple(txt.winfo_children()):
            try:
                w.bind("<Enter>", on_enter)
                w.bind("<Leave>", on_leave)
                w.bind("<Button-1>", on_click)
            except Exception:
                pass

    def _open_perdcomp(self):
        win = tk.Toplevel(self.root)
        App(win)

    def _open_darf(self):
        DarfApp(self.root)

    def _open_dctf(self):
        DctfApp(self.root)

    def _open_conciliacao(self):
        ConciliacaoApp(self.root)


# =============================================================================
# ConciliacaoApp — Cruzamento DARF × DCOMP
# =============================================================================

# Colunas do resultado da conciliação
CONC_COLS = [
    # Identificação
    ("cnpj",              "CNPJ",                    130),
    ("razao_social",      "Razão Social",             210),
    ("codigo_receita",    "Cód. Receita",              80),
    ("descricao_tributo", "Descrição Tributo",         220),
    ("periodo",           "Período de Apuração",       130),
    # Lado DARF
    ("darf_numero",       "Nº Documento DARF",        170),
    ("darf_dt_arrec",     "Dt. Arrecadação",          100),
    ("darf_banco",        "Banco (DARF)",             180),
    ("darf_principal",    "Vl. Principal DARF",       125),
    ("darf_multa",        "Vl. Multa DARF",           100),
    ("darf_juros",        "Vl. Juros DARF",           100),
    ("darf_total",        "Vl. Total DARF",           125),
    # Lado DCOMP
    ("dcomp_numero",      "Nº PERDCOMP",              175),
    ("dcomp_tipo_pedido", "Tipo Pedido",              160),
    ("dcomp_dt_transm",   "Dt. Transmissão",          100),
    ("dcomp_principal",   "Vl. Comp. Principal",      125),
    ("dcomp_multa",       "Vl. Comp. Multa",          100),
    ("dcomp_juros",       "Vl. Comp. Juros",          100),
    ("dcomp_total",       "Vl. Total DCOMP",          125),
    # Análise
    ("situacao",          "Situação",                 160),
    ("diferenca",         "Diferença (DARF − DCOMP)", 160),
    ("obs",               "Observação",               260),
]
CONC_KEYS  = [k for k,_,_ in CONC_COLS]
CONC_MONEY = {"darf_principal","darf_multa","darf_juros","darf_total",
               "dcomp_principal","dcomp_multa","dcomp_juros","dcomp_total","diferenca"}

# Situações possíveis
SIT_DUPLO      = "⚠ Duplo Pagamento"   # DARF pago E DCOMP compensou o mesmo débito
SIT_SO_DARF    = "✓ Só DARF"           # débito quitado apenas via DARF
SIT_SO_DCOMP   = "✓ Só DCOMP"          # débito quitado apenas via compensação
SIT_DIVERGENTE = "⚠ Valores Divergentes"  # ambos existem mas valores diferem
SIT_OK         = "✓ Conciliado"        # DARF + DCOMP somam o mesmo valor declarado

# ─────────────────────────────────────────────────────────────────────────────
# Colunas da aba TRIPLO (DCTF × DARF × DCOMP)
# ─────────────────────────────────────────────────────────────────────────────
TRIPLO_COLS = [
    # Identificação
    ("cnpj",              "CNPJ",                     130),
    ("razao_social",      "Razão Social",             210),
    ("codigo_receita",    "Cód. Receita",              80),
    ("grupo_tributo",     "Grupo do Tributo",         220),
    ("periodo",           "Período",                  100),
    ("competencia_teste", "Competência Teste",        120),
    # Lado DCTF clássica — só o que a empresa DECLAROU devendo
    ("dctf_num_decl",     "Nº Declaração DCTF",       190),
    ("dctf_periodicidade","Periodicidade",             90),
    ("dctf_debito",       "Débito Apurado (DCTF)",    140),
    # Lado DCTFWeb — débitos previdenciários (CP) + IRRF folha
    ("dctfweb_num_recibo","Nº Recibo DCTFWeb",        160),
    ("dctfweb_categoria", "Categoria DCTFWeb",        130),
    ("dctfweb_debito",    "Débito a Pagar (DCTFWeb)", 145),
    # Total do declarado (DCTF + DCTFWeb)
    ("total_declarado",   "Total Declarado",          135),
    # Lado DARF (o que foi PAGO de fato — baixa, somente principal)
    ("darf_numero",       "Nº Documento DARF",        170),
    ("darf_dt_arrec",     "Dt. Arrecadação",          100),
    ("darf_total",        "Princ. Pago DARF",         130),
    # Lado DCOMP ATIVA (sem canceladas/retificadas — baixa por compensação, principal)
    ("dcomp_numero",      "Nº PERDCOMP Ativa",        175),
    ("dcomp_dt_transm",   "Dt. Transmissão",          100),
    ("dcomp_total",       "Princ. Compensado DCOMP",  150),
    # Análise contábil — Saldo = Total Declarado − (DARF + DCOMP ativa)
    ("saldo_final",       "Saldo Final",              130),
    ("situacao_triplo",   "Situação",                 180),
    ("obs",               "Observação",               260),
]
TRIPLO_KEYS  = [k for k,_,_ in TRIPLO_COLS]
TRIPLO_MONEY = {"dctf_debito", "dctfweb_debito", "total_declarado",
                 "darf_total", "dcomp_total", "saldo_final"}

# Situações da aba triplo (contábil)
SIT_T_QUITADO   = "✓ Quitado"                        # total declarado = DARF + DCOMP
SIT_T_SALDO     = "⚠ Saldo a Pagar"                  # total declarado > DARF + DCOMP
SIT_T_A_MAIOR   = "⚠ Pagto/Compens. a Maior"         # total declarado < DARF + DCOMP
SIT_T_SEM_DECL  = "⚠ Sem Declaração"                 # há DARF/DCOMP mas sem DCTF nem DCTFWeb


def _conc_brl(v) -> float:
    """Converte string BRL ou float → float."""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _conc_norm_periodo(p: str) -> str:
    """
    Normaliza período para comparação canônica no formato 'MM/AAAA' (ou 'TN/AAAA' para
    créditos trimestrais, 'AAAA' para anuais).

    Formatos de entrada suportados e resultado:
      '28/02/2026'            → '02/2026'   (DARF: DD/MM/AAAA — descarta o dia)
      '02/2026'               → '02/2026'   (DAS: MM/AAAA já canônico)
      'Fevereiro de 2026'     → '02/2026'   (DCOMP débito: "Mês de AAAA")
      'Fevereiro/2026'        → '02/2026'
      '1º Trimestre/2024'     → 'T1/2024'   (PERDCOMP crédito trimestral)
      '3º Trimestre 2024'     → 'T3/2024'
      '3º Trimestre'          → 'T3'        (sem ano — raro)
      'Trimestre 3/2024'      → 'T3/2024'
      '2024'                  → '2024'      (crédito anual)

    Retorna string vazia se `p` for vazio/None.
    """
    import re as _re
    if not p:
        return ""

    s = str(p).strip().upper()
    if not s:
        return ""

    # Remove a preposição "DE" solta entre palavras (ex.: "FEVEREIRO DE 2026")
    s = _re.sub(r"\s+DE\s+", " ", s)
    # Unifica separadores: barras, traços, pontos → espaço temporário
    s_clean = s

    # ── Caso 1: DD/MM/AAAA (DARF) → MM/AAAA ───────────────────────────────────
    m = _re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s_clean)
    if m:
        return f"{int(m.group(2)):02d}/{m.group(3)}"

    # ── Caso 2: MM/AAAA já canônico (DAS) ─────────────────────────────────────
    m = _re.fullmatch(r"(\d{1,2})/(\d{4})", s_clean)
    if m:
        return f"{int(m.group(1)):02d}/{m.group(2)}"

    # ── Caso 3: Nome do mês por extenso ───────────────────────────────────────
    MESES = {
        "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "MARCO": "03",
        "ABRIL":   "04", "MAIO":      "05", "JUNHO":  "06",
        "JULHO":   "07", "AGOSTO":    "08", "SETEMBRO": "09",
        "OUTUBRO": "10", "NOVEMBRO":  "11", "DEZEMBRO": "12",
    }
    for nome, num in MESES.items():
        if nome in s_clean:
            m_ano = _re.search(r"(\d{4})", s_clean)
            if m_ano:
                return f"{num}/{m_ano.group(1)}"
            return num  # sem ano (muito raro)

    # ── Caso 3b: Abreviação do mês (DCTF diário/decendial/quinzenal/semanal) ──
    # "21° DIA/DEZ/2023", "3° DECENDIO/DEZ/2023", "2° QUINZENA/JAN/2024", etc.
    MESES_ABREV = {
        "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04", "MAI": "05", "JUN": "06",
        "JUL": "07", "AGO": "08", "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12",
    }
    for ab, num in MESES_ABREV.items():
        # O mês abreviado precisa ser uma palavra inteira (boundary)
        if _re.search(rf"\b{ab}\b", s_clean):
            m_ano = _re.search(r"(\d{4})", s_clean)
            if m_ano:
                return f"{num}/{m_ano.group(1)}"
            return num

    # ── Caso 4: Trimestre ─────────────────────────────────────────────────────
    # "1º TRIMESTRE/2024", "3º TRIMESTRE 2024", "3 TRIMESTRE 2024"
    m = _re.search(r"(\d)[ºO°]?\s*TRIMESTRE\s*[/\s]*(\d{4})", s_clean)
    if m:
        return f"T{m.group(1)}/{m.group(2)}"
    # "TRIMESTRE 3/2024" ou "TRIMESTRE 3 2024"
    m = _re.search(r"TRIMESTRE\s+(\d)\s*[/\s]*(\d{4})", s_clean)
    if m:
        return f"T{m.group(1)}/{m.group(2)}"
    # Sem ano: "3º TRIMESTRE"
    m = _re.search(r"(\d)[ºO°]?\s*TRIMESTRE\b", s_clean)
    if m:
        return f"T{m.group(1)}"

    # ── Caso 5: apenas o ano ──────────────────────────────────────────────────
    m = _re.fullmatch(r"(\d{4})", s_clean)
    if m:
        return s_clean

    # Fallback — retorna a string limpa
    return s_clean


def _conc_norm_codigo(s: str) -> str:
    """
    Extrai apenas os 4 dígitos do código de receita para comparação.
    '8109 - PIS...' → '8109'  |  '8109-01' → '8109'
    """
    m = re.search(r"(\d{4})", str(s))
    return m.group(1) if m else s.strip()


def _conc_extract_codigo_from_tipo(tipo_debito: str) -> str:
    """
    Extrai código de receita de strings como:
    'COFINS - Contribuição... | 2172 - COFINS...'
    'PIS/PASEP - ... | 8109 - PIS...'
    '0561 - IRRF...'
    """
    # Tenta padrão "NNNN - " (código de 4 dígitos)
    m = re.search(r"\|\s*(\d{4})\b", tipo_debito)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4})\s*[-–]", tipo_debito)
    if m:
        return m.group(1)
    return ""


def run_conciliacao(darf_rows: list, dcomp_rows: list,
                     status_map: dict = None) -> list:
    """
    Cruza linhas do DARF Extractor com linhas do PERDCOMP Extractor (apenas DCOMPs).

    Chave de cruzamento: (CNPJ normalizado, código receita 4 dígitos, período normalizado)

    Se `status_map` for fornecido (planilha de status do eCAC), ignora DCOMPs
    com situação "Cancelado" ou "Retificado" — essas compensações foram
    revogadas pelo contribuinte e não devem aparecer no confronto.

    Retorna lista de dicts para exibição e exportação.
    """
    from collections import defaultdict

    if status_map is None:
        status_map = {}

    # ── Indexa DARFs por chave ─────────────────────────────────────────────
    darf_idx: dict = defaultdict(list)
    for r in darf_rows:
        cnpj  = re.sub(r"\D", "", r.get("cnpj", ""))
        cod   = _conc_norm_codigo(r.get("codigo", ""))
        per   = _conc_norm_periodo(r.get("periodo", ""))
        if cnpj and cod and per:
            darf_idx[(cnpj, cod, per)].append(r)

    # ── Indexa DCOMPs por chave — filtra apenas DCOMPs (não PERs isolados) ─
    # Tipo de pedido contém: "DCOMP", "Declaração de Compensação", etc.
    # Também exclui DCOMPs com status Cancelado ou Retificado na planilha.
    dcomp_idx: dict = defaultdict(list)
    for r in dcomp_rows:
        tipo = r.get("tipo_pedido", "").upper()
        # Inclui somente se for DCOMP (tem débito associado)
        if not re.search(r"DCOMP|COMPENSA[CÇ]", tipo, re.IGNORECASE):
            continue
        # Filtro por status: canceladas/retificadas não contam no confronto
        num_dcomp = r.get("numero_perdcomp", "").strip()
        if _is_cancelled(num_dcomp, status_map):
            continue
        if _is_retified(num_dcomp, status_map):
            continue
        cnpj  = re.sub(r"\D", "", r.get("cnpj", ""))
        cod   = _conc_extract_codigo_from_tipo(r.get("tipo", ""))
        per   = _conc_norm_periodo(r.get("periodo_apuracao", ""))
        if cnpj and cod and per:
            dcomp_idx[(cnpj, cod, per)].append(r)

    # ── Coleta todas as chaves únicas ──────────────────────────────────────
    all_keys = set(darf_idx.keys()) | set(dcomp_idx.keys())

    result = []
    for key in sorted(all_keys):
        cnpj_raw, cod, per = key
        darfs  = darf_idx.get(key, [])
        dcomps = dcomp_idx.get(key, [])

        # Usa o primeiro registro de cada lado para dados descritivos
        d0 = darfs[0]  if darfs  else {}
        c0 = dcomps[0] if dcomps else {}

        # Valores agregados (soma de todos os documentos na chave)
        darf_princ = sum(_conc_brl(r.get("principal", 0)) for r in darfs)
        darf_multa = sum(_conc_brl(r.get("multa", 0))     for r in darfs)
        darf_juros = sum(_conc_brl(r.get("juros", 0))     for r in darfs)
        darf_total = sum(_conc_brl(r.get("total_item", 0)) for r in darfs)

        dcomp_princ = sum(_conc_brl(r.get("valor_original", 0)) for r in dcomps)
        dcomp_multa = sum(_conc_brl(r.get("valor_multa", 0))     for r in dcomps)
        dcomp_juros = sum(_conc_brl(r.get("valor_juros", 0))     for r in dcomps)
        dcomp_total = sum(_conc_brl(r.get("valor_total", 0)) for r in dcomps)

        # ── Diferença e situação consideram SOMENTE O PRINCIPAL ───────────
        # Multas e juros não fazem parte do tributo declarado — são acréscimos
        # legais por atraso. Para validar se o débito foi quitado, comparamos
        # apenas o principal de cada lado. O total (com multa/juros) fica nas
        # colunas para visibilidade, mas não entra na classificação.
        diferenca = round(darf_princ - dcomp_princ, 2)

        # ── Situação ──────────────────────────────────────────────────────
        tem_darf  = bool(darfs)
        tem_dcomp = bool(dcomps)

        if tem_darf and tem_dcomp:
            tol = 0.05  # tolerância de R$ 0,05 para arredondamentos
            if abs(diferenca) <= tol:
                situacao = SIT_DUPLO       # pagou DARF E compensou o mesmo débito!
                obs = (f"DARF pago em {d0.get('dt_arrecadacao','')} via "
                       f"{d0.get('banco','')} | DCOMP {c0.get('numero_perdcomp','')}")
            else:
                situacao = SIT_DIVERGENTE
                obs = (f"DARF princ. R$ {darf_princ:,.2f} | DCOMP princ. R$ {dcomp_princ:,.2f} | "
                       f"Dif. R$ {diferenca:,.2f}").replace(",","X").replace(".",",").replace("X",".")
        elif tem_darf:
            situacao = SIT_SO_DARF
            obs = f"Pago via DARF em {d0.get('dt_arrecadacao','')} | Banco: {d0.get('banco','')}"
        else:
            situacao = SIT_SO_DCOMP
            obs = f"Compensado via DCOMP {c0.get('numero_perdcomp','')} em {c0.get('data_transmissao','')}"

        # ── Monta linha de resultado ───────────────────────────────────────
        cnpj_fmt = d0.get("cnpj", "") or c0.get("cnpj", "")
        razao    = d0.get("razao_social", "") or c0.get("razao_social", "")
        desc     = d0.get("descricao", "") or c0.get("tipo_debito", "")

        row = {
            "cnpj":              cnpj_fmt,
            "razao_social":      razao,
            "codigo_receita":    cod,
            "descricao_tributo": desc[:80] if desc else "",
            # Período canônico (chave de match) — sempre no formato MM/AAAA | TN/AAAA | AAAA.
            # Evita confusão visual quando o DARF trazia DD/MM/AAAA e a DCOMP "Mês de AAAA".
            "periodo":           per,
            # DARF
            "darf_numero":       " / ".join(r.get("numero_doc","") for r in darfs if r.get("numero_doc")),
            "darf_dt_arrec":     d0.get("dt_arrecadacao",""),
            "darf_banco":        d0.get("banco",""),
            "darf_principal":    darf_princ,
            "darf_multa":        darf_multa,
            "darf_juros":        darf_juros,
            "darf_total":        darf_total,
            # DCOMP
            "dcomp_numero":      " / ".join(r.get("numero_perdcomp","") for r in dcomps if r.get("numero_perdcomp")),
            "dcomp_tipo_pedido": c0.get("tipo_pedido",""),
            "dcomp_dt_transm":   c0.get("data_transmissao",""),
            "dcomp_principal":   dcomp_princ,
            "dcomp_multa":       dcomp_multa,
            "dcomp_juros":       dcomp_juros,
            "dcomp_total":       dcomp_total,
            # Análise
            "situacao":  situacao,
            "diferenca": diferenca,
            "obs":       obs,
        }
        result.append(row)

    # Ordena: primeiro duplos/divergentes, depois só DARF, depois só DCOMP
    _ord = {SIT_DUPLO:0, SIT_DIVERGENTE:1, SIT_SO_DCOMP:2, SIT_SO_DARF:3, SIT_OK:4}
    result.sort(key=lambda r: (_ord.get(r["situacao"],9), r.get("cnpj",""), r.get("periodo","")))
    return result


def run_triplo_dctf_darf_dcomp(dctf_rows: list, darf_rows: list,
                                 dcomp_rows: list,
                                 status_map: dict = None,
                                 dctfweb_rows: list = None) -> list:
    """
    Quádruplo cruzamento simétrico: (DCTF + DCTFWeb) × DARF × DCOMP.

    Para cada chave (CNPJ + código de receita + período):
      - Lado DECLARAÇÃO: soma os débitos apurados na DCTF clássica + DCTFWeb
      - Lado QUITAÇÃO: DARFs efetivamente recolhidos + DCOMPs ATIVAS
        (DCOMPs canceladas/retificadas pelo eCAC são EXCLUÍDAS via status_map)

    Saldo Final = (Débito DCTF + Débito DCTFWeb) − (DARF + DCOMP ativa)
      > 0  → Saldo a Pagar (declarado mas não quitado)
      = 0  → Quitado (tolerância R$ 0,05)
      < 0  → Pago/Compensado a maior
      caso especial: sem DCTF nem DCTFWeb mas há DARF/DCOMP → Sem Declaração

    Args:
      dctf_rows: linhas da DCTF clássica (PIS/COFINS/IRRF, etc.)
      darf_rows: DARFs/DAS arrecadados
      dcomp_rows: DCOMPs (apenas linhas tipo Débito do PERDCOMP)
      status_map: planilha do eCAC com situações (cancelada/retificada/ativa)
      dctfweb_rows: linhas da DCTFWeb (CP previdenciária + IRRF folha) — opcional

    Mantém o nome run_triplo_dctf_darf_dcomp para compatibilidade — ainda gera
    o cruzamento mesmo sem DCTFWeb.
    """
    from collections import defaultdict

    if status_map is None:
        status_map = {}
    if dctfweb_rows is None:
        dctfweb_rows = []

    # ── Indexa DCTFs clássicas (linha por tributo) ─────────────────────────
    dctf_idx: dict = defaultdict(list)
    for r in dctf_rows:
        if not r.get("codigo_receita", "").strip():
            continue
        if "nenhum" in str(r.get("grupo_tributo", "")).lower():
            continue
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(r.get("codigo_receita", ""))
        per  = _conc_norm_periodo(r.get("periodo_apuracao", ""))
        if cnpj and cod and per:
            dctf_idx[(cnpj, cod, per)].append(r)

    # ── Indexa DCTFWebs (linha por tributo, ignora "Sem fatos geradores") ──
    dctfweb_idx: dict = defaultdict(list)
    for r in dctfweb_rows:
        cod_raw = r.get("codigo_receita", "").strip()
        if not cod_raw:
            continue   # ignora declarações vazias (Ausência de Fatos Geradores)
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(cod_raw)
        per  = _conc_norm_periodo(r.get("periodo", ""))
        if cnpj and cod and per:
            dctfweb_idx[(cnpj, cod, per)].append(r)

    # ── Indexa DARFs ───────────────────────────────────────────────────────
    darf_idx: dict = defaultdict(list)
    for r in darf_rows:
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_norm_codigo(r.get("codigo", ""))
        per  = _conc_norm_periodo(r.get("periodo", ""))
        if cnpj and cod and per:
            darf_idx[(cnpj, cod, per)].append(r)

    # ── Indexa DCOMPs ATIVAS (filtra canceladas/retificadas via eCAC) ──────
    dcomp_idx: dict = defaultdict(list)
    for r in dcomp_rows:
        tipo = r.get("tipo_pedido", "").upper()
        if not re.search(r"DCOMP|COMPENSA[CÇ]", tipo, re.IGNORECASE):
            continue
        num_dcomp = r.get("numero_perdcomp", "").strip()
        if _is_cancelled(num_dcomp, status_map):
            continue
        if _is_retified(num_dcomp, status_map):
            continue
        cnpj = re.sub(r"\D", "", r.get("cnpj", ""))
        cod  = _conc_extract_codigo_from_tipo(r.get("tipo", ""))
        per  = _conc_norm_periodo(r.get("periodo_apuracao", ""))
        if cnpj and cod and per:
            dcomp_idx[(cnpj, cod, per)].append(r)

    # ── União de TODAS as chaves dos quatro lados ──────────────────────────
    all_keys = (set(dctf_idx.keys()) | set(dctfweb_idx.keys())
                | set(darf_idx.keys()) | set(dcomp_idx.keys()))

    result = []
    for key in sorted(all_keys):
        cnpj_raw, cod, per = key
        dctfs    = dctf_idx.get(key, [])
        dctfwebs = dctfweb_idx.get(key, [])
        darfs    = darf_idx.get(key, [])
        dcomps   = dcomp_idx.get(key, [])

        d0   = dctfs[0]    if dctfs    else {}
        dw0  = dctfwebs[0] if dctfwebs else {}
        df0  = darfs[0]    if darfs    else {}
        dc0  = dcomps[0]   if dcomps   else {}

        # ── Débitos apurados (declarado pela empresa) ──────────────────────
        # DCTF clássica: usa débito apurado (não há deduções nesse formato)
        dctf_debito = sum(_conc_brl(r.get("debito_apurado", 0)) for r in dctfs)
        # DCTFWeb: usa SALDO A PAGAR (já líquido de deduções de Salário Família).
        # Deduções como Salário Família são abatidas na origem e o saldo é o que
        # de fato precisa ser quitado via DARF/DCOMP. Isso evita falsos positivos
        # de "Saldo a Pagar" causados pelas deduções.
        dctfweb_debito = sum(_conc_brl(r.get("saldo_pagar", 0)) for r in dctfwebs)
        total_declarado = dctf_debito + dctfweb_debito

        # ── Quitação efetiva (DARF + DCOMP ativa) — SOMENTE PRINCIPAL ──────
        # Multas e juros são acréscimos legais por atraso, não fazem parte do
        # tributo declarado. Para validar quitação do débito, comparamos contra
        # o principal de cada lado. Isso evita falsos "pagto a maior" quando há
        # multa/juros acumulados.
        darf_total  = sum(_conc_brl(r.get("principal", 0))      for r in darfs)
        dcomp_total = sum(_conc_brl(r.get("valor_original", 0)) for r in dcomps)

        # ── CNPJ, Razão, Grupo do Tributo ──────────────────────────────────
        cnpj_fmt = (d0.get("cnpj", "") or dw0.get("cnpj", "")
                    or df0.get("cnpj", "") or dc0.get("cnpj", ""))
        razao = (d0.get("nome_empresarial", "")
                 or dw0.get("razao_social", "")
                 or df0.get("razao_social", "")
                 or dc0.get("razao_social", ""))
        grupo = (d0.get("grupo_tributo", "")
                 or dw0.get("grupo_tributo", "")
                 or df0.get("descricao", "")
                 or dc0.get("tipo_debito", ""))

        # ── Campos da DCTF clássica ────────────────────────────────────────
        dctf_num_decl = " / ".join(
            r.get("numero_declaracao", "")
            for r in dctfs if r.get("numero_declaracao"))
        dctf_periodic = d0.get("periodicidade", "") if dctfs else ""

        # ── Campos da DCTFWeb ──────────────────────────────────────────────
        dctfweb_num_recibo = " / ".join(
            r.get("numero_recibo", "")
            for r in dctfwebs if r.get("numero_recibo"))
        dctfweb_categoria = dw0.get("categoria", "") if dctfwebs else ""

        # ── Observação descritiva ──────────────────────────────────────────
        partes = []
        if dctfs:    partes.append(f"DCTF: R$ {dctf_debito:,.2f}")
        if dctfwebs: partes.append(f"DCTFWeb: R$ {dctfweb_debito:,.2f}")
        if darfs:    partes.append(f"{len(darfs)} DARF(s) princ. R$ {darf_total:,.2f}")
        if dcomps:   partes.append(f"{len(dcomps)} DCOMP(s) ativa(s) princ. R$ {dcomp_total:,.2f}")
        obs = " | ".join(partes).replace(",", "X").replace(".", ",").replace("X", ".")

        # ── Saldo Final e Situação ─────────────────────────────────────────
        TOL = 0.05
        tem_declaracao = bool(dctfs or dctfwebs)
        tem_quitacao   = bool(darfs or dcomps)

        if not tem_declaracao and tem_quitacao:
            # DARF/DCOMP sem nenhuma declaração → tudo é "a maior"
            saldo_final    = -(darf_total + dcomp_total)
            situacao_triplo = SIT_T_SEM_DECL
        else:
            saldo_final = round(total_declarado - darf_total - dcomp_total, 2)
            if abs(saldo_final) <= TOL:
                situacao_triplo = SIT_T_QUITADO
            elif saldo_final > 0:
                situacao_triplo = SIT_T_SALDO
            else:
                situacao_triplo = SIT_T_A_MAIOR

        # ── Competência Teste (formato canônico) ───────────────────────────
        periodo_raw = (d0.get("periodo_apuracao", "")
                        or dw0.get("periodo", "")
                        or df0.get("periodo", "")
                        or dc0.get("periodo_apuracao", ""))
        comp_teste = format_competencia_teste(periodo_raw)

        row = {
            "cnpj":              cnpj_fmt,
            "razao_social":      razao,
            "codigo_receita":    cod,
            "grupo_tributo":     (grupo or "")[:80],
            "periodo":           per,
            "competencia_teste": comp_teste,
            # DCTF clássica
            "dctf_num_decl":      dctf_num_decl,
            "dctf_periodicidade": dctf_periodic,
            "dctf_debito":        dctf_debito,
            # DCTFWeb
            "dctfweb_num_recibo": dctfweb_num_recibo,
            "dctfweb_categoria":  dctfweb_categoria,
            "dctfweb_debito":     dctfweb_debito,
            # Total declarado (somatório)
            "total_declarado":    total_declarado,
            # DARF
            "darf_numero":   " / ".join(r.get("numero_doc","") for r in darfs if r.get("numero_doc")),
            "darf_dt_arrec": df0.get("dt_arrecadacao", ""),
            "darf_total":    darf_total,
            # DCOMP
            "dcomp_numero":     " / ".join(r.get("numero_perdcomp","") for r in dcomps if r.get("numero_perdcomp")),
            "dcomp_dt_transm":  dc0.get("data_transmissao", ""),
            "dcomp_total":      dcomp_total,
            # Análise
            "saldo_final":      saldo_final,
            "situacao_triplo":  situacao_triplo,
            "obs":              obs,
        }
        result.append(row)

    # Ordena: Saldo a Pagar, A Maior, Sem Declaração, Quitado
    _ord = {SIT_T_SALDO:0, SIT_T_A_MAIOR:1, SIT_T_SEM_DECL:2, SIT_T_QUITADO:3}
    result.sort(key=lambda r: (_ord.get(r.get("situacao_triplo",""),9),
                                 r.get("cnpj",""), r.get("competencia_teste",""),
                                 r.get("codigo_receita","")))
    return result



def export_conciliacao_excel(rows: list, path: str, triplo_rows: list = None):
    """Exporta resultado da conciliação para Excel com formatação visual.

    Se `triplo_rows` for fornecido, adiciona uma terceira aba com o cruzamento
    triplo DCTF × DARF × DCOMP.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação DARF x DCOMP"

    # Paleta de cores por situação
    COR_DUPLO      = "FDEDEC"   # vermelho claro
    COR_DIVERGENTE = "FEF3DC"   # amarelo claro
    COR_SO_DCOMP   = "EFF6FF"   # azul claro
    COR_SO_DARF    = "F2F4F0"   # cinza claro
    COR_OK         = "EAF4D3"   # verde claro

    def _sit_bg(sit):
        if SIT_DUPLO      in sit: return COR_DUPLO
        if SIT_DIVERGENTE in sit: return COR_DIVERGENTE
        if SIT_SO_DCOMP   in sit: return COR_SO_DCOMP
        if SIT_SO_DARF    in sit: return COR_SO_DARF
        return COR_OK

    # Linha de título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CONC_COLS))
    tc = ws.cell(row=1, column=1, value="CONCILIAÇÃO DARF × DCOMP  |  AgriTax Audit")
    tc.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    tc.fill = PatternFill("solid", fgColor="3D3D3D")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Cabeçalhos de seção (linha 2)
    SEC = [
        ("Identificação",    1,  5,  "3D3D3D"),
        ("DARF (pagamento)", 6,  12, "5A8A1E"),
        ("DCOMP (compensação)", 13, 19, "2563EB"),
        ("Análise",          20, 22, "C0392B"),
    ]
    for label, c1, c2, cor in SEC:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        sc = ws.cell(row=2, column=c1, value=label)
        sc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        sc.fill = PatternFill("solid", fgColor=cor)
        sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Cabeçalhos de coluna (linha 3)
    for ci, (key, label, width) in enumerate(CONC_COLS, 1):
        hc = ws.cell(row=3, column=ci, value=label)
        hc.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        hc.fill = PatternFill("solid", fgColor="7AB82E")
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width / 7
    ws.row_dimensions[3].height = 28

    thin = Side(style="thin", color="D0DDB8")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Dados
    for ri, row in enumerate(rows, 4):
        bg = _sit_bg(row.get("situacao",""))
        for ci, (key, _, _w) in enumerate(CONC_COLS, 1):
            v = row.get(key, "")
            if key in CONC_MONEY:
                v = float(v or 0)
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="right" if key in CONC_MONEY else "left",
                vertical="center")
            c.border = brd
            if key in CONC_MONEY:
                c.number_format = '#.##0,00'

    # Linha de totais
    tr = len(rows) + 4
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    ws.cell(row=tr, column=1, value="TOTAIS").font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    ws.cell(row=tr, column=1).fill = PatternFill("solid", fgColor="3D3D3D")
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    for ci, (key, _, _w) in enumerate(CONC_COLS, 1):
        if key in CONC_MONEY:
            v = sum(float(r.get(key,0) or 0) for r in rows)
            c = ws.cell(row=tr, column=ci, value=v)
            c.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="3D3D3D")
            c.border = brd
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = '#.##0,00'

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(CONC_COLS))}3"

    # Segunda aba: resumo por situação
    ws2 = wb.create_sheet("Resumo por Situação")
    from collections import Counter
    contagem = Counter(r.get("situacao","") for r in rows)
    ws2.cell(row=1, column=1, value="Situação").font = Font(bold=True, name="Calibri")
    ws2.cell(row=1, column=2, value="Qtd").font = Font(bold=True, name="Calibri")
    ws2.cell(row=1, column=3, value="Total DARF (R$)").font = Font(bold=True, name="Calibri")
    ws2.cell(row=1, column=4, value="Total DCOMP (R$)").font = Font(bold=True, name="Calibri")
    ws2.cell(row=1, column=5, value="Diferença (R$)").font = Font(bold=True, name="Calibri")
    for ci2, w2 in zip(range(1,6),[220,60,130,130,130]):
        ws2.column_dimensions[get_column_letter(ci2)].width = w2 / 7
    for ri2, (sit, qtd) in enumerate(sorted(contagem.items()), 2):
        sit_rows = [r for r in rows if r.get("situacao","") == sit]
        ws2.cell(row=ri2, column=1, value=sit)
        ws2.cell(row=ri2, column=2, value=qtd)
        ws2.cell(row=ri2, column=3, value=sum(float(r.get("darf_total",0) or 0) for r in sit_rows)).number_format = '#.##0,00'
        ws2.cell(row=ri2, column=4, value=sum(float(r.get("dcomp_total",0) or 0) for r in sit_rows)).number_format = '#.##0,00'
        ws2.cell(row=ri2, column=5, value=sum(float(r.get("diferenca",0) or 0) for r in sit_rows)).number_format = '#.##0,00'
        bg2 = _sit_bg(sit)
        for ci2 in range(1,6):
            ws2.cell(row=ri2, column=ci2).fill = PatternFill("solid", fgColor=bg2)
            ws2.cell(row=ri2, column=ci2).font = Font(size=9, name="Calibri")

    # ── Terceira aba (opcional): DCTF × DARF × DCOMP — triplo simétrico ───
    if triplo_rows:
        ws3 = wb.create_sheet("DCTF + DCTFWeb x DARF x DCOMP"[:31])

        # Linha de título
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TRIPLO_COLS))
        tc3 = ws3.cell(row=1, column=1,
            value="DCTF + DCTFWeb × DARF × DCOMP  |  Quádruplo Cruzamento  |  AgriTax Audit")
        tc3.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        tc3.fill = PatternFill("solid", fgColor="3D3D3D")
        tc3.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 24

        # Cabeçalhos de seção (linha 2) — 3 blocos lado a lado + análise
        SEC3 = [
            ("Identificação",         1,  6,  "3D3D3D"),
            ("DCTF (declarado)",      7,  9,  "F5A623"),
            ("DCTFWeb (declarado)",   10, 12, "8B5CF6"),
            ("Total Declarado",       13, 13, "C0392B"),
            ("DARF (pago)",           14, 16, "5A8A1E"),
            ("DCOMP Ativa (compens.)", 17, 19, "2563EB"),
            ("Análise Contábil",      20, 22, "C0392B"),
        ]
        for label, c1, c2, cor in SEC3:
            ws3.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
            sc = ws3.cell(row=2, column=c1, value=label)
            sc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            sc.fill = PatternFill("solid", fgColor=cor)
            sc.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[2].height = 18

        # Cabeçalhos de coluna (linha 3)
        for ci, (key, label, width) in enumerate(TRIPLO_COLS, 1):
            hc = ws3.cell(row=3, column=ci, value=label)
            hc.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            hc.fill = PatternFill("solid", fgColor="7AB82E")
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws3.column_dimensions[get_column_letter(ci)].width = width / 7
        ws3.row_dimensions[3].height = 28

        # Dados — cor por situação (mais informativo que linhas alternadas)
        COR_T_QUITADO  = "EAF4D3"   # verde claro
        COR_T_SALDO    = "FDEDEC"   # vermelho claro
        COR_T_A_MAIOR  = "FEF3DC"   # amarelo claro (possível restituição — atenção!)
        COR_T_SEM_DECL = "EFF6FF"   # azul claro (DARF/DCOMP sem DCTF nem DCTFWeb)
        def _sit_t_bg(sit):
            if SIT_T_QUITADO  in sit: return COR_T_QUITADO
            if SIT_T_SALDO    in sit: return COR_T_SALDO
            if SIT_T_A_MAIOR  in sit: return COR_T_A_MAIOR
            if SIT_T_SEM_DECL in sit: return COR_T_SEM_DECL
            return "FFFFFF"

        for ri, row in enumerate(triplo_rows, 4):
            bg = _sit_t_bg(row.get("situacao_triplo", ""))
            for ci, (key, _, _w) in enumerate(TRIPLO_COLS, 1):
                v = row.get(key, "")
                if key in TRIPLO_MONEY:
                    v = float(v or 0)
                c = ws3.cell(row=ri, column=ci, value=v)
                c.font = Font(size=9, name="Calibri")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal="right" if key in TRIPLO_MONEY else "left",
                    vertical="center")
                c.border = brd
                if key in TRIPLO_MONEY:
                    c.number_format = '#.##0,00'

        # Linha de totais
        tr3 = len(triplo_rows) + 4
        ws3.merge_cells(start_row=tr3, start_column=1, end_row=tr3, end_column=6)
        ws3.cell(row=tr3, column=1, value="TOTAIS").font = Font(
            bold=True, color="FFFFFF", size=9, name="Calibri")
        ws3.cell(row=tr3, column=1).fill = PatternFill("solid", fgColor="3D3D3D")
        ws3.cell(row=tr3, column=1).alignment = Alignment(horizontal="center")
        for ci, (key, _, _w) in enumerate(TRIPLO_COLS, 1):
            if key in TRIPLO_MONEY:
                v = sum(float(r.get(key,0) or 0) for r in triplo_rows)
                c = ws3.cell(row=tr3, column=ci, value=v)
                c.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="3D3D3D")
                c.border = brd
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#.##0,00'

        ws3.freeze_panes = "A4"
        ws3.auto_filter.ref = f"A3:{get_column_letter(len(TRIPLO_COLS))}3"

    wb.save(path)


class ConciliacaoApp(tk.Toplevel):
    """Módulo de Conciliação DARF × DCOMP."""

    _SIT_COLORS = {
        SIT_DUPLO:      C_RED_LIGHT,
        SIT_DIVERGENTE: C_YELLOW_LIGHT,
        SIT_SO_DCOMP:   C_BLUE_LIGHT,
        SIT_SO_DARF:    C_GRAY_LIGHT,
        SIT_OK:         C_GREEN_LIGHT,
    }

    def __init__(self, master):
        super().__init__(master)
        self.title("Conciliação DARF × DCOMP — AgriTax Audit")
        self.configure(bg=C_GRAY_LIGHT)
        self.geometry("1400x820")
        self.minsize(1000, 600)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self._logo_img = None

        self.darf_files:  list = []
        self.dcomp_files: list = []
        self.dctf_files:  list = []        # NOVO — DCTFs para o triplo cruzamento
        self.status_file: str  = ""        # planilha de status do eCAC
        self.status_map:  dict = {}        # {numero_perdcomp: {situacao, ...}}
        self.result_rows: list = []        # aba 1 — DARF × DCOMP
        self.triplo_rows: list = []        # aba 2 — DCTF × DARF × DCOMP
        self._processing = False
        # Filtros tipo Excel por coluna do resultado: {col_key: set_valores | None}
        self.col_filt: dict = {}           # filtros da aba DARF × DCOMP
        self.col_filt_t: dict = {}         # filtros da aba triplo

        self._build_ui()

    def _build_ui(self):
        # Cabeçalho
        top = tk.Frame(self, bg=C_GREEN_DARK, height=56)
        top.pack(fill="x"); top.pack_propagate(False)
        try:
            import base64 as _b64, io as _io
            from PIL import Image, ImageTk
            img = Image.open(_io.BytesIO(_b64.b64decode(LOGO_B64))).resize((42,42), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(top, image=self._logo_img, bg=C_GREEN_DARK).pack(side="left", padx=12)
        except Exception:
            pass
        tk.Label(top, text="AgriTax", font=("Segoe UI",16,"bold"),
                 fg=C_WHITE, bg=C_GREEN_DARK).pack(side="left", pady=10)
        tk.Label(top, text="  Conciliação DARF × DCOMP",
                 font=("Segoe UI",11), fg=C_GREEN_MID,
                 bg=C_GREEN_DARK).pack(side="left", pady=10)
        tk.Label(top, text="  Cruzamento débitos pagos × compensações PERDCOMP",
                 font=("Segoe UI",8), fg="#8fb84a",
                 bg=C_GREEN_DARK).pack(side="left", pady=10)

        body = tk.Frame(self, bg=C_GRAY_LIGHT)
        body.pack(fill="both", expand=True, padx=10, pady=8)

        # Painel esquerdo — controles (rolável por dentro de um Canvas)
        left_outer = tk.Frame(body, bg=C_WHITE, width=278,
                              highlightbackground=C_BORDER, highlightthickness=1)
        left_outer.pack(side="left", fill="y", padx=(0,8))
        left_outer.pack_propagate(False)

        canvas = tk.Canvas(left_outer, bg=C_WHITE, highlightthickness=0, width=260)
        vsb_l  = ttk.Scrollbar(left_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb_l.set)
        vsb_l.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        left = tk.Frame(canvas, bg=C_WHITE)
        win_id = canvas.create_window((0, 0), window=left, anchor="nw")

        def _on_configure(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Garante que o frame interno ocupe toda a largura do canvas
            canvas.itemconfig(win_id, width=canvas.winfo_width())
        left.bind("<Configure>", _on_configure)
        canvas.bind("<Configure>", _on_configure)

        # Rolagem com roda do mouse — ativada ao entrar no painel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        def _bind_wheel(e=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        def _unbind_wheel(e=None):
            canvas.unbind_all("<MouseWheel>")
        left_outer.bind("<Enter>", _bind_wheel)
        left_outer.bind("<Leave>", _unbind_wheel)

        self._build_left(left)

        # Painel direito — resultado
        right = tk.Frame(body, bg=C_GRAY_LIGHT)
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

        # Status bar
        sb = tk.Frame(self, bg=C_GREEN_DARK, height=24)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Adicione os PDFs de DARF e PERDCOMP para iniciar.")
        tk.Label(sb, textvariable=self.status_var, font=("Segoe UI",8),
                 fg=C_GREEN_LIGHT, bg=C_GREEN_DARK, anchor="w").pack(fill="x", padx=12, pady=3)

    def _build_left(self, parent):
        cfg_lbl = dict(bg=C_WHITE, fg=C_GREEN_DARK, font=("Segoe UI",9,"bold"), anchor="w")
        cfg_btn = dict(relief="flat", cursor="hand2", pady=4, font=("Segoe UI",8,"bold"))

        # DARF
        tk.Label(parent, text="1. Arquivos DARF (PDFs)", **cfg_lbl).pack(fill="x", padx=12, pady=(14,2))
        f1 = tk.Frame(parent, bg=C_GRAY_LIGHT, height=90)
        f1.pack(fill="x", padx=8, pady=(0,4)); f1.pack_propagate(False)
        sb1 = ttk.Scrollbar(f1); sb1.pack(side="right", fill="y")
        self.lb_darf = tk.Listbox(f1, yscrollcommand=sb1.set, font=("Segoe UI",8),
                                   bg=C_GRAY_LIGHT, fg=C_GRAY_DARK,
                                   selectbackground=C_GREEN_MID, relief="flat")
        self.lb_darf.pack(fill="both", expand=True)
        sb1.config(command=self.lb_darf.yview)
        tk.Button(parent, text="Adicionar PDFs DARF", bg=C_GREEN, fg=C_WHITE,
                  command=self._add_darf, **cfg_btn).pack(fill="x", padx=8, pady=(0,8))

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8)

        # PERDCOMP
        tk.Label(parent, text="2. Arquivos PERDCOMP (PDFs)", **cfg_lbl).pack(fill="x", padx=12, pady=(10,2))
        f2 = tk.Frame(parent, bg=C_GRAY_LIGHT, height=90)
        f2.pack(fill="x", padx=8, pady=(0,4)); f2.pack_propagate(False)
        sb2 = ttk.Scrollbar(f2); sb2.pack(side="right", fill="y")
        self.lb_dcomp = tk.Listbox(f2, yscrollcommand=sb2.set, font=("Segoe UI",8),
                                    bg=C_GRAY_LIGHT, fg=C_GRAY_DARK,
                                    selectbackground=C_GREEN_MID, relief="flat")
        self.lb_dcomp.pack(fill="both", expand=True)
        sb2.config(command=self.lb_dcomp.yview)
        tk.Button(parent, text="Adicionar PDFs PERDCOMP", bg=C_GREEN, fg=C_WHITE,
                  command=self._add_dcomp, **cfg_btn).pack(fill="x", padx=8, pady=(0,8))

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8)

        # DCTF (para cruzamento triplo)
        tk.Label(parent, text="3. Arquivos DCTF (PDFs) — opcional",
                 **cfg_lbl).pack(fill="x", padx=12, pady=(10,2))
        f3 = tk.Frame(parent, bg=C_GRAY_LIGHT, height=70)
        f3.pack(fill="x", padx=8, pady=(0,4)); f3.pack_propagate(False)
        sb3 = ttk.Scrollbar(f3); sb3.pack(side="right", fill="y")
        self.lb_dctf = tk.Listbox(f3, yscrollcommand=sb3.set, font=("Segoe UI",8),
                                    bg=C_GRAY_LIGHT, fg=C_GRAY_DARK,
                                    selectbackground=C_GREEN_MID, relief="flat")
        self.lb_dctf.pack(fill="both", expand=True)
        sb3.config(command=self.lb_dctf.yview)
        tk.Button(parent, text="Adicionar PDFs DCTF", bg=C_GREEN, fg=C_WHITE,
                  command=self._add_dctf, **cfg_btn).pack(fill="x", padx=8, pady=(0,8))

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8)

        # Planilha de status (eCAC) — filtra canceladas/retificadas
        tk.Label(parent, text="4. Planilha de Status (eCAC) — opcional",
                 **cfg_lbl).pack(fill="x", padx=12, pady=(10,2))
        self.lbl_status_file = tk.Label(parent, text="(nenhuma planilha carregada)",
                                        bg=C_GRAY_LIGHT, fg=C_GRAY,
                                        font=("Segoe UI",8), anchor="w", padx=6, pady=3,
                                        relief="flat")
        self.lbl_status_file.pack(fill="x", padx=8, pady=(0,4))
        tk.Button(parent, text="Carregar Planilha de Status", bg=C_GREEN, fg=C_WHITE,
                  command=self._load_status, **cfg_btn).pack(fill="x", padx=8, pady=(0,8))

        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8)

        # Estatísticas
        sf = tk.Frame(parent, bg=C_GREEN_LIGHT, highlightbackground=C_BORDER, highlightthickness=1)
        sf.pack(fill="x", padx=8, pady=8)
        self.lbl_stats = tk.Label(sf,
            text="DARFs carregados : 0\nDCOMPs carregados: 0\nDCTFs carregados : 0\n"
                 "DCOMPs filtradas : 0\nLinhas conciliação: 0\nLinhas triplo    : 0\n"
                 "Duplo pagamento  : 0\nDivergentes      : 0",
            font=("Courier",8), fg=C_GREEN_DARK, bg=C_GREEN_LIGHT, justify="left")
        self.lbl_stats.pack(anchor="w", padx=8, pady=6)

        # Botões de ação
        self.btn_run = tk.Button(parent, text="Executar Conciliação",
                                  bg=C_BLUE, fg=C_WHITE,
                                  command=self._run, state="disabled", **cfg_btn)
        self.btn_run.pack(fill="x", padx=8, pady=(0,4))
        self.btn_export = tk.Button(parent, text="Exportar Excel",
                                     bg=C_YELLOW, fg=C_WHITE,
                                     command=self._export, state="disabled", **cfg_btn)
        self.btn_export.pack(fill="x", padx=8, pady=(0,4))
        self.btn_clear = tk.Button(parent, text="Limpar Tudo",
                                    bg=C_RED, fg=C_WHITE,
                                    command=self._clear, **cfg_btn)
        self.btn_clear.pack(fill="x", padx=8, pady=(0,8))
        self.progress = ttk.Progressbar(parent, mode="indeterminate", length=240)
        self.progress.pack(fill="x", padx=8, pady=(0,8))

        # Legenda
        tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8)
        tk.Label(parent, text="Legenda:", bg=C_WHITE, fg=C_GRAY_DARK,
                 font=("Segoe UI",8,"bold"), anchor="w").pack(fill="x", padx=12, pady=(6,2))
        legenda = [
            (SIT_DUPLO,      C_RED_LIGHT),
            (SIT_DIVERGENTE, C_YELLOW_LIGHT),
            (SIT_SO_DCOMP,   C_BLUE_LIGHT),
            (SIT_SO_DARF,    C_GRAY_LIGHT),
        ]
        for txt, cor in legenda:
            row = tk.Frame(parent, bg=C_WHITE)
            row.pack(fill="x", padx=12, pady=1)
            tk.Frame(row, bg=cor, width=12, height=12,
                     highlightbackground="#AAAAAA", highlightthickness=1).pack(side="left", padx=(0,5))
            tk.Label(row, text=txt, bg=C_WHITE, fg=C_GRAY_DARK,
                     font=("Segoe UI",7)).pack(side="left")

    def _build_right(self, parent):
        # Notebook com 2 abas: DARF × DCOMP   e   DCTF × DARF × DCOMP
        self.nb = ttk.Notebook(parent)
        self.nb.pack(fill="both", expand=True)

        # ── Aba 1: DARF × DCOMP (original) ─────────────────────────────────
        tab1 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
        self.nb.add(tab1, text="  DARF × DCOMP  ")

        # Filtro rápido por situação
        bar = tk.Frame(tab1, bg=C_WHITE,
                       highlightbackground=C_BORDER, highlightthickness=1)
        bar.pack(fill="x", pady=(0,6))
        tk.Label(bar, text="Filtrar:", bg=C_WHITE, fg=C_GRAY_DARK,
                 font=("Segoe UI",8,"bold")).pack(side="left", padx=(10,4), pady=6)
        self.filt_var = tk.StringVar(value="Todos")
        opts = ["Todos", SIT_DUPLO, SIT_DIVERGENTE, SIT_SO_DCOMP, SIT_SO_DARF]
        filt_menu = ttk.Combobox(bar, textvariable=self.filt_var,
                                  values=opts, state="readonly", width=28,
                                  font=("Segoe UI",8))
        filt_menu.pack(side="left", padx=4)
        filt_menu.bind("<<ComboboxSelected>>", lambda e: self._refresh())

        # Treeview de resultado (DARF × DCOMP)
        frm = tk.Frame(tab1, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        self.tree = ttk.Treeview(frm, columns=CONC_KEYS, show="headings",
                                   yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for key, label, width in CONC_COLS:
            def _make_cmd(k=key, l=label):
                return lambda: self._open_col_filter(k, l)
            self.tree.heading(key, text=f"{label}  ▼", command=_make_cmd())
            self.tree.column(key, width=width, minwidth=50,
                             anchor="e" if key in CONC_MONEY else "w")
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        self.tree.tag_configure("duplo",      background=C_RED_LIGHT)
        self.tree.tag_configure("divergente", background=C_YELLOW_LIGHT)
        self.tree.tag_configure("so_dcomp",   background=C_BLUE_LIGHT)
        self.tree.tag_configure("so_darf",    background=C_GRAY_LIGHT)
        self.tree.tag_configure("ok",         background=C_GREEN_LIGHT)

        # ── Aba 2: DCTF × DARF × DCOMP (triplo simétrico) ──────────────────
        tab2 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
        self.nb.add(tab2, text="  DCTF + DCTFWeb × DARF × DCOMP  ")

        # Barra de informação (sem filtro por situação — é triplo simétrico)
        bar2 = tk.Frame(tab2, bg=C_WHITE,
                        highlightbackground=C_BORDER, highlightthickness=1)
        bar2.pack(fill="x", pady=(0,6))
        tk.Label(bar2, text="Triplo cruzamento simétrico — 3 colunas lado a lado",
                 bg=C_WHITE, fg=C_GRAY_DARK,
                 font=("Segoe UI",8,"bold")).pack(side="left", padx=10, pady=6)
        tk.Label(bar2, text="  |  DCTFs ignoradas se status = Cancelado/Retificado",
                 bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI",8)).pack(side="left", pady=6)

        # Treeview do triplo
        frm2 = tk.Frame(tab2, bg=C_GRAY_LIGHT)
        frm2.pack(fill="both", expand=True)
        vsb2 = ttk.Scrollbar(frm2, orient="vertical")
        hsb2 = ttk.Scrollbar(frm2, orient="horizontal")
        self.tree_t = ttk.Treeview(frm2, columns=TRIPLO_KEYS, show="headings",
                                    yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        for key, label, width in TRIPLO_COLS:
            def _make_cmd_t(k=key, l=label):
                return lambda: self._open_col_filter_t(k, l)
            self.tree_t.heading(key, text=f"{label}  ▼", command=_make_cmd_t())
            self.tree_t.column(key, width=width, minwidth=50,
                                anchor="e" if key in TRIPLO_MONEY else "w")
        vsb2.config(command=self.tree_t.yview)
        hsb2.config(command=self.tree_t.xview)
        vsb2.pack(side="right", fill="y")
        hsb2.pack(side="bottom", fill="x")
        self.tree_t.pack(fill="both", expand=True)

        # Linhas coloridas por situação contábil (igual aos módulos anteriores)
        self.tree_t.tag_configure("t_quitado",  background=C_GREEN_LIGHT)
        self.tree_t.tag_configure("t_saldo",    background=C_RED_LIGHT)
        self.tree_t.tag_configure("t_amaior",   background=C_YELLOW_LIGHT)
        self.tree_t.tag_configure("t_sem_decl", background=C_BLUE_LIGHT)

    def _tag_for(self, sit: str) -> str:
        if SIT_DUPLO      in sit: return "duplo"
        if SIT_DIVERGENTE in sit: return "divergente"
        if SIT_SO_DCOMP   in sit: return "so_dcomp"
        if SIT_SO_DARF    in sit: return "so_darf"
        return "ok"

    def _fmt(self, v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v is not None else ""

    def _add_darf(self):
        paths = filedialog.askopenfilenames(parent=self,
            title="PDFs de DARF / DAS",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        new = [p for p in paths if p not in self.darf_files]
        self.darf_files.extend(new)
        for p in new: self.lb_darf.insert("end", Path(p).name)
        self._check_ready()
        self._set_status(f"{len(new)} DARF(s) adicionado(s).")

    def _add_dcomp(self):
        paths = filedialog.askopenfilenames(parent=self,
            title="PDFs de PERDCOMP (DCOMP)",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        new = [p for p in paths if p not in self.dcomp_files]
        self.dcomp_files.extend(new)
        for p in new: self.lb_dcomp.insert("end", Path(p).name)
        self._check_ready()
        self._set_status(f"{len(new)} PERDCOMP(s) adicionado(s).")

    def _add_dctf(self):
        paths = filedialog.askopenfilenames(parent=self,
            title="PDFs de DCTF",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        new = [p for p in paths if p not in self.dctf_files]
        self.dctf_files.extend(new)
        for p in new: self.lb_dctf.insert("end", Path(p).name)
        self._check_ready()
        self._set_status(f"{len(new)} DCTF(s) adicionado(s).")

    def _check_ready(self):
        if self.darf_files and self.dcomp_files:
            self.btn_run.config(state="normal")

    def _run(self):
        if self._processing: return
        self._processing = True
        self.btn_run.config(state="disabled")
        self.progress.start(12)
        threading.Thread(target=self._run_thread, daemon=True).start()

    def _run_thread(self):
        erros_parse = []
        try:
            self._set_status("Processando DARFs...")
            darf_rows = []
            for p in self.darf_files:
                try:
                    darf_rows.extend(parse_darf_pdf(p))
                except Exception as e:
                    erros_parse.append(f"DARF  {Path(p).name}: {type(e).__name__}: {e}")

            self._set_status("Processando PERDCOMPs...")
            dcomp_rows = []
            for p in self.dcomp_files:
                try:
                    raw = parse_pdf(p)
                    flat = flatten_rows(raw, Path(p).name)
                    # Filtra apenas linhas de DÉBITO (tipo_pedido DCOMP + tipo_registro Débito)
                    for r in flat:
                        tp = r.get("tipo_pedido","").upper()
                        tr = r.get("tipo_registro","")
                        if re.search(r"DCOMP|COMPENSA", tp, re.IGNORECASE) and tr == "Débito":
                            dcomp_rows.append(r)
                except Exception as e:
                    erros_parse.append(f"PERDCOMP {Path(p).name}: {type(e).__name__}: {e}")

            # Processa DCTFs (opcional — só roda triplo se houver pelo menos 1 DCTF)
            # MODO SERIAL: a paralelização causou crash do aplicativo no Windows
            # (provavelmente Tesseract/Poppler/pdfplumber.to_image não são
            # 100% thread-safe). Voltamos a processar 1 por vez, mas com:
            #   - Log de cada etapa num arquivo-texto (para diagnosticar se crashear de novo)
            #   - try/except em volta de CADA PDF (um PDF ruim não derruba os demais)
            dctf_rows = []
            if self.dctf_files:
                import tempfile
                log_path = Path(tempfile.gettempdir()) / "agritax_dctf_log.txt"
                total = len(self.dctf_files)
                self._set_status(f"Processando {total} DCTF(s) em série (modo seguro)...")

                with open(log_path, "w", encoding="utf-8") as logf:
                    logf.write(f"=== AgriTax — Log de processamento DCTF ===\n")
                    logf.write(f"Início: {datetime.now().isoformat()}\n")
                    logf.write(f"Total DCTFs: {total}\n\n")
                    logf.flush()

                    for i, p in enumerate(self.dctf_files, 1):
                        pdf_name = Path(p).name
                        logf.write(f"[{i}/{total}] INICIANDO: {pdf_name}\n")
                        logf.flush()
                        self.after(0, self._set_status,
                            f"DCTF {i}/{total} — {pdf_name}")
                        try:
                            rows_one = extract_dctf(p)
                            dctf_rows.extend(rows_one)
                            logf.write(f"[{i}/{total}] OK: {pdf_name} — {len(rows_one)} tributo(s)\n")
                        except Exception as e:
                            import traceback as _tb
                            msg = f"{type(e).__name__}: {e}"
                            erros_parse.append(f"DCTF     {pdf_name}: {msg}")
                            logf.write(f"[{i}/{total}] ERRO: {pdf_name} — {msg}\n")
                            logf.write(_tb.format_exc() + "\n")
                        logf.flush()

                    logf.write(f"\nFim: {datetime.now().isoformat()}\n")
                    logf.write(f"Total de tributos extraídos: {len(dctf_rows)}\n")
                    logf.write(f"Erros: {len(erros_parse)}\n")

                # Se o sistema crashar no meio, o usuário pode ler o log para saber qual PDF deu problema
                self._dctf_log_path = str(log_path)

            # Conta DCOMPs únicas carregadas e filtradas por status
            dcomp_unicas = {r.get("numero_perdcomp","").strip() for r in dcomp_rows}
            dcomp_unicas.discard("")
            self._dcomp_filtradas = sum(
                1 for num in dcomp_unicas
                if _is_cancelled(num, self.status_map) or _is_retified(num, self.status_map)
            )

            self._set_status("Executando conciliação DARF × DCOMP...")
            result = run_conciliacao(darf_rows, dcomp_rows, self.status_map)
            self.result_rows = result
            self.col_filt = {}

            # Triplo cruzamento (só se tiver DCTF)
            if dctf_rows:
                self._set_status("Executando triplo DCTF × DARF × DCOMP...")
                self.triplo_rows = run_triplo_dctf_darf_dcomp(
                    dctf_rows, darf_rows, dcomp_rows, self.status_map)
                self.col_filt_t = {}
            else:
                self.triplo_rows = []
                self.col_filt_t = {}

            self.after(0, self._refresh)
            self.after(0, self._refresh_triplo)
            self.after(0, self._update_stats)

            n_duplo = sum(1 for r in result if SIT_DUPLO in r.get("situacao",""))
            n_div   = sum(1 for r in result if SIT_DIVERGENTE in r.get("situacao",""))
            status_parts = [f"{len(result)} registro(s) DARF×DCOMP",
                            f"{n_duplo} duplo(s)",
                            f"{n_div} divergente(s)"]
            if self.triplo_rows:
                status_parts.append(f"{len(self.triplo_rows)} linha(s) triplo")
            if self._dcomp_filtradas:
                status_parts.append(f"{self._dcomp_filtradas} DCOMP(s) ignorada(s) por status")
            if erros_parse:
                status_parts.append(f"{len(erros_parse)} erro(s) de leitura")

            self.after(0, self._set_status,
                "Conciliação concluída — " + " | ".join(status_parts))

            if result or self.triplo_rows:
                self.after(0, lambda: self.btn_export.config(state="normal"))
            # Mostra os erros acumulados em popup (não mais engolidos silenciosamente)
            if erros_parse:
                amostra = "\n".join(erros_parse[:15])
                sufixo = f"\n\n... e mais {len(erros_parse)-15} erro(s)." if len(erros_parse) > 15 else ""
                self.after(0, lambda: messagebox.showwarning(
                    "Avisos de leitura",
                    f"Alguns PDFs tiveram problemas na extração:\n\n{amostra}{sufixo}",
                    parent=self))
        except Exception as e:
            import traceback as _tb
            tb = _tb.format_exc()
            self.after(0, self._set_status, f"Erro: {type(e).__name__}: {e}")
            self.after(0, lambda: messagebox.showerror(
                "Erro na conciliação", f"{type(e).__name__}: {e}\n\n{tb}",
                parent=self))
        finally:
            self._processing = False
            self.after(0, self.progress.stop)
            self.after(0, lambda: self.btn_run.config(state="normal"))

    def _refresh(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        filt = self.filt_var.get()
        for row in self.result_rows:
            sit = row.get("situacao","")
            # Filtro por situação (combobox rápido)
            if filt != "Todos" and sit != filt:
                continue
            # Filtros de coluna estilo Excel (acumulativos)
            if not self._col_match(row):
                continue
            vals = []
            for key,_,_ in CONC_COLS:
                v = row.get(key,"")
                vals.append(self._fmt(v) if key in CONC_MONEY else (v or ""))
            self.tree.insert("","end", values=vals, tags=(self._tag_for(sit),))

    def _refresh_triplo(self):
        """Popula a Treeview do triplo cruzamento DCTF × DARF × DCOMP."""
        for i in self.tree_t.get_children(): self.tree_t.delete(i)
        for idx, row in enumerate(self.triplo_rows):
            # Filtros de coluna estilo Excel (acumulativos)
            if not self._col_match_t(row):
                continue
            vals = []
            for key,_,_ in TRIPLO_COLS:
                v = row.get(key,"")
                vals.append(self._fmt(v) if key in TRIPLO_MONEY else (v or ""))
            # Tag conforme situação
            sit = row.get("situacao_triplo", "")
            if   SIT_T_QUITADO  in sit: tag = "t_quitado"
            elif SIT_T_SALDO    in sit: tag = "t_saldo"
            elif SIT_T_A_MAIOR  in sit: tag = "t_amaior"
            elif SIT_T_SEM_DECL in sit: tag = "t_sem_decl"
            else: tag = ""
            self.tree_t.insert("","end", values=vals, tags=(tag,) if tag else ())

    def _update_stats(self):
        n_total = len(self.result_rows)
        n_duplo = sum(1 for r in self.result_rows if SIT_DUPLO      in r.get("situacao",""))
        n_div   = sum(1 for r in self.result_rows if SIT_DIVERGENTE in r.get("situacao",""))
        n_filt  = getattr(self, "_dcomp_filtradas", 0)
        n_triplo= len(self.triplo_rows)
        self.lbl_stats.config(text=(
            f"DARFs carregados : {len(self.darf_files)}\n"
            f"DCOMPs carregados: {len(self.dcomp_files)}\n"
            f"DCTFs carregados : {len(self.dctf_files)}\n"
            f"DCOMPs filtradas : {n_filt}\n"
            f"Linhas conciliação: {n_total}\n"
            f"Linhas triplo    : {n_triplo}\n"
            f"Duplo pagamento  : {n_duplo}\n"
            f"Divergentes      : {n_div}"
        ))

    def _export(self):
        if not self.result_rows and not self.triplo_rows:
            messagebox.showinfo("Sem dados","Execute a conciliação primeiro.", parent=self)
            return
        path = filedialog.asksaveasfilename(parent=self,
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"Conciliacao_DARF_DCOMP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            export_conciliacao_excel(self.result_rows, path, triplo_rows=self.triplo_rows)
            self._set_status(f"Excel exportado: {Path(path).name}")
            if messagebox.askyesno("Exportado!", f"Arquivo salvo:\n{path}\n\nDeseja abrir agora?", parent=self):
                import subprocess, platform
                if platform.system()=="Windows": os.startfile(path)
                elif platform.system()=="Darwin": subprocess.call(["open",path])
                else: subprocess.call(["xdg-open",path])
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e), parent=self)

    def _clear(self):
        self.darf_files.clear(); self.dcomp_files.clear(); self.dctf_files.clear()
        self.result_rows.clear(); self.triplo_rows.clear()
        self.status_file = ""; self.status_map.clear()
        self.col_filt.clear(); self.col_filt_t.clear()
        self._dcomp_filtradas = 0
        self.lb_darf.delete(0,"end"); self.lb_dcomp.delete(0,"end"); self.lb_dctf.delete(0,"end")
        self.lbl_status_file.config(text="(nenhuma planilha carregada)", fg=C_GRAY)
        for i in self.tree.get_children(): self.tree.delete(i)
        for i in self.tree_t.get_children(): self.tree_t.delete(i)
        self.btn_run.config(state="disabled"); self.btn_export.config(state="disabled")
        self._update_stats(); self._set_status("Dados limpos.")
        self._update_headings()
        self._update_headings_t()

    def _load_status(self):
        """Carrega planilha de status do eCAC (xlsx) para filtrar canceladas/retificadas."""
        path = filedialog.askopenfilename(parent=self,
            title="Planilha de status do eCAC (xlsx)",
            filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not path:
            return
        try:
            sm = parse_status_excel(path)
            if not sm:
                messagebox.showwarning("Planilha vazia",
                    "Não foi possível identificar PERDCOMPs na planilha.", parent=self)
                return
            self.status_file = path
            self.status_map  = sm
            self.lbl_status_file.config(
                text=f"✓ {Path(path).name}  ({len(sm)} registros)",
                fg=C_GREEN_DARK)
            # Conta quantos seriam filtrados na próxima execução
            cnt_c = sum(1 for v in sm.values() if "cancelad" in v.get("situacao","").lower())
            cnt_r = sum(1 for v in sm.values() if "retificad" in v.get("situacao","").lower())
            self._set_status(
                f"Planilha carregada — {len(sm)} PERDCOMPs | "
                f"{cnt_c} cancelado(s) | {cnt_r} retificado(s) serão ignorados")
        except Exception as e:
            messagebox.showerror("Erro ao ler planilha", str(e), parent=self)

    # ── Filtros de coluna estilo Excel ─────────────────────────────────────────
    def _update_headings(self):
        """Atualiza ícones ▼ / 🔽 no cabeçalho conforme filtro ativo."""
        if not hasattr(self, "tree"):
            return
        for key, label, _ in CONC_COLS:
            icon = " 🔽" if key in self.col_filt else "  ▼"
            self.tree.heading(key, text=f"{label}{icon}")

    def _open_col_filter(self, col_key: str, col_label: str, event=None):
        """Abre popup de filtro (com busca substring) para a coluna."""
        # Valores únicos da coluna no conjunto atual de resultados
        vals = set()
        for r in self.result_rows:
            v = r.get(col_key, "")
            if col_key in CONC_MONEY:
                try: v = self._fmt(v)
                except Exception: v = str(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)
        active = self.col_filt.get(col_key)
        px, py = (event.x_root, event.y_root) if event else (self.winfo_pointerx(), self.winfo_pointery())

        def on_apply(selected):
            if selected is None:
                self.col_filt.pop(col_key, None)
            else:
                self.col_filt[col_key] = selected
            self._refresh()
            self._update_headings()

        ColFilterPopup(self, col_label, vals, active, on_apply, px, py)

    def _col_match(self, row: dict) -> bool:
        """Retorna True se row atende TODOS os filtros de coluna ativos."""
        for key, val_set in self.col_filt.items():
            if not val_set:
                continue
            v = row.get(key, "")
            if key in CONC_MONEY:
                try: v = self._fmt(v)
                except Exception: v = str(v)
            else:
                v = str(v) if v is not None else ""
            if v not in val_set:
                return False
        return True

    # ── Filtros de coluna do TRIPLO ────────────────────────────────────────────
    def _update_headings_t(self):
        if not hasattr(self, "tree_t"):
            return
        for key, label, _ in TRIPLO_COLS:
            icon = " 🔽" if key in self.col_filt_t else "  ▼"
            self.tree_t.heading(key, text=f"{label}{icon}")

    def _open_col_filter_t(self, col_key: str, col_label: str, event=None):
        vals = set()
        for r in self.triplo_rows:
            v = r.get(col_key, "")
            if col_key in TRIPLO_MONEY:
                try: v = self._fmt(v)
                except Exception: v = str(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)
        active = self.col_filt_t.get(col_key)
        px, py = (event.x_root, event.y_root) if event else (self.winfo_pointerx(), self.winfo_pointery())

        def on_apply(selected):
            if selected is None:
                self.col_filt_t.pop(col_key, None)
            else:
                self.col_filt_t[col_key] = selected
            self._refresh_triplo()
            self._update_headings_t()

        ColFilterPopup(self, col_label, vals, active, on_apply, px, py)

    def _col_match_t(self, row: dict) -> bool:
        for key, val_set in self.col_filt_t.items():
            if not val_set:
                continue
            v = row.get(key, "")
            if key in TRIPLO_MONEY:
                try: v = self._fmt(v)
                except Exception: v = str(v)
            else:
                v = str(v) if v is not None else ""
            if v not in val_set:
                return False
        return True

    def _set_status(self, msg):
        self.status_var.set(msg)
        try: self.update_idletasks()
        except Exception: pass


# =============================================================================
# Views — Frames embarcáveis que consomem o DataStore
# =============================================================================

class CentralImportView:
    """Aba 'Central de Importação': único local onde importações acontecem.

    Seções:
      1. PERDCOMPs  (PDFs)
      2. DARFs/DAS  (PDFs)
      3. DCTFs      (PDFs — OCR)
      4. Planilha de Status (eCAC — xlsx)
    """

    def __init__(self, parent, datastore: DataStore, mainapp):
        self.parent = parent
        self.ds = datastore
        self.mainapp = mainapp
        self._processing = False
        # Filas temporárias de arquivos ADICIONADOS mas ainda NÃO processados
        self._pending_perdcomp: list = []
        self._pending_darf:     list = []
        self._pending_dctf:     list = []
        self._pending_dctfweb:  list = []
        self._pending_efd:      list = []
        self._build_ui()
        # Subscribe pra redesenhar contadores se algo mudar
        self.ds.subscribe("any", self._refresh_counters)

    def _build_ui(self):
        # Canvas com scroll para caber tudo
        canvas = tk.Canvas(self.parent, bg=C_GRAY_LIGHT, highlightthickness=0)
        vsb = ttk.Scrollbar(self.parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        body = tk.Frame(canvas, bg=C_GRAY_LIGHT)
        win_id = canvas.create_window((0, 0), window=body, anchor="nw")

        def _cfg(_=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=canvas.winfo_width())
        body.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Título
        tk.Label(body, text="Central de Importação",
                 bg=C_GRAY_LIGHT, fg=C_GREEN_DARK,
                 font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(10, 4))
        tk.Label(body,
                 text="Todos os documentos processados aqui ficam disponíveis "
                      "automaticamente para os módulos de Auditoria.",
                 bg=C_GRAY_LIGHT, fg=C_GRAY,
                 font=("Segoe UI", 9)).pack(anchor="w", padx=16, pady=(0, 10))

        # ── Barra de ação global ───────────────────────────────────────────
        # Botão "Processar Tudo" — processa todas as filas pendentes numa única ação
        action_bar = tk.Frame(body, bg=C_GREEN_LIGHT,
                               highlightbackground=C_GREEN_DARK, highlightthickness=1)
        action_bar.pack(fill="x", padx=16, pady=(0, 14))

        tk.Label(action_bar,
            text="Depois de adicionar os PDFs em cada seção abaixo, clique aqui para processar todos de uma vez:",
            bg=C_GREEN_LIGHT, fg=C_GREEN_DARK,
            font=("Segoe UI", 9),
            anchor="w").pack(side="left", fill="x", expand=True, padx=12, pady=8)

        self.btn_process_all = tk.Button(action_bar,
            text="▶  Processar Tudo",
            bg=C_GREEN_DARK, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 10, "bold"), padx=20, pady=8,
            command=self._process_all)
        self.btn_process_all.pack(side="right", padx=10, pady=6)

        # ── Seção 1: PERDCOMPs ─────────────────────────────────────────────
        self._build_section(body, "1. PERDCOMPs (PDFs)",
            kind="perdcomp",
            hint="Pedidos de Restituição, Ressarcimento e Declarações de Compensação do eCAC.")

        # ── Seção 2: DARFs/DAS ─────────────────────────────────────────────
        self._build_section(body, "2. DARFs / DAS (PDFs)",
            kind="darf",
            hint="Comprovantes de arrecadação da Receita Federal.")

        # ── Seção 3: DCTFs ─────────────────────────────────────────────────
        self._build_section(body, "3. DCTFs (PDFs)",
            kind="dctf",
            hint="Declarações de Débitos e Créditos Tributários Federais. "
                 "⚠ Usa OCR — processamento mais lento.")

        # ── Seção 4: DCTFWeb ───────────────────────────────────────────────
        self._build_section(body, "4. DCTFWeb Extractor",
            kind="dctfweb",
            hint="Declarações DCTFWeb (Declaração Completa) — "
                 "tributos previdenciários (CP) e IRRF. "
                 "PDF de texto puro, sem necessidade de OCR.")

        # ── Seção 5: EFD Contribuições ─────────────────────────────────────
        self._build_section(body, "5. EFD Contribuições (SPED)",
            kind="efd",
            hint="Escrituração Fiscal Digital das Contribuições para o "
                 "PIS/Pasep e COFINS. Arquivos .txt no padrão SPED. "
                 "Permite confronto entre o apurado na escrituração e o "
                 "declarado em DCTF + DCTFWeb.")

        # ── Seção 6: Planilha de Status ────────────────────────────────────
        self._build_status_section(body)

        # ── Resumo geral ───────────────────────────────────────────────────
        rs = tk.Frame(body, bg=C_GREEN_LIGHT,
                      highlightbackground=C_BORDER, highlightthickness=1)
        rs.pack(fill="x", padx=16, pady=12)
        tk.Label(rs, text="Resumo da Sessão",
                 bg=C_GREEN_LIGHT, fg=C_GREEN_DARK,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10, pady=(8, 2))
        self.lbl_summary = tk.Label(rs, text="",
                                     bg=C_GREEN_LIGHT, fg=C_GRAY_DARK,
                                     font=("Courier", 9), justify="left")
        self.lbl_summary.pack(anchor="w", padx=10, pady=(0, 8))

        # Botão Limpar Tudo
        btn_clear_all = tk.Button(body, text="🗑  Limpar Todos os Dados da Sessão",
            bg=C_RED, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=6,
            command=self._clear_all)
        btn_clear_all.pack(padx=16, pady=(0, 20))

        self._refresh_counters()

    def _build_section(self, parent, title, kind, hint):
        """Constrói uma seção de importação (PERDCOMP/DARF/DCTF)."""
        sec = tk.Frame(parent, bg=C_WHITE,
                       highlightbackground=C_BORDER, highlightthickness=1)
        sec.pack(fill="x", padx=16, pady=(0, 10))

        hdr = tk.Frame(sec, bg=C_GREEN_DARK)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"  {title}",
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 10, "bold"),
                 anchor="w").pack(side="left", fill="x", expand=True, pady=6)

        contador_var = tk.StringVar(value="0 arquivos")
        tk.Label(hdr, textvariable=contador_var,
                 bg=C_GREEN_DARK, fg=C_GREEN_MID,
                 font=("Segoe UI", 9)).pack(side="right", padx=10, pady=6)
        setattr(self, f"_cnt_var_{kind}", contador_var)

        tk.Label(sec, text=hint,
                 bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI", 8), anchor="w",
                 wraplength=1000, justify="left").pack(fill="x", padx=10, pady=(6, 4))

        # Barra de botões
        bf = tk.Frame(sec, bg=C_WHITE)
        bf.pack(fill="x", padx=10, pady=4)

        btn_add = tk.Button(bf, text="📄  Adicionar PDFs...",
            bg=C_GREEN, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=5,
            command=lambda k=kind: self._add_files(k))
        btn_add.pack(side="left", padx=(0, 6))

        btn_proc = tk.Button(bf, text="▶  Processar Pendentes",
            bg=C_BLUE, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=5,
            command=lambda k=kind: self._process(k),
            state="disabled")
        btn_proc.pack(side="left", padx=(0, 6))
        setattr(self, f"_btn_proc_{kind}", btn_proc)

        btn_clear = tk.Button(bf, text="🗑  Limpar",
            bg=C_RED, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9), padx=10, pady=5,
            command=lambda k=kind: self._clear_kind(k))
        btn_clear.pack(side="left")

        # Progressbar
        pb = ttk.Progressbar(bf, mode="indeterminate", length=180)
        pb.pack(side="right", padx=10)
        setattr(self, f"_pb_{kind}", pb)

        # Listbox dos arquivos (pendentes + processados)
        lf = tk.Frame(sec, bg=C_WHITE, height=130)
        lf.pack(fill="x", padx=10, pady=(4, 10))
        lf.pack_propagate(False)
        sb = ttk.Scrollbar(lf)
        sb.pack(side="right", fill="y")
        lb = tk.Listbox(lf, yscrollcommand=sb.set,
                         font=("Segoe UI", 8), bg=C_GRAY_LIGHT,
                         fg=C_GRAY_DARK, relief="flat",
                         selectbackground=C_GREEN_MID)
        lb.pack(fill="both", expand=True)
        sb.config(command=lb.yview)
        setattr(self, f"_lb_{kind}", lb)

    def _build_status_section(self, parent):
        sec = tk.Frame(parent, bg=C_WHITE,
                       highlightbackground=C_BORDER, highlightthickness=1)
        sec.pack(fill="x", padx=16, pady=(0, 10))

        hdr = tk.Frame(sec, bg=C_GREEN_DARK)
        hdr.pack(fill="x")
        tk.Label(hdr, text="  6. Planilha de Status (eCAC) — opcional",
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 10, "bold"),
                 anchor="w").pack(side="left", fill="x", expand=True, pady=6)

        tk.Label(sec,
                 text="Exportada do eCAC, permite filtrar DCOMPs canceladas e "
                      "retificadas no Controle de Créditos e na Conciliação.",
                 bg=C_WHITE, fg=C_GRAY,
                 font=("Segoe UI", 8), anchor="w",
                 wraplength=1000, justify="left").pack(fill="x", padx=10, pady=(6, 4))

        bf = tk.Frame(sec, bg=C_WHITE)
        bf.pack(fill="x", padx=10, pady=(4, 10))

        btn_load = tk.Button(bf, text="📂  Carregar / Substituir Planilha",
            bg=C_GREEN, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=5,
            command=self._load_status)
        btn_load.pack(side="left", padx=(0, 6))

        btn_clear = tk.Button(bf, text="🗑  Limpar",
            bg=C_RED, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9), padx=10, pady=5,
            command=self._clear_status)
        btn_clear.pack(side="left")

        self._lbl_status = tk.Label(bf,
            text="  Nenhuma planilha carregada",
            bg=C_WHITE, fg=C_GRAY,
            font=("Segoe UI", 9, "italic"))
        self._lbl_status.pack(side="left", padx=14)

    # ── Ações ──────────────────────────────────────────────────────────────
    def _add_files(self, kind: str):
        title = {"perdcomp": "PDFs de PERDCOMP",
                 "darf":     "PDFs de DARF/DAS",
                 "dctf":     "PDFs de DCTF",
                 "dctfweb":  "PDFs de DCTFWeb (Declaração Completa)",
                 "efd":      "Arquivos SPED da EFD Contribuições (.txt)"}[kind]
        if kind == "efd":
            ftypes = [("SPED EFD (TXT)", "*.txt"), ("Todos", "*.*")]
        else:
            ftypes = [("PDF", "*.pdf"), ("Todos", "*.*")]
        paths = filedialog.askopenfilenames(parent=self.parent,
            title=title, filetypes=ftypes)
        if not paths:
            return

        pending = getattr(self, f"_pending_{kind}")
        existing = set(pending)
        # Também ignora arquivos já processados
        already_done = {f["path"] for f in getattr(self.ds, f"{kind}_files")}
        new = [p for p in paths if p not in existing and p not in already_done]
        pending.extend(new)

        self._refresh_listbox(kind)
        # Habilita botão processar se há pendentes
        btn = getattr(self, f"_btn_proc_{kind}")
        btn.config(state="normal" if pending else "disabled")

        dup = len(paths) - len(new)
        self.mainapp.set_status(
            f"{len(new)} arquivo(s) adicionado(s) a {kind.upper()}"
            + (f"  |  {dup} duplicado(s) ignorado(s)" if dup else ""))

        # Avisa se houve duplicados ignorados (mesmo arquivo já na fila ou processado)
        if dup > 0:
            from os.path import basename
            duplicados_paths = [p for p in paths
                                 if p in existing or p in already_done]
            detalhe = "\n".join(f"  • {basename(p)}" for p in duplicados_paths[:8])
            if dup > 8:
                detalhe += f"\n  ... e mais {dup-8}"
            messagebox.showwarning("Arquivos duplicados ignorados",
                f"⚠ {dup} arquivo(s) já estavam na fila ou foram processados "
                f"anteriormente e foram IGNORADOS:\n\n{detalhe}",
                parent=self.parent)

    def _process_all(self):
        """Processa todas as filas pendentes (PERDCOMP + DARF + DCTF) em sequência.

        Executa numa única thread, um tipo por vez (serial), para não sobrecarregar
        o sistema nem causar os crashes de paralelismo com OCR no Windows. Mostra
        progresso consolidado e desabilita todos os botões enquanto processa.

        Antes de iniciar, verifica se há tipos de documento SEM dados (nem na fila
        pendente, nem já processados no DataStore) e alerta o usuário para que ele
        possa confirmar se quer seguir mesmo assim ou cancelar para adicionar.
        """
        if self._processing:
            messagebox.showinfo("Aguarde",
                "Já há um processamento em andamento.", parent=self.parent)
            return

        # Coleta todas as filas pendentes
        kinds_to_process = []
        for kind in ("perdcomp", "darf", "dctf", "dctfweb", "efd"):
            pending = getattr(self, f"_pending_{kind}")
            if pending:
                kinds_to_process.append((kind, list(pending)))

        if not kinds_to_process:
            messagebox.showinfo("Sem pendentes",
                "Não há arquivos pendentes para processar.\n"
                "Adicione PDFs em cada seção primeiro.",
                parent=self.parent)
            return

        # ── Validação de completude ─────────────────────────────────────────
        # Verifica quais tipos NÃO têm dados (nem pendentes, nem já no DataStore).
        # Se algum faltar, avisa o usuário — ele pode ter esquecido de adicionar.
        # Regras:
        #   - PERDCOMP: recomendado para maioria das análises
        #   - DARF:     recomendado (Conciliação DARF×DCOMP precisa)
        #   - DCTF:     opcional, mas Triplo Cruzamento precisa dela
        #   - Status:   opcional (filtro de canceladas/retificadas)
        labels = {"perdcomp": "PERDCOMPs",
                  "darf":     "DARFs / DAS",
                  "dctf":     "DCTFs",
                  "dctfweb":  "DCTFWeb",
                  "efd":      "EFD Contribuições"}
        faltando = []
        for kind in ("perdcomp", "darf", "dctf", "dctfweb", "efd"):
            pending = getattr(self, f"_pending_{kind}")
            ja_processados = getattr(self.ds, f"{kind}_files")
            if not pending and not ja_processados:
                faltando.append(labels[kind])

        # Status é um caso à parte — só avisa se não está carregado
        status_faltando = not self.ds.status_map

        if faltando or status_faltando:
            partes_msg = []
            if faltando:
                partes_msg.append(
                    "Os seguintes tipos de documento NÃO foram adicionados:\n\n"
                    + "\n".join(f"   • {nome}" for nome in faltando))
            if status_faltando:
                partes_msg.append(
                    "A Planilha de Status do eCAC também não foi carregada.\n"
                    "Sem ela, DCOMPs canceladas/retificadas não serão filtradas\n"
                    "na Conciliação nem no Controle de Créditos.")

            # Impactos práticos de cada faltante, para informar a decisão
            impactos = []
            if "PERDCOMPs" in faltando:
                impactos.append(
                    "• Sem PERDCOMPs: não haverá análise de Detalhamento, Controle\n"
                    "  de Créditos, Ressarcimentos, nem cruzamento DARF×DCOMP.")
            if "DARFs / DAS" in faltando:
                impactos.append(
                    "• Sem DARFs: Conciliação (DARF×DCOMP) e Triplo (DCTF×DARF×DCOMP)\n"
                    "  ficarão incompletos ou vazios.")
            if "DCTFs" in faltando:
                impactos.append(
                    "• Sem DCTFs: aba Triplo Cruzamento (DCTF×DARF×DCOMP) ficará vazia.")
            if "DCTFWeb" in faltando:
                impactos.append(
                    "• Sem DCTFWeb: análises de tributos previdenciários (CP) e\n"
                    "  IRRF ficarão sem visualização dedicada.")
            if "EFD Contribuições" in faltando:
                impactos.append(
                    "• Sem EFD Contribuições: confronto EFD × DCTF/DCTFWeb\n"
                    "  (validação de PIS/COFINS) ficará vazio.")

            msg = ("⚠  Validação antes do processamento\n\n"
                   + "\n\n".join(partes_msg))
            if impactos:
                msg += "\n\nImpactos das ausências:\n" + "\n".join(impactos)
            msg += "\n\nDeseja processar assim mesmo (SIM) ou cancelar para\nadicionar os documentos faltantes (NÃO)?"

            if not messagebox.askyesno(
                    "Documentos faltantes — confirmar processamento", msg,
                    parent=self.parent, icon="warning"):
                # Usuário optou por cancelar para completar a importação
                self.mainapp.set_status(
                    "Processamento cancelado — adicione os documentos faltantes.")
                return

        # Confirma se há muito trabalho a fazer (principalmente DCTF que é lento)
        total_files = sum(len(pdfs) for _, pdfs in kinds_to_process)
        dctf_count = sum(len(pdfs) for k, pdfs in kinds_to_process if k == "dctf")
        if dctf_count > 10:
            estimated_min = dctf_count * 0.3  # ~18s por DCTF em 200 DPI
            if not messagebox.askyesno("Confirmar processamento",
                    f"Serão processados {total_files} arquivo(s) no total.\n\n"
                    f"Inclui {dctf_count} DCTFs — o OCR pode levar "
                    f"cerca de {estimated_min:.0f} minuto(s).\n\n"
                    f"Deseja continuar?",
                    parent=self.parent):
                return

        self._processing = True
        self.btn_process_all.config(state="disabled")
        # Desabilita também os botões individuais
        for k in ("perdcomp", "darf", "dctf", "dctfweb", "efd"):
            btn = getattr(self, f"_btn_proc_{k}", None)
            if btn: btn.config(state="disabled")

        # Inicia as barras de progresso dos tipos que serão processados
        for kind, _ in kinds_to_process:
            pb = getattr(self, f"_pb_{kind}")
            pb.start(12)

        threading.Thread(target=self._process_all_thread,
                          args=(kinds_to_process,), daemon=True).start()

    def _process_all_thread(self, kinds_to_process: list):
        """Thread que processa cada tipo em sequência, reaproveitando _process_thread."""
        try:
            total_types = len(kinds_to_process)
            all_ok = 0
            all_err = 0
            all_erros_detail = []
            # Acumuladores de duplicidade (preenchidos na thread principal via _apply)
            all_dup_summary = {"total_ignoradas": 0, "por_kind": {}, "chaves_dup": []}

            for idx, (kind, pdfs) in enumerate(kinds_to_process, 1):
                self.mainapp.set_status(
                    f"[{idx}/{total_types}] Processando {len(pdfs)} {kind.upper()}(s)...")

                new_rows = []
                file_records = []
                total = len(pdfs)
                for i, path in enumerate(pdfs, 1):
                    nome = Path(path).name
                    self.mainapp.set_status(
                        f"[{idx}/{total_types}] {kind.upper()} {i}/{total} — {nome}")
                    try:
                        if kind == "perdcomp":
                            # parse_pdf retorna LISTA de dicts (geralmente 1 elemento)
                            parsed_list = parse_pdf(path)
                            for raw in parsed_list:
                                raw["_source"] = nome
                                new_rows.append(raw)
                        elif kind == "darf":
                            rows = parse_darf_pdf(path)
                            new_rows.extend(rows)
                        elif kind == "dctf":
                            rows = extract_dctf(path)
                            new_rows.extend(rows)
                        elif kind == "dctfweb":
                            rows = extract_dctfweb(path)
                            new_rows.extend(rows)
                        elif kind == "efd":
                            rows = extract_efd_contribuicoes(path)
                            new_rows.extend(rows)
                        file_records.append({
                            "path": path, "nome": nome,
                            "status": "✓", "erro": "",
                        })
                        all_ok += 1
                    except Exception as e:
                        err = f"{type(e).__name__}: {e}"
                        file_records.append({
                            "path": path, "nome": nome,
                            "status": "✗", "erro": err,
                        })
                        all_err += 1
                        all_erros_detail.append(f"• [{kind.upper()}] {nome}: {err}")

                # Aplica no DataStore (thread-safe via .after)
                def _apply(k=kind, nr=new_rows, fr=file_records):
                    info = None
                    if k == "perdcomp":
                        info = self.ds.add_perdcomps(nr, fr)
                    elif k == "darf":
                        info = self.ds.add_darfs(nr, fr)
                    elif k == "dctf":
                        info = self.ds.add_dctfs(nr, fr)
                    elif k == "dctfweb":
                        info = self.ds.add_dctfwebs(nr, fr)
                    elif k == "efd":
                        info = self.ds.add_efds(nr, fr)
                    getattr(self, f"_pending_{k}").clear()
                    self._refresh_listbox(k)
                    # Acumula info de duplicados
                    if info and info.get("ignoradas", 0) > 0:
                        all_dup_summary["total_ignoradas"] += info["ignoradas"]
                        all_dup_summary["por_kind"][k] = (
                            all_dup_summary["por_kind"].get(k, 0) + info["ignoradas"])
                        for c in info.get("chaves_dup", []):
                            all_dup_summary["chaves_dup"].append(f"[{k.upper()}] {c}")
                self.parent.after(0, _apply)

            # Finaliza — status consolidado
            def _finalize():
                msg = f"✓ Processamento concluído — {all_ok} OK"
                if all_dup_summary["total_ignoradas"]:
                    msg += f", {all_dup_summary['total_ignoradas']} duplicado(s) ignorado(s)"
                if all_err:
                    msg += f", {all_err} com erro"
                self.mainapp.set_status(msg)

                # Aviso consolidado de duplicados (se houver)
                if all_dup_summary["total_ignoradas"]:
                    por_kind = " | ".join(
                        f"{k.upper()}: {n}" for k, n in all_dup_summary["por_kind"].items())
                    chaves = all_dup_summary["chaves_dup"][:8]
                    detalhe = "\n".join(f"  • {c}" for c in chaves)
                    if len(all_dup_summary["chaves_dup"]) > 8:
                        detalhe += f"\n  ... e mais {len(all_dup_summary['chaves_dup'])-8}"
                    messagebox.showwarning(
                        "Duplicidades detectadas e ignoradas",
                        f"⚠ {all_dup_summary['total_ignoradas']} item(ns) já estavam "
                        f"importados e foram IGNORADOS para evitar contagem em duplicidade.\n\n"
                        f"Por tipo: {por_kind}\n\n"
                        f"Exemplos:\n{detalhe}",
                        parent=self.parent)

                if all_erros_detail:
                    messagebox.showwarning(
                        f"Processamento concluído com {all_err} erro(s)",
                        "Alguns arquivos não puderam ser processados:\n\n"
                        + "\n".join(all_erros_detail[:20])
                        + (f"\n\n... e mais {len(all_erros_detail)-20}." if len(all_erros_detail) > 20 else ""),
                        parent=self.parent)
                elif not all_dup_summary["total_ignoradas"]:
                    messagebox.showinfo("Sucesso",
                        f"Todos os {all_ok} arquivos foram processados com êxito!\n\n"
                        "Vá à aba '🔍 Auditoria' para ver os resultados.",
                        parent=self.parent)

            self.parent.after(0, _finalize)

        except Exception as e:
            import traceback as _tb
            tb = _tb.format_exc()
            self.parent.after(0, lambda: messagebox.showerror(
                "Erro no processamento em lote",
                f"{type(e).__name__}: {e}\n\n{tb}",
                parent=self.parent))
        finally:
            def _done():
                self._processing = False
                self.btn_process_all.config(state="normal")
                # Para todas as progressbars e reavalia estado dos botões individuais
                for k in ("perdcomp", "darf", "dctf", "dctfweb", "efd"):
                    pb = getattr(self, f"_pb_{k}", None)
                    if pb: pb.stop()
                    btn = getattr(self, f"_btn_proc_{k}", None)
                    pending = getattr(self, f"_pending_{k}", [])
                    if btn: btn.config(state="normal" if pending else "disabled")
            self.parent.after(0, _done)

    def _process(self, kind: str):
        if self._processing:
            messagebox.showinfo("Aguarde",
                "Já há um processamento em andamento.", parent=self.parent)
            return

        pending = getattr(self, f"_pending_{kind}")
        if not pending:
            return

        self._processing = True
        btn = getattr(self, f"_btn_proc_{kind}")
        btn.config(state="disabled")
        pb = getattr(self, f"_pb_{kind}")
        pb.start(12)

        pdfs = list(pending)
        threading.Thread(target=self._process_thread,
                         args=(kind, pdfs), daemon=True).start()

    def _process_thread(self, kind: str, pdfs: list):
        try:
            new_rows = []
            file_records = []
            total = len(pdfs)
            for i, path in enumerate(pdfs, 1):
                nome = Path(path).name
                self.mainapp.set_status(
                    f"Processando {kind.upper()} {i}/{total} — {nome}")
                try:
                    if kind == "perdcomp":
                        # parse_pdf retorna uma LISTA de dicts (geralmente com 1 elemento)
                        parsed_list = parse_pdf(path)
                        for raw in parsed_list:
                            raw["_source"] = nome
                            new_rows.append(raw)
                    elif kind == "darf":
                        rows = parse_darf_pdf(path)
                        new_rows.extend(rows)
                    elif kind == "dctf":
                        rows = extract_dctf(path)
                        new_rows.extend(rows)
                    elif kind == "dctfweb":
                        rows = extract_dctfweb(path)
                        new_rows.extend(rows)
                    elif kind == "efd":
                        rows = extract_efd_contribuicoes(path)
                        new_rows.extend(rows)
                    file_records.append({
                        "path": path, "nome": nome,
                        "status": "✓", "erro": "",
                    })
                except Exception as e:
                    file_records.append({
                        "path": path, "nome": nome,
                        "status": "✗", "erro": f"{type(e).__name__}: {e}",
                    })

            # Atualiza o DataStore (executa na thread principal)
            def _apply():
                info = None
                if kind == "perdcomp":
                    info = self.ds.add_perdcomps(new_rows, file_records)
                elif kind == "darf":
                    info = self.ds.add_darfs(new_rows, file_records)
                elif kind == "dctf":
                    info = self.ds.add_dctfs(new_rows, file_records)
                elif kind == "dctfweb":
                    info = self.ds.add_dctfwebs(new_rows, file_records)
                elif kind == "efd":
                    info = self.ds.add_efds(new_rows, file_records)
                getattr(self, f"_pending_{kind}").clear()
                self._refresh_listbox(kind)
                ok = sum(1 for f in file_records if f["status"] == "✓")
                err = len(file_records) - ok
                # Conta arquivos marcados como duplicados (status "⚠")
                dup_files = info.get("arquivos_dup", 0) if info else 0
                msg = f"✓ {kind.upper()} processados — {ok} OK"
                if dup_files: msg += f", {dup_files} duplicado(s) ignorado(s)"
                if err:       msg += f", {err} com erro"
                self.mainapp.set_status(msg)
                # Avisa o usuário sobre duplicados
                if info and info["ignoradas"] > 0:
                    chaves = info["chaves_dup"][:5]
                    detalhe = "\n".join(f"  • {c}" for c in chaves)
                    if info["ignoradas"] > 5:
                        detalhe += f"\n  ... e mais {info['ignoradas']-5}"
                    messagebox.showwarning("Documentos duplicados ignorados",
                        f"⚠ {info['ignoradas']} item(ns) já estavam importados "
                        f"e foram IGNORADOS para evitar duplicidade.\n\n"
                        f"Exemplos das chaves duplicadas:\n{detalhe}",
                        parent=self.parent)
                if err:
                    erros = [f"• {f['nome']}: {f['erro']}"
                             for f in file_records if f["status"] == "✗"]
                    messagebox.showwarning("Erros na importação",
                        "Alguns PDFs não puderam ser processados:\n\n"
                        + "\n".join(erros[:15])
                        + (f"\n\n... e mais {len(erros)-15}." if len(erros) > 15 else ""),
                        parent=self.parent)

            self.parent.after(0, _apply)
        except Exception as e:
            import traceback as _tb
            tb = _tb.format_exc()
            self.parent.after(0, lambda: messagebox.showerror(
                f"Erro no processamento de {kind.upper()}",
                f"{type(e).__name__}: {e}\n\n{tb}",
                parent=self.parent))
        finally:
            def _done():
                self._processing = False
                pb = getattr(self, f"_pb_{kind}")
                pb.stop()
                btn = getattr(self, f"_btn_proc_{kind}")
                pending = getattr(self, f"_pending_{kind}")
                btn.config(state="normal" if pending else "disabled")
            self.parent.after(0, _done)

    def _refresh_listbox(self, kind: str):
        lb = getattr(self, f"_lb_{kind}")
        lb.delete(0, "end")
        # Primeiro os pendentes (⏳), depois os processados (✓ ou ✗)
        for p in getattr(self, f"_pending_{kind}"):
            lb.insert("end", f"  ⏳  {Path(p).name}")
        for f in getattr(self.ds, f"{kind}_files"):
            lb.insert("end", f"  {f['status']}  {f['nome']}")

    def _refresh_counters(self):
        s = self.ds.summary()
        for k in ("perdcomp", "darf", "dctf", "dctfweb", "efd"):
            var = getattr(self, f"_cnt_var_{k}", None)
            if var is not None:
                processed = s[f"{k}_files"]
                pending = len(getattr(self, f"_pending_{k}"))
                txt = f"{processed} processados"
                if pending:
                    txt += f"  |  {pending} pendentes"
                var.set(txt)
            # Refresh a listbox
            if hasattr(self, f"_lb_{k}"):
                self._refresh_listbox(k)

        if s["status_loaded"]:
            self._lbl_status.config(
                text=f"  ✓  {Path(self.ds.status_path).name}  "
                     f"({s['status_registros']} registros)",
                fg=C_GREEN_DARK)
        else:
            self._lbl_status.config(
                text="  Nenhuma planilha carregada",
                fg=C_GRAY)

        # Resumo geral
        self.lbl_summary.config(text=(
            f"PERDCOMPs      : {s['perdcomp_files']:>4} arquivos  |  {s['perdcomp_rows']:>5} linhas\n"
            f"DARFs / DAS    : {s['darf_files']:>4} arquivos  |  {s['darf_rows']:>5} linhas\n"
            f"DCTFs          : {s['dctf_files']:>4} arquivos  |  {s['dctf_rows']:>5} linhas\n"
            f"DCTFWeb        : {s['dctfweb_files']:>4} arquivos  |  {s['dctfweb_rows']:>5} linhas\n"
            f"EFD Contribuições: {s['efd_files']:>2} arquivos  |  {s['efd_rows']:>5} linhas\n"
            f"Planilha Status: {('✓ carregada ('+str(s['status_registros'])+' regs)') if s['status_loaded'] else '✗ não carregada'}"
        ))

    def _load_status(self):
        path = filedialog.askopenfilename(parent=self.parent,
            title="Planilha de status do eCAC",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not path:
            return
        try:
            sm = parse_status_excel(path)
            if not sm:
                messagebox.showwarning("Planilha vazia",
                    "Não foi possível identificar PERDCOMPs na planilha.",
                    parent=self.parent)
                return
            self.ds.set_status_map(sm, path)
            self.mainapp.set_status(
                f"✓ Planilha de status carregada — {len(sm)} registros")
        except Exception as e:
            messagebox.showerror("Erro ao carregar", str(e), parent=self.parent)

    def _clear_status(self):
        self.ds.set_status_map({}, "")
        self.mainapp.set_status("Planilha de status removida.")

    def _clear_kind(self, kind: str):
        if not messagebox.askyesno("Limpar",
                f"Remover todos os {kind.upper()}s importados desta sessão?",
                parent=self.parent):
            return
        getattr(self, f"_pending_{kind}").clear()
        if kind == "perdcomp":
            self.ds.clear_perdcomps()
        elif kind == "darf":
            self.ds.clear_darfs()
        elif kind == "dctf":
            self.ds.clear_dctfs()
        elif kind == "dctfweb":
            self.ds.clear_dctfwebs()
        elif kind == "efd":
            self.ds.clear_efds()

    def _clear_all(self):
        if not messagebox.askyesno("Limpar Tudo",
                "Remover TODOS os dados importados nesta sessão?\n"
                "Esta ação não pode ser desfeita.",
                parent=self.parent):
            return
        self._pending_perdcomp.clear()
        self._pending_darf.clear()
        self._pending_dctf.clear()
        self._pending_dctfweb.clear()
        self._pending_efd.clear()
        self.ds.clear_all()
        self.mainapp.set_status("✓ Todos os dados foram limpos.")


# =============================================================================
# PerdcompView, DarfView, DctfView — Views de auditoria (Read-only do DataStore)
# =============================================================================

class _BaseAuditView:
    """Base comum para views de auditoria."""
    def __init__(self, parent, datastore: DataStore, mainapp):
        self.parent = parent
        self.ds = datastore
        self.mainapp = mainapp
        self._already_loaded = False
        self._build_ui()

    def _build_header(self, title: str, subtitle: str, refresh_cmd, export_cmd):
        hdr = tk.Frame(self.parent, bg=C_GREEN_DARK, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=title,
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=12, pady=8)
        tk.Label(hdr, text=f"  {subtitle}",
                 bg=C_GREEN_DARK, fg="#C5E08A",
                 font=("Segoe UI", 8)).pack(side="left", pady=8)

        # Botão Exportar (desta view)
        tk.Button(hdr, text="⬇  Exportar Excel",
            bg=C_YELLOW, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 8, "bold"), padx=10, pady=3,
            command=export_cmd).pack(side="right", padx=8, pady=8)

        # Botão Atualizar
        tk.Button(hdr, text="↻  Atualizar Análise",
            bg=C_BLUE, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 8, "bold"), padx=10, pady=3,
            command=refresh_cmd).pack(side="right", padx=2, pady=8)


class PerdcompView(_BaseAuditView):
    """View do PERDCOMP Extractor — exibe as 6 abas clássicas."""

    def _build_ui(self):
        self._build_header("🧾  PERDCOMP Extractor",
            "Análise de PERDCOMPs importados",
            self._refresh, self._export)

        # Área de conteúdo — lazy load: só constrói tudo ao clicar pela primeira vez
        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nNenhum PERDCOMP importado ainda.\n\n"
                 "Vá à Central de Importação para adicionar PDFs.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        # Notebook interno (criado sob demanda na primeira atualização)
        self.nb = None
        self.trees = {}
        self.col_filt = {1:{}, 2:{}, 3:{}, 4:{}, 5:{}, 6:{}}

        # Híbrido: primeira carga automática quando dados estão disponíveis
        self.ds.subscribe("perdcomp", self._auto_refresh_once)
        self.ds.subscribe("status", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded and self.ds.perdcomp_rows:
            self._refresh()

    def _refresh(self):
        if not self.ds.perdcomp_rows:
            if self.nb is not None:
                self.nb.destroy()
                self.nb = None
                self.info_label.pack(expand=True)
            return

        # Primeira vez: esconde o label de info e cria o notebook
        if self.nb is None:
            self.info_label.pack_forget()
            self.nb = ttk.Notebook(self.container, style="Sub.TNotebook")
            self.nb.pack(fill="both", expand=True, padx=4, pady=4)
            self._build_treeviews()

        self._populate_all()
        self._already_loaded = True
        self.mainapp.set_status("✓ PERDCOMP — análise atualizada")

    def _build_treeviews(self):
        # 6 abas: Detalhamento, Controle de Créditos, Sem Vínculo, Na Planilha, Sem Planilha, Ressarcimentos
        tabs_config = [
            ("📋 Detalhamento",       ABA1_COLS, 1),
            ("💰 Controle Créditos",  CTRL_COLS, 2),
            ("⚠ Comp. Sem Vínculo",  DETAIL_COLS, 3),
            ("📝 Na Planilha s/ PDF", ABA4_COLS, 4),
            ("📄 PDF s/ Planilha",    ABA5_COLS, 5),
            ("🌿 Ressarcimentos",     ABA6_COLS, 6),
        ]
        for label, cols, idx in tabs_config:
            frm = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm, text=f"  {label}  ")
            tree = self._make_tree(frm, cols, idx)
            self.trees[idx] = (tree, cols)

    def _make_tree(self, parent, cols, tab_idx):
        frm = tk.Frame(parent, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [c[0] for c in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
                             yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        money_keys = {"valor_total_credito","valor_utilizado_doc",
                       "valor_principal","valor_multa","valor_juros","valor_total_debito",
                       "vl_credito","vl_correcao","vl_compensado","saldo_disponivel",
                       "vl_total","valor_total"}
        for key, label, width in cols:
            anchor = "e" if key in money_keys else "w"
            tree.heading(key, text=f"{label}  ▼",
                command=lambda k=key, l=label, t=tab_idx: self._open_filter(k, l, t))
            tree.column(key, width=width, minwidth=50, anchor=anchor)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        tree.tag_configure("credito", background="#E8F5D5")
        tree.tag_configure("debito",  background="#FDE8E5")
        tree.tag_configure("warn",    background=C_YELLOW_LIGHT)
        return tree

    def _populate_all(self):
        # Aba 1 — Detalhamento (combine_rows_for_aba1)
        aba1 = combine_rows_for_aba1(self.ds.perdcomp_rows, self.ds.status_map)
        self._fill_tree(1, aba1, ABA1_COLS)

        # Aba 2 — Controle de Créditos
        aba2 = build_credit_control(self.ds.perdcomp_rows, self.ds.status_map)
        self._fill_tree(2, aba2, CTRL_COLS)

        # Aba 3 — Comp. Sem Vínculo
        aba3 = build_unlinked_compensations(self.ds.perdcomp_rows, self.ds.status_map)
        self._fill_tree(3, aba3, DETAIL_COLS)

        # Abas 4 e 5 — Na Planilha s/ PDF e PDF s/ Planilha
        if self.ds.status_map:
            aba4 = build_missing_from_excel(self.ds.status_map, self.ds.perdcomp_rows)
            self._fill_tree(4, aba4, ABA4_COLS)
            aba5 = build_missing_from_pdfs(self.ds.status_map, self.ds.perdcomp_rows)
            self._fill_tree(5, aba5, ABA5_COLS)
        else:
            self._fill_tree(4, [], ABA4_COLS)
            self._fill_tree(5, [], ABA5_COLS)

        # Aba 6 — Ressarcimentos
        aba6 = build_ressarcimento_aba6(self.ds.perdcomp_rows, self.ds.status_map)
        self._fill_tree(6, aba6, ABA6_COLS)

    def _fill_tree(self, idx, rows, cols):
        tree, _ = self.trees[idx]
        for i in tree.get_children():
            tree.delete(i)
        setattr(self, f"_data_{idx}", rows)
        self._apply_filter(idx)

    def _apply_filter(self, idx):
        tree, cols = self.trees[idx]
        rows = getattr(self, f"_data_{idx}", [])
        filt = self.col_filt.get(idx, {})
        for i in tree.get_children():
            tree.delete(i)
        money_keys = {"valor_total_credito","valor_utilizado_doc",
                       "valor_principal","valor_multa","valor_juros","valor_total_debito",
                       "vl_credito","vl_correcao","vl_compensado","saldo_disponivel",
                       "vl_total","valor_total"}
        for r in rows:
            # Aplica filtros
            ok = True
            for k, vset in filt.items():
                v = r.get(k, "")
                if k in money_keys:
                    try: v = self._fmt_money(v)
                    except Exception: v = str(v)
                else:
                    v = str(v) if v is not None else ""
                if v and v not in vset:
                    ok = False; break
            if not ok:
                continue
            vals = []
            for k, _, _ in cols:
                v = r.get(k, "")
                if k in money_keys:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v or "")
            tag = ""
            tr = r.get("tipo_registro", "")
            if tr == "Crédito": tag = "credito"
            elif tr == "Débito": tag = "debito"
            if "cancel" in str(r.get("situacao_perdcomp", "")).lower():
                tag = "warn"
            tree.insert("", "end", values=vals, tags=(tag,) if tag else ())

    @staticmethod
    def _fmt_money(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v else ""

    def _open_filter(self, col_key, col_label, tab_idx):
        tree, cols = self.trees[tab_idx]
        rows = getattr(self, f"_data_{tab_idx}", [])
        money_keys = {"valor_total_credito","valor_utilizado_doc",
                       "valor_principal","valor_multa","valor_juros","valor_total_debito",
                       "vl_credito","vl_correcao","vl_compensado","saldo_disponivel"}
        vals = set()
        for r in rows:
            v = r.get(col_key, "")
            if col_key in money_keys:
                try: v = self._fmt_money(v)
                except Exception: v = str(v)
            else:
                v = str(v) if v is not None else ""
            if v: vals.add(v)
        active = self.col_filt[tab_idx].get(col_key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                self.col_filt[tab_idx].pop(col_key, None)
            else:
                self.col_filt[tab_idx][col_key] = selected
            self._apply_filter(tab_idx)

        ColFilterPopup(self.parent.winfo_toplevel(),
                       col_label, vals, active, on_apply, px, py)

    def _export(self):
        if not self.ds.perdcomp_rows:
            messagebox.showinfo("Sem dados",
                "Nenhum PERDCOMP foi importado ainda.", parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"PERDCOMP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            # export_excel recebe as 6 listas de abas já calculadas
            aba1_rows = combine_rows_for_aba1(self.ds.perdcomp_rows, self.ds.status_map)
            ctrl_rows = build_credit_control(self.ds.perdcomp_rows, self.ds.status_map)
            unlinked  = build_unlinked_compensations(self.ds.perdcomp_rows, self.ds.status_map)
            if self.ds.status_map:
                aba4_rows = build_missing_from_excel(self.ds.status_map, self.ds.perdcomp_rows)
                aba5_rows = build_missing_from_pdfs(self.ds.status_map, self.ds.perdcomp_rows)
            else:
                aba4_rows, aba5_rows = [], []
            aba6_rows = build_ressarcimento_aba6(self.ds.perdcomp_rows, self.ds.status_map)
            export_excel(aba1_rows, ctrl_rows, unlinked, aba4_rows, aba5_rows, aba6_rows,
                          path, status_map=self.ds.status_map)
            if messagebox.askyesno("Exportado!",
                    f"Arquivo:\n{path}\n\nAbrir agora?",
                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


class DarfView(_BaseAuditView):
    """View do DARF Extractor — Treeview simples."""

    def _build_ui(self):
        self._build_header("📋  DARF Extractor",
            "Comprovantes de Arrecadação (DARF / DAS)",
            self._refresh, self._export)

        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nNenhum DARF importado ainda.\n\n"
                 "Vá à Central de Importação para adicionar PDFs.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        self.tree = None
        self.col_filt = {}

        self.ds.subscribe("darf", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded and self.ds.darf_rows:
            self._refresh()

    def _refresh(self):
        if not self.ds.darf_rows:
            return

        if self.tree is None:
            self.info_label.pack_forget()
            frm = tk.Frame(self.container, bg=C_GRAY_LIGHT)
            frm.pack(fill="both", expand=True, padx=4, pady=4)
            vsb = ttk.Scrollbar(frm, orient="vertical")
            hsb = ttk.Scrollbar(frm, orient="horizontal")
            keys = [c[0] for c in DARF_COLS]
            self.tree = ttk.Treeview(frm, columns=keys, show="headings",
                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            for key, label, width in DARF_COLS:
                anchor = "e" if key in DARF_MONEY_KEYS else "w"
                self.tree.heading(key, text=f"{label}  ▼",
                    command=lambda k=key, l=label: self._open_filter(k, l))
                self.tree.column(key, width=width, minwidth=50, anchor=anchor)
            vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
            vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
            self.tree.pack(fill="both", expand=True)

        for i in self.tree.get_children():
            self.tree.delete(i)
        for r in self.ds.darf_rows:
            if not self._match_filter(r): continue
            vals = []
            for key, _, _ in DARF_COLS:
                v = r.get(key, "")
                if key in DARF_MONEY_KEYS:
                    try: vals.append(f"{float(v or 0):,.2f}".replace(",","X").replace(".",",").replace("X","."))
                    except: vals.append(str(v))
                else:
                    vals.append(v or "")
            self.tree.insert("", "end", values=vals)

        self._already_loaded = True
        self.mainapp.set_status(f"✓ DARF — {len(self.ds.darf_rows)} linhas")

    def _match_filter(self, r):
        for k, vset in self.col_filt.items():
            v = str(r.get(k, ""))
            if k in DARF_MONEY_KEYS:
                try: v = f"{float(r.get(k,0) or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
                except: pass
            if v and v not in vset:
                return False
        return True

    def _open_filter(self, key, label):
        vals = set()
        for r in self.ds.darf_rows:
            v = r.get(key, "")
            if key in DARF_MONEY_KEYS:
                try: v = f"{float(v or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
                except: v = str(v)
            else: v = str(v) if v else ""
            if v: vals.add(v)
        active = self.col_filt.get(key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                self.col_filt.pop(key, None)
            else:
                self.col_filt[key] = selected
            self._refresh()

        ColFilterPopup(self.parent.winfo_toplevel(),
                       label, vals, active, on_apply, px, py)

    def _export(self):
        if not self.ds.darf_rows:
            messagebox.showinfo("Sem dados", "Nenhum DARF importado.", parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"DARF_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            export_darf_excel(self.ds.darf_rows, path)
            if messagebox.askyesno("Exportado!", f"Arquivo:\n{path}\n\nAbrir?",
                                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


class DctfView(_BaseAuditView):
    """View do DCTF Extractor — Detalhamento + Resumo."""

    def _build_ui(self):
        self._build_header("📊  DCTF Extractor",
            "Declarações de Débitos e Créditos Tributários Federais",
            self._refresh, self._export)

        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nNenhuma DCTF importada ainda.\n\n"
                 "Vá à Central de Importação para adicionar PDFs.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        self.nb = None
        self.trees = {}
        # Filtros estilo Excel (um dict por aba)
        self.col_filt_1: dict = {}   # Detalhamento
        self.col_filt_2: dict = {}   # Resumo por Tributo
        # Dados atuais de cada aba (usados para re-popular ao filtrar sem recalcular)
        self._data_1: list = []
        self._data_2: list = []

        self.ds.subscribe("dctf", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded and self.ds.dctf_rows:
            self._refresh()

    def _refresh(self):
        if not self.ds.dctf_rows:
            return

        if self.nb is None:
            self.info_label.pack_forget()
            self.nb = ttk.Notebook(self.container, style="Sub.TNotebook")
            self.nb.pack(fill="both", expand=True, padx=4, pady=4)

            # Aba 1 — Detalhamento
            frm1 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm1, text="  📋  Detalhamento  ")
            self.trees[1] = self._make_tree(frm1, DCTF_DETAIL_COLS, tab_idx=1)

            # Aba 2 — Resumo
            frm2 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm2, text="  📊  Resumo por Tributo  ")
            self.trees[2] = self._make_tree(frm2, DCTF_RESUMO_COLS, tab_idx=2)

        # Atualiza os dados e popula (aplicando filtros se houver)
        self._data_1 = list(self.ds.dctf_rows)
        self._data_2 = build_dctf_resumo(self.ds.dctf_rows)
        self._populate(1)
        self._populate(2)

        self._already_loaded = True
        self.mainapp.set_status(f"✓ DCTF — {len(self.ds.dctf_rows)} linhas")

    def _make_tree(self, parent, cols, tab_idx):
        frm = tk.Frame(parent, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [c[0] for c in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for key, label, width in cols:
            anchor = "e" if key in DCTF_MONEY_KEYS else "w"
            # Heading clicável para abrir popup de filtro (estilo Excel)
            tree.heading(key, text=f"{label}  ▼",
                command=lambda k=key, l=label, t=tab_idx: self._open_filter(k, l, t))
            tree.column(key, width=width, minwidth=50, anchor=anchor)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    @staticmethod
    def _fmt_money(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v else ""

    def _row_matches(self, r, filt):
        """Testa se uma linha passa por todos os filtros ativos da aba."""
        for k, vset in filt.items():
            v = r.get(k, "")
            if k in DCTF_MONEY_KEYS:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v and v not in vset:
                return False
        return True

    def _populate(self, tab_idx):
        tree = self.trees[tab_idx]
        cols = DCTF_DETAIL_COLS if tab_idx == 1 else DCTF_RESUMO_COLS
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")

        for i in tree.get_children():
            tree.delete(i)
        shown = 0
        for r in rows:
            if not self._row_matches(r, filt):
                continue
            vals = []
            for key, _, _ in cols:
                v = r.get(key, "")
                if key in DCTF_MONEY_KEYS:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v or "")
            # Zebra + alerta visual para saldo a pagar (aba 1)
            saldo = r.get("saldo_pagar", 0)
            try: saldo_f = float(saldo or 0)
            except (TypeError, ValueError): saldo_f = 0.0
            if tab_idx == 1 and saldo_f > 0:
                tag = "alert"
            else:
                tag = "even" if shown % 2 == 0 else "odd"
            tree.insert("", "end", values=vals, tags=(tag,))
            shown += 1

        # Configura tags (linhas alternadas + alerta para saldo a pagar)
        tree.tag_configure("even",  background=C_WHITE)
        tree.tag_configure("odd",   background="#F7FBED")
        tree.tag_configure("alert", background=C_RED_LIGHT)

    def _open_filter(self, col_key, col_label, tab_idx):
        """Abre popup de filtro para a coluna da aba especificada."""
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")

        # Coleta valores únicos da coluna
        vals = set()
        for r in rows:
            v = r.get(col_key, "")
            if col_key in DCTF_MONEY_KEYS:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)

        active = filt.get(col_key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                filt.pop(col_key, None)
            else:
                filt[col_key] = selected
            self._populate(tab_idx)

        ColFilterPopup(self.parent.winfo_toplevel(),
                       col_label, vals, active, on_apply, px, py)

    def _export(self):
        if not self.ds.dctf_rows:
            messagebox.showinfo("Sem dados", "Nenhuma DCTF importada.", parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"DCTF_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            resumo = build_dctf_resumo(self.ds.dctf_rows)
            export_dctf_excel(self.ds.dctf_rows, resumo, path)
            if messagebox.askyesno("Exportado!", f"Arquivo:\n{path}\n\nAbrir?",
                                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


class DctfWebView(_BaseAuditView):
    """View do DCTFWeb Extractor — Detalhamento + Resumo por Tributo."""

    def _build_ui(self):
        self._build_header("🌐  DCTFWeb Extractor",
            "Declaração Completa — Tributos Previdenciários (CP) e IRRF",
            self._refresh, self._export)

        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nNenhuma DCTFWeb importada ainda.\n\n"
                 "Vá à Central de Importação para adicionar PDFs.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        self.nb = None
        self.trees = {}
        # Filtros estilo Excel (um dict por aba)
        self.col_filt_1: dict = {}   # Detalhamento
        self.col_filt_2: dict = {}   # Resumo
        self._data_1: list = []
        self._data_2: list = []

        self.ds.subscribe("dctfweb", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded and self.ds.dctfweb_rows:
            self._refresh()

    def _refresh(self):
        if not self.ds.dctfweb_rows:
            return

        if self.nb is None:
            self.info_label.pack_forget()
            self.nb = ttk.Notebook(self.container, style="Sub.TNotebook")
            self.nb.pack(fill="both", expand=True, padx=4, pady=4)

            frm1 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm1, text="  📋  Detalhamento  ")
            self.trees[1] = self._make_tree(frm1, DCTFWEB_DETAIL_COLS, tab_idx=1)

            frm2 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm2, text="  📊  Resumo por Tributo  ")
            self.trees[2] = self._make_tree(frm2, DCTFWEB_RESUMO_COLS, tab_idx=2)

        self._data_1 = list(self.ds.dctfweb_rows)
        self._data_2 = build_dctfweb_resumo(self.ds.dctfweb_rows)
        self._populate(1)
        self._populate(2)

        self._already_loaded = True
        self.mainapp.set_status(
            f"✓ DCTFWeb — {len(self.ds.dctfweb_rows)} linhas detalhe, "
            f"{len(self._data_2)} tributos no resumo")

    def _make_tree(self, parent, cols, tab_idx):
        frm = tk.Frame(parent, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [c[0] for c in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        money_keys = DCTFWEB_MONEY_KEYS if tab_idx == 1 else DCTFWEB_RESUMO_MONEY
        for key, label, width in cols:
            anchor = "e" if key in money_keys else "w"
            tree.heading(key, text=f"{label}  ▼",
                command=lambda k=key, l=label, t=tab_idx: self._open_filter(k, l, t))
            tree.column(key, width=width, minwidth=50, anchor=anchor)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    @staticmethod
    def _fmt_money(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v else ""

    def _row_matches(self, r, filt, money_keys):
        for k, vset in filt.items():
            v = r.get(k, "")
            if k in money_keys:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v and v not in vset:
                return False
        return True

    def _populate(self, tab_idx):
        tree = self.trees[tab_idx]
        cols = DCTFWEB_DETAIL_COLS if tab_idx == 1 else DCTFWEB_RESUMO_COLS
        money = DCTFWEB_MONEY_KEYS if tab_idx == 1 else DCTFWEB_RESUMO_MONEY
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")

        for i in tree.get_children():
            tree.delete(i)
        shown = 0
        for r in rows:
            if not self._row_matches(r, filt, money):
                continue
            vals = []
            for key, _, _ in cols:
                v = r.get(key, "")
                if key in money:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v if v != 0 else (v or ""))
            # Coloração por situação (aba 1 detalhe)
            if tab_idx == 1:
                saldo = r.get("saldo_pagar", 0)
                comp  = r.get("cred_compensacao", 0)
                try: saldo_f = float(saldo or 0)
                except (TypeError, ValueError): saldo_f = 0.0
                try: comp_f = float(comp or 0)
                except (TypeError, ValueError): comp_f = 0.0
                if saldo_f > 0.01:
                    tag = "alert"
                elif comp_f > 0.01:
                    tag = "compensado"
                else:
                    tag = "even" if shown % 2 == 0 else "odd"
            else:
                tag = "even" if shown % 2 == 0 else "odd"
            tree.insert("", "end", values=vals, tags=(tag,))
            shown += 1

        tree.tag_configure("even",       background=C_WHITE)
        tree.tag_configure("odd",        background="#F7FBED")
        tree.tag_configure("alert",      background=C_RED_LIGHT)
        tree.tag_configure("compensado", background=C_GREEN_LIGHT)

    def _open_filter(self, col_key, col_label, tab_idx):
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")
        money = DCTFWEB_MONEY_KEYS if tab_idx == 1 else DCTFWEB_RESUMO_MONEY

        vals = set()
        for r in rows:
            v = r.get(col_key, "")
            if col_key in money:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)

        active = filt.get(col_key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                filt.pop(col_key, None)
            else:
                filt[col_key] = selected
            self._populate(tab_idx)

        ColFilterPopup(self.parent.winfo_toplevel(),
                       col_label, vals, active, on_apply, px, py)

    def _export(self):
        if not self.ds.dctfweb_rows:
            messagebox.showinfo("Sem dados", "Nenhuma DCTFWeb importada.",
                                 parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"DCTFWeb_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            resumo = build_dctfweb_resumo(self.ds.dctfweb_rows)
            export_dctfweb_excel(self.ds.dctfweb_rows, resumo, path)
            if messagebox.askyesno("Exportado!", f"Arquivo:\n{path}\n\nAbrir?",
                                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


class EfdView(_BaseAuditView):
    """View da EFD Contribuições — 3 abas: Detalhamento, Resumo, Confronto."""

    def _build_ui(self):
        self._build_header("🌾  EFD Contribuições",
            "PIS/COFINS apurados na escrituração + Confronto com DCTF/DCTFWeb",
            self._refresh, self._export)

        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nNenhuma EFD Contribuições importada ainda.\n\n"
                 "Vá à Central de Importação para adicionar arquivos .txt SPED.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        self.nb = None
        self.trees = {}
        # Filtros estilo Excel (um dict por aba)
        self.col_filt_1: dict = {}   # Detalhamento
        self.col_filt_2: dict = {}   # Resumo
        self.col_filt_3: dict = {}   # Confronto
        # Dados atuais
        self._data_1: list = []
        self._data_2: list = []
        self._data_3: list = []  # Confronto EFD × DCTF + DCTFWeb

        # Re-roda quando dados de EFD, DCTF ou DCTFWeb mudarem
        self.ds.subscribe("efd",     self._auto_refresh_once)
        self.ds.subscribe("dctf",    self._auto_refresh_once)
        self.ds.subscribe("dctfweb", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded and self.ds.efd_rows:
            self._refresh()

    def _refresh(self):
        if not self.ds.efd_rows:
            return

        if self.nb is None:
            self.info_label.pack_forget()
            self.nb = ttk.Notebook(self.container, style="Sub.TNotebook")
            self.nb.pack(fill="both", expand=True, padx=4, pady=4)

            frm1 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm1, text="  📋  Detalhamento  ")
            self.trees[1] = self._make_tree(frm1, EFD_DETAIL_COLS, tab_idx=1)

            frm2 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm2, text="  📊  Resumo por Código  ")
            self.trees[2] = self._make_tree(frm2, EFD_RESUMO_COLS, tab_idx=2)

            frm3 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
            self.nb.add(frm3, text="  ⚖  Confronto EFD × DCTF + DCTFWeb  ")
            self.trees[3] = self._make_tree(frm3, CONFRONTO_EFD_COLS, tab_idx=3)

        self._data_1 = list(self.ds.efd_rows)
        self._data_2 = build_efd_resumo(self.ds.efd_rows)
        self._data_3 = run_confronto_efd_dctf(
            self.ds.efd_rows, self.ds.dctf_rows, self.ds.dctfweb_rows)

        self._populate(1)
        self._populate(2)
        self._populate(3)

        self._already_loaded = True
        self.mainapp.set_status(
            f"✓ EFD Contribuições — {len(self.ds.efd_rows)} linhas"
            + f", {len(self._data_3)} confrontos")

    def _make_tree(self, parent, cols, tab_idx):
        frm = tk.Frame(parent, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [c[0] for c in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        # Money keys depende da aba
        if tab_idx == 1:    money_keys = EFD_MONEY_KEYS
        elif tab_idx == 2:  money_keys = EFD_RESUMO_MONEY
        else:               money_keys = CONFRONTO_EFD_MONEY
        for key, label, width in cols:
            anchor = "e" if key in money_keys else "w"
            tree.heading(key, text=f"{label}  ▼",
                command=lambda k=key, l=label, t=tab_idx: self._open_filter(k, l, t))
            tree.column(key, width=width, minwidth=50, anchor=anchor)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        # Configura tags de cor
        if tab_idx == 1:
            tree.tag_configure("pis",    background="#EFF6FF")
            tree.tag_configure("cofins", background="#FEF3DC")
        elif tab_idx == 3:
            tree.tag_configure("ok",       background=C_GREEN_LIGHT)
            tree.tag_configure("diverg",   background=C_RED_LIGHT)
            tree.tag_configure("so_efd",   background=C_YELLOW_LIGHT)
            tree.tag_configure("so_decl",  background=C_BLUE_LIGHT)
        return tree

    @staticmethod
    def _fmt_money(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v else ""

    def _row_matches(self, r, filt, money_keys):
        for k, vset in filt.items():
            v = r.get(k, "")
            if k in money_keys:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v and v not in vset:
                return False
        return True

    def _populate(self, tab_idx):
        tree = self.trees[tab_idx]
        if tab_idx == 1:
            cols = EFD_DETAIL_COLS;     money = EFD_MONEY_KEYS
        elif tab_idx == 2:
            cols = EFD_RESUMO_COLS;     money = EFD_RESUMO_MONEY
        else:
            cols = CONFRONTO_EFD_COLS;  money = CONFRONTO_EFD_MONEY
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")

        for i in tree.get_children():
            tree.delete(i)
        for r in rows:
            if not self._row_matches(r, filt, money):
                continue
            vals = []
            for key, _, _ in cols:
                v = r.get(key, "")
                if key in money:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v if v != 0 else (v or ""))
            # Tag por situação
            if tab_idx == 1:
                tag = "pis" if r.get("tributo") == "PIS" else "cofins"
            elif tab_idx == 3:
                sit = r.get("situacao", "")
                if   SIT_E_OK      in sit: tag = "ok"
                elif SIT_E_DIVERG  in sit: tag = "diverg"
                elif SIT_E_SO_EFD  in sit: tag = "so_efd"
                elif SIT_E_SO_DECL in sit: tag = "so_decl"
                else: tag = ""
            else:
                tag = ""
            tree.insert("", "end", values=vals, tags=(tag,) if tag else ())

    def _open_filter(self, col_key, col_label, tab_idx):
        rows = getattr(self, f"_data_{tab_idx}")
        filt = getattr(self, f"col_filt_{tab_idx}")
        if tab_idx == 1:    money = EFD_MONEY_KEYS
        elif tab_idx == 2:  money = EFD_RESUMO_MONEY
        else:               money = CONFRONTO_EFD_MONEY

        vals = set()
        for r in rows:
            v = r.get(col_key, "")
            if col_key in money:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)

        active = filt.get(col_key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                filt.pop(col_key, None)
            else:
                filt[col_key] = selected
            self._populate(tab_idx)

        ColFilterPopup(self.parent.winfo_toplevel(),
                       col_label, vals, active, on_apply, px, py)

    def _export(self):
        if not self.ds.efd_rows:
            messagebox.showinfo("Sem dados", "Nenhuma EFD importada.",
                                 parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"EFD_Contribuicoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            resumo = build_efd_resumo(self.ds.efd_rows)
            export_efd_excel(self.ds.efd_rows, resumo, path)
            if messagebox.askyesno("Exportado!", f"Arquivo:\n{path}\n\nAbrir?",
                                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


class ConciliacaoView(_BaseAuditView):
    """View da Conciliação — 2 abas: DARF×DCOMP e DCTF×DARF×DCOMP."""

    def _build_ui(self):
        self._build_header("⚖  Conciliação",
            "DARF × DCOMP  e  DCTF × DARF × DCOMP",
            self._refresh, self._export)

        self.container = tk.Frame(self.parent, bg=C_GRAY_LIGHT)
        self.container.pack(fill="both", expand=True)

        self.info_label = tk.Label(self.container,
            text="\n\nA Conciliação requer PERDCOMPs + DARFs importados.\n"
                 "Para o triplo cruzamento, importe também DCTFs.\n\n"
                 "Vá à Central de Importação.",
            bg=C_GRAY_LIGHT, fg=C_GRAY,
            font=("Segoe UI", 10), justify="center")
        self.info_label.pack(expand=True)

        self.nb = None
        self.result_rows = []
        self.triplo_rows = []
        # Filtros estilo Excel por coluna (uma dict por aba)
        self.col_filt_c: dict = {}   # DARF × DCOMP
        self.col_filt_t: dict = {}   # Triplo DCTF × DARF × DCOMP

        self.ds.subscribe("perdcomp", self._auto_refresh_once)
        self.ds.subscribe("darf", self._auto_refresh_once)
        self.ds.subscribe("dctf", self._auto_refresh_once)
        self.ds.subscribe("dctfweb", self._auto_refresh_once)
        self.ds.subscribe("status", self._auto_refresh_once)

    def _auto_refresh_once(self):
        if not self._already_loaded:
            # Só roda se houver dados suficientes
            if self.ds.darf_rows and self.ds.perdcomp_rows:
                self._refresh()

    def get_result_rows(self): return self.result_rows
    def get_triplo_rows(self): return self.triplo_rows

    def _refresh(self):
        # Filtra DCOMPs dos perdcomp_rows
        dcomp_rows = [r for r in self.ds.perdcomp_rows
                      if re.search(r"DCOMP|COMPENSA", r.get("tipo_pedido", ""), re.IGNORECASE)
                      and r.get("tipo_registro") == "Débito"]

        if not self.ds.darf_rows and not dcomp_rows:
            return

        if self.nb is None:
            self.info_label.pack_forget()
            self.nb = ttk.Notebook(self.container, style="Sub.TNotebook")
            self.nb.pack(fill="both", expand=True, padx=4, pady=4)
            self._build_tabs()

        # Roda as conciliações
        self.result_rows = run_conciliacao(
            self.ds.darf_rows, dcomp_rows, self.ds.status_map)
        self._populate_conc()

        # Roda o cruzamento DCTF + DCTFWeb × DARF × DCOMP
        # Triplo roda se houver QUALQUER lado declarativo (DCTF ou DCTFWeb)
        if self.ds.dctf_rows or self.ds.dctfweb_rows:
            self.triplo_rows = run_triplo_dctf_darf_dcomp(
                self.ds.dctf_rows, self.ds.darf_rows, dcomp_rows,
                self.ds.status_map,
                dctfweb_rows=self.ds.dctfweb_rows)
        else:
            self.triplo_rows = []
        self._populate_triplo()

        self._already_loaded = True
        self.mainapp.set_status(
            f"✓ Conciliação — {len(self.result_rows)} linhas DARF×DCOMP"
            + (f", {len(self.triplo_rows)} linhas triplo" if self.triplo_rows else ""))

    def _build_tabs(self):
        # Tab 1 — DARF × DCOMP
        frm1 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
        self.nb.add(frm1, text="  DARF × DCOMP  ")
        self.tree_c = self._make_tree(frm1, CONC_COLS, CONC_MONEY, which="c")
        self.tree_c.tag_configure("duplo",      background=C_RED_LIGHT)
        self.tree_c.tag_configure("divergente", background=C_YELLOW_LIGHT)
        self.tree_c.tag_configure("so_dcomp",   background=C_BLUE_LIGHT)
        self.tree_c.tag_configure("so_darf",    background=C_GRAY_LIGHT)

        # Tab 2 — DCTF × DARF × DCOMP
        frm2 = tk.Frame(self.nb, bg=C_GRAY_LIGHT)
        self.nb.add(frm2, text="  DCTF + DCTFWeb × DARF × DCOMP  ")
        self.tree_t = self._make_tree(frm2, TRIPLO_COLS, TRIPLO_MONEY, which="t")
        self.tree_t.tag_configure("t_quitado",  background=C_GREEN_LIGHT)
        self.tree_t.tag_configure("t_saldo",    background=C_RED_LIGHT)
        self.tree_t.tag_configure("t_amaior",   background=C_YELLOW_LIGHT)
        self.tree_t.tag_configure("t_sem_decl", background=C_BLUE_LIGHT)

    def _make_tree(self, parent, cols, money_set, which="c"):
        frm = tk.Frame(parent, bg=C_GRAY_LIGHT)
        frm.pack(fill="both", expand=True, padx=2, pady=2)
        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        keys = [c[0] for c in cols]
        tree = ttk.Treeview(frm, columns=keys, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for key, label, width in cols:
            anchor = "e" if key in money_set else "w"
            # Heading clicável para abrir popup de filtro (estilo Excel)
            tree.heading(key, text=f"{label}  ▼",
                command=lambda k=key, l=label, w=which: self._open_filter(k, l, w))
            tree.column(key, width=width, minwidth=50, anchor=anchor)
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    @staticmethod
    def _fmt_money(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except (TypeError, ValueError):
            return str(v) if v else ""

    def _row_matches(self, r, filt, money_keys):
        """Testa se uma linha passa por todos os filtros ativos."""
        for k, vset in filt.items():
            v = r.get(k, "")
            if k in money_keys:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v and v not in vset:
                return False
        return True

    def _open_filter(self, col_key, col_label, which):
        """Abre popup de filtro para a coluna.  `which` = 'c' ou 't'."""
        if which == "c":
            rows = self.result_rows
            money = CONC_MONEY
            filt = self.col_filt_c
        else:
            rows = self.triplo_rows
            money = TRIPLO_MONEY
            filt = self.col_filt_t

        vals = set()
        for r in rows:
            v = r.get(col_key, "")
            if col_key in money:
                v = self._fmt_money(v)
            else:
                v = str(v) if v is not None else ""
            if v:
                vals.add(v)

        active = filt.get(col_key)
        px, py = self.parent.winfo_pointerxy()

        def on_apply(selected):
            if selected is None:
                filt.pop(col_key, None)
            else:
                filt[col_key] = selected
            # Re-popula a aba correta
            if which == "c":
                self._populate_conc()
            else:
                self._populate_triplo()

        ColFilterPopup(self.parent.winfo_toplevel(),
                       col_label, vals, active, on_apply, px, py)

    def _populate_conc(self):
        for i in self.tree_c.get_children():
            self.tree_c.delete(i)
        for r in self.result_rows:
            # Aplica filtros ativos
            if not self._row_matches(r, self.col_filt_c, CONC_MONEY):
                continue
            sit = r.get("situacao", "")
            tag = ""
            if SIT_DUPLO in sit: tag = "duplo"
            elif SIT_DIVERGENTE in sit: tag = "divergente"
            elif SIT_SO_DCOMP in sit: tag = "so_dcomp"
            elif SIT_SO_DARF in sit: tag = "so_darf"
            vals = []
            for key, _, _ in CONC_COLS:
                v = r.get(key, "")
                if key in CONC_MONEY:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v or "")
            self.tree_c.insert("", "end", values=vals, tags=(tag,) if tag else ())

    def _populate_triplo(self):
        for i in self.tree_t.get_children():
            self.tree_t.delete(i)
        for r in self.triplo_rows:
            # Aplica filtros ativos
            if not self._row_matches(r, self.col_filt_t, TRIPLO_MONEY):
                continue
            sit = r.get("situacao_triplo", "")
            tag = ""
            if SIT_T_QUITADO in sit: tag = "t_quitado"
            elif SIT_T_SALDO in sit: tag = "t_saldo"
            elif SIT_T_A_MAIOR in sit: tag = "t_amaior"
            elif SIT_T_SEM_DECL in sit: tag = "t_sem_decl"
            vals = []
            for key, _, _ in TRIPLO_COLS:
                v = r.get(key, "")
                if key in TRIPLO_MONEY:
                    vals.append(self._fmt_money(v))
                else:
                    vals.append(v or "")
            self.tree_t.insert("", "end", values=vals, tags=(tag,) if tag else ())

    def _export(self):
        if not self.result_rows and not self.triplo_rows:
            messagebox.showinfo("Sem dados",
                "Atualize a análise primeiro.", parent=self.parent)
            return
        path = filedialog.asksaveasfilename(parent=self.parent,
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"Conciliacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            export_conciliacao_excel(self.result_rows, path, triplo_rows=self.triplo_rows)
            if messagebox.askyesno("Exportado!", f"Arquivo:\n{path}\n\nAbrir?",
                                    parent=self.parent):
                if sys.platform == "win32": os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.parent)


# =============================================================================
# MainApp — Janela única com abas (Central de Importação + Auditoria)
# =============================================================================

class MainApp:
    """Janela única do AgriTax Audit.

    Estrutura:
      ┌────────────────────────────────────────────────────┐
      │ [Logo] AgriTax Audit                               │
      ├────────────────────────────────────────────────────┤
      │ 📥 Central de Importação | 🔍 Auditoria           │
      ├────────────────────────────────────────────────────┤
      │  (conteúdo da aba)                                 │
      └────────────────────────────────────────────────────┘
    """

    def __init__(self, root: tk.Tk):
        self.root = root
        self.ds = get_datastore()
        self._logo_img = None
        self._setup_window()
        self._apply_styles()
        self._build_ui()

    def _setup_window(self):
        self.root.title("AgriTax Audit v7.0")
        self.root.configure(bg=C_GRAY_LIGHT)
        self.root.geometry("1400x840")
        self.root.minsize(1100, 680)
        try:
            self.root.state("zoomed")  # Maximizado no Windows
        except Exception:
            pass

    def _apply_styles(self):
        s = ttk.Style()
        try: s.theme_use("clam")
        except Exception: pass
        # Notebook principal (abas grandes)
        s.configure("Main.TNotebook", background=C_GRAY_LIGHT, borderwidth=0,
                    tabmargins=[0, 0, 0, 0])
        s.configure("Main.TNotebook.Tab",
                    background=C_GREEN_DARK, foreground=C_WHITE,
                    padding=[24, 10], font=("Segoe UI", 11, "bold"),
                    borderwidth=0)
        s.map("Main.TNotebook.Tab",
              background=[("selected", C_WHITE)],
              foreground=[("selected", C_GREEN_DARK)])
        # Notebook secundário (sub-abas da Auditoria)
        s.configure("Sub.TNotebook", background=C_WHITE, borderwidth=0,
                    tabmargins=[0, 4, 0, 0])
        s.configure("Sub.TNotebook.Tab",
                    background="#D8E8B8", foreground=C_GRAY_DARK,
                    padding=[16, 7], font=("Segoe UI", 9),
                    borderwidth=0)
        s.map("Sub.TNotebook.Tab",
              background=[("selected", C_WHITE)],
              foreground=[("selected", C_GREEN_DARK)])
        # Treeview base
        s.configure("Treeview", background=C_WHITE, foreground=C_GRAY_DARK,
                    fieldbackground=C_WHITE, rowheight=24, font=("Segoe UI", 8))
        s.configure("Treeview.Heading", background=C_GREEN_DARK, foreground=C_WHITE,
                    font=("Segoe UI", 9, "bold"), relief="flat")
        s.map("Treeview", background=[("selected", C_GREEN_MID)])

    def _build_ui(self):
        # ── Cabeçalho com logo AgriTax ─────────────────────────────────────
        hdr = tk.Frame(self.root, bg=C_GREEN_DARK, height=62)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        try:
            import base64 as _b64, io as _io
            from PIL import Image, ImageTk
            img = Image.open(_io.BytesIO(_b64.b64decode(LOGO_B64))).resize((48, 48), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(hdr, image=self._logo_img, bg=C_GREEN_DARK).pack(side="left", padx=14)
        except Exception:
            pass

        tk.Label(hdr, text="AgriTax Audit",
                 bg=C_GREEN_DARK, fg=C_WHITE,
                 font=("Segoe UI", 18, "bold")).pack(side="left", pady=10)
        tk.Label(hdr, text="  Plataforma de Auditoria Tributária  |  v7.0",
                 bg=C_GREEN_DARK, fg="#C5E08A",
                 font=("Segoe UI", 10)).pack(side="left", pady=10)

        # Botão global "Exportar Tudo" no canto direito
        self.btn_export_all = tk.Button(hdr, text="⬇  Exportar Tudo",
            bg=C_YELLOW, fg=C_WHITE, relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=4,
            command=self._export_all)
        self.btn_export_all.pack(side="right", padx=12, pady=14)

        # ── Notebook principal ─────────────────────────────────────────────
        self.nb_main = ttk.Notebook(self.root, style="Main.TNotebook")
        self.nb_main.pack(fill="both", expand=True, padx=0, pady=0)

        # Aba 1 — Central de Importação
        self.frame_central = tk.Frame(self.nb_main, bg=C_GRAY_LIGHT)
        self.nb_main.add(self.frame_central, text="  📥  Central de Importação  ")
        self.central = CentralImportView(self.frame_central, self.ds, self)

        # Aba 2 — Auditoria (com sub-abas)
        self.frame_audit = tk.Frame(self.nb_main, bg=C_WHITE)
        self.nb_main.add(self.frame_audit, text="  🔍  Auditoria  ")
        self._build_audit_area(self.frame_audit)

        # ── Status bar ─────────────────────────────────────────────────────
        sb = tk.Frame(self.root, bg=C_GREEN_DARK, height=24)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Pronto — Use a Central de Importação para adicionar PDFs.")
        tk.Label(sb, textvariable=self.status_var, font=("Segoe UI", 8),
                 fg=C_GREEN_LIGHT, bg=C_GREEN_DARK, anchor="w").pack(fill="x", padx=12, pady=3)

        self.ds.subscribe("any", self._update_status_summary)

    def _build_audit_area(self, parent):
        # Sub-Notebook dentro da Auditoria
        self.nb_sub = ttk.Notebook(parent, style="Sub.TNotebook")
        self.nb_sub.pack(fill="both", expand=True, padx=6, pady=6)

        # Sub-aba 1 — PERDCOMP Extractor
        frm1 = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm1, text="  🧾  PERDCOMP Extractor  ")
        self.view_perdcomp = PerdcompView(frm1, self.ds, self)

        # Sub-aba 2 — DARF Extractor
        frm2 = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm2, text="  📋  DARF Extractor  ")
        self.view_darf = DarfView(frm2, self.ds, self)

        # Sub-aba 3 — DCTF Extractor
        frm3 = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm3, text="  📊  DCTF Extractor  ")
        self.view_dctf = DctfView(frm3, self.ds, self)

        # Sub-aba 4 — DCTFWeb Extractor
        frm_web = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm_web, text="  🌐  DCTFWeb Extractor  ")
        self.view_dctfweb = DctfWebView(frm_web, self.ds, self)

        # Sub-aba 5 — EFD Contribuições
        frm_efd = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm_efd, text="  🌾  EFD Contribuições  ")
        self.view_efd = EfdView(frm_efd, self.ds, self)

        # Sub-aba 6 — Conciliação
        frm4 = tk.Frame(self.nb_sub, bg=C_GRAY_LIGHT)
        self.nb_sub.add(frm4, text="  ⚖  Conciliação  ")
        self.view_conc = ConciliacaoView(frm4, self.ds, self)

    def set_status(self, msg: str):
        self.status_var.set(msg)
        try: self.root.update_idletasks()
        except Exception: pass

    def _update_status_summary(self):
        s = self.ds.summary()
        self.set_status(
            f"PERDCOMPs: {s['perdcomp_files']} arquivos / {s['perdcomp_rows']} linhas  |  "
            f"DARFs: {s['darf_files']}/{s['darf_rows']}  |  "
            f"DCTFs: {s['dctf_files']}/{s['dctf_rows']}  |  "
            f"DCTFWeb: {s['dctfweb_files']}/{s['dctfweb_rows']}  |  "
            f"EFD: {s['efd_files']}/{s['efd_rows']}  |  "
            f"Status: {'✓ '+str(s['status_registros']) if s['status_loaded'] else 'não carregado'}"
        )

    def _export_all(self):
        """Exporta um Excel único com todas as abas de todos os módulos."""
        from tkinter import filedialog, messagebox

        s = self.ds.summary()
        if not (s["perdcomp_rows"] or s["darf_rows"] or s["dctf_rows"]
                or s["dctfweb_rows"] or s["efd_rows"]):
            messagebox.showinfo("Sem dados",
                "Nenhum dado foi importado ainda. Use a Central de Importação.",
                parent=self.root)
            return

        path = filedialog.asksaveasfilename(parent=self.root,
            title="Exportar todas as análises",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"AgriTax_Analise_Completa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path:
            return

        self.set_status("Gerando planilha unificada...")
        try:
            from openpyxl import Workbook
            wb_final = Workbook()
            # Remove a aba default
            wb_final.remove(wb_final.active)

            # 1) Abas do PERDCOMP
            if self.ds.perdcomp_rows:
                import tempfile
                tmp1 = Path(tempfile.gettempdir()) / "_agritax_p.xlsx"
                aba1_rows = combine_rows_for_aba1(self.ds.perdcomp_rows, self.ds.status_map)
                ctrl_rows = build_credit_control(self.ds.perdcomp_rows, self.ds.status_map)
                unlinked  = build_unlinked_compensations(self.ds.perdcomp_rows, self.ds.status_map)
                if self.ds.status_map:
                    aba4_rows = build_missing_from_excel(self.ds.status_map, self.ds.perdcomp_rows)
                    aba5_rows = build_missing_from_pdfs(self.ds.status_map, self.ds.perdcomp_rows)
                else:
                    aba4_rows, aba5_rows = [], []
                aba6_rows = build_ressarcimento_aba6(self.ds.perdcomp_rows, self.ds.status_map)
                export_excel(aba1_rows, ctrl_rows, unlinked, aba4_rows, aba5_rows, aba6_rows,
                              str(tmp1), status_map=self.ds.status_map)
                self._merge_workbook(str(tmp1), wb_final, prefix="PERDCOMP")

            # 2) Abas do DARF
            if self.ds.darf_rows:
                import tempfile
                tmp2 = Path(tempfile.gettempdir()) / "_agritax_d.xlsx"
                export_darf_excel(self.ds.darf_rows, str(tmp2))
                self._merge_workbook(str(tmp2), wb_final, prefix="DARF")

            # 3) Abas da DCTF
            if self.ds.dctf_rows:
                import tempfile
                tmp3 = Path(tempfile.gettempdir()) / "_agritax_t.xlsx"
                resumo = build_dctf_resumo(self.ds.dctf_rows)
                export_dctf_excel(self.ds.dctf_rows, resumo, str(tmp3))
                self._merge_workbook(str(tmp3), wb_final, prefix="DCTF")

            # 3b) Abas da DCTFWeb
            if self.ds.dctfweb_rows:
                import tempfile
                tmp3w = Path(tempfile.gettempdir()) / "_agritax_tw.xlsx"
                resumo_web = build_dctfweb_resumo(self.ds.dctfweb_rows)
                export_dctfweb_excel(self.ds.dctfweb_rows, resumo_web, str(tmp3w))
                self._merge_workbook(str(tmp3w), wb_final, prefix="DCTFWeb")

            # 3c) Abas da EFD Contribuições
            if self.ds.efd_rows:
                import tempfile
                tmp3e = Path(tempfile.gettempdir()) / "_agritax_efd.xlsx"
                resumo_efd = build_efd_resumo(self.ds.efd_rows)
                export_efd_excel(self.ds.efd_rows, resumo_efd, str(tmp3e))
                self._merge_workbook(str(tmp3e), wb_final, prefix="EFD")

            # 4) Abas da Conciliação
            conc_rows = self.view_conc.get_result_rows()
            triplo_rows = self.view_conc.get_triplo_rows()
            if conc_rows or triplo_rows:
                import tempfile
                tmp4 = Path(tempfile.gettempdir()) / "_agritax_c.xlsx"
                export_conciliacao_excel(conc_rows, str(tmp4),
                                           triplo_rows=triplo_rows)
                self._merge_workbook(str(tmp4), wb_final, prefix="Concil.")

            wb_final.save(path)
            self.set_status(f"✓ Exportado: {Path(path).name}")
            if messagebox.askyesno("Exportado!",
                    f"Arquivo salvo:\n{path}\n\nDeseja abrir agora?",
                    parent=self.root):
                import subprocess, platform
                if platform.system() == "Windows":
                    os.startfile(path)
                elif platform.system() == "Darwin":
                    subprocess.call(["open", path])
                else:
                    subprocess.call(["xdg-open", path])
        except Exception as e:
            import traceback as _tb
            messagebox.showerror("Erro na exportação",
                f"{type(e).__name__}: {e}\n\n{_tb.format_exc()}",
                parent=self.root)

    def _merge_workbook(self, src_path: str, dst_wb, prefix: str):
        """Copia todas as abas de src_path para dst_wb, prefixando o nome."""
        from openpyxl import load_workbook
        src = load_workbook(src_path)
        for sheet_name in src.sheetnames:
            src_ws = src[sheet_name]
            # Nome da nova aba (Excel limita a 31 chars)
            new_name = f"{prefix}_{sheet_name}"[:31]
            dst_ws = dst_wb.create_sheet(title=new_name)
            # Copia células com valor + formatação básica
            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = dst_ws.cell(row=cell.row, column=cell.column,
                                            value=cell.value)
                    if cell.has_style:
                        try:
                            new_cell.font = cell.font.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.alignment = cell.alignment.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.number_format = cell.number_format
                        except Exception:
                            pass
            # Copia larguras de coluna
            for col_letter, col_dim in src_ws.column_dimensions.items():
                dst_ws.column_dimensions[col_letter].width = col_dim.width
            # Copia merges
            for mr in src_ws.merged_cells.ranges:
                dst_ws.merge_cells(str(mr))
            # Freeze panes
            if src_ws.freeze_panes:
                dst_ws.freeze_panes = src_ws.freeze_panes


def main():
    root = tk.Tk()
    MainApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
