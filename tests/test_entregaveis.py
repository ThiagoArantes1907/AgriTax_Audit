"""Testes do M7 (parcial): matriz decisória de achados em Excel."""
import json

import openpyxl
import pytest

from audit.core import db
from audit.core.modelo import Achado
from audit.entregaveis.matriz import gerar_matriz

CNPJ = "04461884000132"


def _achado(ref, risco, prioridade="ALTA", diferenca=100.0, **kw):
    return Achado(ref=ref, cnpj=CNPJ, competencia="2026.02", tributo="PIS",
                  titulo=f"Teste {ref}", descricao="evidência",
                  valores={"escriturado": 1000.0, "declarado": 900.0},
                  diferenca=diferenca, risco=risco, base_legal="IN 2.005/2021",
                  acao_proposta="agir", prioridade=prioridade, **kw)


@pytest.fixture
def engaj(tmp_path):
    (tmp_path / "achados").mkdir()
    (tmp_path / "achados" / "CR-04.json").write_text(json.dumps([
        {"cnpj": CNPJ, "codigo_receita": "6912", "competencia": "2026.02",
         "tributo": "PIS", "efd_debito": 1000.0, "dctf_debito": 900.0,
         "dctfweb_debito": 0.0, "total_declarado": 900.0, "diferenca": 100.0,
         "situacao": "Divergente"}]), encoding="utf-8")
    (tmp_path / "achados" / "SN_apuracoes.json").write_text(json.dumps([
        {"cnpj": CNPJ, "competencia": "2026.02", "rbt12": 1200000.0,
         "rpa": 100000.0, "anexo": "I", "fator_r": None, "total_debito": 9067.0,
         "num_declaracao": "AP1", "debitos": {"PIS": 270.0, "ICMS": 3400.0},
         "das_pago": 9067.0}]), encoding="utf-8")
    return tmp_path


def test_gerar_matriz_completa(engaj):
    con = db.conectar(engaj)
    db.inserir_achados(con, [
        _achado("CR-04", "R1"),
        _achado("CR-05", "R2", diferenca=400.0),
        _achado("SN-11", "R7", prioridade="MEDIA", diferenca=-433.0),
    ])
    destino = gerar_matriz(engaj, con, {"cliente": "TESTE", "cnpj": CNPJ})
    con.close()
    assert destino.exists()

    wb = openpyxl.load_workbook(destino)
    assert wb.sheetnames[:4] == ["1. Resumo", "2. Achados", "3. CR-04", "4. SN Apurações"]
    # achado R7 no seed → mapa de créditos e plano de regularização presentes
    assert "5. Mapa de Créditos" in wb.sheetnames
    assert "6. Plano de Regularização" in wb.sheetnames

    resumo = wb["1. Resumo"]
    assert "MATRIZ DE ACHADOS" in resumo.cell(2, 2).value
    assert "TESTE" in resumo.cell(3, 2).value

    achados = wb["2. Achados"]
    cab = [c.value for c in achados[1]]
    assert cab[0] == "Cruzamento" and "SUA DECISÃO" in cab
    # ALTA primeiro (ordenação por prioridade)
    assert achados.cell(2, 3).value == "ALTA"
    assert achados.cell(4, 1).value == "SN-11"
    # valores do dict abrem em colunas
    col_escr = cab.index("Escriturado (R$)") + 1
    assert achados.cell(2, col_escr).value == 1000.0
    # dropdown de decisão presente
    assert len(achados.data_validations.dataValidation) == 1

    cr04 = wb["3. CR-04"]
    assert cr04.cell(2, 2).value == "6912"

    sn = wb["4. SN Apurações"]
    cab_sn = [c.value for c in sn[1]]
    assert any("Deb Icms" in (v or "") for v in cab_sn)   # débitos achatados


def test_gerar_matriz_sem_achados(engaj):
    con = db.conectar(engaj)
    destino = gerar_matriz(engaj, con, {})
    con.close()
    wb = openpyxl.load_workbook(destino)
    assert "2. Achados" in wb.sheetnames        # só cabeçalho, sem linhas
