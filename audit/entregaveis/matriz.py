"""Matriz de cruzamentos/achados (entregável nº 1 da seção 9 do PT-AF-003).

Um achado por linha: competência, valores (escriturado × declarado × pago),
cruzamento de origem, risco, base legal, ação proposta e o campo de DECISÃO
do contador com justificativa — mesmo desenho da planilha de decisão do
sistema PIS/COFINS: o sistema instrui e evidencia; a decisão é do contador.
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from audit.core.dominio import RISCOS
from audit.core.modelo import DECISOES

VERDE_AGRITAX = "1F6E43"
VERDE_CLARO = "E8F3EC"
CINZA = "F2F2F2"
VERMELHO = "C0392B"
AMARELO = "F4D03F"

_FILL_CAB = PatternFill("solid", fgColor=VERDE_AGRITAX)
_FONT_CAB = Font(bold=True, color="FFFFFF", size=10)
_BORDA = Border(*[Side(style="thin", color="D0D0D0")] * 4)

COLUNAS_ACHADOS = [
    ("ref", "Cruzamento", 11), ("risco", "Risco", 7),
    ("prioridade", "Prioridade", 11), ("competencia", "Competência", 12),
    ("tributo", "Tributo", 12), ("titulo", "Achado", 42),
    ("escriturado", "Escriturado (R$)", 15), ("declarado", "Declarado (R$)", 15),
    ("pago", "Pago (R$)", 14), ("compensado", "Compensado (R$)", 15),
    ("diferenca", "Diferença (R$)", 14), ("descricao", "Evidência", 60),
    ("base_legal", "Base legal", 32), ("acao_proposta", "Ação proposta", 45),
    ("decadencia", "Decadência", 11), ("decisao_cliente", "SUA DECISÃO", 14),
    ("justificativa", "Justificativa", 30),
]
_MONEY = {"escriturado", "declarado", "pago", "compensado", "diferenca"}


def gerar_matriz(engaj_dir: str | Path, con: sqlite3.Connection,
                 parametros: dict | None = None) -> Path:
    engaj_dir = Path(engaj_dir)
    parametros = parametros or {}
    achados = [dict(r) for r in con.execute(
        "SELECT * FROM achados ORDER BY "
        "CASE prioridade WHEN 'ALTA' THEN 0 WHEN 'MEDIA' THEN 1 ELSE 2 END, "
        "ref, competencia")]
    for a in achados:
        a["valores"] = json.loads(a.get("valores") or "{}")

    wb = openpyxl.Workbook()
    _aba_resumo(wb.active, achados, parametros)
    _aba_achados(wb.create_sheet("2. Achados"), achados)
    for ref in ("CR-04", "CR-05"):
        arq = engaj_dir / "achados" / f"{ref}.json"
        if arq.exists():
            _aba_matriz_json(wb.create_sheet(f"3. {ref}"),
                             json.loads(arq.read_text(encoding="utf-8")))
    sn = engaj_dir / "achados" / "SN_apuracoes.json"
    if sn.exists():
        _aba_matriz_json(wb.create_sheet("4. SN Apurações"),
                         _achata_sn(json.loads(sn.read_text(encoding="utf-8"))))

    from .complementos import mapa_creditos, plano_regularizacao
    mapa = mapa_creditos(con)
    if mapa:
        _aba_matriz_json(wb.create_sheet("5. Mapa de Créditos"), mapa)
    plano = plano_regularizacao(con)
    if plano:
        _aba_matriz_json(wb.create_sheet("6. Plano de Regularização"), plano)

    destino = engaj_dir / "entregaveis" / "Matriz_Achados.xlsx"
    destino.parent.mkdir(exist_ok=True)
    wb.save(destino)
    return destino


def _cab(ws, colunas):
    for j, (_, titulo, largura) in enumerate(colunas, 1):
        c = ws.cell(1, j, titulo)
        c.fill, c.font, c.border = _FILL_CAB, _FONT_CAB, _BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = largura
    ws.freeze_panes = "A2"


def _aba_resumo(ws, achados: list[dict], parametros: dict):
    ws.title = "1. Resumo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in (("B", 16), ("C", 52), ("D", 14), ("E", 18)):
        ws.column_dimensions[col].width = w

    c = ws.cell(2, 2, "AGRITAX AUDIT — MATRIZ DE ACHADOS (PT-AF-003)")
    c.font = Font(bold=True, size=14, color=VERDE_AGRITAX)
    cliente = parametros.get("cliente", "")
    cnpj = parametros.get("cnpj", "")
    ws.cell(3, 2, f"Cliente: {cliente}   CNPJ: {cnpj}").font = Font(size=10)
    ws.cell(4, 2, "Diagnóstico de conformidade declaratória — fontes exclusivas "
                  "e-CAC e ReceitaNetBX. A decisão tributária é do contador."
            ).font = Font(size=9, italic=True, color="666666")

    linha = 6
    for titulo, chave in (("Por risco", "risco"), ("Por cruzamento", "ref")):
        ws.cell(linha, 2, titulo).font = Font(bold=True, size=11)
        linha += 1
        for cel, txt in zip("BCDE", ("Código", "Descrição", "Achados", "Σ Diferença (R$)")):
            c = ws.cell(linha, "BCDE".index(cel) + 2, txt)
            c.fill, c.font, c.border = _FILL_CAB, _FONT_CAB, _BORDA
        linha += 1
        grupos: dict = {}
        for a in achados:
            k = a.get(chave) or "—"
            g = grupos.setdefault(k, {"n": 0, "soma": 0.0})
            g["n"] += 1
            g["soma"] += abs(a.get("diferenca") or 0)
        for k in sorted(grupos):
            g = grupos[k]
            desc = RISCOS.get(k, "") if chave == "risco" else ""
            for j, v in enumerate((k, desc, g["n"], g["soma"]), 2):
                c = ws.cell(linha, j, v)
                c.border = _BORDA
                if j == 5:
                    c.number_format = "#,##0.00"
            linha += 1
        linha += 1


def _aba_achados(ws, achados: list[dict]):
    _cab(ws, COLUNAS_ACHADOS)
    dv = DataValidation(type="list", formula1=f'"{",".join(DECISOES)}"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    fill_alta = PatternFill("solid", fgColor="FDECEA")
    for i, a in enumerate(achados, 2):
        for j, (chave, _, _) in enumerate(COLUNAS_ACHADOS, 1):
            if chave in ("escriturado", "declarado", "pago", "compensado"):
                v = a["valores"].get(chave)
            else:
                v = a.get(chave)
            c = ws.cell(i, j, v if v not in ("", None) else None)
            c.border = _BORDA
            c.alignment = Alignment(vertical="top", wrap_text=chave in (
                "titulo", "descricao", "base_legal", "acao_proposta", "justificativa"))
            if chave in _MONEY:
                c.number_format = "#,##0.00"
            if a.get("prioridade") == "ALTA" and chave == "prioridade":
                c.fill = fill_alta
                c.font = Font(bold=True, color=VERMELHO)
        dv.add(ws.cell(i, 16))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS_ACHADOS))}{max(len(achados) + 1, 2)}"


def _aba_matriz_json(ws, linhas: list[dict]):
    """Aba genérica a partir das linhas completas salvas pelo executor."""
    if not linhas:
        ws.cell(1, 1, "(sem linhas)")
        return
    chaves = list(linhas[0].keys())
    colunas = [(k, k.replace("_", " ").title(), max(14, min(len(k) + 6, 24)))
               for k in chaves]
    _cab(ws, colunas)
    for i, ln in enumerate(linhas, 2):
        for j, k in enumerate(chaves, 1):
            v = ln.get(k)
            c = ws.cell(i, j, v if not isinstance(v, (dict, list)) else
                        json.dumps(v, ensure_ascii=False))
            c.border = _BORDA
            if isinstance(v, float):
                c.number_format = "#,##0.00"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(chaves))}{len(linhas) + 1}"


def _achata_sn(apuracoes: list[dict]) -> list[dict]:
    saida = []
    for ap in apuracoes:
        plano = {k: v for k, v in ap.items() if k != "debitos"}
        for trib, val in (ap.get("debitos") or {}).items():
            plano[f"deb_{trib.lower()}"] = val
        saida.append(plano)
    return saida
